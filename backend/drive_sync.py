"""Public Google Drive folder crawling for active-client attendance workbooks."""
from __future__ import annotations

import io
import re
import urllib.request
from html import unescape
from typing import Any, Dict, List, Optional, Tuple

UA = "Mozilla/5.0 BoostGrowthSync/1.0"
ACTIVE_CLIENTS_FOLDER_ID = "1iMDwfucwzsEIl9WxwhJi_h6tg2vVtAFr"

_DRIVE_LINK_RE = re.compile(
    r'href="(https://(?:drive\.google\.com/(?:drive/folders/|file/d/)[^"]+|docs\.google\.com/(?:spreadsheets|document|presentation)/d/[^"]+))"'
    r'.*?flip-entry-title">([^<]+)</div>',
    re.S,
)


def _fetch_bytes(url: str, timeout: int = 45) -> bytes:
    import httpx
    resp = httpx.get(url, headers={"User-Agent": UA}, follow_redirects=True, timeout=timeout)
    resp.raise_for_status()
    return resp.content


def _embedded_folder_html(folder_id: str) -> str:
    return _fetch_bytes(
        f"https://drive.google.com/embeddedfolderview?id={folder_id}"
    ).decode("utf-8", errors="replace")


def _clean_title(raw: str) -> str:
    return unescape(raw.replace("&#39;", "'")).strip()


def _classify_url(url: str) -> Tuple[str, Optional[str]]:
    """Return (kind, id) for a Drive/Docs URL."""
    if not url:
        return "unknown", None
    if "spreadsheets" in url:
        m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
        return "sheet", m.group(1) if m else None
    if "document" in url:
        m = re.search(r"/document/d/([a-zA-Z0-9-_]+)", url)
        return "doc", m.group(1) if m else None
    if "presentation" in url:
        m = re.search(r"/presentation/d/([a-zA-Z0-9-_]+)", url)
        return "presentation", m.group(1) if m else None
    if "/folders/" in url:
        m = re.search(r"/folders/([^?]+)", url)
        return "folder", m.group(1) if m else None
    if "/file/d/" in url:
        m = re.search(r"/file/d/([a-zA-Z0-9-_]+)", url)
        return "file", m.group(1) if m else None
    return "unknown", None


def parse_embedded_folder(html: str) -> List[Dict[str, Any]]:
    """Parse Google embedded folder view into folders, spreadsheets, docs, and files."""
    entries: List[Dict[str, Any]] = []
    seen: set = set()
    for m in _DRIVE_LINK_RE.finditer(html):
        url = m.group(1)
        if url in seen:
            continue
        seen.add(url)
        title = _clean_title(m.group(2))
        kind, item_id = _classify_url(url)
        entries.append({
            "title": title,
            "url": url,
            "kind": kind,
            "id": item_id,
        })
    return entries


def parse_file_no_from_title(title: str) -> Optional[str]:
    m = re.match(r"^(\d{3})\s*\|", (title or "").strip())
    return m.group(1) if m else None


def extract_folder_id(drive_url: str) -> Optional[str]:
    url = (drive_url or "").strip()
    m = re.search(r"/folders/([a-zA-Z0-9_-]+)", url) or re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    return m.group(1) if m else None


def extract_sheet_id(drive_url: str) -> Optional[str]:
    url = (drive_url or "").strip()
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url) or re.search(r"[?&]id=([a-zA-Z0-9-_]+)", url)
    return m.group(1) if m else None


def extract_doc_id(drive_url: str) -> Optional[str]:
    url = (drive_url or "").strip()
    m = re.search(r"/document/d/([a-zA-Z0-9-_]+)", url) or re.search(r"[?&]id=([a-zA-Z0-9-_]+)", url)
    return m.group(1) if m else None


def sheet_export_url(sheet_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def doc_export_url(doc_id: str, fmt: str = "txt") -> str:
    return f"https://docs.google.com/document/d/{doc_id}/export?format={fmt}"


def fetch_workbook_from_url(drive_url: str):
    """Download a Google Sheet as openpyxl workbook (data_only=True)."""
    import openpyxl

    sheet_id = extract_sheet_id(drive_url)
    if not sheet_id:
        raise ValueError("Could not extract Google Sheets ID from URL")
    content = _fetch_bytes(sheet_export_url(sheet_id))
    return openpyxl.load_workbook(io.BytesIO(content), data_only=True)


def fetch_doc_text(drive_url: str) -> str:
    """Download plain text from a shared Google Doc."""
    doc_id = extract_doc_id(drive_url)
    if not doc_id:
        raise ValueError("Could not extract Google Doc ID from URL")
    raw = _fetch_bytes(doc_export_url(doc_id, "txt"))
    return raw.decode("utf-8", errors="replace").strip()


_PHONE_RE = re.compile(
    r"(?:\+966[\s\-]?|966[\s\-]?0?)(5\d{8})\b|0(5\d{8})\b"
)


def extract_parent_phone_from_text(text: str) -> Optional[str]:
    """Find Saudi mobile number in free text (intake docs)."""
    if not text:
        return None
    compact = re.sub(r"[\s\-()]", "", text)
    for m in _PHONE_RE.finditer(compact):
        digits = m.group(1) or m.group(2)
        if digits and len(digits) == 9 and digits.startswith("5"):
            return f"0{digits}"
    return None


def _fetch_doc_text_by_id(doc_id: str) -> str:
    raw = _fetch_bytes(doc_export_url(doc_id, "txt"))
    return raw.decode("utf-8", errors="replace").strip()


def fetch_intake_parent_phone(drive_url: str) -> Optional[str]:
    """Extract guardian phone from intake Google Doc or Sheet."""
    url = (drive_url or "").strip()
    if not url:
        return None
    if "document" in url or "/file/d/" in url:
        doc_id = extract_doc_id(url)
        if not doc_id:
            m = re.search(r"/file/d/([a-zA-Z0-9-_]+)", url)
            doc_id = m.group(1) if m else None
        if doc_id:
            try:
                text = _fetch_doc_text_by_id(doc_id)
                ph = extract_parent_phone_from_text(text)
                if ph:
                    return ph
            except Exception:
                pass
    if "spreadsheets" in url:
        import openpyxl
        wb = fetch_workbook_from_url(url)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=120, values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                joined = " ".join(cells).lower()
                if any(k in joined for k in ("phone", "mobile", "جوال", "هاتف", "ولي", "parent", "guardian", "contact")):
                    for c in cells:
                        ph = extract_parent_phone_from_text(c)
                        if ph:
                            return ph
                    ph = extract_parent_phone_from_text(" ".join(cells))
                    if ph:
                        return ph
            for row in ws.iter_rows(min_row=1, max_row=120, values_only=True):
                for c in row:
                    if c is None:
                        continue
                    ph = extract_parent_phone_from_text(str(c))
                    if ph:
                        return ph
    return None


def fetch_case_summary_content(drive_url: str) -> Dict[str, Any]:
    """Parse case summary from Google Doc or Sheet."""
    url = (drive_url or "").strip()
    if not url:
        return {"sections": [], "raw_preview": ""}
    if "document" in url or "/file/d/" in url:
        try:
            text = fetch_doc_text(url)
            if text.strip():
                return parse_case_summary_text(text)
        except Exception:
            pass
        doc_id = extract_doc_id(url)
        if not doc_id:
            m = re.search(r"/file/d/([a-zA-Z0-9-_]+)", url)
            doc_id = m.group(1) if m else None
        if doc_id:
            try:
                text = _fetch_doc_text_by_id(doc_id)
                if text.strip():
                    return parse_case_summary_text(text)
            except Exception:
                pass
    if "spreadsheets" in url:
        wb = fetch_workbook_from_url(url)
        rows: List[List[str]] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=120, values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    rows.append(cells)
            if rows:
                break
        if not rows:
            return {"sections": [], "raw_preview": ""}
        sections: List[Dict[str, Any]] = []
        current: Optional[Dict[str, Any]] = None
        for cells in rows:
            if len(cells) == 1 and len(cells[0]) < 60:
                if current:
                    sections.append(current)
                current = {"heading": cells[0].rstrip(":"), "paragraphs": [], "bullets": [], "tables": []}
            elif len(cells) >= 2:
                if current is None:
                    current = {"heading": "Details", "paragraphs": [], "bullets": [], "tables": []}
                current.setdefault("tables", []).append(cells)
            elif current is not None:
                current["paragraphs"].append(cells[0])
        if current:
            sections.append(current)
        return {"sections": sections, "raw_preview": ""}
    return {"sections": [], "raw_preview": ""}


def _year_score(title: str) -> int:
    years = [int(y) for y in re.findall(r"(20\d{2})", title or "")]
    return max(years) if years else 0


def _is_attendance_related(title: str) -> bool:
    """True for attendance workbook folders/sheets (excluded from client link lists)."""
    t = (title or "").lower().strip()
    if not t:
        return False
    if "attendance" in t and ("sheet" in t or "sheets" in t or t.endswith("attendance")):
        return True
    if t in ("attendance sheets", "attendance sheet"):
        return True
    if re.search(r"attendance\s*(sheet|20\d{2})", t):
        return True
    return False


def _is_copy_of(title: str) -> bool:
    return (title or "").lower().startswith("copy of ")


def pick_attendance_sheet(entries: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    sheets = [e for e in entries if e.get("kind") == "sheet" and not _is_copy_of(e.get("title", ""))]
    if not sheets:
        sheets = [e for e in entries if e.get("kind") == "sheet"]
    if not sheets:
        return None
    sheets.sort(key=lambda e: (_year_score(e.get("title", "")), e.get("title", "")), reverse=True)
    for s in sheets:
        t = (s.get("title") or "").lower()
        if "attendance" in t:
            return s
    return sheets[0]


def resolve_attendance_sheet_url(client_folder_id: str) -> Optional[str]:
    """Find the best attendance spreadsheet inside a client Drive folder."""
    items = parse_embedded_folder(_embedded_folder_html(client_folder_id))
    att_folder = next(
        (i for i in items if i.get("kind") == "folder" and _is_attendance_related(i.get("title", ""))),
        None,
    )
    if att_folder and att_folder.get("id"):
        inner = parse_embedded_folder(_embedded_folder_html(att_folder["id"]))
        picked = pick_attendance_sheet(inner)
        if picked:
            return picked.get("url")
    picked = pick_attendance_sheet([i for i in items if i.get("kind") == "sheet"])
    return picked.get("url") if picked else None


_IMAGE_EXT = re.compile(r"\.(jpe?g|png|gif|webp|heic|bmp|tiff?|svg)$", re.I)
_PHOTO_FOLDER_RE = re.compile(
    r"photo|picture|image|pic|صور|مرفق|attachments?|gallery",
    re.I,
)


def _is_image_file(title: str, kind: str) -> bool:
    if kind != "file":
        return False
    t = (title or "").lower()
    if _IMAGE_EXT.search(t):
        return True
    if any(k in t for k in ("img_", "dsc", "screenshot", "photo")):
        return True
    return False


def _is_photos_folder(title: str) -> bool:
    return bool(_PHOTO_FOLDER_RE.search(title or ""))


def resolve_parent_phone_from_links(link_meta: dict, client: Optional[dict] = None) -> Optional[str]:
    """Try intake, case summary, then other docs/sheets for a guardian phone."""
    client = client or {}
    tried: set = set()
    candidates: List[str] = []
    for key in ("intake_file_url", "case_summary_url"):
        u = (link_meta.get(key) or client.get(key) or "").strip()
        if u and u not in tried:
            tried.add(u)
            candidates.append(u)
    for entry in link_meta.get("links") or []:
        if entry.get("kind") not in ("doc", "sheet", "file", "document"):
            continue
        u = (entry.get("url") or "").strip()
        if not u or u in tried:
            continue
        tried.add(u)
        title = (entry.get("title") or "").lower()
        if _case_summary_title_match(entry.get("title") or ""):
            candidates.insert(min(2, len(candidates)), u)
        elif _intake_title_match(entry.get("title") or ""):
            candidates.insert(0, u)
        else:
            candidates.append(u)
    for url in candidates[:12]:
        try:
            ph = fetch_intake_parent_phone(url)
            if ph:
                return ph
        except Exception:
            continue
    return None


def _intake_title_match(title: str) -> bool:
    tl = (title or "").lower()
    if "intake" in tl or "انتيك" in (title or "") or "الانتيك" in (title or "") or "انتاك" in (title or ""):
        return True
    if "client information" in tl or "client info" in tl or "initial assessment" in tl:
        return True
    if "تقييم" in (title or "") or "معلومات" in (title or ""):
        return True
    return False


def _case_summary_title_match(title: str) -> bool:
    tl = (title or "").lower()
    return "case summary" in tl or "ملخص الحالة" in (title or "") or "ملخص حالة" in (title or "")


def list_client_folder_links(client_folder_id: str) -> Dict[str, Any]:
    """Crawl a client folder (and shallow subfolders) for Case Summary, Intake, etc."""
    folder_url = f"https://drive.google.com/drive/folders/{client_folder_id}"
    links: List[Dict[str, Any]] = []
    case_summary_url: Optional[str] = None
    intake_file_url: Optional[str] = None
    photos_folder_url: Optional[str] = None
    seen_urls: set = set()
    root_image_count = 0

    def add_item(item: Dict[str, Any], *, group: Optional[str] = None) -> None:
        nonlocal case_summary_url, intake_file_url
        title = item.get("title") or ""
        if _is_attendance_related(title):
            return
        if _is_image_file(title, item.get("kind") or ""):
            return
        url = item.get("url") or ""
        if not url or url in seen_urls:
            return
        seen_urls.add(url)
        kind = item.get("kind") or "unknown"
        entry: Dict[str, Any] = {"title": title, "url": url, "kind": kind}
        if group:
            entry["group"] = group
        links.append(entry)
        if not case_summary_url and _case_summary_title_match(title):
            case_summary_url = url
        if not intake_file_url and _intake_title_match(title) and kind in ("doc", "sheet", "file", "document"):
            intake_file_url = url

    def scan_folder(fid: str, depth: int = 0) -> None:
        nonlocal photos_folder_url
        if depth > 2 or not fid:
            return
        try:
            items = parse_embedded_folder(_embedded_folder_html(fid))
        except Exception:
            return
        for item in items:
            kind = item.get("kind") or ""
            title = item.get("title") or ""
            if kind == "folder":
                if _is_attendance_related(title):
                    continue
                if _is_photos_folder(title):
                    url = item.get("url") or f"https://drive.google.com/drive/folders/{item.get('id')}"
                    if url not in seen_urls:
                        seen_urls.add(url)
                        photos_folder_url = url
                        links.append({
                            "title": f"Attached Photos — {title}",
                            "url": url,
                            "kind": "folder",
                            "group": "photos",
                        })
                    continue
                scan_folder(item.get("id") or "", depth + 1)
            else:
                if _is_image_file(title, kind):
                    if depth == 0:
                        root_image_count += 1
                    continue
                add_item(item)

    scan_folder(client_folder_id)

    if not photos_folder_url and root_image_count > 0:
        photos_folder_url = folder_url
        if folder_url not in seen_urls:
            seen_urls.add(folder_url)
            links.append({
                "title": f"Attached Photos ({root_image_count} files)",
                "url": folder_url,
                "kind": "folder",
                "group": "photos",
            })

    if not intake_file_url:
        for entry in links:
            if entry.get("group") == "photos":
                continue
            if entry["kind"] in ("doc", "sheet", "file") and not _case_summary_title_match(entry["title"]):
                if _intake_title_match(entry["title"]):
                    intake_file_url = entry["url"]
                    break
        if not intake_file_url:
            for entry in links:
                if entry.get("group") == "photos":
                    continue
                if entry["kind"] in ("doc", "sheet") and not _case_summary_title_match(entry["title"]):
                    intake_file_url = entry["url"]
                    break

    doc_links = [l for l in links if l.get("group") != "photos"]
    photo_links = [l for l in links if l.get("group") == "photos"]
    ordered = sorted(doc_links, key=lambda x: (x.get("title") or "").lower()) + photo_links
    return {
        "folder_id": client_folder_id,
        "folder_url": folder_url,
        "links": ordered,
        "case_summary_url": case_summary_url,
        "intake_file_url": intake_file_url,
        "photos_folder_url": photos_folder_url,
    }


def parse_case_summary_text(text: str) -> Dict[str, Any]:
    """Turn exported Google Doc plain text into structured sections for the portal."""
    lines = [ln.rstrip() for ln in (text or "").splitlines()]
    sections: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None
    table_rows: List[List[str]] = []

    def flush_table():
        nonlocal table_rows, current
        if table_rows and current is not None:
            current.setdefault("tables", []).append(table_rows)
            table_rows = []

    def flush_section():
        nonlocal current
        flush_table()
        if current and (current.get("paragraphs") or current.get("tables") or current.get("bullets")):
            sections.append(current)
        current = None

    heading_re = re.compile(
        r"^("
        r"diagnosis|goals?|objectives?|strengths?|challenges?|"
        r"background|history|summary|recommendations?|"
        r"services?|interventions?|medications?|allergies?|"
        r"family|parent|school|notes?|plan|targets?"
        r")\s*:?\s*$",
        re.I,
    )

    for raw in lines:
        line = raw.strip()
        if not line:
            flush_table()
            continue
        if "\t" in line:
            cells = [c.strip() for c in line.split("\t") if c.strip()]
            if len(cells) >= 2:
                if current is None:
                    current = {"heading": "Details", "paragraphs": [], "bullets": [], "tables": []}
                table_rows.append(cells)
                continue
        if heading_re.match(line) or (line.endswith(":") and len(line) < 60 and line[0].isupper()):
            flush_section()
            heading = line.rstrip(":").strip()
            current = {"heading": heading, "paragraphs": [], "bullets": [], "tables": []}
            continue
        if line.startswith(("-", "•", "·", "*")) and len(line) > 2:
            if current is None:
                current = {"heading": "Overview", "paragraphs": [], "bullets": [], "tables": []}
            current["bullets"].append(line.lstrip("-•·* ").strip())
            continue
        if current is None:
            current = {"heading": "Overview", "paragraphs": [], "bullets": [], "tables": []}
        current["paragraphs"].append(line)

    flush_section()

    if not sections and text.strip():
        sections = [{
            "heading": "Case Summary",
            "paragraphs": [p.strip() for p in text.split("\n\n") if p.strip()],
            "bullets": [],
            "tables": [],
        }]

    return {"sections": sections, "raw_preview": text[:2000] if text else ""}


def list_active_client_folders(parent_folder_id: str = ACTIVE_CLIENTS_FOLDER_ID) -> List[Dict[str, str]]:
    items = parse_embedded_folder(_embedded_folder_html(parent_folder_id))
    out: List[Dict[str, str]] = []
    for item in items:
        if item.get("kind") != "folder" or not item.get("id"):
            continue
        file_no = parse_file_no_from_title(item.get("title", ""))
        if file_no:
            out.append({
                "file_no": file_no,
                "folder_id": item["id"],
                "title": item.get("title", ""),
                "folder_url": item.get("url") or f"https://drive.google.com/drive/folders/{item['id']}",
            })
    out.sort(key=lambda x: x["file_no"])
    return out
