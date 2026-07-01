#!/usr/bin/env python3
"""Batch recovery: import Active Clients Excel + sync attendance/invoices from Drive.

The shared Drive folder (ACTIVE_CLIENTS_FOLDER_ID) holds per-child subfolders with
attendance Google Sheets — not the master Active Clients roster. You still need the
master workbook (e.g. ``Clients' Info.xlsx`` with an ``Active Clients`` tab) for
client records; Drive sync imports invoices/sessions after clients exist.

Usage — direct MongoDB (production .env):
  cd backend
  MONGO_URL='mongodb+srv://...' python scripts/import_active_clients_batch.py \\
    --clients-excel "/path/Clients' Info-4.xlsx" --recover --sync-drive

Usage — production API (no local MONGO_URL):
  python scripts/import_active_clients_batch.py --api \\
    --api-base https://boost-growth-main-production-7283.up.railway.app/api \\
    --clients-excel brand-assets/Clients-Info.xlsx --sync-drive --recover

Import every .xlsx in a folder (master list + per-child attendance workbooks):
  python scripts/import_active_clients_batch.py --import-dir brand-assets/client-files/ \\
    --recover --sync-drive

Dry-run Drive sync (lists sheets without writing):
  python scripts/import_active_clients_batch.py --sync-drive --dry-run
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import subprocess
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

DEFAULT_API = "https://boost-growth-main-production-7283.up.railway.app/api"
DEFAULT_DRIVE_FOLDER = "https://drive.google.com/drive/folders/1iMDwfucwzsEIl9WxwhJi_h6tg2vVtAFr"


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--clients-excel", type=Path, help="Master Active Clients workbook or CSV")
    p.add_argument("--import-dir", type=Path, help="Folder of .xlsx/.csv client files")
    p.add_argument("--sync-drive", action="store_true", help="Bulk sync attendance/invoices from Drive")
    p.add_argument("--drive-folder", default=DEFAULT_DRIVE_FOLDER, help="Drive folder URL")
    p.add_argument("--recover", action="store_true", help="Run auto-recover (prep relink, dedupe)")
    p.add_argument("--dry-run", action="store_true", help="Preview Drive sync only")
    p.add_argument("--replace-missing", action="store_true", help="Soft-delete clients not in import file")
    p.add_argument("--api", action="store_true", help="Use HTTP API instead of MONGO_URL")
    p.add_argument("--api-base", default=os.environ.get("API_BASE", DEFAULT_API))
    p.add_argument("--email", default=os.environ.get("ADMIN_EMAIL", "admin@boostgrowthsa.com"))
    p.add_argument("--password", default=os.environ.get("ADMIN_PASSWORD", "Admin123"))
    return p.parse_args()


class ApiClient:
    def __init__(self, base: str, email: str, password: str):
        out = subprocess.check_output(
            [
                "curl", "-s", "-X", "POST", f"{base}/auth/login",
                "-H", "Content-Type: application/json",
                "-d", json.dumps({"email": email, "password": password}),
            ]
        )
        data = json.loads(out)
        if "token" not in data:
            raise RuntimeError(f"Login failed: {data}")
        self.base = base.rstrip("/")
        self.token = data["token"]
        self._h = ["-H", f"Authorization: Bearer {self.token}"]

    def post_json(self, path: str, body: dict | None = None) -> dict:
        cmd = ["curl", "-s", "-X", "POST", f"{self.base}{path}"] + self._h
        if body is not None:
            cmd += ["-H", "Content-Type: application/json", "-d", json.dumps(body)]
        return json.loads(subprocess.check_output(cmd))

    def upload_clients_and_sync(self, path: Path, replace_missing: bool) -> dict:
        cmd = [
            "curl", "-s", "-X", "POST",
            f"{self.base}/admin/import-clients-and-sync",
        ] + self._h + [
            "-F", f"file=@{path}",
            "-F", f"replace_missing={'true' if replace_missing else 'false'}",
        ]
        return json.loads(subprocess.check_output(cmd))

    def sync_drive(self, folder_url: str, dry_run: bool) -> dict:
        return self.post_json("/admin/sync-active-clients-from-drive", {
            "folder_url": folder_url or None,
            "dry_run": dry_run,
        })

    def auto_recover(self) -> dict:
        return self.post_json("/admin/auto-recover")


def _collect_files(import_dir: Path) -> list[Path]:
    files: list[Path] = []
    for ext in ("*.xlsx", "*.xls", "*.csv"):
        files.extend(sorted(import_dir.glob(ext)))
    return files


def _is_master_clients_file(path: Path) -> bool:
    """Heuristic: workbook with Active Clients sheet."""
    if path.suffix.lower() not in (".xlsx", ".xls"):
        return path.suffix.lower() == ".csv"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = [s.strip().lower() for s in wb.sheetnames]
        wb.close()
        return any(n.startswith("active client") for n in names)
    except Exception:
        return False


async def _run_mongo(args: argparse.Namespace) -> int:
    url = os.environ.get("MONGO_URL")
    if not url:
        print("ERROR: MONGO_URL not set — use backend/.env or --api mode")
        return 1

    from server import (  # noqa: E402
        _import_clients_from_rows,
        _ingest_workbook_for_client,
        _read_clients_import_rows,
        _run_auto_recover,
        _find_client_by_file_no,
        db,
    )
    from drive_sync import (  # noqa: E402
        ACTIVE_CLIENTS_FOLDER_ID,
        extract_folder_id,
        fetch_workbook_from_url,
        list_active_client_folders,
        list_client_folder_links,
        resolve_attendance_sheet_url,
    )

    admin = await db.users.find_one({"role": "admin"}, {"_id": 0, "id": 1})
    admin_id = (admin or {}).get("id") or "batch-import"

    summary: dict = {"imports": [], "drive": None, "recover": None}

    paths: list[Path] = []
    if args.clients_excel:
        paths.append(args.clients_excel)
    if args.import_dir:
        paths.extend(_collect_files(args.import_dir))

    for path in paths:
        if not path.is_file():
            print(f"  SKIP missing: {path}")
            continue
        content = path.read_bytes()
        if _is_master_clients_file(path):
            rows = _read_clients_import_rows(content, path.name)
            result = await _import_clients_from_rows(rows, args.replace_missing)
            summary["imports"].append({"file": str(path), "type": "clients", **result})
            print(f"  Clients from {path.name}: +{result['created']} new, ~{result['updated']} updated, {result['skipped']} skipped")
        else:
            rows = _read_clients_import_rows(content, path.name)
            if rows and all(r.get("file_no") for r in rows[:3]):
                result = await _import_clients_from_rows(rows, False)
                summary["imports"].append({"file": str(path), "type": "clients", **result})
                print(f"  Clients from {path.name}: +{result['created']} new, ~{result['updated']} updated")
                continue
            import openpyxl
            import io
            file_no = None
            for part in path.stem.replace("_", " ").split():
                if part.isdigit() and len(part) <= 3:
                    file_no = part.zfill(3)
                    break
            if not file_no:
                print(f"  SKIP {path.name}: cannot detect file_no for attendance workbook")
                continue
            client = await _find_client_by_file_no(file_no)
            if not client:
                print(f"  SKIP {path.name}: client {file_no} not in DB — import master list first")
                continue
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ingest = await _ingest_workbook_for_client(client["id"], client, wb, admin_id, origin="batch-excel")
            summary["imports"].append({"file": str(path), "type": "attendance", "file_no": file_no, **ingest})
            print(f"  Attendance {path.name} ({file_no}): invoices + sessions ingested")

    if args.sync_drive:
        parent_id = extract_folder_id(args.drive_folder or "") or ACTIVE_CLIENTS_FOLDER_ID
        folders = list_active_client_folders(parent_id)
        drive_results = []
        synced = meta = skipped = errors = 0
        for entry in folders:
            file_no = entry["file_no"]
            client = await _find_client_by_file_no(file_no)
            if not client:
                drive_results.append({"file_no": file_no, "status": "skipped", "reason": "client not in portal"})
                skipped += 1
                continue
            try:
                link_meta = list_client_folder_links(entry["folder_id"])
                sheet_url = resolve_attendance_sheet_url(entry["folder_id"])
                if args.dry_run:
                    drive_results.append({
                        "file_no": file_no,
                        "status": "dry_run",
                        "sheet_url": sheet_url,
                        "client_name": client.get("name"),
                    })
                    continue
                patch = {
                    "drive_url": entry.get("folder_url"),
                    "drive_folder_id": entry["folder_id"],
                }
                if sheet_url:
                    patch["attendance_sheet_url"] = sheet_url
                await db.clients.update_one({"id": client["id"]}, {"$set": patch})
                if not sheet_url:
                    meta += 1
                    drive_results.append({"file_no": file_no, "status": "meta_synced"})
                    continue
                wb = fetch_workbook_from_url(sheet_url)
                ingest = await _ingest_workbook_for_client(
                    client["id"], client, wb, admin_id, origin="drive-bulk-sync"
                )
                synced += 1
                drive_results.append({"file_no": file_no, "status": "synced", **ingest})
                print(f"  Drive {file_no} {client.get('name')}: synced")
            except Exception as exc:
                errors += 1
                drive_results.append({"file_no": file_no, "status": "error", "error": str(exc)})
                print(f"  Drive {file_no}: ERROR {exc}")
        summary["drive"] = {
            "total_folders": len(folders),
            "synced": synced,
            "meta_synced": meta,
            "skipped": skipped,
            "errors": errors,
            "results": drive_results,
        }
        print(f"Drive sync: {synced} synced · {meta} meta-only · {skipped} skipped · {errors} errors")

    if args.recover and not args.dry_run:
        summary["recover"] = await _run_auto_recover(store_backup=True)
        print(f"Recover: {summary['recover'].get('summary_ar', 'done')}")

    print("\n=== Summary ===")
    print(json.dumps(summary, indent=2, default=str))
    return 0


def _run_api(args: argparse.Namespace) -> int:
    api = ApiClient(args.api_base, args.email, args.password)
    summary: dict = {"imports": [], "drive": None, "recover": None}

    paths: list[Path] = []
    if args.clients_excel:
        paths.append(args.clients_excel)
    if args.import_dir:
        paths.extend(_collect_files(args.import_dir))

    for path in paths:
        if not path.is_file():
            print(f"  SKIP missing: {path}")
            continue
        if _is_master_clients_file(path) or path.suffix.lower() == ".csv":
            if args.recover:
                data = api.upload_clients_and_sync(path, args.replace_missing)
                summary["imports"].append({"file": str(path), **data.get("import", data)})
                print(f"  {path.name}: {data.get('summary_ar', data)}")
            else:
                cmd = [
                    "curl", "-s", "-X", "POST", f"{api.base}/import/clients",
                ] + api._h + [
                    "-F", f"file=@{path}",
                    "-F", f"replace_missing={'true' if args.replace_missing else 'false'}",
                ]
                data = json.loads(subprocess.check_output(cmd))
                summary["imports"].append({"file": str(path), **data})
                print(f"  {path.name}: +{data.get('created', 0)} new, ~{data.get('updated', 0)} updated")
        else:
            print(f"  SKIP {path.name}: per-child attendance upload needs MONGO_URL or manual Import per client")

    if args.sync_drive:
        data = api.sync_drive(args.drive_folder, args.dry_run)
        summary["drive"] = data
        print(data.get("message") or data)

    if args.recover and not args.dry_run and not paths:
        summary["recover"] = api.auto_recover()
        print(summary["recover"].get("summary_ar") or "Recover done")

    print("\n=== Summary ===")
    print(json.dumps(summary, indent=2, default=str))
    return 0


def main() -> int:
    args = _parse_args()
    if not any([args.clients_excel, args.import_dir, args.sync_drive, args.recover]):
        print("Nothing to do — pass --clients-excel, --import-dir, --sync-drive, and/or --recover")
        return 1

    if args.api or not os.environ.get("MONGO_URL"):
        if not os.environ.get("MONGO_URL"):
            print("Note: MONGO_URL not set — using API mode")
        return _run_api(args)
    return asyncio.run(_run_mongo(args))


if __name__ == "__main__":
    raise SystemExit(main())
