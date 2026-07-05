from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import asyncio
import base64
import mimetypes
import os
import re
import uuid
import logging
from urllib.parse import quote
from datetime import datetime, timezone, timedelta, date
import calendar
from typing import Dict, List, Optional

import bcrypt
import jwt
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Form, Query, Body
from fastapi.responses import FileResponse, Response
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(
    mongo_url,
    serverSelectionTimeoutMS=15000,
    connectTimeoutMS=15000,
)
db = client[os.environ['DB_NAME']]

def _active_client_filter(extra: Optional[dict] = None) -> dict:
    """Exclude soft-deleted clients from normal queries."""
    q = {"deleted": {"$ne": True}}
    if extra:
        q.update(extra)
    return q


def _billing_active_client_filter(extra: Optional[dict] = None) -> dict:
    """Non-deleted clients still on active billing / audit lists."""
    return _active_client_filter({"status": {"$ne": "Inactive"}, **(extra or {})})


INACTIVE_CLIENT_FILE_NOS = frozenset({"018", "023", "030", "037", "047", "063"})


async def _client_data_score(client_id: str) -> int:
    """Higher = richer record; used when picking which duplicate client to keep."""
    inv_count = await db.invoices.count_documents({"client_id": client_id})
    sess_count = await db.sessions.count_documents({"client_id": client_id})
    client = await db.clients.find_one(
        {"id": client_id},
        {"_id": 0, "attendance_sheet_url": 1, "drive_links": 1, "case_summary_url": 1, "file_no": 1},
    ) or {}
    score = inv_count * 100 + sess_count * 10
    if client.get("attendance_sheet_url"):
        score += 50
    if client.get("drive_links"):
        score += 20
    if client.get("case_summary_url"):
        score += 10
    if client.get("file_no"):
        score += 5
    return score


async def _dedupe_duplicate_clients() -> dict:
    """Soft-delete duplicate client rows; keep the record with invoices/sessions."""
    clients = await db.clients.find({"deleted": {"$ne": True}}, {"_id": 0}).to_list(500)
    by_name: dict = {}
    by_file_no: dict = {}
    for c in clients:
        name_key = re.sub(r"\s+", " ", (c.get("name") or "").strip().lower())
        if name_key:
            by_name.setdefault(name_key, []).append(c)
        fn = str(c.get("file_no") or "").strip()
        if fn:
            by_file_no.setdefault(fn.zfill(3), []).append(c)

    to_delete: set = set()
    actions: List[dict] = []

    async def pick_winner(group: List[dict], reason: str):
        if len(group) < 2:
            return
        scored = []
        for c in group:
            if c["id"] in to_delete:
                continue
            scored.append((await _client_data_score(c["id"]), c))
        if len(scored) < 2:
            return
        scored.sort(key=lambda x: (-x[0], x[1].get("created_at") or ""))
        winner = scored[0][1]
        for _, loser in scored[1:]:
            if loser["id"] == winner["id"] or loser["id"] in to_delete:
                continue
            to_delete.add(loser["id"])
            actions.append({
                "kept": winner.get("name"),
                "kept_id": winner["id"],
                "kept_file_no": winner.get("file_no"),
                "removed": loser.get("name"),
                "removed_id": loser["id"],
                "removed_file_no": loser.get("file_no"),
                "reason": reason,
            })

    for name_key, group in by_name.items():
        await pick_winner(group, f"duplicate name: {name_key}")
    for file_no, group in by_file_no.items():
        active = [c for c in group if c["id"] not in to_delete]
        await pick_winner(active, f"duplicate file_no: {file_no}")

    for cid in to_delete:
        await db.clients.update_one(
            {"id": cid},
            {"$set": {"deleted": True, "deleted_at": now_iso(), "dedupe_note": "auto-deduped duplicate"}},
        )
    return {"removed": len(to_delete), "actions": actions}


def _therapist_canonical_email(email: Optional[str]) -> str:
    em = (email or "").strip().lower()
    return THERAPIST_LOGIN_EMAIL_ALIASES.get(em, em) if em else ""


def _is_uuid_therapist_id(tid: Optional[str]) -> bool:
    try:
        uuid.UUID(str(tid or ""))
        return True
    except (ValueError, AttributeError, TypeError):
        return False


def _is_legacy_ms_numeric_key(key: Optional[str]) -> bool:
    return bool(re.match(r"^ms\d+$", (key or "").strip(), re.I))


def _therapist_record_score(t: dict) -> int:
    """Higher = canonical therapist row; prefer UUID accounts and user-chosen passwords."""
    score = 0
    if t.get("password_hash") and not t.get("must_change_password"):
        score += 1000
    if _is_uuid_therapist_id(t.get("id")):
        score += 200
    key = (t.get("key") or "").strip()
    if key and not _is_legacy_ms_numeric_key(key):
        score += 100
    elif key:
        score += 10
    if t.get("password_hash"):
        score += 50
    if t.get("role"):
        score += 10
    if t.get("email"):
        score += 5
    return score


def _pick_canonical_therapist(group: List[dict]) -> dict:
    scored = [(_therapist_record_score(t), t.get("created_at") or "", t) for t in group]
    scored.sort(key=lambda x: (-x[0], x[1]))
    return scored[0][2]


async def _find_therapist_by_email(email: str) -> Optional[dict]:
    """Resolve therapist by email; when duplicates exist, keep the richest record."""
    email_l = email.lower().strip()
    matches = await db.therapists.find(
        {"email": {"$regex": f"^{re.escape(email_l)}$", "$options": "i"}},
        {"_id": 0},
    ).to_list(20)
    if not matches:
        return None
    if len(matches) == 1:
        return matches[0]
    scored = [(_therapist_record_score(m), m.get("created_at") or "", m) for m in matches]
    scored.sort(key=lambda x: (-x[0], x[1]))
    return scored[0][2]


# Old/simple emails → canonical work email (login accepts either form).
THERAPIST_LOGIN_EMAIL_ALIASES: Dict[str, str] = {
    "maha@boostgrowthsa.com": "msalthunayan@boostgrowthsa.com",
    "fahda@boostgrowthsa.com": "falghadeeb@boostgrowthsa.com",
    "fahdah@boostgrowthsa.com": "falghadeeb@boostgrowthsa.com",
    "jenan@boostgrowthsa.com": "jsalmuhaisin@boostgrowthsa.com",
    "genan@boostgrowthsa.com": "jsalmuhaisin@boostgrowthsa.com",
    "razan@boostgrowthsa.com": "ralshatery@boostgrowthsa.com",
    "ralshatri@boostgrowthsa.com": "ralshatery@boostgrowthsa.com",
    "shatha@boostgrowthsa.com": "shalhammami@boostgrowthsa.com",
    "salhammamy@boostgrowthsa.com": "shalhammami@boostgrowthsa.com",
    "manal@boostgrowthsa.com": "maldosery@boostgrowthsa.com",
    "hajer@boostgrowthsa.com": "halfulaij@boostgrowthsa.com",
    "hajar@boostgrowthsa.com": "halfulaij@boostgrowthsa.com",
}


def _login_email_variants(email: str) -> List[str]:
    """All login emails that should resolve to the same account."""
    email_l = email.lower().strip()
    variants = [email_l]
    alias = THERAPIST_LOGIN_EMAIL_ALIASES.get(email_l)
    if alias:
        variants.append(alias)
    for old, new in THERAPIST_LOGIN_EMAIL_ALIASES.items():
        if new == email_l:
            variants.append(old)
    if email_l in WALAA_LOGIN_EMAILS:
        variants.extend(sorted(WALAA_LOGIN_EMAILS))
    return list(dict.fromkeys(variants))


def _is_client_lead_login_email(email: str) -> bool:
    for em in _login_email_variants(email):
        if em in CLIENT_LEAD_EMAILS:
            return True
    return False


async def _find_user_by_login_email(email: str) -> Optional[dict]:
    for em in _login_email_variants(email):
        user = await db.users.find_one({"email": em})
        if user:
            return user
    if email.lower().strip() in WALAA_LOGIN_EMAILS:
        return await db.users.find_one(
            {"email": {"$regex": r"^(wabuissa|walaa)@boostgrowthsa\.com$", "$options": "i"}},
        )
    return None


async def _find_therapist_by_login_email(email: str) -> Optional[dict]:
    for em in _login_email_variants(email):
        t = await _find_therapist_by_email(em)
        if t:
            return t
    if email.lower().strip() in WALAA_LOGIN_EMAILS:
        t = await db.therapists.find_one({"key": "msWalaa"}, {"_id": 0})
        if t:
            return t
        return await db.therapists.find_one(
            {"name": {"$regex": r"^ms\.?\s*walaa\b", "$options": "i"}},
            {"_id": 0},
        )
    return None


def _therapist_has_nationality(t: dict) -> bool:
    """True when therapist has a family/nationality label (DB field, key map, or multi-word name)."""
    nat = (t.get("nationality") or "").strip()
    if nat:
        return True
    key = (t.get("key") or "").strip()
    if key and key in THERAPIST_FAMILY_NAMES:
        return True
    raw = re.sub(r"^Ms\.?\s*", "", (t.get("name") or ""), flags=re.I).strip()
    parts = [p for p in raw.split() if p]
    return len(parts) >= 2


async def _remove_therapists_without_nationality() -> dict:
    """Delete duplicate therapist rows missing nationality/family (e.g. Ms. Hajer vs Ms. Hajar)."""
    therapists = await db.therapists.find({}, {"_id": 0}).to_list(500)
    to_delete: List[str] = []
    actions: List[dict] = []

    canonical_hajar = next(
        (t for t in therapists
         if (t.get("email") or "").lower() == "halfulaij@boostgrowthsa.com"
         or (t.get("key") or "").lower() == "mshajer"),
        None,
    )
    for t in therapists:
        if canonical_hajar and t["id"] == canonical_hajar["id"]:
            continue
        em = (t.get("email") or "").lower().strip()
        name = (t.get("name") or "").strip()
        if not canonical_hajar:
            continue
        if em == "hajer@boostgrowthsa.com" or (
            name.lower() == "ms. hajer" and not _therapist_has_nationality(t)
        ):
            to_delete.append(t["id"])
            actions.append({
                "removed_id": t["id"],
                "removed_name": name,
                "removed_email": em,
                "kept_id": canonical_hajar["id"],
                "kept_name": canonical_hajar.get("name"),
                "reason": "duplicate without nationality",
            })

    for tid in set(to_delete):
        await db.schedule_cells.delete_many({"therapist_id": tid})
        await db.users.delete_many({"therapist_id": tid})
        await db.therapists.delete_one({"id": tid})
    return {"removed": len(set(to_delete)), "actions": actions}


async def _dedupe_duplicate_therapists() -> dict:
    """Delete duplicate therapist rows that share the same email; keep user-chosen passwords."""
    therapists = await db.therapists.find({}, {"_id": 0}).to_list(500)
    by_email: dict = {}
    for t in therapists:
        em = _therapist_canonical_email(t.get("email"))
        if em:
            by_email.setdefault(em, []).append(t)

    to_delete: List[tuple] = []
    actions: List[dict] = []
    for em, group in by_email.items():
        if len(group) < 2:
            continue
        scored = [(_therapist_record_score(t), t.get("created_at") or "", t) for t in group]
        scored.sort(key=lambda x: (-x[0], x[1]))
        winner = scored[0][2]

        user_ready = [
            t for t in group
            if t.get("password_hash") and not t.get("must_change_password")
        ]
        if user_ready and (
            not winner.get("password_hash")
            or winner.get("must_change_password")
        ):
            best = max(user_ready, key=lambda t: _therapist_record_score(t))
            if best["id"] != winner["id"]:
                await db.therapists.update_one({"id": winner["id"]}, {"$set": {
                    "password_hash": best["password_hash"],
                    "must_change_password": False,
                    "temp_password_set_at": None,
                    "launch_credentials_generated_at": best.get("launch_credentials_generated_at"),
                }})
                actions.append({
                    "email": em,
                    "action": "merged_user_password",
                    "kept_id": winner["id"],
                    "from_id": best["id"],
                })

        for _, _, loser in scored[1:]:
            if loser["id"] == winner["id"] or any(loser["id"] == lid for lid, _ in to_delete):
                continue
            to_delete.append((loser["id"], winner["id"]))
            actions.append({
                "email": em,
                "kept_id": winner["id"],
                "kept_name": winner.get("name"),
                "removed_id": loser["id"],
                "removed_name": loser.get("name"),
            })

    for loser_id, winner_id in to_delete:
        await db.schedule_cells.update_many(
            {"therapist_id": loser_id},
            {"$set": {"therapist_id": winner_id}},
        )
        await db.users.delete_many({"therapist_id": loser_id})
        await db.therapists.delete_one({"id": loser_id})
    return {"removed": len(to_delete), "actions": actions}


THERAPIST_FIRST_NAME_ALIASES = {
    "hajer": "hajar", "hajar": "hajar",
    "shrooq": "shroug", "shroug": "shroug",
    "bodoor": "bodour", "bodour": "bodour",
    "genan": "jenan", "jenan": "jenan",
    "fahdah": "fahda", "fahda": "fahda",
}

_therapist_alias_map_cache: Optional[Dict[str, List[str]]] = None


def _therapist_identity_token(t: dict) -> Optional[str]:
    """Stable grouping key for duplicate therapist rows (display name, email, or key)."""
    disp = therapist_schedule_display_name(t).strip().lower()
    if disp:
        return f"display:{disp}"
    canon = _therapist_canonical_email(t.get("email"))
    if canon:
        return f"email:{canon}"
    key = (t.get("key") or "").strip().lower()
    if key and not _is_legacy_ms_numeric_key(key):
        return f"key:{key}"
    raw = re.sub(r"^Ms\.?\s*", "", (t.get("name") or "").strip(), flags=re.I)
    parts = [p for p in raw.split() if p]
    if parts:
        first = THERAPIST_FIRST_NAME_ALIASES.get(parts[0].lower(), parts[0].lower())
        return f"name:{first}"
    return None


def _therapist_dedupe_tokens(t: dict) -> List[str]:
    """All grouping tokens for one therapist row (union-find links duplicates)."""
    tokens: List[str] = []
    disp = therapist_schedule_display_name(t).strip().lower()
    if disp:
        tokens.append(f"display:{disp}")
    canon = _therapist_canonical_email(t.get("email"))
    if canon:
        tokens.append(f"email:{canon}")
    key = (t.get("key") or "").strip().lower()
    if key and not _is_legacy_ms_numeric_key(key):
        tokens.append(f"key:{key}")
    raw = re.sub(r"^Ms\.?\s*", "", (t.get("name") or "").strip(), flags=re.I)
    parts = [p for p in raw.split() if p]
    if parts:
        first = THERAPIST_FIRST_NAME_ALIASES.get(parts[0].lower(), parts[0].lower())
        tokens.append(f"name:{first}")
    tok = _therapist_identity_token(t)
    if tok:
        tokens.append(tok)
    return list(dict.fromkeys(tokens))


def _cluster_therapist_rows(rows: List[dict]) -> List[List[dict]]:
    """Group therapist rows that refer to the same person."""
    parent: Dict[str, str] = {}
    by_id: Dict[str, dict] = {}

    def find(i: str) -> str:
        parent.setdefault(i, i)
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i

    def union(a: str, b: str) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    token_to_id: Dict[str, str] = {}
    for t in rows:
        tid = t.get("id")
        if not tid:
            continue
        by_id[tid] = t
        for tok in _therapist_dedupe_tokens(t):
            if tok in token_to_id:
                union(tid, token_to_id[tok])
            else:
                token_to_id[tok] = tid

    clusters: Dict[str, List[dict]] = {}
    for tid in by_id:
        root = find(tid)
        clusters.setdefault(root, []).append(by_id[tid])
    return list(clusters.values())


def _dedupe_therapist_rows_for_display(rows: List[dict]) -> List[dict]:
    """Return one canonical row per therapist for list endpoints."""
    deduped = [_pick_canonical_therapist(group) for group in _cluster_therapist_rows(rows)]
    return sorted(deduped, key=lambda x: therapist_schedule_display_name(x).lower())


async def _build_therapist_id_alias_map() -> Dict[str, List[str]]:
    """Map each therapist row id to equivalent ids (duplicate logins / email aliases)."""
    global _therapist_alias_map_cache
    if _therapist_alias_map_cache is not None:
        return _therapist_alias_map_cache
    therapists = await db.therapists.find(
        {}, {"_id": 0, "id": 1, "name": 1, "key": 1, "email": 1}
    ).to_list(500)
    groups: Dict[str, List[str]] = {}
    for t in therapists:
        tid = t.get("id")
        if not tid:
            continue
        tok = _therapist_identity_token(t)
        if tok:
            groups.setdefault(tok, []).append(tid)
    alias_map: Dict[str, List[str]] = {}
    for ids in groups.values():
        unique = list(dict.fromkeys(ids))
        for tid in unique:
            alias_map[tid] = unique
    _therapist_alias_map_cache = alias_map
    return alias_map


def _invalidate_therapist_alias_map_cache() -> None:
    global _therapist_alias_map_cache
    _therapist_alias_map_cache = None


async def _expand_therapist_ids(therapist_id: Optional[str]) -> List[str]:
    """All therapist row ids equivalent to therapist_id (for prep/suppression matching)."""
    if not therapist_id:
        return []
    alias_map = await _build_therapist_id_alias_map()
    return alias_map.get(therapist_id, [therapist_id])


def _therapist_ids_overlap(a: Optional[str], b: Optional[str], alias_map: Dict[str, List[str]]) -> bool:
    if not a or not b:
        return False
    if a == b:
        return True
    return b in alias_map.get(a, [a])


async def _dedupe_therapists_by_identity() -> dict:
    """Merge duplicate therapist rows that share display name, email alias, or first-name identity."""
    therapists = await db.therapists.find({}, {"_id": 0}).to_list(500)
    cell_counts: Dict[str, int] = {}
    for t in therapists:
        cell_counts[t["id"]] = await db.schedule_cells.count_documents({"therapist_id": t["id"]})

    to_delete: List[tuple] = []
    actions: List[dict] = []
    for group in _cluster_therapist_rows(therapists):
        if len(group) < 2:
            continue
        scored = [
            (_therapist_record_score(t) + cell_counts.get(t["id"], 0), t.get("created_at") or "", t)
            for t in group
        ]
        scored.sort(key=lambda x: (-x[0], x[1]))
        winner = scored[0][2]
        for _, _, loser in scored[1:]:
            if loser["id"] == winner["id"]:
                continue
            to_delete.append((loser["id"], winner["id"]))
            actions.append({
                "token": therapist_schedule_display_name(winner).lower() or _therapist_identity_token(winner),
                "kept_id": winner["id"],
                "kept_name": winner.get("name"),
                "removed_id": loser["id"],
                "removed_name": loser.get("name"),
            })

    for loser_id, winner_id in to_delete:
        await db.schedule_cells.update_many(
            {"therapist_id": loser_id},
            {"$set": {"therapist_id": winner_id}},
        )
        async for sess in db.sessions.find({"therapist_ids": loser_id}, {"_id": 0, "id": 1, "therapist_ids": 1}):
            ids = [winner_id if x == loser_id else x for x in (sess.get("therapist_ids") or [])]
            await db.sessions.update_one(
                {"id": sess["id"]},
                {"$set": {"therapist_ids": list(dict.fromkeys(ids))}},
            )
        await db.prep_history.update_many(
            {"therapist_id": loser_id},
            {"$set": {"therapist_id": winner_id}},
        )
        await db.schedule_preparations.update_many(
            {"therapist_id": loser_id},
            {"$set": {"therapist_id": winner_id}},
        )
        await db.schedule_prep_suppressions.update_many(
            {"therapist_id": loser_id},
            {"$set": {"therapist_id": winner_id}},
        )
        await db.users.delete_many({"therapist_id": loser_id})
        await db.therapists.delete_one({"id": loser_id})
    _invalidate_therapist_alias_map_cache()
    return {"removed": len(to_delete), "actions": actions}


# Known therapist passwords for first-time email login (only when password_hash is missing).
THERAPIST_BOOTSTRAP_PASSWORDS = {
    "asma@boostgrowthsa.com": "Asma@123",
}

LAUNCH_PASSWORD_SUFFIX = "Launch2026"
UNIFIED_LAUNCH_PASSWORD = "growth2026"
# Clients whose invoices stay partial (half paid) during bulk mark-paid rollout.
PARTIAL_PAYMENT_CLIENT_FILE_NOS = frozenset({"079"})  # Fahad Suliman — half paid (not Fahad Alyahya #011)


async def _migrate_bootstrap_therapist_passwords() -> int:
    """Set bootstrap passwords only when a therapist has no password yet — never overwrite on deploy."""
    updated = 0
    for email, password in THERAPIST_BOOTSTRAP_PASSWORDS.items():
        t = await _find_therapist_by_email(email)
        if not t or t.get("password_hash") or t.get("launch_credentials_generated_at"):
            continue
        await db.therapists.update_one(
            {"id": t["id"]},
            {"$set": {
                "password_hash": hash_password(password),
                "must_change_password": True,
                "temp_password_set_at": now_iso(),
            }},
        )
        updated += 1
    return updated


def _launch_temp_password(name: str) -> str:
    """Predictable launch password: Firstname@Launch2026 (from Ms. Asma → Asma@Launch2026)."""
    first = re.sub(r"^Ms\.?\s*", "", (name or "").strip(), flags=re.I).split()[0] or "User"
    return f"{first[0].upper()}{first[1:]}@{LAUNCH_PASSWORD_SUFFIX}" if first else f"User@{LAUNCH_PASSWORD_SUFFIX}"


async def _migrate_hr_password_once() -> bool:
    """One-time: set HR password to Boost@2026 without re-applying on future deploys."""
    meta_key = "hr_password_boost2026"
    if await db.meta.find_one({"key": meta_key, "done": True}):
        return False
    hr = await db.users.find_one({"email": HR_OPS_EMAIL})
    if not hr:
        return False
    await db.users.update_one({"email": HR_OPS_EMAIL}, {"$set": {
        "password_hash": hash_password(UNIFIED_LAUNCH_PASSWORD),
        "must_change_password": False,
        "is_hr_ops": True,
    }})
    await db.meta.update_one(
        {"key": meta_key},
        {"$set": {"done": True, "updated_at": now_iso()}},
        upsert=True,
    )
    return True


async def _ensure_walaa_ops_login_once() -> bool:
    """One-time: link Walaa ops login to wabuissa@ and restore known password."""
    meta_key = "walaa_ops_login_restore_v1"
    if await db.meta.find_one({"key": meta_key, "done": True}):
        return False
    restore_pw = "Walaa@12345"
    pw_hash = hash_password(restore_pw)

    t = await db.therapists.find_one({"key": "msWalaa"}, {"_id": 0})
    if not t:
        t = await db.therapists.find_one(
            {"name": {"$regex": r"^ms\.?\s*walaa\b", "$options": "i"}},
            {"_id": 0},
        )
    if not t:
        for em in ("wabuissa@boostgrowthsa.com", "walaa@boostgrowthsa.com"):
            t = await _find_therapist_by_email(em)
            if t:
                break
    if t:
        await db.therapists.update_one({"id": t["id"]}, {"$set": {
            "email": "wabuissa@boostgrowthsa.com",
            "key": "msWalaa",
            "password_hash": pw_hash,
            "must_change_password": False,
            "temp_password_set_at": None,
        }})

    user = await db.users.find_one(
        {"email": {"$regex": r"^(wabuissa|walaa)@boostgrowthsa\.com$", "$options": "i"}},
    )
    if user:
        await db.users.update_one({"id": user["id"]}, {"$set": {
            "email": "wabuissa@boostgrowthsa.com",
            "password_hash": pw_hash,
            "must_change_password": False,
            "name": user.get("name") or "Walaa",
            "role": "admin",
        }})
        await db.users.delete_many({
            "email": {"$regex": r"^walaa@boostgrowthsa\.com$", "$options": "i"},
            "id": {"$ne": user["id"]},
        })
    else:
        uid = t["id"] if t else str(uuid.uuid4())
        await db.users.insert_one({
            "id": uid,
            "email": "wabuissa@boostgrowthsa.com",
            "password_hash": pw_hash,
            "name": "Walaa",
            "role": "admin",
            "must_change_password": False,
            "created_at": now_iso(),
        })

    await db.meta.update_one(
        {"key": meta_key},
        {"$set": {"done": True, "updated_at": now_iso()}},
        upsert=True,
    )
    return True


async def _apply_inactive_client_status() -> int:
    """Ensure known inactive clients stay out of active billing/prep lists."""
    n = 0
    for file_no in INACTIVE_CLIENT_FILE_NOS:
        res = await db.clients.update_many(
            {"file_no": file_no, "deleted": {"$ne": True}},
            {"$set": {"status": "Inactive"}},
        )
        n += res.modified_count
    return n

UPLOAD_DIR = ROOT_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)

# Binary payloads stored in MongoDB so uploads survive Railway ephemeral disk.
FILE_DATA_FIELDS = frozenset({
    "attachment_file_data", "document_file_data", "file_data",
})


def _b64_encode(data: bytes) -> str:
    return base64.b64encode(data).decode("ascii")


def _b64_decode(data_b64: str) -> bytes:
    return base64.b64decode(data_b64)


def _persist_upload(stored: str, content: bytes) -> str:
    """Best-effort disk write; always return base64 for MongoDB."""
    try:
        (UPLOAD_DIR / stored).write_bytes(content)
    except OSError:
        pass
    return _b64_encode(content)


def _load_upload(stored: Optional[str], file_data_b64: Optional[str] = None) -> Optional[bytes]:
    if stored:
        fp = UPLOAD_DIR / stored
        if fp.exists():
            return fp.read_bytes()
    if file_data_b64:
        try:
            return _b64_decode(file_data_b64)
        except Exception:
            pass
    return None


def _strip_file_data(doc: Optional[dict]) -> Optional[dict]:
    if not doc:
        return doc
    for key in FILE_DATA_FIELDS:
        doc.pop(key, None)
    return doc


def _bytes_file_response(content: bytes, filename: str) -> Response:
    media_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    safe = quote(filename)
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'inline; filename="{safe}"; filename*=UTF-8\'\'{safe}'},
    )


FILE_UNAVAILABLE_DETAIL = (
    "Attachment unavailable on server — please ask the therapist to upload the file again"
)


app = FastAPI(title="Boost Growth Portal API")
api = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ["JWT_SECRET"]

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_password(p: str, h: str) -> bool:
    return bcrypt.checkpw(p.encode(), h.encode())

def create_token(data: dict, hours: int = 24) -> str:
    payload = {**data, "exp": datetime.now(timezone.utc) + timedelta(hours=hours)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = decode_token(token)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    if payload.get("role") == "admin":
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    else:
        user = await db.therapists.find_one({"id": payload["sub"]}, {"_id": 0, "pin_hash": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    user["role"] = payload.get("role", user.get("role", "therapist"))
    return user

async def admin_only(user: dict = Depends(get_current_user)) -> dict:
    if _is_walaa_ops(user):
        return user
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    if _is_client_lead(user):
        raise HTTPException(status_code=403, detail="Admin access required")
    if _is_hr_ops(user):
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

FULL_CLIENT_ACCESS_KEYS = frozenset({"mswalaa", "msmaha", "msjenan", "msfahda"})
FULL_CLIENT_NAME_TOKENS = frozenset({"walaa", "maha", "jenan", "fahda"})
WALAA_LOGIN_EMAILS = frozenset({
    "wabuissa@boostgrowthsa.com",
    "walaa@boostgrowthsa.com",
})
WALAA_CANONICAL_EMAIL = "wabuissa@boostgrowthsa.com"
CLIENT_LEAD_EMAILS = frozenset({
    "wabuissa@boostgrowthsa.com",
    "walaa@boostgrowthsa.com",
    "msalthunayan@boostgrowthsa.com",
    "falghadeeb@boostgrowthsa.com",
    "jsalmuhaisin@boostgrowthsa.com",
})
HR_OPS_EMAIL = "hr@boostgrowthsa.com"
HR_OPS_PASSWORD = UNIFIED_LAUNCH_PASSWORD  # Boost@2026 — set on first seed only; never overwritten on deploy
JENAN_EMAIL = "jsalmuhaisin@boostgrowthsa.com"
PENDING_MANAGER_STATUSES = frozenset({"pending", "pending_manager"})
LEAVE_DOC_REQUIRED_TYPES = frozenset({"Sickleave", "Absence", "Permission"})
PENDING_MANAGER_REQUEST_STATUSES = frozenset({"pending", "pending_manager"})
MANAGER_ACTIVE_REQUEST_STATUSES = frozenset({"pending", "pending_manager", "in_progress"})
MANAGER_FORWARD_HR_LEAVE_SOURCES = PENDING_MANAGER_STATUSES | frozenset({"pending_attachment"})
OPEN_LEAVE_STATUSES = frozenset({"pending", "pending_manager", "pending_hr", "pending_attachment", "in_progress"})


def _can_staff_request_scope(user: dict) -> bool:
    """Manager Hub / HR staff queue — Jenan, HR, portal admin, Walaa ops."""
    return _is_portal_admin(user) or _is_hr_ops(user) or _is_jenan(user) or _is_walaa_ops(user)


def _can_view_all_leaves(user: dict, scope_norm: str) -> bool:
    if _is_portal_admin(user) or _is_hr_ops(user):
        return True
    return scope_norm == "staff" and (_is_jenan(user) or _is_walaa_ops(user))


def _coerce_manager_approve_to_hr(status: str, notify_hr: Optional[bool] = None) -> str:
    """Manager approve → pending_hr when HR follow-up requested, else final approved."""
    if status in ("approved", "manager_approve"):
        if notify_hr is False:
            return "approved"
        return "pending_hr"
    return status


def _is_client_lead(user: dict) -> bool:
    email = (user.get("email") or "").lower().strip()
    if email in CLIENT_LEAD_EMAILS:
        return True
    key = (user.get("key") or "").lower()
    if key in FULL_CLIENT_ACCESS_KEYS:
        return True
    name = (user.get("name") or "").lower().replace("ms.", "").replace("ms ", "").strip()
    first = name.split()[0] if name else ""
    return first in FULL_CLIENT_NAME_TOKENS


def _has_full_client_access(user: dict) -> bool:
    if user.get("role") == "admin":
        return True
    return _is_client_lead(user)


def _session_editable_by_user(user: dict, session: dict) -> bool:
    """Therapists may edit/delete only within 1 hour of logging; ops/admin always."""
    if _has_full_client_access(user) or _is_hr_ops(user):
        return True
    if _is_portal_admin(user) and not _is_client_lead(user):
        return True
    if _is_walaa_ops(user):
        return True
    if user.get("role") != "therapist":
        return False
    created = session.get("created_at") or ""
    if not created:
        return False
    try:
        ts = datetime.fromisoformat(created.replace("Z", "+00:00"))
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - ts) <= timedelta(hours=1)
    except Exception:
        return False


def _normalize_client_status(status: Optional[str]) -> str:
    s = (status or "Active").strip()
    if s.upper() == "OK":
        return "Active"
    return s or "Active"


def _can_supervisor_review_purchases(user: dict) -> bool:
    return (
        _is_portal_admin(user)
        or _is_hr_ops(user)
        or _is_walaa_ops(user)
        or (_is_client_lead(user) and not _is_jenan(user))
    )


def _can_manager_finalize_purchases(user: dict) -> bool:
    return _is_jenan(user) or _is_portal_admin(user)


def _is_hr_ops(user: dict) -> bool:
    if user.get("is_hr_ops"):
        return True
    email = (user.get("email") or "").lower().strip()
    return email == HR_OPS_EMAIL


def _is_walaa_ops(user: dict) -> bool:
    email = (user.get("email") or "").lower().strip()
    if email in WALAA_LOGIN_EMAILS:
        return True
    key = (user.get("key") or "").lower()
    if key == "mswalaa":
        return True
    name = (user.get("name") or "").lower().replace("ms.", "").replace("ms ", "").strip()
    return name.startswith("walaa")


def _can_parent_cancellation_ops(user: dict) -> bool:
    return _is_portal_admin(user) or _is_hr_ops(user) or _is_walaa_ops(user)


def _is_portal_admin(user: dict) -> bool:
    return user.get("role") == "admin" and not _is_client_lead(user) and not _is_hr_ops(user)


def _is_staff_admin(user: dict) -> bool:
    """Admin login or client-lead team (Walaa, Maha, Jenan, Fahda) — full client access."""
    return _is_portal_admin(user) or _has_full_client_access(user)


def _is_jenan(user: dict) -> bool:
    email = (user.get("email") or "").lower().strip()
    if email in (
        "jsalmuhaisin@boostgrowthsa.com",
        "jenan@boostgrowthsa.com",
        "genan@boostgrowthsa.com",
    ):
        return True
    key = (user.get("key") or "").lower()
    if key == "msjenan":
        return True
    name = (user.get("name") or "").lower().replace("ms.", "").replace("ms ", "").strip()
    return name.startswith("jenan")


def _can_delete_staff_submission(user: dict, owner_id: str | None) -> bool:
    if _is_portal_admin(user) or _is_walaa_ops(user) or _is_jenan(user):
        return True
    return bool(owner_id and owner_id == user.get("id"))


async def leave_manager(user: dict = Depends(get_current_user)) -> dict:
    if _is_portal_admin(user) or _is_jenan(user):
        return user
    raise HTTPException(status_code=403, detail="Leave management access required")


async def manager_reports_access(user: dict = Depends(get_current_user)) -> dict:
    """Reports dashboard — portal admin, Walaa ops, HR ops, and Jenan."""
    if _is_portal_admin(user) or _is_walaa_ops(user) or _is_hr_ops(user) or _is_jenan(user):
        return user
    raise HTTPException(status_code=403, detail="Reports access required")


async def hr_manager_access(user: dict = Depends(get_current_user)) -> dict:
    if _can_staff_request_scope(user):
        return user
    raise HTTPException(status_code=403, detail="HR manager access required")


async def import_access(user: dict = Depends(get_current_user)) -> dict:
    if _is_portal_admin(user) or _is_walaa_ops(user):
        return user
    raise HTTPException(status_code=403, detail="Import access required")


async def client_lead_or_admin(user: dict = Depends(get_current_user)) -> dict:
    if _is_portal_admin(user) or _is_client_lead(user) or _is_hr_ops(user):
        return user
    raise HTTPException(status_code=403, detail="Admin access required")


async def schedule_edit_or_admin(user: dict = Depends(get_current_user)) -> dict:
    """Portal admin or ops leads (Walaa, Maha, Fahda, Jenan) may edit schedule cells."""
    if _is_portal_admin(user) or _is_client_lead(user) or _is_hr_ops(user):
        return user
    raise HTTPException(status_code=403, detail="Admin access required")


async def ops_or_admin(user: dict = Depends(get_current_user)) -> dict:
    if _is_portal_admin(user) or _is_hr_ops(user) or _is_walaa_ops(user):
        return user
    raise HTTPException(status_code=403, detail="Admin access required")


async def billing_view_or_ops(user: dict = Depends(get_current_user)) -> dict:
    """View client invoices — ops team + Jenan (read-only)."""
    if _is_portal_admin(user) or _is_hr_ops(user) or _is_walaa_ops(user) or _is_jenan(user):
        return user
    raise HTTPException(status_code=403, detail="Admin access required")

def _actor_display(user: dict) -> str:
    name = (user.get("name") or "").strip()
    if name:
        return name.replace("Ms. ", "", 1) if name.startswith("Ms. ") else name
    email = (user.get("email") or "").strip()
    if email:
        return email.split("@")[0].replace(".", " ").title()
    return "Staff"

def set_auth_cookie(response: Response, token: str):
    response.set_cookie(key="access_token", value=token, httponly=True,
                        secure=True, samesite="none", max_age=86400, path="/")

# ------------------- Models -------------------
class LoginIn(BaseModel):
    email: EmailStr
    password: str

class TherapistPinLogin(BaseModel):
    therapist_id: str
    pin: str

class TherapistEmailLogin(BaseModel):
    email: EmailStr
    password: str

class ChangePasswordIn(BaseModel):
    old_password: str
    new_password: str

class LaunchCredentialsIn(BaseModel):
    force: bool = False

class UnifiedLaunchPasswordIn(BaseModel):
    force_change: bool = True

class TherapistIn(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    color: Optional[str] = "#7A8A6A"
    pin: str

class TherapistUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    color: Optional[str] = None
    pin: Optional[str] = None
    show_on_schedule: Optional[bool] = None

class ScheduleCellIn(BaseModel):
    therapist_id: str
    day: int
    time_slot: str
    service_code: Optional[str] = "SS"
    child_name: Optional[str] = None
    note: Optional[str] = None
    cover_child_name: Optional[str] = None  # ops: specialist covered at another client (no session log required)
    custom_time: Optional[str] = None
    state: Optional[str] = "normal"
    color: Optional[str] = None
    duration: Optional[float] = 1  # hours spanned (supports 1.5, 2.5, etc.)
    week_start: str

class SchedulePreparationIn(BaseModel):
    therapist_id: str
    client_id: str
    session_date: str  # ISO yyyy-mm-dd
    time_slot: Optional[str] = None
    schedule_cell_id: Optional[str] = None
    week_start: Optional[str] = None
    day: Optional[int] = None
    notes: Optional[str] = None
    internal_note: Optional[str] = None
    cell_child_name: Optional[str] = None

class SchedulePreparationNoteIn(BaseModel):
    therapist_id: str
    client_id: str
    session_date: str
    schedule_cell_id: Optional[str] = None
    time_slot: Optional[str] = None
    internal_note: Optional[str] = None

class SchedulePreparationClearIn(BaseModel):
    therapist_id: str
    client_id: str
    session_date: str
    schedule_cell_id: Optional[str] = None
    time_slot: Optional[str] = None
    """When true, hide the green badge even if a session is still logged (schedule-only fix)."""
    suppress_badge: bool = True
    """When true, also remove matching rows from Session Preparation history."""
    delete_prep_history: bool = False
    """When true, delete logged sessions for this therapist + client + date."""
    delete_sessions: bool = False

class PrepHistoryInvoiceLinkIn(BaseModel):
    invoice_id: Optional[str] = None
    notes: Optional[str] = None

class LocationIn(BaseModel):
    service: str
    address: str

class ClientIn(BaseModel):
    name: str
    file_no: Optional[str] = None
    birth_date: Optional[str] = None  # ISO yyyy-mm-dd
    age: Optional[str] = None
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    package_hours: Optional[float] = 24
    billing_mode: Optional[str] = "hours"   # "hours" (count by hours used) or "weeks" (4-week school cycle)
    cycle_weeks: Optional[int] = 4           # for billing_mode="weeks"
    cycle_start_date: Optional[str] = None   # ISO yyyy-mm-dd; first day of current billing cycle
    package_end_date: Optional[str] = None   # ISO yyyy-mm-dd; package expiry / end-of-cycle date (manual)
    payment_status: Optional[str] = "pending"  # "complete" or "pending"
    package_reset_at: Optional[str] = None    # ISO timestamp; sessions before this are excluded from current cycle (manual reset)
    notes: Optional[str] = None
    main_therapist_id: Optional[str] = None
    co_therapist_ids: Optional[List[str]] = []
    supervisor: Optional[str] = None
    locations: Optional[List[LocationIn]] = []
    color: Optional[str] = None
    drive_url: Optional[str] = None
    schedule_color: Optional[str] = None  # custom color for all schedule cells for this client
    # Optional client status & service type & attachment URLs (Change 4)
    status: Optional[str] = "Active"                  # Active / Inactive
    service_type: Optional[str] = None                # HS / SS / HS+SS / AVC
    address: Optional[str] = None                     # general address (also via locations)
    intake_file_url: Optional[str] = None
    attendance_sheet_url: Optional[str] = None
    progress_reports_url: Optional[str] = None
    case_summary_url: Optional[str] = None
    drive_folder_id: Optional[str] = None
    drive_links: Optional[List[dict]] = None
    record_files: Optional[List[dict]] = None
    case_summary_sections: Optional[dict] = None

class InvoiceIn(BaseModel):
    invoice_number: str  # manual entry, e.g. "INV 4042"
    notes: Optional[str] = None
    amount: Optional[float] = None
    period_from: Optional[str] = None  # ISO date (cycle start)
    period_to: Optional[str] = None    # ISO date (package end date)
    package_size: Optional[float] = None         # number of sessions or hours
    payment_status: Optional[str] = "pending"    # "complete" | "partial" | "pending"
    start_date: Optional[str] = None             # ISO date - invoice cycle start (for filtering sessions)
    service_type: Optional[str] = None           # "Home Session" | "School Support"
    is_closed: Optional[bool] = False            # whether the invoice is closed
    close_date: Optional[str] = None             # ISO date when closed
    week_overrides: Optional[dict] = None        # {"1": "excluded"|"completed"} manual SS weeks
    ss_week_count: Optional[int] = 4             # SS billing weeks (4 default; admin may extend to 5+)
    installment_percent: Optional[float] = None  # e.g. 50 = paid 50% of invoice amount
    amount_paid: Optional[float] = None          # partial payments received so far
    next_payment_reminder_at: Optional[str] = None  # ISO date — remind parents about next installment
    payment_notes: Optional[str] = None

class InvoicePaymentIn(BaseModel):
    payment_status: Optional[str] = None
    amount: Optional[float] = None
    amount_paid: Optional[float] = None
    installment_percent: Optional[float] = None
    next_payment_reminder_at: Optional[str] = None
    payment_notes: Optional[str] = None


class InvoiceCalendarManualIn(BaseModel):
    title: str
    date: str  # ISO yyyy-mm-dd
    client_id: Optional[str] = None
    invoice_id: Optional[str] = None
    notes: Optional[str] = None

class ScheduleClosureIn(BaseModel):
    date: str   # ISO yyyy-mm-dd
    label: str  # e.g. National Day, Eid
    therapist_ids: Optional[List[str]] = None  # empty / omitted = all therapists

class CenterUpdateIn(BaseModel):
    title: str
    body: Optional[str] = None
    date: Optional[str] = None  # ISO yyyy-mm-dd display date
    is_important: bool = False
    requires_ack: bool = False
    send_to_specialists: bool = False

class SessionIn(BaseModel):
    client_id: str
    session_date: str  # ISO date
    start_time: Optional[str] = None  # "14:00"
    end_time: Optional[str] = None
    hours: float = 0
    status: str = "Completed"  # Completed, No Service, Cancelled, No Show
    therapist_ids: List[str] = []
    note: Optional[str] = None
    location: Optional[str] = None  # which location used (HS / SS)
    service_type: Optional[str] = None  # HS / SS — links session to open invoice


def _session_hours_from_times(start_time: Optional[str], end_time: Optional[str]) -> Optional[float]:
    """Compute billable hours from therapist-edited HH:MM times."""
    if not start_time or not end_time:
        return None
    try:
        h1, m1 = [int(x) for x in start_time.split(":")[:2]]
        h2, m2 = [int(x) for x in end_time.split(":")[:2]]
    except (TypeError, ValueError):
        return None
    diff = (h2 * 60 + m2) - (h1 * 60 + m1)
    if diff < 0:
        diff += 24 * 60
    return round(diff / 30) / 2


def _apply_session_time_edits(payload: SessionIn) -> SessionIn:
    """Honor therapist-edited start/end on create and update."""
    computed = _session_hours_from_times(payload.start_time, payload.end_time)
    if computed is not None:
        payload.hours = computed
    return payload


def _require_session_log_fields(user: dict, payload: SessionIn) -> None:
    """Therapists must record time and notes when logging/preparing a session."""
    if _has_full_client_access(user) or _is_hr_ops(user):
        return
    if not (payload.start_time or "").strip() or not (payload.end_time or "").strip():
        raise HTTPException(
            status_code=400,
            detail="Start and end time are required when logging a session.",
        )
    if not (payload.note or "").strip():
        raise HTTPException(
            status_code=400,
            detail="Session notes are required when logging a session.",
        )

class RequestIn(BaseModel):
    title: str
    description: Optional[str] = ""
    request_type: str = "general"
    priority: str = "normal"
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    reward_type: Optional[str] = None
    extra_notes: Optional[str] = None
    requires_attachment: bool = False

class RequestStatusUpdate(BaseModel):
    status: str
    admin_note: Optional[str] = None
    notify_hr: Optional[bool] = None
    notify_therapist: Optional[bool] = None

PURCHASE_CATEGORIES = [
    "Events & Celebrations",
    "Training & Workshops",
    "Catering & Hospitality",
    "Supplies & Materials",
    "Services",
    "Transportations",
    "Software & Subscriptions",
    "Marketing & Media",
    "Decoration",
    "Miscellaneous",
]

PURCHASE_STATUSES = (
    "pending",
    "supervisor_approved",
    "supervisor_rejected",
    "pending_manager",
    "manager_approved",
    "manager_rejected",
    "approved",
    "rejected",
    "reimbursed",
)

class PurchaseLineItemIn(BaseModel):
    item: str
    qty: Optional[str] = "1"
    unit_price: Optional[str] = ""
    total: Optional[float] = None

class PurchaseIn(BaseModel):
    item: str
    category: str
    description: Optional[str] = ""
    qty: Optional[str] = "1"
    unit_price: Optional[str] = ""
    total: Optional[float] = None
    purchase_date: Optional[str] = None
    notes: Optional[str] = None
    therapist_id: Optional[str] = None
    line_items: Optional[List[PurchaseLineItemIn]] = None

class PurchaseUpdate(BaseModel):
    item: Optional[str] = None
    category: Optional[str] = None
    description: Optional[str] = None
    qty: Optional[str] = None
    unit_price: Optional[str] = None
    total: Optional[float] = None
    status: Optional[str] = None
    reimbursement_date: Optional[str] = None
    purchase_date: Optional[str] = None
    notes: Optional[str] = None
    supervisor_note: Optional[str] = None
    forward_to_manager: Optional[bool] = None

class PurchaseReminderSettingsIn(BaseModel):
    day_of_month: int = 25
    enabled: bool = True
    therapist_ids: List[str] = []

class PersonalEventIn(BaseModel):
    date: str
    title: str
    notes: Optional[str] = ""
    time_label: Optional[str] = None

class DirectoryContactIn(BaseModel):
    name: str
    role: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None

class IntakeIn(BaseModel):
    child_name: str
    parent_name: Optional[str] = None
    phone: Optional[str] = None
    intake_type: str = "pre"
    notes: Optional[str] = None
    status: Optional[str] = "new"
    intake_date: Optional[str] = None
    birth_date: Optional[str] = None  # ISO yyyy-mm-dd
    age: Optional[str] = None
    service: Optional[str] = None          # HS / SS / HS / SS
    district: Optional[str] = None          # Dis column
    time_pref: Optional[str] = None         # Morning / Evening / Any
    diagnosis: Optional[str] = None
    language: Optional[str] = None          # Post-intake only
    priority: Optional[bool] = False

class ResourceIn(BaseModel):
    title: str
    description: Optional[str] = None
    url: str
    category: Optional[str] = "drive"       # drive / file / link
    visibility: str = "all"                 # all / admin / therapist
    icon: Optional[str] = "Folders"
    bg: Optional[str] = "#E5EBE1"
    color: Optional[str] = "#3D4F35"
    sort_order: Optional[int] = 100

class DirectoryContactUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None

class LeaveIn(BaseModel):
    therapist_id: str
    start_date: str               # ISO yyyy-mm-dd
    end_date: str
    days: float = 1
    leave_type: Optional[str] = "Annual"   # Annual / Unpaid / Sickleave / Exam / Emergency
    status: Optional[str] = "pending"      # pending / approved / done / rejected / cancelled
    start_time: Optional[str] = None       # Permission: "14:00"
    end_time: Optional[str] = None         # Permission: "16:00"
    notes: Optional[str] = None
    admin_note: Optional[str] = None

class LeaveStatusUpdate(BaseModel):
    status: str
    admin_note: Optional[str] = None
    deduct_balance: Optional[bool] = True
    is_paid: Optional[bool] = True
    notify_hr: Optional[bool] = None
    notify_therapist: Optional[bool] = None


class TherapistHrProfileUpdate(BaseModel):
    probation_end: Optional[str] = None
    trial_end: Optional[str] = None
    annual_contract_end: Optional[str] = None
    meeting_date: Optional[str] = None
    meeting_notes: Optional[str] = None

class MarkAbsentIn(BaseModel):
    cancel_sessions: bool = True

class MarkAbsenceIn(BaseModel):
    therapist_id: str
    date_from: str
    date_to: str
    leave_type: Optional[str] = "Absence"
    notes: Optional[str] = None
    cancel_sessions: bool = True

class LeaveDocumentVerifyIn(BaseModel):
    verified: bool = True

class CancelNotifyIn(BaseModel):
    cell_id: str
    state: Optional[str] = None           # cancel_therapist / cancel_child (optional)
    message: str
    send_email: Optional[bool] = False
    send_in_app: Optional[bool] = True
    recipient_ids: Optional[List[str]] = None
    extra_email: Optional[str] = None     # legacy single email override

class ScheduleNotifyIn(BaseModel):
    cell_id: str
    message: str
    recipient_ids: List[str] = []
    send_email: Optional[bool] = False
    send_in_app: Optional[bool] = True

class ParentWhatsAppSentIn(BaseModel):
    message: Optional[str] = None

# ------------------- Auth -------------------
@api.post("/auth/login")
async def admin_login(payload: LoginIn, response: Response):
    email = payload.email.lower().strip()
    # Primary: admin users (Walaa: wabuissa@ and walaa@ are aliases)
    user = await _find_user_by_login_email(email)
    if user and verify_password(payload.password, user["password_hash"]):
        login_email = (user.get("email") or email).lower()
        token = create_token({"sub": user["id"], "role": "admin", "email": login_email})
        set_auth_cookie(response, token)
        return {"id": user["id"], "email": login_email, "name": user.get("name"), "role": "admin", "token": token}

    # Secondary: allow ops/supervisors to sign in from the same form using therapist email+password
    t = await _find_therapist_by_login_email(email)
    if t and t.get("password_hash") and verify_password(payload.password, t["password_hash"]):
        token = create_token({"sub": t["id"], "role": "therapist", "name": t["name"], "email": t.get("email")})
        set_auth_cookie(response, token)
        return {"id": t["id"], "name": t["name"], "color": t.get("color"), "email": t.get("email"),
                "key": t.get("key"), "role": "therapist", "token": token,
                "must_change_password": bool(t.get("must_change_password"))}

    raise HTTPException(status_code=401, detail="Invalid credentials")

@api.get("/auth/therapists-list")
async def therapists_list_public():
    rows = await db.therapists.find({}, {"_id": 0, "pin_hash": 0, "password_hash": 0}).sort("name", 1).to_list(500)
    return _dedupe_therapist_rows_for_display(rows)

@api.post("/auth/therapist-login")
async def therapist_login(payload: TherapistPinLogin, response: Response):
    t = await db.therapists.find_one({"id": payload.therapist_id})
    if not t or not verify_password(payload.pin, t["pin_hash"]):
        raise HTTPException(status_code=401, detail="Incorrect PIN")
    token = create_token({"sub": t["id"], "role": "therapist", "name": t["name"]})
    set_auth_cookie(response, token)
    return {"id": t["id"], "name": t["name"], "color": t.get("color"), "key": t.get("key"),
            "role": "therapist", "token": token,
            "must_change_password": bool(t.get("must_change_password"))}

@api.post("/auth/therapist-email-login")
async def therapist_email_login(payload: TherapistEmailLogin, response: Response):
    """Login a therapist using their email + password (new flow). PIN flow remains available."""
    email = payload.email.lower().strip()
    # Ops / supervisors / HR should use the Admin & Supervisor login entry
    if _is_client_lead_login_email(email) or email == HR_OPS_EMAIL:
        raise HTTPException(status_code=403, detail="Please use Admin / Supervisor login for this account")
    t = await _find_therapist_by_login_email(email)
    if not t or not t.get("password_hash") or not verify_password(payload.password, t["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_token({"sub": t["id"], "role": "therapist", "name": t["name"]})
    set_auth_cookie(response, token)
    return {"id": t["id"], "name": t["name"], "color": t.get("color"), "key": t.get("key"),
            "email": t.get("email"), "role": "therapist", "token": token,
            "must_change_password": bool(t.get("must_change_password"))}

@api.post("/auth/change-password")
async def change_password(payload: ChangePasswordIn, user=Depends(get_current_user)):
    """Change password for the currently logged-in user (admin or therapist)."""
    if not payload.new_password or len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    if user.get("role") == "admin":
        u = await db.users.find_one({"id": user["id"]})
        if not u or not verify_password(payload.old_password, u["password_hash"]):
            raise HTTPException(status_code=401, detail="Current password is incorrect")
        await db.users.update_one({"id": user["id"]},
                                  {"$set": {"password_hash": hash_password(payload.new_password),
                                            "must_change_password": False}})
    else:
        t = await db.therapists.find_one({"id": user["id"]})
        if not t:
            raise HTTPException(status_code=404, detail="Therapist not found")
        # Allow change with either current password OR current PIN (for first-time migration from PIN)
        ok = False
        if t.get("password_hash") and verify_password(payload.old_password, t["password_hash"]):
            ok = True
        elif t.get("pin_hash") and verify_password(payload.old_password, t["pin_hash"]):
            ok = True
        if not ok:
            raise HTTPException(status_code=401, detail="Current password is incorrect")
        await db.therapists.update_one({"id": user["id"]},
                                       {"$set": {"password_hash": hash_password(payload.new_password),
                                                 "must_change_password": False,
                                                 "temp_password_set_at": None}})
    return {"ok": True}

@api.post("/therapists/{tid}/reset-password")
async def reset_therapist_password(tid: str, _=Depends(admin_only)):
    """Admin generates a temporary password for a therapist.
    Therapist must change it on next login. Password stays until changed or admin resets again."""
    import secrets
    t = await db.therapists.find_one({"id": tid})
    if not t:
        raise HTTPException(status_code=404, detail="Therapist not found")
    temp = secrets.token_urlsafe(6)[:8]
    ts = now_iso()
    await db.therapists.update_one({"id": tid}, {"$set": {
        "password_hash": hash_password(temp),
        "must_change_password": True,
        "temp_password_set_at": ts,
    }})
    return {"ok": True, "therapist_id": tid, "email": t.get("email"), "temp_password": temp, "temp_password_set_at": ts}

@api.post("/admin/generate-launch-credentials")
async def generate_launch_credentials(body: LaunchCredentialsIn, _=Depends(admin_only)):
    """Bulk-set stable launch passwords for all therapists with email.
    Skips therapists who already have launch credentials unless force=true.
    Passwords remain valid until the therapist changes them or admin regenerates."""
    therapists = await db.therapists.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    generated = []
    skipped = []
    no_email = []
    ts = now_iso()

    for t in therapists:
        email = (t.get("email") or "").strip()
        if not email:
            no_email.append({"id": t["id"], "name": t.get("name")})
            continue
        if not body.force and t.get("temp_password_set_at") and t.get("password_hash"):
            skipped.append({
                "id": t["id"], "name": t.get("name"), "email": email,
                "temp_password_set_at": t.get("temp_password_set_at"),
            })
            continue
        temp = _launch_temp_password(t.get("name") or "")
        await db.therapists.update_one({"id": t["id"]}, {"$set": {
            "password_hash": hash_password(temp),
            "must_change_password": True,
            "temp_password_set_at": ts,
        }})
        generated.append({
            "id": t["id"], "name": t.get("name"), "email": email,
            "temp_password": temp, "temp_password_set_at": ts,
        })

    return {
        "ok": True,
        "generated_at": ts,
        "generated": generated,
        "skipped": skipped,
        "no_email": no_email,
        "message": (
            f"Generated {len(generated)} credential(s). "
            f"Skipped {len(skipped)} with existing launch passwords (use force to regenerate). "
            f"{len(no_email)} without email."
        ),
    }

@api.post("/admin/set-unified-launch-password")
async def set_unified_launch_password(body: UnifiedLaunchPasswordIn = UnifiedLaunchPasswordIn(), _=Depends(admin_only)):
    """Set the same launch password for every therapist with email. Admin-triggered only — never on deploy."""
    therapists = await db.therapists.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    ts = now_iso()
    updated = []
    no_email = []
    pw_hash = hash_password(UNIFIED_LAUNCH_PASSWORD)
    force_change = bool(body.force_change)

    for t in therapists:
        email = (t.get("email") or "").strip()
        if not email:
            no_email.append({"id": t["id"], "name": t.get("name")})
            continue
        await db.therapists.update_one({"id": t["id"]}, {"$set": {
            "password_hash": pw_hash,
            "must_change_password": force_change,
            "launch_credentials_generated_at": ts,
            "temp_password_set_at": ts,
        }})
        updated.append({"id": t["id"], "name": t.get("name"), "email": email})

    change_note = (
        "They must set a new password on next login."
        if force_change
        else "No forced password change on login."
    )
    return {
        "ok": True,
        "password": UNIFIED_LAUNCH_PASSWORD,
        "force_change": force_change,
        "generated_at": ts,
        "updated_count": len(updated),
        "updated": updated,
        "no_email": no_email,
        "message": (
            f"{len(updated)} therapists updated. "
            f"Password: {UNIFIED_LAUNCH_PASSWORD}. {change_note}"
        ),
    }


@api.post("/admin/force-therapist-password-change")
async def force_therapist_password_change(_=Depends(admin_only)):
    """Require all therapists with email to change password on next login (does not reset passwords)."""
    therapists = await db.therapists.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    ts = now_iso()
    updated = []
    no_email = []
    for t in therapists:
        email = (t.get("email") or "").strip()
        if not email:
            no_email.append({"id": t["id"], "name": t.get("name")})
            continue
        await db.therapists.update_one({"id": t["id"]}, {"$set": {
            "must_change_password": True,
            "password_change_required_at": ts,
        }})
        updated.append({"id": t["id"], "name": t.get("name"), "email": email})
    return {
        "ok": True,
        "updated_at": ts,
        "updated_count": len(updated),
        "updated": updated,
        "no_email": no_email,
        "message": (
            f"{len(updated)} therapist(s) will be prompted to set a new password on next login."
        ),
    }


@api.post("/admin/reset-hr-password")
async def admin_reset_hr_password(_=Depends(admin_only)):
    """Reset HR ops login to Boost@2026 (admin-triggered; not applied on every deploy)."""
    hr = await db.users.find_one({"email": HR_OPS_EMAIL})
    if not hr:
        raise HTTPException(status_code=404, detail=f"HR user not found: {HR_OPS_EMAIL}")
    await db.users.update_one({"email": HR_OPS_EMAIL}, {"$set": {
        "password_hash": hash_password(UNIFIED_LAUNCH_PASSWORD),
        "must_change_password": False,
        "is_hr_ops": True,
    }})
    return {
        "ok": True,
        "email": HR_OPS_EMAIL,
        "password": UNIFIED_LAUNCH_PASSWORD,
        "must_change_password": False,
        "message": f"HR password reset for {HR_OPS_EMAIL}",
    }

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    tid = await _resolve_user_therapist_id(user)
    if tid:
        therapist = await db.therapists.find_one(
            {"id": tid}, {"_id": 0, "pin_hash": 0, "password_hash": 0}
        )
        if therapist:
            user["therapist_id"] = tid
            if therapist.get("key"):
                user["key"] = therapist["key"]
            if therapist.get("name"):
                user["name"] = therapist["name"]
            user["must_change_password"] = bool(therapist.get("must_change_password"))
    elif user.get("role") == "therapist":
        user["must_change_password"] = bool(user.get("must_change_password"))
    user["ops_access"] = _is_portal_admin(user) or _is_hr_ops(user) or bool(_is_client_lead(user))
    user["client_lead"] = _is_client_lead(user)
    user["hr_ops"] = _is_hr_ops(user)
    user["portal_admin"] = _is_portal_admin(user)
    user["staff_admin"] = user["portal_admin"] or _is_walaa_ops(user)
    user["jenan_hr"] = _is_jenan(user)
    user["can_manage_leaves"] = _is_jenan(user)
    user["can_hr_review_leaves"] = _is_hr_ops(user)
    user["can_edit_staff_requests"] = _is_jenan(user) or _is_hr_ops(user)
    user["can_access_manager_hub"] = _is_jenan(user) or _is_portal_admin(user) or _is_walaa_ops(user)
    user["can_import"] = _is_portal_admin(user) or _is_walaa_ops(user)
    user["can_edit_intake"] = _is_portal_admin(user) or _is_client_lead(user) or _is_hr_ops(user)
    user["schedule_lead"] = _is_client_lead(user) and not _is_portal_admin(user)
    user["walaa_ops"] = _is_walaa_ops(user)
    user["can_parent_cancellation_ops"] = _can_parent_cancellation_ops(user)
    user["can_view_billing"] = _is_portal_admin(user) or _is_hr_ops(user) or _is_walaa_ops(user) or _is_jenan(user)
    return user

@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

# ------------------- Therapists -------------------
@api.get("/therapists")
async def list_therapists(user=Depends(get_current_user)):
    rows = await db.therapists.find({}, {"_id": 0, "pin_hash": 0, "password_hash": 0}).sort("name", 1).to_list(500)
    return _dedupe_therapist_rows_for_display(rows)

@api.post("/therapists")
async def create_therapist(payload: TherapistIn, _=Depends(admin_only)):
    tid = str(uuid.uuid4())
    doc = {"id": tid, "name": payload.name, "email": payload.email, "phone": payload.phone,
           "color": payload.color or "#7A8A6A", "pin_hash": hash_password(payload.pin),
           "created_at": now_iso()}
    await db.therapists.insert_one(doc)
    return {k: v for k, v in doc.items() if k not in ("_id", "pin_hash")}

@api.put("/therapists/{tid}")
async def update_therapist(tid: str, payload: TherapistUpdate, _=Depends(admin_only)):
    update = {k: v for k, v in payload.model_dump().items() if v is not None and k != "pin"}
    if payload.pin:
        update["pin_hash"] = hash_password(payload.pin)
    if not update:
        raise HTTPException(status_code=400, detail="No fields")
    await db.therapists.update_one({"id": tid}, {"$set": update})
    return await db.therapists.find_one({"id": tid}, {"_id": 0, "pin_hash": 0, "password_hash": 0})

@api.delete("/therapists/{tid}")
async def delete_therapist(tid: str, _=Depends(admin_only)):
    t = await db.therapists.find_one({"id": tid}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Therapist not found")
    sc = await db.schedule_cells.delete_many({"therapist_id": tid})
    uc = await db.users.delete_many({"$or": [{"therapist_id": tid}, {"name": t.get("name")}]})
    await db.therapists.delete_one({"id": tid})
    return {
        "ok": True,
        "name": t.get("name"),
        "schedule_cells_deleted": sc.deleted_count,
        "users_deleted": uc.deleted_count,
    }

# ------------------- Full Database Backup (admin only) -------------------
BACKUP_COLLECTIONS = [
    "users", "therapists", "clients", "sessions", "invoices",
    "leaves", "requests", "progress_reports", "schedule_cells", "schedule_preparations",
    "schedule_prep_suppressions", "prep_history", "schedule_weeks",
    "intake", "intake_pre", "intake_post", "notifications", "attendance_sheets",
    "email_settings", "email_queue",
]
BACKUP_SENSITIVE_FIELDS = {"pin_hash", "password_hash"}
BACKUP_RETENTION_DAYS = int(os.environ.get("BACKUP_RETENTION_DAYS", "30"))
BACKUP_MAX_STORED = int(os.environ.get("BACKUP_MAX_STORED", "30"))
BACKUP_CRON_SECRET = os.environ.get("BACKUP_CRON_SECRET", "")


def _json_safe(obj):
    """Recursively convert BSON / non-JSON-native types to plain JSON values."""
    from bson import ObjectId
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items() if k != "_id"}
    if isinstance(obj, list):
        return [_json_safe(x) for x in obj]
    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, ObjectId):
        return str(obj)
    return obj


async def _build_backup_dump() -> dict:
    dump = {
        "exported_at": now_iso(),
        "db_name": os.environ.get("DB_NAME"),
        "collections": {},
    }
    for cname in BACKUP_COLLECTIONS:
        try:
            docs = await db[cname].find({}, {"_id": 0}).to_list(20000)
        except Exception:
            docs = []
        clean = []
        for d in docs:
            d2 = _json_safe(d)
            for k in list(d2.keys()):
                if k in BACKUP_SENSITIVE_FIELDS:
                    d2[k] = "[REDACTED]"
            clean.append(d2)
        dump["collections"][cname] = clean
    dump["totals"] = {k: len(v) for k, v in dump["collections"].items()}
    return dump


async def _prune_old_backups() -> int:
    """Drop stored backups older than retention window; cap total count."""
    removed = 0
    cutoff = (datetime.now(timezone.utc) - timedelta(days=BACKUP_RETENTION_DAYS)).isoformat()
    r = await db.backups.delete_many({"created_at": {"$lt": cutoff}})
    removed += int(r.deleted_count or 0)
    excess = await db.backups.count_documents({}) - BACKUP_MAX_STORED
    if excess > 0:
        oldest = await db.backups.find({}, {"_id": 0, "id": 1}).sort("created_at", 1).limit(excess).to_list(excess)
        ids = [b["id"] for b in oldest if b.get("id")]
        if ids:
            r2 = await db.backups.delete_many({"id": {"$in": ids}})
            removed += int(r2.deleted_count or 0)
    return removed


async def _store_backup(source: str = "manual") -> dict:
    dump = await _build_backup_dump()
    backup_id = str(uuid.uuid4())
    doc = {
        "id": backup_id,
        "created_at": dump["exported_at"],
        "source": source,
        "db_name": dump.get("db_name"),
        "totals": dump["totals"],
        "data": dump,
    }
    await db.backups.insert_one(doc)
    pruned = await _prune_old_backups()
    return {
        "id": backup_id,
        "created_at": doc["created_at"],
        "source": source,
        "totals": dump["totals"],
        "pruned_old": pruned,
    }


async def _restore_collections_from_dump(
    dump: dict,
    *,
    dry_run: bool,
    only: Optional[List[str]] = None,
    mode: str = "merge",
) -> dict:
    """Restore collections from backup JSON. merge=upsert by id; replace=clear collection first."""
    collections = dump.get("collections") or {}
    if only:
        wanted = {c.strip() for c in only if c.strip()}
        collections = {k: v for k, v in collections.items() if k in wanted}
    report: dict = {}
    for cname, docs in collections.items():
        if not isinstance(docs, list):
            continue
        existing = await db[cname].count_documents({})
        would_insert = 0
        would_update = 0
        for doc in docs:
            if not isinstance(doc, dict):
                continue
            doc_id = doc.get("id")
            if doc_id and await db[cname].find_one({"id": doc_id}, {"_id": 0, "id": 1}):
                would_update += 1
            else:
                would_insert += 1
        entry = {
            "existing_before": existing,
            "would_insert": would_insert,
            "would_update": would_update,
            "docs_in_backup": len(docs),
        }
        if dry_run:
            report[cname] = entry
            continue
        inserted = updated = 0
        if mode == "replace":
            await db[cname].delete_many({})
        for doc in docs:
            if not isinstance(doc, dict):
                continue
            doc_id = doc.get("id")
            if doc_id:
                r = await db[cname].update_one({"id": doc_id}, {"$set": doc}, upsert=True)
                if r.matched_count:
                    updated += 1
                else:
                    inserted += 1
            else:
                await db[cname].insert_one(doc)
                inserted += 1
        entry["inserted"] = inserted
        entry["updated"] = updated
        entry["after"] = await db[cname].count_documents({})
        report[cname] = entry
    return report


def _verify_backup_cron_token(request: Request) -> None:
    token = (
        request.headers.get("X-Backup-Token")
        or request.headers.get("X-Cron-Secret")
        or ""
    ).strip()
    if not BACKUP_CRON_SECRET or token != BACKUP_CRON_SECRET:
        raise HTTPException(status_code=403, detail="Invalid backup token")


@api.get("/admin/full-backup")
async def full_backup(_=Depends(admin_only)):
    """Return a JSON dump of every collection in the DB. Admin-only."""
    import json
    dump = await _build_backup_dump()
    body = json.dumps(dump, ensure_ascii=False, indent=2)
    stamp = now_iso().replace(":", "-")[:19]
    fname = f"boost-growth-backup-{stamp}.json"
    return Response(
        content=body,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@api.post("/admin/scheduled-backup")
async def scheduled_backup(request: Request):
    """Daily cron endpoint — requires BACKUP_CRON_SECRET via X-Backup-Token header."""
    _verify_backup_cron_token(request)
    result = await _store_backup(source="scheduled")
    return {"ok": True, **result}


@api.post("/admin/store-backup")
async def store_backup_now(_=Depends(admin_only)):
    """Store a timestamped backup in MongoDB (Admin UI / manual)."""
    result = await _store_backup(source="manual")
    return {"ok": True, **result}


@api.get("/admin/backups")
async def list_backups(_=Depends(admin_only)):
    """List stored backups (metadata only, no payload)."""
    rows = await db.backups.find(
        {},
        {"_id": 0, "id": 1, "created_at": 1, "source": 1, "totals": 1, "db_name": 1},
    ).sort("created_at", -1).limit(BACKUP_MAX_STORED).to_list(BACKUP_MAX_STORED)
    return {"backups": rows, "retention_days": BACKUP_RETENTION_DAYS, "max_stored": BACKUP_MAX_STORED}


class RestoreBackupIn(BaseModel):
    dry_run: bool = True
    collections: Optional[List[str]] = None
    mode: str = "merge"


@api.post("/admin/restore-backup/{backup_id}")
async def restore_stored_backup(backup_id: str, body: RestoreBackupIn, _=Depends(admin_only)):
    """Restore from a stored backup by id. Default dry_run=true — set false to apply."""
    row = await db.backups.find_one({"id": backup_id}, {"_id": 0})
    if not row or not row.get("data"):
        raise HTTPException(status_code=404, detail="Backup not found")
    if body.mode not in ("merge", "replace"):
        raise HTTPException(status_code=400, detail="mode must be merge or replace")
    report = await _restore_collections_from_dump(
        row["data"],
        dry_run=body.dry_run,
        only=body.collections,
        mode=body.mode,
    )
    return {
        "ok": True,
        "backup_id": backup_id,
        "backup_created_at": row.get("created_at"),
        "dry_run": body.dry_run,
        "collections": report,
    }


@api.post("/admin/restore-from-backup")
async def restore_from_uploaded_backup(
    file: UploadFile = File(...),
    dry_run: str = Form("true"),
    collections: Optional[str] = Form(None),
    mode: str = Form("merge"),
    _=Depends(admin_only),
):
    """Upload a JSON backup file and restore selected collections (dry-run by default)."""
    import json
    raw = await file.read()
    try:
        dump = json.loads(raw.decode("utf-8"))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")
    if not dump.get("collections"):
        raise HTTPException(status_code=400, detail="Missing collections in backup file")
    only = [c.strip() for c in (collections or "").split(",") if c.strip()] or None
    is_dry = str(dry_run).lower() in ("1", "true", "yes")
    if mode not in ("merge", "replace"):
        raise HTTPException(status_code=400, detail="mode must be merge or replace")
    report = await _restore_collections_from_dump(dump, dry_run=is_dry, only=only, mode=mode)
    return {
        "ok": True,
        "exported_at": dump.get("exported_at"),
        "dry_run": is_dry,
        "collections": report,
        "recovery_note_ar": (
            "تجربة جافة — لم يُكتب شيء. عيّن dry_run=false للتطبيق."
            if is_dry else
            "تمت الاستعادة. راجع الأعداد أدناه ثم أعد تحميل البوابة."
        ),
    }


TRIAL_WEEK_START = "2026-06-28"
TRIAL_WEEK_END = "2026-07-02"
AUTO_BACKUP_HOURS = 24
MASTER_CLIENTS_SHEET_URL = os.environ.get(
    "GOOGLE_DRIVE_CLIENTS_SHEET_URL",
    "https://docs.google.com/spreadsheets/d/1D2DQX0M4ieeKz4Z7c-QdO67XbDl1llnlXolLOrDXopk/edit",
)
SCHEDULE_MASTER_SHEET_URL = os.environ.get(
    "GOOGLE_DRIVE_SCHEDULE_URL",
    "https://docs.google.com/spreadsheets/d/1nObLcjV0btqOcPJhZu4fP5S42qHUxAzxdQLdPO-QyUk/edit",
)


async def _get_data_health_snapshot() -> dict:
    """Collection counts and duplicate-therapist groups for Admin health panel."""
    therapists = await db.therapists.find({}, {"_id": 0}).to_list(500)
    dup_groups = sum(1 for g in _cluster_therapist_rows(therapists) if len(g) > 1)

    schedule_weeks: List[dict] = []
    pipeline = [
        {"$group": {"_id": "$week_start", "n": {"$sum": 1}}},
        {"$sort": {"n": -1}},
        {"$limit": 20},
    ]
    async for row in db.schedule_cells.aggregate(pipeline):
        ws = row["_id"]
        meta = await db.schedule_weeks.find_one({"week_start": ws}, {"_id": 0, "status": 1})
        schedule_weeks.append({
            "week_start": ws,
            "cells": row["n"],
            "status": (meta or {}).get("status") or "(none)",
        })

    last_backup = await db.backups.find_one(
        {},
        {"_id": 0, "created_at": 1, "id": 1, "source": 1},
        sort=[("created_at", -1)],
    )
    clients_total = await db.clients.count_documents(_active_client_filter())
    clients_billing = await db.clients.count_documents(_billing_active_client_filter())
    clients_inactive = clients_total - clients_billing

    missing_invoices: List[str] = []
    async for c in db.clients.find(_billing_active_client_filter(), {"_id": 0, "id": 1, "file_no": 1, "name": 1}):
        fn = str(c.get("file_no") or "").zfill(3)
        if fn not in OFFICIAL_CLIENT_FILE_NOS:
            continue
        n_inv = await db.invoices.count_documents({"client_id": c["id"]})
        if n_inv == 0:
            missing_invoices.append(f"#{fn} {c.get('name')}")

    expected_billing = len(OFFICIAL_CLIENT_FILE_NOS - INACTIVE_CLIENT_FILE_NOS)
    return {
        "clients": clients_billing,
        "clients_total": clients_total,
        "clients_inactive": clients_inactive,
        "invoices": await db.invoices.count_documents({}),
        "sessions": await db.sessions.count_documents({}),
        "prep_history": await db.prep_history.count_documents({}),
        "therapists": len(therapists),
        "duplicate_therapist_groups": dup_groups,
        "schedule_weeks": schedule_weeks,
        "schedule_cells_total": await db.schedule_cells.count_documents({}),
        "stored_backups": await db.backups.count_documents({}),
        "last_backup_at": (last_backup or {}).get("created_at"),
        "last_backup_id": ((last_backup or {}).get("id") or "")[:8] or None,
        "last_backup_source": (last_backup or {}).get("source"),
        "missing_invoices": missing_invoices,
        "ok": clients_billing >= expected_billing and dup_groups == 0 and not missing_invoices,
    }


def _arabic_recover_summary(results: dict) -> str:
    lines: List[str] = []
    td = results.get("therapist_dedupe") or {}
    if td.get("removed"):
        lines.append(f"دمج {td['removed']} معالج مكرر (بريد)")
    idd = results.get("identity_dedupe") or {}
    if idd.get("removed"):
        lines.append(f"دمج {idd['removed']} حساب بنفس الهوية")
    prep = results.get("prep_recovery") or {}
    if any((prep.get(k) or 0) for k in prep):
        lines.append("إصلاح تواريخ التحضير للأسبوع التجريبي")
    if results.get("prep_relink"):
        lines.append("إعادة ربط التحضير (الحالي + التجريبي)")
    if results.get("schedule_order_fix"):
        lines.append(f"إصلاح صفوف مكررة في الجدول ({results['schedule_order_fix']} أسبوع)")
    seed = results.get("seed_master")
    if seed:
        created = len((seed.get("clients") or {}).get("created") or [])
        updated = len((seed.get("clients") or {}).get("updated") or [])
        if created or updated:
            lines.append(f"تحديث البيانات الرئيسية (+{created} طفل / ~{updated} محدّث)")
    if results.get("backup"):
        lines.append("حفظ نسخة احتياطية")
    if not lines:
        return "لا توجد مشاكل — البيانات تبدو سليمة ✓"
    return " · ".join(lines)


async def _auto_backup_if_stale(hours: int = AUTO_BACKUP_HOURS) -> Optional[dict]:
    """Store backup when none exists in the last N hours (startup / recovery)."""
    last = await db.backups.find_one({}, {"_id": 0, "created_at": 1}, sort=[("created_at", -1)])
    if last and last.get("created_at"):
        try:
            ts = datetime.fromisoformat(str(last["created_at"]).replace("Z", "+00:00"))
            if ts.tzinfo is None:
                ts = ts.replace(tzinfo=timezone.utc)
            age_h = (datetime.now(timezone.utc) - ts).total_seconds() / 3600
            if age_h < hours:
                return None
        except ValueError:
            pass
    return await _store_backup(source="startup")


async def _run_auto_recover(*, store_backup: bool = False) -> dict:
    """Therapist dedupe, prep recovery, relink, optional seed — used by Admin and startup."""
    from datetime import date as date_cls

    results: dict = {}
    results["therapist_dedupe"] = await _dedupe_duplicate_therapists()
    results["identity_dedupe"] = await _dedupe_therapists_by_identity()
    results["schedule_order_fix"] = await _fix_schedule_therapist_order_duplicates()
    await _migrate_schedule_week_therapist_orders()

    results["prep_recovery"] = await _recover_misdated_week_prep(TRIAL_WEEK_START, TRIAL_WEEK_END)

    current_week = _normalize_week_start(date_cls.today().isoformat())
    current_end = (
        datetime.fromisoformat(current_week) + timedelta(days=4)
    ).strftime("%Y-%m-%d")
    results["prep_relink"] = {
        "trial": await _sync_schedule_preparations_for_week(TRIAL_WEEK_START, TRIAL_WEEK_END),
        "current_week": current_week,
        "current": await _sync_schedule_preparations_for_week(current_week, current_end),
    }

    clients = await db.clients.count_documents(_billing_active_client_filter())
    expected_billing = len(OFFICIAL_CLIENT_FILE_NOS - INACTIVE_CLIENT_FILE_NOS)
    if clients < expected_billing:
        results["seed_master"] = await _seed_master_data_impl()

    if store_backup:
        results["backup"] = await _store_backup(source="recovery")

    results["health_after"] = await _get_data_health_snapshot()
    results["summary_ar"] = _arabic_recover_summary(results)
    results["ok"] = True
    return results


@api.get("/admin/data-health")
async def admin_data_health(_=Depends(admin_only)):
    """One-click portal health snapshot for Admin recovery panel."""
    snap = await _get_data_health_snapshot()
    return {"ok": True, **snap}


@api.post("/admin/auto-recover")
async def admin_auto_recover(_=Depends(admin_only)):
    """Run therapist dedupe, prep recovery, relink, and safe seed if clients < 20."""
    results = await _run_auto_recover(store_backup=False)
    return results


class LeaveBalanceIn(BaseModel):
    leave_balance: float
    annual_balance: Optional[float] = None

@api.put("/therapists/{tid}/leave-balance")
async def set_leave_balance(tid: str, payload: LeaveBalanceIn, _=Depends(leave_manager)):
    bal = float(payload.leave_balance)
    annual = float(payload.annual_balance) if payload.annual_balance is not None else bal
    await db.therapists.update_one(
        {"id": tid},
        {"$set": {"leave_balance": bal, "annual_balance": annual}},
    )
    return await db.therapists.find_one({"id": tid}, {"_id": 0, "pin_hash": 0, "password_hash": 0})

# ------------------- Master Data Seed (idempotent) -------------------
# Source of truth for therapists & clients. Idempotent: matches by name-token
# (therapists) and file_no (clients), updates missing fields, never deletes.

MASTER_THERAPISTS = [
    # (key,        first_name_token,  display_email,                    role,        leave_balance, join_date)
    ("msMaha",     "Maha",     "msalthunayan@boostgrowthsa.com",  "therapist", None, None),
    ("msFahda",    "Fahda",    "falghadeeb@boostgrowthsa.com",    "therapist", 19,   None),
    ("msRazan",    "Razan",    "ralshatery@boostgrowthsa.com",    "therapist", 17,   None),
    ("msManal",    "Manal",    "maldosery@boostgrowthsa.com",     "therapist", 7,    None),
    ("msAsma",     "Asma",     "asma@boostgrowthsa.com",          "therapist", None, None),
    ("msHajer",    "Hajar",    "halfulaij@boostgrowthsa.com",     "therapist", 11,   None),
    ("msRahaf",    "Rahaf",    "raljuhani@boostgrowthsa.com",     "therapist", 7,    None),
    ("msShatha",   "Shatha",   "shalhammami@boostgrowthsa.com",   "therapist", 21,   "2025-04-06"),
    ("msAlhanouf", "Alhanouf", "a.alromman@boostgrowthsa.com",    "therapist", 0,    "2025-07-14"),
    ("msWaad",     "Waad",     "walhamed@boostgrowthsa.com",      "therapist", 0,    "2025-08-24"),
    ("msBodoor",   "Bodour",   "baalkhlifah@boostgrowthsa.com",   "therapist", 28,   "2025-10-21"),
    ("msFatimah",  "Fatimah",  "falkhater@boostgrowthsa.com",     "therapist", 26,   "2025-11-09"),
    ("msShrooq",   "Shroug",   "shalamri@boostgrowthsa.com",      "therapist", 18,   "2026-02-08"),
    ("msAbeer",    "Abeer",    "a.alshareef@boostgrowthsa.com",   "therapist", 4,    None),
    ("msJenan",    "Jenan",    "jsalmuhaisin@boostgrowthsa.com",  "therapist", None, None),
    ("msWalaa",    "Walaa",    "wabuissa@boostgrowthsa.com",      "operations", None, None),
]

MASTER_CLIENTS = [
    # (file_no, name,                 main_key,  co_keys,                   pkg, supervisor_key, service, address)
    ("009", "Saleh Ahusainy",        "msWaad",     ["msManal", "msFahda"],     24, "msFahda", "SS/HS", "Alnakeel"),
    ("011", "Fahad Alyahya",         "msAlhanouf", ["msFahda"],                24, "msFahda", "SS",    "Alyasmin"),
    ("018", "Layan AlSaud",          "msJenan",    [],                         24, "msJenan", "ABA",   "Alaqiq"),
    ("023", "Yahya Alqahtani",       "msHajer",    ["msManal"],                24, "msFahda", "HS",    "Alaarid"),
    ("024", "Abdulaziz Alrasheed",   "msShatha",   ["msManal", "msHajer"],     40, "msFahda", "HS",    "Alnada Bldg 26"),
    ("027", "Mohammed Alaqel",       "msRahaf",    ["msFahda"],                24, "msFahda", "HS",    "AlMalqa"),
    ("030", "Husam Alturaigy",       "msManal",    ["msShatha"],               24, "msFahda", "SS/HS", "Whales daycare"),
    ("034", "Aljouhrah Alduailij",   "msFahda",    [],                         24, "msFahda", "SS",    "Alnakheel Talat"),
    ("037", "Suzan Alsultan",        "msAsma",     [],                         24, "msMaha",  "SS",    "King Fahad Villa"),
    ("038", "Salman Alrasheed",      "msManal",    ["msFahda"],                24, "msMaha",  "HS",    "Stars of Knowledge"),
    ("040", "Abdulaziz AlAbdulwahab","msFatimah",  ["msFahda", "msHajer"],     40, "msMaha",  "HS",    "Alraed"),
    ("041", "Ameerah Alshehri",      "msFahda",    ["msFatimah"],              24, "msMaha",  "HS",    "Roshen"),
    ("042", "Sultan Aldamer",        "msShrooq",   ["msRahaf", "msManal"],     40, "msMaha",  "SS/HS", "Bright Mind"),
    ("047", "Alwaleed Alotaibi",     "msHajer",    ["msAlhanouf"],             20, "msMaha",  "HS/SS", "Alqairawan"),
    ("052", "Sulaiman Alkhurashi",   "msRahaf",    ["msMaha"],                 24, "msMaha",  "HS",    "Alsulaimanyah"),
    ("054", "Omar Alkhurashi",       "msManal",    ["msMaha"],                 16, "msMaha",  "HS",    "Alsulaimanyah"),
    ("060", "Mohammed Albedayea",    "msBodoor",   ["msShatha"],               40, "msMaha",  "HS/SS", "Alyasmin"),
    ("061", "Ibrahim Alnasir",       "msRahaf",    ["msFahda"],                24, "msFahda", "HS",    "Alyasmin"),
    ("062", "Lulu Almutair",         "msRazan",    ["msFahda"],                24, "msFahda", "HS/SS", "Almuroj"),
    ("065", "Aser Alharbi",          "msMaha",     ["msMaha"],                 24, "msMaha",  "HS",    "Al Izdihar"),
    ("068", "Abdulrahman Alshawi",   "msRazan",    ["msFahda"],                24, "msFahda", "HS",    "AR Rayan"),
    ("070", "Abdulelah Almuhana",    "msAbeer",    ["msMaha"],                 32, "msMaha",  "HS",    "Al-Manziliyah"),
    ("072", "Khalid Bin Shuael",     "msShatha",   ["msFahda"],                24, "msFahda", "HS",    "AlMursalat"),
    ("076", "Sultan Abalkhail",      "msShatha",   [],                         24, "msFahda", "HS/SS", "Al-Mursalat"),
    ("079", "Fahad Suliman",         "msFahda",    ["msFahda"],                40, "msFahda", "HS",    "Al-Sahafa"),
    ("053", "Ahmad Alshalfan",       "msHajer",    ["msFahda"],                24, "msFahda", "HS/SS", "Almalqa"),
    ("080", "Faisal Alzughaibi",     "msFatimah",  [],                         24, "msFahda", "HS",    "Alyasmeen"),
]

async def _resolve_therapist_id(key_to_id: dict, key: str) -> Optional[str]:
    return key_to_id.get(key)

async def _seed_master_data_impl() -> dict:
    """Idempotently seed/update therapists and clients with the canonical master list."""
    results = {"therapists": {"updated": [], "created": [], "skipped": []},
               "clients": {"updated": [], "created": [], "skipped": []}}

    # 1) Therapists
    existing_ts = await db.therapists.find({}, {"_id": 0}).to_list(500)
    key_to_id: dict = {}
    for (key, first, email, role, leave_balance, join_date) in MASTER_THERAPISTS:
        match = next((t for t in existing_ts if first.lower() in (t.get("name") or "").lower()), None)
        update = {"key": key, "role": role}
        if leave_balance is not None:
            update["leave_balance"] = leave_balance
        if join_date is not None:
            update["join_date"] = join_date
        if match:
            await db.therapists.update_one({"id": match["id"]}, {"$set": update})
            key_to_id[key] = match["id"]
            results["therapists"]["updated"].append({"key": key, "id": match["id"], "name": match.get("name")})
        else:
            tid = str(uuid.uuid4())
            doc = {"id": tid, "name": f"Ms. {first}", "email": email,
                   "color": "#7A8A6A", "pin_hash": hash_password("0000"),
                   "must_change_password": True,
                   "created_at": now_iso(), **update}
            await db.therapists.insert_one(doc)
            key_to_id[key] = tid
            results["therapists"]["created"].append({"key": key, "id": tid, "name": doc["name"]})

    # Build supplementary key->id map by also probing first-name tokens
    # (catches therapists already in DB but not in MASTER_THERAPISTS like msJenan, msWalaa)
    for t in await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "key": 1}).to_list(500):
        if t.get("key") and t["key"] not in key_to_id:
            key_to_id[t["key"]] = t["id"]
    # Also probe by first-name match for any keys still missing (e.g. msJenan)
    for client_row in MASTER_CLIENTS:
        for k in [client_row[2]] + list(client_row[3]) + [client_row[5]]:
            if k and k not in key_to_id:
                token = k[2:] if k.startswith("ms") else k
                hit = await db.therapists.find_one({"name": {"$regex": token, "$options": "i"}}, {"_id": 0, "id": 1})
                if hit:
                    key_to_id[k] = hit["id"]
                    await db.therapists.update_one({"id": hit["id"]}, {"$set": {"key": k}})

    # 2) Clients
    for (file_no, name, main_k, co_ks, pkg, sup_k, service, address) in MASTER_CLIENTS:
        match = await db.clients.find_one({"file_no": file_no})
        if match and match.get("deleted"):
            results["clients"]["skipped"].append({"file_no": file_no, "name": name, "reason": "soft-deleted"})
            continue
        main_id = key_to_id.get(main_k)
        co_ids = [key_to_id[k] for k in co_ks if k in key_to_id]
        sup_id = key_to_id.get(sup_k)
        sup_name = None
        if sup_id:
            sup_doc = next((t for t in existing_ts if t.get("id") == sup_id), None)
            if not sup_doc:
                sup_doc = await db.clients.database.therapists.find_one({"id": sup_id}, {"_id": 0, "name": 1})
            sup_name = sup_doc.get("name") if sup_doc else sup_k
        update = {
            "name": name,
            "package_hours": pkg,
            "service_type": service,
            "address": address,
        }
        if main_id:
            update["main_therapist_id"] = main_id
        if co_ids:
            update["co_therapist_ids"] = co_ids
        if sup_name:
            update["supervisor"] = sup_name
        if file_no in INACTIVE_CLIENT_FILE_NOS:
            update["status"] = "Inactive"
        if match:
            await db.clients.update_one({"file_no": file_no}, {"$set": update})
            results["clients"]["updated"].append({"file_no": file_no, "name": name})
        else:
            cid = str(uuid.uuid4())
            if file_no not in INACTIVE_CLIENT_FILE_NOS:
                update.setdefault("status", "Active")
            doc = {"id": cid, "file_no": file_no, "color": "#7A8A6A",
                   "billing_mode": "hours", "payment_status": "pending",
                   "created_at": now_iso(), **update}
            await db.clients.insert_one(doc)
            results["clients"]["created"].append({"file_no": file_no, "name": name, "id": cid})

    return results


@api.post("/admin/seed-master-data")
async def seed_master_data(_=Depends(admin_only)):
    """Idempotently seed/update therapists and clients with the canonical master list.
    - Therapists: match by first-name token (case-insensitive) inside existing DB name.
      If found -> update (key, role, leave_balance, join_date) WITHOUT touching name/email.
      If not found -> create new therapist with display_email and default PIN 0000.
    - Clients: match by file_no. If found -> patch missing/new fields. If not found -> create.
    - Never deletes any record. Sessions/invoices remain intact.
    """
    return await _seed_master_data_impl()

# ------------------- Schedule -------------------
@api.get("/schedule/week-status")
async def schedule_week_status(week_start: str, user=Depends(get_current_user)):
    week_start = _normalize_week_start(week_start)
    doc = await db.schedule_weeks.find_one({"week_start": week_start}, {"_id": 0})
    status = (doc or {}).get("status") or "published"
    therapist_order = (doc or {}).get("therapist_order") or []
    return {
        "week_start": week_start,
        "status": status,
        "published_at": (doc or {}).get("published_at"),
        "therapist_order": therapist_order,
    }

@api.post("/schedule/publish")
async def publish_schedule_week(body: dict, admin=Depends(ops_or_admin)):
    week_start = (body.get("week_start") or "").strip()
    if not week_start:
        raise HTTPException(status_code=400, detail="week_start required")
    await db.schedule_weeks.update_one(
        {"week_start": week_start},
        {"$set": {"status": "published", "published_at": now_iso(), "published_by": admin.get("name") or "Admin"}},
        upsert=True,
    )
    wanted_ids = body.get("therapist_ids")
    if wanted_ids is not None and not isinstance(wanted_ids, list):
        wanted_ids = None
    if wanted_ids is not None:
        wanted_ids = [str(x).strip() for x in wanted_ids if str(x).strip()]
    therapists = await db.therapists.find({"email": {"$exists": True, "$ne": None}}, {"_id": 0, "id": 1, "email": 1, "name": 1}).to_list(200)
    sent = 0
    emailed: List[str] = []
    for t in therapists:
        if wanted_ids is not None and t.get("id") not in wanted_ids:
            continue
        if t.get("email"):
            r = await _send_email_stub(
                t["email"],
                f"[Boost Growth] New Schedule Published — Week of {week_start}",
                f"Dear {t.get('name', '')},\n\nThe schedule for the week of {week_start} has been published.\nPlease review your sessions for the coming week.\n\n— Boost Growth Portal",
            )
            if r.get("status") == "sent":
                sent += 1
            emailed.append(t.get("email") or "")
    await _push_center_update(
        f"Schedule published — week of {week_start}",
        "Your schedule may have changed. Please review your sessions.",
    )
    return {"ok": True, "week_start": week_start, "emails_sent": sent, "recipients": [e for e in emailed if e]}

@api.post("/schedule/set-draft")
async def set_schedule_draft(body: dict, _=Depends(ops_or_admin)):
    week_start = (body.get("week_start") or "").strip()
    if not week_start:
        raise HTTPException(status_code=400, detail="week_start required")
    await db.schedule_weeks.update_one(
        {"week_start": week_start},
        {"$set": {"status": "draft", "updated_at": now_iso()}},
        upsert=True,
    )
    return {"ok": True, "week_start": week_start, "status": "draft"}

@api.get("/schedule/closures")
async def list_schedule_closures(from_date: str = "", to_date: str = "", user=Depends(get_current_user)):
    q = {}
    if from_date and to_date:
        q["date"] = {"$gte": from_date, "$lte": to_date}
    elif from_date:
        q["date"] = {"$gte": from_date}
    return await db.schedule_closures.find(q, {"_id": 0}).sort("date", 1).to_list(500)

@api.post("/schedule/closures")
async def create_schedule_closure(body: ScheduleClosureIn, _=Depends(ops_or_admin)):
    date = (body.date or "").strip()[:10]
    label = (body.label or "").strip()
    if not date or not label:
        raise HTTPException(status_code=400, detail="date and label required")
    tids = [t for t in (body.therapist_ids or []) if t]
    doc = {
        "id": str(uuid.uuid4()),
        "date": date,
        "label": label,
        "therapist_ids": tids,
        "created_at": now_iso(),
    }
    await db.schedule_closures.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.delete("/schedule/closures/{cid}")
async def delete_schedule_closure(cid: str, _=Depends(ops_or_admin)):
    await db.schedule_closures.delete_one({"id": cid})
    return {"ok": True}

def _center_update_flags(body: CenterUpdateIn) -> dict:
    is_important = bool(body.is_important)
    requires_ack = bool(body.requires_ack)
    send_to_specialists = bool(body.send_to_specialists)
    if requires_ack and not is_important:
        is_important = True
    if send_to_specialists and not is_important:
        is_important = True
    return {
        "is_important": is_important,
        "requires_ack": requires_ack,
        "send_to_specialists": send_to_specialists,
    }


async def _therapist_recipient_rows() -> List[dict]:
    return await db.therapists.find(
        {"email": {"$exists": True, "$nin": [None, ""]}},
        {"_id": 0, "id": 1, "email": 1, "name": 1},
    ).to_list(300)


async def _broadcast_important_center_update(doc: dict):
    """Email + in-app alert for important / specialist-targeted center updates."""
    title = doc.get("title") or "Center update"
    body = (doc.get("body") or "").strip()
    date = doc.get("date") or now_iso()[:10]
    portal = _portal_base_url()
    email_body = f"Dear team member,\n\nAn important update has been posted:\n\n{title}\n"
    if body:
        email_body += f"\n{body}\n"
    email_body += f"\nDate: {date}\n"
    if doc.get("requires_ack"):
        email_body += "\nPlease open the portal Home page and confirm you have read this update.\n"
    if portal:
        email_body += f"\nOpen portal: {portal}/\n"
    email_body += "\n— Boost Growth Portal"
    msg = body[:240] if body else title
    for t in await _therapist_recipient_rows():
        tid = t.get("id")
        if tid:
            await _notify(
                tid,
                "center_update_important",
                title,
                msg,
                link="/",
                update_id=doc.get("id"),
                requires_ack=bool(doc.get("requires_ack")),
            )
        email = (t.get("email") or "").strip()
        if email:
            await _send_urgent_email(email, title, email_body)


def _enrich_center_updates(items: List[dict], user: dict, therapist_names: Optional[dict] = None) -> List[dict]:
    tid = None
    if user.get("role") == "therapist":
        tid = user.get("id")
    show_ack_admin = user.get("role") == "admin" or _is_walaa_ops(user)
    for item in items:
        acks = item.get("acknowledged_by") or []
        if tid:
            item["acked_by_me"] = any(a.get("therapist_id") == tid for a in acks)
        if show_ack_admin and therapist_names is not None:
            ack_ids = {a.get("therapist_id") for a in acks if a.get("therapist_id")}
            item["ack_read"] = [
                {
                    "therapist_id": a.get("therapist_id"),
                    "name": a.get("name") or therapist_names.get(a.get("therapist_id"), "Unknown"),
                    "at": a.get("at"),
                }
                for a in acks
            ]
            item["ack_pending"] = [
                {"therapist_id": t_id, "name": name}
                for t_id, name in therapist_names.items()
                if t_id not in ack_ids
            ]
    return items


@api.get("/center-updates")
async def list_center_updates(user=Depends(get_current_user)):
    items = await db.center_updates.find({}, {"_id": 0}).sort("date", -1).to_list(50)
    therapist_names = None
    if user.get("role") == "admin" or _is_walaa_ops(user):
        therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
        therapist_names = {t["id"]: therapist_schedule_display_name(t) or "Unknown" for t in therapists if t.get("id")}
    return _enrich_center_updates(items, user, therapist_names)


@api.post("/center-updates")
async def create_center_update(body: CenterUpdateIn, _=Depends(admin_only)):
    title = (body.title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="title required")
    flags = _center_update_flags(body)
    doc = {
        "id": str(uuid.uuid4()),
        "title": title,
        "body": (body.body or "").strip(),
        "date": (body.date or now_iso()[:10]),
        "created_at": now_iso(),
        "acknowledged_by": [],
        **flags,
    }
    await db.center_updates.insert_one(doc)
    doc.pop("_id", None)
    if flags["is_important"] or flags["send_to_specialists"]:
        await _broadcast_important_center_update(doc)
    return doc

@api.put("/center-updates/{uid}")
async def update_center_update(uid: str, body: CenterUpdateIn, _=Depends(admin_only)):
    title = (body.title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="title required")
    existing = await db.center_updates.find_one({"id": uid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Update not found")
    flags = _center_update_flags(body)
    update = {
        "title": title,
        "body": (body.body or "").strip(),
        "date": (body.date or now_iso()[:10]),
        "updated_at": now_iso(),
        **flags,
    }
    was_important = bool(existing.get("is_important") or existing.get("send_to_specialists"))
    await db.center_updates.update_one({"id": uid}, {"$set": update})
    doc = await db.center_updates.find_one({"id": uid}, {"_id": 0})
    if (flags["is_important"] or flags["send_to_specialists"]) and not was_important:
        await _broadcast_important_center_update(doc)
    return doc

@api.post("/center-updates/{uid}/acknowledge")
async def acknowledge_center_update(uid: str, user=Depends(get_current_user)):
    tid = await _resolve_user_therapist_id(user) or user.get("id")
    if not tid:
        raise HTTPException(status_code=403, detail="Therapist profile required")
    doc = await db.center_updates.find_one({"id": uid}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Update not found")
    if not doc.get("requires_ack"):
        return {"ok": True, "skipped": True}
    acks = doc.get("acknowledged_by") or []
    if any(a.get("therapist_id") == tid for a in acks):
        return {"ok": True, "already": True}
    entry = {"therapist_id": tid, "at": now_iso(), "name": user.get("name") or ""}
    await db.center_updates.update_one({"id": uid}, {"$push": {"acknowledged_by": entry}})
    return {"ok": True}

@api.delete("/center-updates/{uid}")
async def delete_center_update(uid: str, _=Depends(admin_only)):
    res = await db.center_updates.delete_one({"id": uid})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Update not found")
    return {"ok": True}

async def _push_center_update(title: str, body: str = "", *, is_important: bool = False, requires_ack: bool = False):
    """Broadcast a platform update visible on therapist home."""
    doc = {
        "id": str(uuid.uuid4()),
        "title": (title or "").strip(),
        "body": (body or "").strip(),
        "date": now_iso()[:10],
        "created_at": now_iso(),
        "is_important": is_important,
        "requires_ack": requires_ack,
        "send_to_specialists": is_important,
        "acknowledged_by": [],
    }
    if doc["title"]:
        await db.center_updates.insert_one(doc)
        if is_important:
            await _broadcast_important_center_update(doc)

@api.get("/calendar/personal")
async def list_personal_events(from_date: str = "", to_date: str = "", user=Depends(get_current_user)):
    tid = await _resolve_user_therapist_id(user) or user.get("id")
    if not tid:
        raise HTTPException(status_code=403, detail="Therapist profile required")
    q = {"therapist_id": tid}
    if from_date and to_date:
        q["date"] = {"$gte": from_date[:10], "$lte": to_date[:10]}
    elif from_date:
        q["date"] = {"$gte": from_date[:10]}
    return await db.therapist_personal_events.find(q, {"_id": 0}).sort("date", 1).to_list(200)

@api.post("/calendar/personal")
async def create_personal_event(body: PersonalEventIn, user=Depends(get_current_user)):
    tid = await _resolve_user_therapist_id(user) or user.get("id")
    if not tid:
        raise HTTPException(status_code=403, detail="Therapist profile required")
    title = (body.title or "").strip()
    date = (body.date or "").strip()[:10]
    if not title or not date:
        raise HTTPException(status_code=400, detail="title and date required")
    doc = {
        "id": str(uuid.uuid4()),
        "therapist_id": tid,
        "date": date,
        "title": title,
        "notes": (body.notes or "").strip(),
        "time_label": (body.time_label or "").strip() or None,
        "created_at": now_iso(),
    }
    await db.therapist_personal_events.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.delete("/calendar/personal/{eid}")
async def delete_personal_event(eid: str, user=Depends(get_current_user)):
    tid = await _resolve_user_therapist_id(user) or user.get("id")
    ev = await db.therapist_personal_events.find_one({"id": eid}, {"_id": 0})
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    if ev.get("therapist_id") != tid and not _is_portal_admin(user):
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.therapist_personal_events.delete_one({"id": eid})
    return {"ok": True}

def _schedule_week_start_variants(week_start: str) -> List[str]:
    """Sunday-normalized week_start values to try (handles wrong-year imports)."""
    from datetime import date
    primary = _normalize_week_start(week_start)
    out = [primary]
    try:
        d = date.fromisoformat(primary[:10])
        alt = _normalize_week_start(d.replace(year=d.year - 1).isoformat())
        if alt not in out:
            out.append(alt)
    except ValueError:
        pass
    return out


def _week_start_variants_for_session_date(session_date: str) -> List[str]:
    """Week-start keys to search schedule_cells for a session/prep date."""
    sd = (session_date or "")[:10]
    if not sd:
        return []
    return _schedule_week_start_variants(sd)


def _session_date_iso(value) -> Optional[str]:
    """Normalize session_date from str, datetime, or date to yyyy-mm-dd."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "isoformat") and not isinstance(value, str):
        try:
            return value.isoformat()[:10]
        except Exception:
            pass
    text = str(value).strip()
    if not text:
        return None
    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        return text[:10]
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", text)
    return m.group(1) if m else None


def _session_date_query(session_date: str) -> dict:
    """Mongo filter for a calendar day (handles yyyy-mm-dd and ISO datetimes)."""
    sd = _session_date_iso(session_date)
    if not sd:
        return {}
    return {"session_date": {"$regex": f"^{re.escape(sd)}"}}


def _session_date_range_query(start: str, end: str) -> dict:
    """Mongo filter for session_date within an inclusive yyyy-mm-dd range."""
    from datetime import date as date_cls
    start = _session_date_iso(start) or ""
    end = _session_date_iso(end) or ""
    if not start or not end:
        return {}
    clauses: list = [
        {"session_date": {"$gte": start, "$lte": end}},
        {"session_date": {"$gte": f"{start}T", "$lte": f"{end}T23:59:59.999Z"}},
        {"session_date": {"$gte": f"{start}T", "$lte": f"{end}T23:59:59.999+00:00"}},
    ]
    try:
        start_dt = datetime.fromisoformat(start).replace(tzinfo=timezone.utc)
        end_dt = datetime.fromisoformat(end).replace(
            hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc
        )
        clauses.append({"session_date": {"$gte": start_dt, "$lte": end_dt}})
    except ValueError:
        pass
    try:
        d0 = date_cls.fromisoformat(start)
        d1 = date_cls.fromisoformat(end)
        day = d0
        while day <= d1:
            clauses.append({"session_date": {"$regex": f"^{re.escape(day.isoformat())}"}})
            day += timedelta(days=1)
    except ValueError:
        pass
    return {"$or": clauses}


def _shift_calendar_year(iso_date: str, delta_years: int) -> Optional[str]:
    """Return iso_date shifted by delta_years (e.g. 2025-06-28 → 2026-06-28)."""
    from datetime import date as date_cls
    sd = _session_date_iso(iso_date)
    if not sd:
        return None
    try:
        d = date_cls.fromisoformat(sd)
        return d.replace(year=d.year + delta_years).isoformat()
    except ValueError:
        return None


def _prep_week_marker_scope_query(start: str, end: str) -> dict:
    """Match schedule_preparation rows by session_date OR week_start for a Sun–Thu week."""
    start = _session_date_iso(start) or ""
    end = _session_date_iso(end) or ""
    week_variants = _schedule_week_start_variants(start) if start else []
    clauses: list = [_session_date_range_query(start, end)]
    if week_variants:
        clauses.append({"week_start": {"$in": week_variants}})
    return {"$or": clauses}


async def _recover_misdated_week_prep(start: str, end: str) -> dict:
    """Repair prep/session rows stored under wrong year or session_date before relink."""
    start = _session_date_iso(start) or ""
    end = _session_date_iso(end) or ""
    stats = {
        "remapped_sessions": 0,
        "remapped_history": 0,
        "fixed_prep_markers": 0,
        "history_from_prepared_at": 0,
    }
    if not start or not end:
        return stats

    primary_hist = await db.prep_history.count_documents(_session_date_range_query(start, end))
    primary_sess = await db.sessions.count_documents({
        **_session_date_range_query(start, end),
        "status": "Completed",
    })
    if not primary_hist and not primary_sess:
        alt_start = _shift_calendar_year(start, -1)
        alt_end = _shift_calendar_year(end, -1)
        if alt_start and alt_end:
            alt_sessions = await db.sessions.find(
                {
                    **_session_date_range_query(alt_start, alt_end),
                    "status": "Completed",
                },
                {"_id": 0},
            ).to_list(5000)
            for sess in alt_sessions:
                old_sd = _session_date_iso(sess.get("session_date"))
                new_sd = _shift_calendar_year(old_sd, 1) if old_sd else None
                if not new_sd or not (start <= new_sd <= end):
                    continue
                await db.sessions.update_one({"id": sess["id"]}, {"$set": {"session_date": new_sd}})
                stats["remapped_sessions"] += 1

            alt_hist = await db.prep_history.find(
                _session_date_range_query(alt_start, alt_end),
                {"_id": 0},
            ).to_list(5000)
            for row in alt_hist:
                old_sd = _session_date_iso(row.get("session_date"))
                new_sd = _shift_calendar_year(old_sd, 1) if old_sd else None
                if not new_sd or not (start <= new_sd <= end):
                    continue
                await db.prep_history.update_one({"id": row["id"]}, {"$set": {"session_date": new_sd}})
                stats["remapped_history"] += 1

    hist_by_prepared = await db.prep_history.find(
        {
            "prepared_at": {
                "$gte": f"{start}T00:00:00",
                "$lte": f"{end}T23:59:59.999Z",
            },
        },
        {"_id": 0},
    ).to_list(5000)
    for row in hist_by_prepared:
        sd = _session_date_iso(row.get("session_date"))
        if sd and start <= sd <= end:
            continue
        pa = _session_date_iso(row.get("prepared_at"))
        if pa and start <= pa <= end:
            await db.prep_history.update_one({"id": row["id"]}, {"$set": {"session_date": pa}})
            stats["history_from_prepared_at"] += 1

    week_variants = _schedule_week_start_variants(start)
    prep_rows = await db.schedule_preparations.find(
        {"week_start": {"$in": week_variants}},
        {"_id": 0},
    ).to_list(5000)
    for row in prep_rows:
        sd = _session_date_iso(row.get("session_date"))
        if sd and start <= sd <= end:
            continue
        corrected = _schedule_cell_date_iso(row)
        if corrected and start <= corrected <= end:
            await db.schedule_preparations.update_one(
                {"id": row["id"]},
                {"$set": {"session_date": corrected}},
            )
            stats["fixed_prep_markers"] += 1
            tid = row.get("therapist_id")
            cid = row.get("client_id")
            if tid and cid:
                key = _prep_history_key(tid, cid, corrected, row.get("time_slot"))
                hit = await db.prep_history.find_one(key, {"_id": 0, "id": 1})
                if not hit:
                    doc = {
                        **key,
                        "id": str(uuid.uuid4()),
                        "client_name": row.get("client_name") or "",
                        "prepared_by": row.get("prepared_by") or "",
                        "prepared_at": row.get("prepared_at") or now_iso(),
                        "notes": "",
                        "source": "schedule",
                        "schedule_cell_id": row.get("schedule_cell_id"),
                    }
                    await db.prep_history.insert_one(doc)

    return stats


async def _prep_week_diagnostics(start: str, end: str) -> dict:
    """Counts for relink troubleshooting (primary range + year-shift + week_start scope)."""
    start = _session_date_iso(start) or ""
    end = _session_date_iso(end) or ""
    alt_start = _shift_calendar_year(start, -1) or ""
    alt_end = _shift_calendar_year(end, -1) or ""
    week_variants = _schedule_week_start_variants(start) if start else []
    return {
        "prep_history": await db.prep_history.count_documents(_session_date_range_query(start, end)),
        "completed_sessions": await db.sessions.count_documents({
            **_session_date_range_query(start, end),
            "status": "Completed",
        }),
        "schedule_preparations": await db.schedule_preparations.count_documents(
            _prep_week_marker_scope_query(start, end)
        ),
        "prep_suppressions": await db.schedule_prep_suppressions.count_documents(
            _session_date_range_query(start, end)
        ),
        "schedule_cells": await db.schedule_cells.count_documents(
            {"week_start": {"$in": week_variants}} if week_variants else {"week_start": start}
        ),
        "prep_history_year_shift": await db.prep_history.count_documents(
            _session_date_range_query(alt_start, alt_end)
        ) if alt_start and alt_end else 0,
        "sessions_year_shift": await db.sessions.count_documents({
            **_session_date_range_query(alt_start, alt_end),
            "status": "Completed",
        }) if alt_start and alt_end else 0,
        "prep_by_week_start": await db.schedule_preparations.count_documents(
            {"week_start": {"$in": week_variants}}
        ) if week_variants else 0,
    }


@api.get("/schedule")
async def list_schedule(week_start: Optional[str] = None, user=Depends(get_current_user)):
    q: dict = {}
    if week_start:
        variants = _schedule_week_start_variants(week_start)
        primary = variants[0]
        meta = await db.schedule_weeks.find_one({"week_start": primary}, {"_id": 0})
        if meta and meta.get("status") == "draft" and not _has_full_client_access(user):
            return []
        for ws in variants:
            cells = await db.schedule_cells.find({"week_start": ws}, {"_id": 0}).to_list(5000)
            if cells:
                return await _enrich_schedule_cells_with_client_colors(cells)
        return []
    cells = await db.schedule_cells.find(q, {"_id": 0}).to_list(5000)
    return await _enrich_schedule_cells_with_client_colors(cells)


async def _enrich_schedule_cells_with_client_colors(cells: list) -> list:
    """Schedule grid uses unified session backgrounds; skip per-client rainbow fills."""
    return cells


_CLIENT_SESSION_CODES = frozenset({"SS", "HS", "OS"})


def _strip_session_cell_color(doc: dict) -> dict:
    """Session cells use client-side shift tints — do not persist legacy per-child fills."""
    code = (doc.get("service_code") or "").strip().upper()
    if code in _CLIENT_SESSION_CODES or (doc.get("child_name") or "").strip():
        doc["color"] = None
    return doc


async def _backfill_schedule_cell_colors_for_week(week_start: str) -> int:
    """Persist client colors onto schedule cells for a week (non-destructive)."""
    cells = await db.schedule_cells.find({"week_start": week_start}, {"_id": 0}).to_list(5000)
    if not cells:
        return 0
    enriched = await _enrich_schedule_cells_with_client_colors(cells)
    updated = 0
    for orig, cell in zip(cells, enriched):
        new_color = cell.get("color")
        if new_color and orig.get("color") != new_color:
            await db.schedule_cells.update_one(
                {"id": orig["id"]},
                {"$set": {"color": new_color}},
            )
            updated += 1
    return updated


def _prep_history_key(therapist_id: str, client_id: str, session_date: str, time_slot: Optional[str] = None) -> dict:
    return {
        "therapist_id": therapist_id,
        "client_id": client_id,
        "session_date": (session_date or "")[:10],
        "time_slot": (time_slot or "").strip(),
    }


async def _upsert_prep_history(
    *,
    therapist_id: str,
    client_id: str,
    session_date: str,
    prepared_by: str,
    time_slot: Optional[str] = None,
    client_name: Optional[str] = None,
    notes: Optional[str] = None,
    internal_note: Optional[str] = None,
    invoice_id: Optional[str] = None,
    session_id: Optional[str] = None,
    schedule_cell_id: Optional[str] = None,
    source: str = "schedule",
) -> dict:
    """Persistent preparation log — visible in Preparation history even without an invoice."""
    session_date = (session_date or "")[:10]
    q = _prep_history_key(therapist_id, client_id, session_date, time_slot)
    if not client_name:
        client = await db.clients.find_one({"id": client_id}, {"_id": 0, "name": 1})
        client_name = (client or {}).get("name") or ""
    existing = await db.prep_history.find_one(q, {"_id": 0})
    therapist = await db.therapists.find_one({"id": therapist_id}, {"_id": 0, "name": 1, "key": 1})
    therapist_name = therapist_schedule_display_name(therapist) if therapist else ""
    doc = {
        **q,
        "client_name": client_name,
        "therapist_name": therapist_name,
        "prepared_by": prepared_by,
        "prepared_at": now_iso(),
        "notes": notes or "",
        "source": source,
    }
    if internal_note is not None:
        doc["internal_note"] = (internal_note or "").strip()
    elif existing and existing.get("internal_note"):
        doc["internal_note"] = existing["internal_note"]
    if schedule_cell_id:
        doc["schedule_cell_id"] = schedule_cell_id
    if session_id:
        doc["session_id"] = session_id
    if invoice_id:
        doc["invoice_id"] = invoice_id
    elif existing and existing.get("invoice_id"):
        doc["invoice_id"] = existing["invoice_id"]
    if existing:
        await db.prep_history.update_one({"id": existing["id"]}, {"$set": doc})
        doc["id"] = existing["id"]
    else:
        doc["id"] = str(uuid.uuid4())
        await db.prep_history.insert_one(doc)
    doc.pop("_id", None)
    return doc


async def _log_therapist_cancel_prep_history(cell: dict, user_id: str) -> None:
    """Auto-log therapist cancellation to prep history with timestamp."""
    if not cell or cell.get("state") != "cancel_therapist":
        return
    child = (cell.get("child_name") or "").strip()
    tid = cell.get("therapist_id")
    slot_date = _schedule_cell_date_iso(cell)
    if not child or not tid or not slot_date:
        return
    client = await _find_client_by_schedule_child_name(child)
    if not client:
        return
    note = f"Therapist cancellation — logged {now_iso()[:16].replace('T', ' ')} UTC"
    await _upsert_prep_history(
        therapist_id=tid,
        client_id=client["id"],
        session_date=slot_date,
        prepared_by=user_id or "",
        time_slot=cell.get("time_slot") or "",
        client_name=client.get("name") or child,
        notes=note,
        schedule_cell_id=cell.get("id"),
        source="therapist_cancel",
    )


async def _sync_schedule_preparations_to_prep_history(client_id: Optional[str] = None) -> None:
    """One-time backfill: legacy schedule_preparations → prep_history."""
    q: dict = {}
    if client_id:
        q["client_id"] = client_id
    rows = await db.schedule_preparations.find(q, {"_id": 0}).to_list(2000)
    for rec in rows:
        tid = rec.get("therapist_id")
        cid = rec.get("client_id")
        if not tid or not cid:
            continue
        key = _prep_history_key(tid, cid, rec.get("session_date", ""), rec.get("time_slot"))
        hit = await db.prep_history.find_one(key, {"_id": 0, "id": 1})
        if hit:
            continue
        client = await db.clients.find_one({"id": cid}, {"_id": 0, "name": 1})
        doc = {
            **key,
            "id": str(uuid.uuid4()),
            "client_name": (client or {}).get("name") or "",
            "therapist_name": therapist_schedule_display_name(
                await db.therapists.find_one({"id": tid}, {"_id": 0, "name": 1, "key": 1})
            ),
            "prepared_by": rec.get("prepared_by") or "",
            "prepared_at": rec.get("prepared_at") or now_iso(),
            "notes": "",
            "source": "schedule",
            "schedule_cell_id": rec.get("schedule_cell_id"),
        }
        await db.prep_history.insert_one(doc)


async def _upsert_schedule_preparation(
    *,
    therapist_id: str,
    client_id: str,
    session_date: str,
    prepared_by: str,
    time_slot: Optional[str] = None,
    schedule_cell_id: Optional[str] = None,
    week_start: Optional[str] = None,
    day: Optional[int] = None,
    notes: Optional[str] = None,
    internal_note: Optional[str] = None,
    invoice_id: Optional[str] = None,
    client_name: Optional[str] = None,
    marker_type: str = "prep",
    session_id: Optional[str] = None,
) -> dict:
    """Mark a schedule slot as preparation-complete for therapist + client + date."""
    slot = (time_slot or "").strip()
    session_date = (session_date or "")[:10]
    q = _prep_history_key(therapist_id, client_id, session_date, slot)
    existing = await db.schedule_preparations.find_one(q, {"_id": 0})
    doc = {
        **q,
        "schedule_cell_id": schedule_cell_id,
        "week_start": week_start,
        "day": day,
        "prepared_by": prepared_by,
        "prepared_at": now_iso(),
        "client_name": client_name,
        "marker_type": marker_type,
        "source": "no_show" if marker_type == "no_show" else "prep",
    }
    if session_id:
        doc["session_id"] = session_id
    if internal_note is not None:
        doc["internal_note"] = (internal_note or "").strip()
    elif existing and existing.get("internal_note"):
        doc["internal_note"] = existing["internal_note"]
    if existing:
        await db.schedule_preparations.update_one({"id": existing["id"]}, {"$set": doc})
        doc["id"] = existing["id"]
    else:
        doc["id"] = str(uuid.uuid4())
        await db.schedule_preparations.insert_one(doc)
    doc.pop("_id", None)
    await _clear_prep_suppressions(
        therapist_id, client_id, session_date, schedule_cell_id,
    )
    await _upsert_prep_history(
        therapist_id=therapist_id,
        client_id=client_id,
        session_date=session_date,
        prepared_by=prepared_by,
        time_slot=slot,
        client_name=client_name,
        notes=notes,
        internal_note=internal_note,
        invoice_id=invoice_id,
        schedule_cell_id=schedule_cell_id,
        source="schedule",
    )
    return doc


def _parse_child_name_from_schedule_note(cell: dict) -> str:
    """Parse child label from note (matches grid display when note is set)."""
    note = (cell.get("note") or "").strip()
    if not note:
        return ""
    if "|" in note:
        part = note.split("|", 1)[1].strip()
        if part:
            return re.sub(r"\s*\([^)]*\)\s*$", "", part).strip()
    upper = note.upper()
    for prefix in ("HS", "SS", "OS"):
        if upper.startswith(prefix):
            rest = re.sub(rf"^{prefix}[\s\-|:]+", "", note, flags=re.I).strip()
            if rest:
                return re.sub(r"\s*\([^)]*\)\s*$", "", rest).strip()
    return re.sub(r"\s*\([^)]*\)\s*$", "", note).strip()


def _split_schedule_child_names(label: str) -> List[str]:
    """Split dual-child labels like 'Lulu / Abdulrahman'."""
    raw = (label or "").strip()
    if not raw:
        return []
    if "/" not in raw:
        return [raw]
    return [p.strip() for p in re.split(r"\s*/\s*", raw) if p.strip()]


def _schedule_cell_child_label(cell: dict) -> str:
    """Effective child label — note before child_name so prep matches what therapists see."""
    from_note = _parse_child_name_from_schedule_note(cell)
    if from_note:
        return from_note
    return (cell.get("child_name") or "").strip()


def _slot_label_to_time24(slot: Optional[str]) -> Optional[str]:
    if not slot:
        return None
    m = re.match(r"(\d{1,2}):(\d{2})\s*(AM|PM)?", str(slot).strip(), re.I)
    if not m:
        return None
    h = int(m.group(1))
    mi = m.group(2)
    ap = (m.group(3) or "").upper()
    if ap == "PM" and h < 12:
        h += 12
    if ap == "AM" and h == 12:
        h = 0
    return f"{h:02d}:{mi}"


def _cell_session_start_time(cell: dict) -> Optional[str]:
    """Derive HH:MM start from schedule cell slot / custom_time (mirrors frontend)."""
    anchor = (cell.get("time_slot") or "").strip()
    custom = (cell.get("custom_time") or "").strip()
    if custom:
        m = re.match(r"(\d{1,2}(?::\d{2})?)\s*[-–]\s*(\d{1,2}(?::\d{2})?)", custom)
        if m:
            start_raw = m.group(1)
            if ":" not in start_raw:
                start_raw = f"{start_raw}:00"
            ref = "PM" if "PM" in anchor.upper() else "AM"
            parsed = _slot_label_to_time24(f"{start_raw} {ref}")
            if parsed:
                return parsed
    return _slot_label_to_time24(anchor)


def _session_time_matches_cell(start_time: Optional[str], cell: Optional[dict]) -> bool:
    if not start_time or not cell:
        return True
    cell_start = _cell_session_start_time(cell)
    if not cell_start:
        return True
    st = _slot_label_to_time24(start_time) or (start_time or "").strip()[:5]
    return st == cell_start


def _session_includes_cell_therapist(sess: dict, cell_tid: str, alias_map: Dict[str, List[str]]) -> bool:
    for st in sess.get("therapist_ids") or []:
        if _therapist_ids_overlap(st, cell_tid, alias_map):
            return True
    return False


async def _validate_prep_client_matches_cell(
    *,
    schedule_cell_id: Optional[str],
    client_id: str,
    therapist_id: Optional[str] = None,
    cell_child_name: Optional[str] = None,
) -> Optional[dict]:
    """Reject prep when the requested client does not match the schedule cell."""
    if not schedule_cell_id:
        return None
    cell = await db.schedule_cells.find_one({"id": schedule_cell_id}, {"_id": 0})
    if not cell:
        return None
    if therapist_id and cell.get("therapist_id") and cell.get("therapist_id") != therapist_id:
        raise HTTPException(status_code=400, detail="Schedule cell belongs to a different therapist.")
    if await _cell_matches_session_client(cell, client_id):
        return cell
    label = _schedule_cell_child_label(cell)
    client = await db.clients.find_one(_active_client_filter({"id": client_id}), {"_id": 0, "name": 1})
    cname = (client or {}).get("name") or client_id
    expected = (cell_child_name or "").strip()
    detail = (
        f'Preparation client mismatch: cell shows "{label}"'
        f' but request is for "{cname}".'
    )
    if expected and expected.lower() != label.lower():
        detail += f' Expected cell child: "{expected}".'
    raise HTTPException(status_code=400, detail=detail)


async def _cell_matches_session_client(cell: dict, client_id: str) -> bool:
    """True when a schedule cell refers to the same client as a logged session."""
    label = _schedule_cell_child_label(cell)
    if label:
        matched = await _find_client_by_schedule_child_name(label)
        if matched:
            return matched.get("id") == client_id
    client = await db.clients.find_one(_active_client_filter({"id": client_id}), {"_id": 0, "name": 1})
    if not client:
        return False
    cname = (client.get("name") or "").strip()
    cfirst = cname.split()[0].lower() if cname else ""
    if not cfirst or len(cfirst) < 3:
        return False
    hay = " ".join(
        x for x in [(cell.get("child_name") or ""), (cell.get("note") or ""), label or ""] if x
    ).lower()
    if cfirst not in hay:
        return False
    by_first = await _find_client_by_schedule_child_name(cfirst)
    if by_first:
        return by_first.get("id") == client_id
    return False


async def _schedule_cells_for_prep_match(
    therapist_ids: List[str],
    session_date: str,
    *,
    stale_cell_id: Optional[str] = None,
) -> List[dict]:
    """Load schedule cells for prep matching — scoped to the session week, not all history."""
    sd = (session_date or "")[:10]
    week_variants = _week_start_variants_for_session_date(sd)
    base_q: dict = {
        "state": {"$nin": ["cancel_therapist"]},
        "service_code": {"$nin": ["LEAVE", "BREAK", "AVC", ""]},
    }
    if week_variants:
        base_q["week_start"] = {"$in": week_variants}
    if therapist_ids:
        base_q["therapist_id"] = {"$in": therapist_ids}
    cells = await db.schedule_cells.find(base_q, {"_id": 0}).to_list(2000)
    if stale_cell_id:
        stale = await db.schedule_cells.find_one({"id": stale_cell_id}, {"_id": 0})
        if stale and stale.get("id") not in {c.get("id") for c in cells}:
            cells.append(stale)
    return cells


async def _upsert_session_prep_markers(
    *,
    therapist_id: str,
    client_id: str,
    session_date: str,
    prepared_by: str,
    client_name: Optional[str],
    cell: Optional[dict] = None,
    marker_type: str = "prep",
    session_id: Optional[str] = None,
) -> None:
    """Write schedule_preparation row for one exact grid cell (session-backed badges only)."""
    if not cell:
        return
    week_start = cell.get("week_start") or _normalize_week_start(session_date)
    day = cell.get("day")
    cell_id = cell.get("id")
    slot = (cell.get("time_slot") or "").strip()
    await _upsert_schedule_preparation(
        therapist_id=therapist_id,
        client_id=client_id,
        session_date=session_date,
        prepared_by=prepared_by,
        time_slot=slot,
        schedule_cell_id=cell_id,
        week_start=week_start,
        day=day,
        client_name=client_name,
        marker_type=marker_type,
        session_id=session_id,
    )


async def _auto_mark_schedule_preparation_for_session(sess: dict, user_id: str) -> None:
    """When a completed or no-attendance session is logged, mark the exact matching schedule cell."""
    status = (sess.get("status") or "").strip()
    if status not in ("Completed", "No Show", "Cancelled"):
        return
    client_id = sess.get("client_id")
    session_date = (sess.get("session_date") or "")[:10]
    if not client_id or not session_date:
        return
    client = await db.clients.find_one(_active_client_filter({"id": client_id}), {"_id": 0, "name": 1})
    client_name = (client or {}).get("name")
    therapist_ids = [t for t in (sess.get("therapist_ids") or []) if t]
    cells = await _schedule_cells_for_prep_match(therapist_ids, session_date)
    marked = set()
    prepared_by = user_id or sess.get("created_by") or ""
    marker_type = "no_show" if status in _NO_ATTENDANCE_SESSION_STATUSES else "prep"
    sess_start = sess.get("start_time")
    alias_map = await _build_therapist_id_alias_map()
    for cell in cells:
        slot_date = _schedule_cell_date_iso(cell)
        if slot_date != session_date:
            continue
        if not await _cell_matches_session_client(cell, client_id):
            continue
        tid = cell.get("therapist_id")
        if not tid or not _session_includes_cell_therapist(sess, tid, alias_map):
            continue
        if not _session_time_matches_cell(sess_start, cell):
            continue
        key = (tid, client_id, session_date, cell.get("time_slot") or "", cell.get("id"))
        if key in marked:
            continue
        marked.add(key)
        await _upsert_session_prep_markers(
            therapist_id=tid,
            client_id=client_id,
            session_date=session_date,
            prepared_by=prepared_by,
            client_name=client_name,
            cell=cell,
            marker_type=marker_type,
            session_id=sess.get("id"),
        )


async def _ensure_session_schedule_prep_markers(
    sess: dict,
    user_id: str,
    *,
    notes: Optional[str] = None,
) -> None:
    """Guarantee schedule_preparations + prep_history stay in sync after logging a session."""
    status = (sess.get("status") or "").strip()
    if status not in ("Completed", "No Show", "Cancelled"):
        return
    client_id = sess.get("client_id")
    session_date = (sess.get("session_date") or "")[:10]
    if not client_id or not session_date:
        return
    client = await db.clients.find_one(_active_client_filter({"id": client_id}), {"_id": 0, "name": 1})
    client_name = (client or {}).get("name")
    prepared_by = user_id or sess.get("created_by") or ""
    source = "no_show" if status in _NO_ATTENDANCE_SESSION_STATUSES else "session"
    marker_type = "no_show" if status in _NO_ATTENDANCE_SESSION_STATUSES else "prep"
    try:
        await _auto_mark_schedule_preparation_for_session(sess, user_id)
    except Exception:
        logger.exception("Auto-mark schedule preparation failed for session %s", sess.get("id"))
    for tid in [t for t in (sess.get("therapist_ids") or []) if t]:
        cell = await _resolve_schedule_cell_for_prep(
            tid, client_id, session_date, client_name=client_name,
        )
        try:
            if cell:
                await _upsert_session_prep_markers(
                    therapist_id=cell.get("therapist_id") or tid,
                    client_id=client_id,
                    session_date=session_date,
                    prepared_by=prepared_by,
                    client_name=client_name,
                    cell=cell,
                    marker_type=marker_type,
                    session_id=sess.get("id"),
                )
        except Exception:
            logger.exception("Schedule prep marker upsert failed for %s/%s", tid, client_id)
        time_slot = sess.get("start_time") or (cell or {}).get("time_slot") or ""
        try:
            await _upsert_prep_history(
                therapist_id=tid,
                client_id=client_id,
                session_date=session_date,
                prepared_by=prepared_by,
                time_slot=time_slot,
                client_name=client_name,
                notes=notes if notes is not None else sess.get("note"),
                invoice_id=sess.get("invoice_id"),
                session_id=sess.get("id"),
                schedule_cell_id=(cell or {}).get("id"),
                source=source,
            )
        except Exception:
            logger.exception("Prep history upsert failed for %s/%s", tid, client_id)


async def _schedule_cells_for_client_day(client_id: str, session_date: str) -> List[dict]:
    """All schedule cells for a client on a calendar day (any therapist row)."""
    sd = (session_date or "")[:10]
    if not client_id or not sd:
        return []
    week_variants = _week_start_variants_for_session_date(sd)
    base_q: dict = {
        "state": {"$nin": ["cancel_therapist"]},
        "service_code": {"$nin": ["LEAVE", "BREAK", "AVC", ""]},
    }
    if week_variants:
        base_q["week_start"] = {"$in": week_variants}
    cells = await db.schedule_cells.find(base_q, {"_id": 0}).to_list(2000)
    out: list = []
    for cell in cells:
        if _schedule_cell_date_iso(cell) != sd:
            continue
        if await _cell_matches_session_client(cell, client_id):
            out.append(cell)
    return out


async def _mark_client_day_schedule_prep_cells(
    client_id: str,
    session_date: str,
    prepared_by: str,
    *,
    client_name: Optional[str] = None,
    anchor_cell: Optional[dict] = None,
    marker_type: str = "prep",
) -> int:
    """Mirror prep onto every schedule cell for this client+day (supervision / co-therapist rows)."""
    sd = (session_date or "")[:10]
    if not client_id or not sd:
        return 0
    if not client_name:
        client = await db.clients.find_one(_active_client_filter({"id": client_id}), {"_id": 0, "name": 1})
        client_name = (client or {}).get("name")
    cells = await _schedule_cells_for_client_day(client_id, sd)
    if anchor_cell and anchor_cell.get("id"):
        if not any(c.get("id") == anchor_cell.get("id") for c in cells):
            cells.append(anchor_cell)
    marked = 0
    seen: set = set()
    for cell in cells:
        tid = cell.get("therapist_id")
        if not tid:
            continue
        key = (tid, cell.get("id"))
        if key in seen:
            continue
        seen.add(key)
        try:
            await _upsert_session_prep_markers(
                therapist_id=tid,
                client_id=client_id,
                session_date=sd,
                prepared_by=prepared_by,
                client_name=client_name,
                cell=cell,
                marker_type=marker_type,
            )
            marked += 1
        except Exception:
            logger.exception(
                "propagate prep marker for %s / %s / %s", tid, client_id, sd,
            )
    return marked


async def _related_therapist_ids_for_prep(
    therapist_id: str,
    client_id: str,
    session_date: str,
) -> List[str]:
    """Therapist ids that may share a dual-specialist prep slot (same client + date)."""
    sd = (session_date or "")[:10]
    ids: set = {therapist_id}
    if client_id and sd:
        sessions = await db.sessions.find(
            {"client_id": client_id, **_session_date_query(sd)},
            {"_id": 0, "therapist_ids": 1},
        ).to_list(100)
        for sess in sessions:
            for tid in sess.get("therapist_ids") or []:
                if tid:
                    ids.add(tid)
        hist_rows = await db.prep_history.find(
            {"client_id": client_id, **_session_date_query(sd)},
            {"_id": 0, "therapist_id": 1},
        ).to_list(100)
        for row in hist_rows:
            tid = row.get("therapist_id")
            if tid:
                ids.add(tid)
        for cell in await _schedule_cells_for_client_day(client_id, sd):
            tid = cell.get("therapist_id")
            if tid:
                ids.add(tid)
    return [t for t in ids if t]


async def _resolve_schedule_cell_for_prep(
    therapist_id: str,
    client_id: str,
    session_date: str,
    client_name: Optional[str] = None,
    stale_cell_id: Optional[str] = None,
    extra_therapist_ids: Optional[List[str]] = None,
) -> Optional[dict]:
    """Match schedule cell by id, client, or child name — survives Excel re-import."""
    sd = (session_date or "")[:10]
    if stale_cell_id:
        cell = await db.schedule_cells.find_one({"id": stale_cell_id}, {"_id": 0})
        if cell and _schedule_cell_date_iso(cell) == sd:
            if await _cell_matches_session_client(cell, client_id):
                return cell

    async def _scan_cells(cells: list) -> Optional[dict]:
        for cell in cells:
            if _schedule_cell_date_iso(cell) != sd:
                continue
            if await _cell_matches_session_client(cell, client_id):
                return cell
        if client_name:
            want = _normalize_intake_name(client_name)
            first = want.split()[0] if want else ""
            for cell in cells:
                if _schedule_cell_date_iso(cell) != sd:
                    continue
                label = _normalize_intake_name(_schedule_cell_child_label(cell))
                if not label:
                    continue
                if label == want or (first and len(first) >= 3 and (label.startswith(first) or first == label.split()[0])):
                    return cell
        return None

    search_ids = list(dict.fromkeys(
        [therapist_id] + [t for t in (extra_therapist_ids or []) if t and t != therapist_id]
    ))
    week_variants = _week_start_variants_for_session_date(sd)
    for tid in search_ids:
        q: dict = {
            "therapist_id": tid,
            "state": {"$nin": ["cancel_therapist"]},
            "service_code": {"$nin": ["LEAVE", "BREAK", "AVC", ""]},
        }
        if week_variants:
            q["week_start"] = {"$in": week_variants}
        cells = await db.schedule_cells.find(q, {"_id": 0}).to_list(500)
        hit = await _scan_cells(cells)
        if hit:
            return hit

    # Dual specialists: client may be scheduled under a co-therapist's column.
    fallback_q: dict = {
        "state": {"$nin": ["cancel_therapist"]},
        "service_code": {"$nin": ["LEAVE", "BREAK", "AVC", ""]},
    }
    if week_variants:
        fallback_q["week_start"] = {"$in": week_variants}
    all_cells = await db.schedule_cells.find(fallback_q, {"_id": 0}).to_list(2000)
    return await _scan_cells(all_cells)


async def _sync_prep_history_to_schedule_markers(start: str, end: str) -> None:
    """Mirror prep_history rows into schedule_preparations for schedule badge display."""
    rows = await db.prep_history.find(
        _session_date_range_query(start, end),
        {"_id": 0},
    ).to_list(5000)
    for row in rows:
        tid = row.get("therapist_id")
        cid = row.get("client_id")
        sd = _session_date_iso(row.get("session_date"))
        if not tid or not cid or not sd:
            continue
        stale_cell_id = row.get("schedule_cell_id")
        related = await _related_therapist_ids_for_prep(tid, cid, sd)
        marked_cell_ids: set = set()
        linked_tids: set = set()
        primary_cell_id = None
        prepared_by = row.get("prepared_by") or ""
        row_source = (row.get("source") or "").strip().lower()
        marker_type = "no_show" if row_source == "no_show" else "prep"
        for rtid in related:
            cell = await _resolve_schedule_cell_for_prep(
                rtid,
                cid,
                sd,
                client_name=row.get("client_name"),
                stale_cell_id=stale_cell_id if rtid == tid else None,
                extra_therapist_ids=related,
            )
            if cell and cell.get("id"):
                if cell["id"] in marked_cell_ids:
                    continue
                marked_cell_ids.add(cell["id"])
                if not primary_cell_id:
                    primary_cell_id = cell["id"]
                marker_tid = cell.get("therapist_id") or rtid
                linked_tids.add(marker_tid)
                try:
                    await _upsert_session_prep_markers(
                        therapist_id=marker_tid,
                        client_id=cid,
                        session_date=sd,
                        prepared_by=prepared_by,
                        client_name=row.get("client_name"),
                        cell=cell,
                        marker_type=marker_type,
                    )
                except Exception:
                    logger.exception("sync prep_history to schedule marker %s", row.get("id"))
            elif rtid not in linked_tids:
                linked_tids.add(rtid)
                try:
                    await _upsert_session_prep_markers(
                        therapist_id=rtid,
                        client_id=cid,
                        session_date=sd,
                        prepared_by=prepared_by,
                        client_name=row.get("client_name"),
                        cell=None,
                        marker_type=marker_type,
                    )
                except Exception:
                    logger.exception("sync prep_history marker without cell %s", row.get("id"))
        if primary_cell_id and primary_cell_id != stale_cell_id:
            await db.prep_history.update_one(
                {"id": row.get("id")},
                {"$set": {"schedule_cell_id": primary_cell_id}},
            )


_LOGGED_PREP_SESSION_STATUSES = {"Completed"}
_NO_ATTENDANCE_SESSION_STATUSES = {"No Show", "Cancelled"}
_NO_SHOW_SESSION_STATUSES = _NO_ATTENDANCE_SESSION_STATUSES
_SESSION_BADGE_STATUSES = _LOGGED_PREP_SESSION_STATUSES | _NO_ATTENDANCE_SESSION_STATUSES


def _require_same_day_session(user: dict, session_date: str) -> None:
    """Therapists may only log/prepare sessions on the session day (not future days)."""
    if _has_full_client_access(user) or _is_hr_ops(user):
        return
    sd = (session_date or "")[:10]
    today = now_iso()[:10]
    if sd != today:
        raise HTTPException(
            status_code=400,
            detail="Preparation is only allowed on the session day until 11:59 PM.",
        )


async def _computed_schedule_preparation_markers(
    start: str, end: str, therapist_id: Optional[str] = None
) -> list:
    """Build prep markers from completed sessions on the exact matching schedule cell only."""
    filter_ids: Optional[set] = None
    if therapist_id:
        filter_ids = set(await _expand_therapist_ids(therapist_id))
    sessions = await db.sessions.find(
        {
            **_session_date_range_query(start, end),
            "status": {"$in": list(_LOGGED_PREP_SESSION_STATUSES)},
        },
        {"_id": 0},
    ).to_list(5000)
    if not sessions:
        return []
    alias_map = await _build_therapist_id_alias_map()
    markers: list = []
    seen: set = set()
    for sess in sessions:
        cid = sess.get("client_id")
        sd = (sess.get("session_date") or "")[:10]
        sid = sess.get("id")
        if not cid or not sd or not sid:
            continue
        client = await db.clients.find_one(
            _active_client_filter({"id": cid}), {"_id": 0, "name": 1},
        )
        client_name = (client or {}).get("name")
        cells = await _schedule_cells_for_client_day(cid, sd)
        for cell in cells:
            tid = cell.get("therapist_id")
            if not tid:
                continue
            if filter_ids and tid not in filter_ids:
                continue
            if not _session_includes_cell_therapist(sess, tid, alias_map):
                continue
            if not await _cell_matches_session_client(cell, cid):
                continue
            if not _session_time_matches_cell(sess.get("start_time"), cell):
                continue
            cell_id = cell.get("id")
            key = (tid, cell_id or "", cid, sd)
            if key in seen:
                continue
            seen.add(key)
            markers.append({
                "therapist_id": tid,
                "client_id": cid,
                "session_date": sd,
                "time_slot": cell.get("time_slot") or "",
                "schedule_cell_id": cell_id,
                "week_start": cell.get("week_start"),
                "day": cell.get("day"),
                "client_name": client_name,
                "source": "session",
                "marker_type": "prep",
                "session_id": sid,
            })
    return markers


async def _computed_schedule_no_show_markers(
    start: str, end: str, therapist_id: Optional[str] = None
) -> list:
    """Build no-show markers for red badges on the exact matching schedule cell only."""
    filter_ids: Optional[set] = None
    if therapist_id:
        filter_ids = set(await _expand_therapist_ids(therapist_id))
    sessions = await db.sessions.find(
        {
            **_session_date_range_query(start, end),
            "status": {"$in": list(_NO_SHOW_SESSION_STATUSES)},
        },
        {"_id": 0},
    ).to_list(5000)
    if not sessions:
        return []
    alias_map = await _build_therapist_id_alias_map()
    markers: list = []
    seen: set = set()
    for sess in sessions:
        cid = sess.get("client_id")
        sd = (sess.get("session_date") or "")[:10]
        sid = sess.get("id")
        if not cid or not sd or not sid:
            continue
        client = await db.clients.find_one(
            _active_client_filter({"id": cid}), {"_id": 0, "name": 1},
        )
        client_name = (client or {}).get("name")
        cells = await _schedule_cells_for_client_day(cid, sd)
        for cell in cells:
            tid = cell.get("therapist_id")
            if not tid:
                continue
            if filter_ids and tid not in filter_ids:
                continue
            if not _session_includes_cell_therapist(sess, tid, alias_map):
                continue
            if not await _cell_matches_session_client(cell, cid):
                continue
            if not _session_time_matches_cell(sess.get("start_time"), cell):
                continue
            cell_id = cell.get("id")
            key = (tid, cell_id or "", cid, sd)
            if key in seen:
                continue
            seen.add(key)
            markers.append({
                "therapist_id": tid,
                "client_id": cid,
                "session_date": sd,
                "time_slot": cell.get("time_slot") or "",
                "schedule_cell_id": cell_id,
                "week_start": cell.get("week_start"),
                "day": cell.get("day"),
                "client_name": client_name,
                "source": "no_show",
                "marker_type": "no_show",
                "session_id": sid,
            })
    return markers


def _merge_schedule_preparation_markers(*groups: list) -> list:
    """Merge marker rows; prefer entries with schedule_cell_id. Keyed per cell, not per day."""
    merged: dict = {}
    for group in groups:
        for item in group or []:
            tid = item.get("therapist_id")
            cid = item.get("client_id")
            sd = (item.get("session_date") or "")[:10]
            if not tid or not cid or not sd:
                continue
            cell_id = (item.get("schedule_cell_id") or "").strip()
            slot = (item.get("time_slot") or "").strip()
            key = (tid, cid, sd, cell_id or slot)
            prev = merged.get(key)
            if not prev:
                merged[key] = dict(item)
            elif item.get("marker_type") == "no_show" or item.get("source") == "no_show":
                merged[key] = {**prev, **item, "marker_type": "no_show", "source": "no_show"}
            elif prev.get("marker_type") == "no_show" or prev.get("source") == "no_show":
                merged[key] = {**item, **prev, "marker_type": "no_show", "source": "no_show"}
            elif item.get("schedule_cell_id") and not prev.get("schedule_cell_id"):
                merged[key] = {**prev, **item}
            elif item.get("session_id") and not prev.get("session_id"):
                merged[key] = {**prev, **item}
            elif item.get("client_name") and not prev.get("client_name"):
                merged[key] = {**prev, "client_name": item["client_name"]}
    return list(merged.values())


def _prep_is_suppressed(
    suppressions: list,
    therapist_id: str,
    client_id: str,
    session_date: str,
    schedule_cell_id: Optional[str] = None,
    alias_map: Optional[Dict[str, List[str]]] = None,
) -> bool:
    sd = (session_date or "")[:10]
    related = (alias_map or {}).get(therapist_id, [therapist_id])
    for s in suppressions or []:
        stid = s.get("therapist_id")
        if stid not in related:
            continue
        if s.get("client_id") != client_id:
            continue
        if (s.get("session_date") or "")[:10] != sd:
            continue
        cell_scope = (s.get("schedule_cell_id") or "").strip() or None
        if cell_scope:
            if schedule_cell_id and cell_scope == schedule_cell_id:
                return True
        else:
            return True
    return False


def _filter_suppressed_markers(markers: list, suppressions: list, alias_map: Optional[Dict[str, List[str]]] = None) -> list:
    out = []
    for m in markers or []:
        if _prep_is_suppressed(
            suppressions,
            m.get("therapist_id"),
            m.get("client_id"),
            m.get("session_date"),
            m.get("schedule_cell_id"),
            alias_map=alias_map,
        ):
            continue
        out.append(m)
    return out


async def _clear_prep_suppressions(
    therapist_id: str,
    client_id: str,
    session_date: str,
    schedule_cell_id: Optional[str] = None,
) -> int:
    """Remove badge suppressions when prep is logged again for the same slot."""
    sd = (session_date or "")[:10]
    if not therapist_id or not client_id or not sd:
        return 0
    deleted = 0
    expanded = await _expand_therapist_ids(therapist_id)
    tid_filter = {"$in": expanded} if len(expanded) > 1 else expanded[0]
    # Drop blanket and cell-scoped suppressions so re-prep always restores the badge.
    result = await db.schedule_prep_suppressions.delete_many({
        "therapist_id": tid_filter,
        "client_id": client_id,
        "session_date": sd,
    })
    deleted += int(result.deleted_count or 0)
    if schedule_cell_id:
        for cell in await _schedule_cells_for_client_day(client_id, sd):
            if cell.get("id") == schedule_cell_id:
                other_tid = cell.get("therapist_id")
                if other_tid and other_tid not in expanded:
                    extra = await _expand_therapist_ids(other_tid)
                    extra_filter = {"$in": extra} if len(extra) > 1 else extra[0]
                    r2 = await db.schedule_prep_suppressions.delete_many({
                        "therapist_id": extra_filter,
                        "client_id": client_id,
                        "session_date": sd,
                    })
                    deleted += int(r2.deleted_count or 0)
                break
    return deleted


async def _reconcile_stale_prep_suppressions(start: str, end: str) -> int:
    """Drop suppressions that block badges even though prep/sessions exist again."""
    cleared = 0
    alias_map = await _build_therapist_id_alias_map()
    suppressions = await _list_prep_suppressions(start, end)
    for s in suppressions:
        tid = s.get("therapist_id")
        cid = s.get("client_id")
        sd = _session_date_iso(s.get("session_date"))
        if not tid or not cid or not sd:
            continue
        related_tids = alias_map.get(tid, [tid])
        has_prep = False
        has_hist = False
        has_session = False
        for rtid in related_tids:
            if not has_prep:
                has_prep = bool(await db.schedule_preparations.find_one(
                    {"therapist_id": rtid, "client_id": cid, **_session_date_query(sd)},
                    {"_id": 0, "id": 1},
                ))
            if not has_hist:
                has_hist = bool(await db.prep_history.find_one(
                    {"therapist_id": rtid, "client_id": cid, **_session_date_query(sd)},
                    {"_id": 0, "id": 1},
                ))
            if not has_session:
                has_session = bool(await db.sessions.find_one(
                    {
                        "client_id": cid,
                        **_session_date_query(sd),
                        "status": {"$in": ["Completed", "No Show"]},
                        "therapist_ids": rtid,
                    },
                    {"_id": 0, "id": 1},
                ))
        if has_prep or has_hist or has_session:
            await db.schedule_prep_suppressions.delete_one({"id": s["id"]})
            cleared += 1
    return cleared


async def _list_prep_suppressions(start: str, end: str, therapist_id: Optional[str] = None) -> list:
    q: dict = _session_date_range_query(start, end)
    if therapist_id:
        expanded = await _expand_therapist_ids(therapist_id)
        q["therapist_id"] = {"$in": expanded} if len(expanded) > 1 else expanded[0]
    return await db.schedule_prep_suppressions.find(q, {"_id": 0}).to_list(2000)


async def _delete_sessions_for_prep_slot(
    therapist_id: str,
    client_id: str,
    session_date: str,
) -> int:
    """Delete logged sessions that drove a green prep badge for this slot."""
    sd = (session_date or "")[:10]
    deleted = 0
    sessions = await db.sessions.find(
        {"client_id": client_id, "session_date": sd},
        {"_id": 0, "id": 1, "therapist_ids": 1},
    ).to_list(200)
    for sess in sessions:
        if therapist_id not in (sess.get("therapist_ids") or []):
            continue
        await db.sessions.delete_one({"id": sess["id"]})
        deleted += 1
    return deleted


async def _clear_schedule_preparation_marker(
    *,
    therapist_id: str,
    client_id: str,
    session_date: str,
    schedule_cell_id: Optional[str] = None,
    time_slot: Optional[str] = None,
    suppress_badge: bool = True,
    delete_prep_history: bool = False,
    delete_sessions: bool = False,
    suppressed_by: str = "",
) -> dict:
    """Remove prep markers and optionally suppress the green schedule badge."""
    sd = (session_date or "")[:10]
    slot = (time_slot or "").strip()
    sessions_deleted = 0
    if schedule_cell_id:
        await db.schedule_preparations.delete_many({
            "therapist_id": therapist_id,
            "client_id": client_id,
            "session_date": sd,
            "schedule_cell_id": schedule_cell_id,
        })
    else:
        await db.schedule_preparations.delete_many({
            "therapist_id": therapist_id,
            "client_id": client_id,
            "session_date": sd,
        })
    if delete_prep_history:
        hist_q = {"therapist_id": therapist_id, "client_id": client_id, "session_date": sd}
        if slot:
            hist_q["time_slot"] = slot
        await db.prep_history.delete_many(hist_q)
    if delete_sessions:
        sessions_deleted = await _delete_sessions_for_prep_slot(therapist_id, client_id, sd)
    if suppress_badge:
        doc = {
            "id": str(uuid.uuid4()),
            "therapist_id": therapist_id,
            "client_id": client_id,
            "session_date": sd,
            "schedule_cell_id": schedule_cell_id,
            "suppressed_by": suppressed_by,
            "suppressed_at": now_iso(),
        }
        existing = await db.schedule_prep_suppressions.find_one(
            {
                "therapist_id": therapist_id,
                "client_id": client_id,
                "session_date": sd,
                "schedule_cell_id": schedule_cell_id,
            },
            {"_id": 0, "id": 1},
        )
        if not existing:
            await db.schedule_prep_suppressions.insert_one(doc)
    return {"ok": True, "sessions_deleted": sessions_deleted}


async def _cleanup_prep_for_deleted_session(sess: dict) -> None:
    cid = sess.get("client_id")
    sd = (sess.get("session_date") or "")[:10]
    if not cid or not sd:
        return
    for tid in sess.get("therapist_ids") or []:
        await db.schedule_preparations.delete_many({
            "therapist_id": tid,
            "client_id": cid,
            "session_date": sd,
        })
    await db.prep_history.delete_many({"client_id": cid, "session_date": sd})


async def _refresh_schedule_preparation_cell_ids(start: str, end: str) -> None:
    """Update stale schedule_cell_id on preparation markers after grid re-import."""
    rows = await db.schedule_preparations.find(
        _session_date_range_query(start, end),
        {"_id": 0},
    ).to_list(5000)
    for row in rows:
        tid = row.get("therapist_id")
        cid = row.get("client_id")
        sd = (row.get("session_date") or "")[:10]
        if not tid or not cid or not sd:
            continue
        stale = row.get("schedule_cell_id")
        cell = await _resolve_schedule_cell_for_prep(
            tid, cid, sd, client_name=row.get("client_name"), stale_cell_id=stale,
        )
        if not cell:
            continue
        new_id = cell.get("id")
        if not new_id or new_id == stale:
            continue
        await db.schedule_preparations.update_one(
            {"id": row["id"]},
            {"$set": {
                "schedule_cell_id": new_id,
                "week_start": cell.get("week_start"),
                "day": cell.get("day"),
                "time_slot": cell.get("time_slot") or row.get("time_slot") or "",
            }},
        )


async def _sync_schedule_preparations_for_week(start: str, end: str) -> dict:
    """Backfill schedule prep markers from completed sessions (idempotent)."""
    recovery = await _recover_misdated_week_prep(start, end)
    sessions = await db.sessions.find(
        {
            **_session_date_range_query(start, end),
            "status": {"$in": list(_SESSION_BADGE_STATUSES)},
        },
        {"_id": 0},
    ).to_list(5000)
    for sess in sessions:
        try:
            await _auto_mark_schedule_preparation_for_session(
                sess, sess.get("created_by") or ""
            )
        except Exception:
            logger.exception("sync schedule preparation for session %s", sess.get("id"))
    no_attendance = await db.sessions.find(
        {
            **_session_date_range_query(start, end),
            "status": {"$in": list(_NO_ATTENDANCE_SESSION_STATUSES)},
        },
        {"_id": 0},
    ).to_list(5000)
    for sess in no_attendance:
        try:
            await _auto_mark_schedule_preparation_for_session(
                sess, sess.get("created_by") or ""
            )
        except Exception:
            logger.exception("sync no-show schedule marker for session %s", sess.get("id"))
    try:
        await _sync_prep_history_to_schedule_markers(start, end)
    except Exception:
        logger.exception("sync prep_history to schedule markers for %s–%s", start, end)
    try:
        await _refresh_schedule_preparation_cell_ids(start, end)
    except Exception:
        logger.exception("refresh schedule preparation cell ids for %s–%s", start, end)
    try:
        recovery["cleared_suppressions"] = await _reconcile_stale_prep_suppressions(start, end)
    except Exception:
        logger.exception("reconcile prep suppressions for %s–%s", start, end)
    return recovery


async def _ensure_fahda_saleh_wed_prep_marker(
    *,
    week_start: str = "2026-06-28",
    session_date: str = "2026-07-01",
    prepared_by: str = "startup",
) -> dict:
    """Ensure Saleh (#009) Wed prep badge exists on Fahda's row and co-therapist rows."""
    saleh = await db.clients.find_one(_active_client_filter({"file_no": "009"}), {"_id": 0})
    if not saleh:
        return {"ok": False, "reason": "client_009_not_found"}
    fahda = await db.therapists.find_one(
        {"email": {"$regex": r"falghadeeb@", "$options": "i"}},
        {"_id": 0},
    )
    if not fahda:
        return {"ok": False, "reason": "fahda_not_found"}
    week_start = _normalize_week_start(week_start)
    sd = (session_date or "")[:10]
    cells = await db.schedule_cells.find(
        {
            "week_start": week_start,
            "therapist_id": fahda["id"],
            "day": 3,
            "state": {"$nin": ["cancel_therapist"]},
        },
        {"_id": 0},
    ).to_list(50)
    fahda_cell = None
    for cell in cells:
        if await _cell_matches_session_client(cell, saleh["id"]):
            fahda_cell = cell
            break
        label = _schedule_cell_child_label(cell).lower()
        if "saleh" in label:
            fahda_cell = cell
            break
    if not fahda_cell:
        return {"ok": False, "reason": "fahda_wed_saleh_cell_not_found"}
    cleared = await _clear_prep_suppressions(
        fahda["id"], saleh["id"], sd, fahda_cell.get("id"),
    )
    await _upsert_session_prep_markers(
        therapist_id=fahda["id"],
        client_id=saleh["id"],
        session_date=sd,
        prepared_by=prepared_by,
        client_name=saleh.get("name"),
        cell=fahda_cell,
    )
    propagated = await _mark_client_day_schedule_prep_cells(
        saleh["id"],
        sd,
        prepared_by,
        client_name=saleh.get("name"),
        anchor_cell=fahda_cell,
    )
    markers = await db.schedule_preparations.find(
        {
            "client_id": saleh["id"],
            **_session_date_query(sd),
            "therapist_id": fahda["id"],
        },
        {"_id": 0},
    ).to_list(20)
    return {
        "ok": True,
        "saleh_id": saleh["id"],
        "fahda_id": fahda["id"],
        "cell_id": fahda_cell.get("id"),
        "time_slot": fahda_cell.get("time_slot"),
        "cleared_suppressions": cleared,
        "propagated_cells": propagated,
        "fahda_markers": len(markers),
    }


@api.get("/schedule/preparations")
async def list_schedule_preparations(
    week_start: str,
    therapist_id: Optional[str] = None,
    include_future: bool = False,
    sync: bool = False,
    user=Depends(get_current_user),
):
    """Prep-complete markers for schedule slots in a week (Sun–Thu)."""
    if not week_start:
        raise HTTPException(status_code=400, detail="week_start required")
    week_start = _normalize_week_start(week_start)
    try:
        base = datetime.fromisoformat(str(week_start)[:10])
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid week_start")
    end = (base + timedelta(days=4)).strftime("%Y-%m-%d")
    start = base.strftime("%Y-%m-%d")
    q: dict = _session_date_range_query(start, end)
    tid = therapist_id
    if user.get("role") == "therapist" and not _has_full_client_access(user):
        tid = await _resolve_user_therapist_id(user) or user.get("id")
    if tid:
        expanded = await _expand_therapist_ids(tid)
        q["therapist_id"] = {"$in": expanded} if len(expanded) > 1 else expanded[0]
    if sync:
        if not _can_manage_schedule_prep(user):
            raise HTTPException(status_code=403, detail="Forbidden")
        await _sync_schedule_preparations_for_week(start, end)
    alias_map = await _build_therapist_id_alias_map()
    db_items = await db.schedule_preparations.find(q, {"_id": 0}).to_list(2000)
    computed = await _computed_schedule_preparation_markers(start, end, tid)
    no_show_markers = await _computed_schedule_no_show_markers(start, end, tid)
    suppressions = await _list_prep_suppressions(start, end, tid)
    # Only session-backed rows from DB (internal notes metadata); badges come from sessions.
    db_session_backed = [it for it in db_items if it.get("session_id")]
    merged = _merge_schedule_preparation_markers(computed, no_show_markers, db_session_backed)
    items = _filter_suppressed_markers(merged, suppressions, alias_map)
    if not include_future:
        today = now_iso()[:10]
        items = [it for it in items if (_session_date_iso(it.get("session_date")) or "") <= today]
    # Also drop computed session markers blocked by suppression (frontend uses week sessions too).
    return {"items": items, "suppressions": suppressions}


@api.post("/schedule/relink-prep")
async def relink_schedule_prep(week_start: str = Query(...), user=Depends(get_current_user)):
    """Admin: force re-sync prep markers for a week from sessions + prep_history."""
    if not _can_manage_schedule_prep(user):
        raise HTTPException(status_code=403, detail="Forbidden")
    week_start = _normalize_week_start(week_start)
    try:
        base = datetime.fromisoformat(str(week_start)[:10])
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid week_start")
    start = base.strftime("%Y-%m-%d")
    end = (base + timedelta(days=4)).strftime("%Y-%m-%d")
    before = await _prep_week_diagnostics(start, end)
    recovery = await _sync_schedule_preparations_for_week(start, end)
    rows = await db.schedule_preparations.find(
        _prep_week_marker_scope_query(start, end),
        {"_id": 0, "therapist_id": 1, "client_id": 1, "session_date": 1},
    ).to_list(5000)
    unique = {
        (r.get("therapist_id"), r.get("client_id"), _session_date_iso(r.get("session_date")))
        for r in rows
        if r.get("therapist_id") and r.get("client_id") and _session_date_iso(r.get("session_date"))
    }
    after = await _prep_week_diagnostics(start, end)
    return {
        "ok": True,
        "week_start": week_start,
        "start": start,
        "end": end,
        "linked_count": len(unique),
        "row_count": len(rows),
        "recovery": recovery,
        "before": before,
        "after": after,
    }


@api.post("/admin/fix-fahda-saleh-prep-badge")
async def admin_fix_fahda_saleh_prep_badge(
    week_start: str = Query("2026-06-28"),
    session_date: str = Query("2026-07-01"),
    user=Depends(admin_only),
):
    """One-click: restore Fahda-row green badge for Saleh on Wed (and mirror co-therapist rows)."""
    result = await _ensure_fahda_saleh_wed_prep_marker(
        week_start=week_start,
        session_date=session_date,
        prepared_by=user["id"],
    )
    if not result.get("ok"):
        raise HTTPException(status_code=404, detail=result.get("reason") or "fix_failed")
    ws = _normalize_week_start(week_start)
    start = ws
    end = (datetime.fromisoformat(ws) + timedelta(days=4)).strftime("%Y-%m-%d")
    await _reconcile_stale_prep_suppressions(start, end)
    return result


def _can_manage_schedule_prep(user: dict) -> bool:
    return (
        _has_full_client_access(user)
        or _is_hr_ops(user)
        or _is_walaa_ops(user)
        or _is_portal_admin(user)
    )


@api.post("/schedule/preparations/clear")
async def clear_schedule_preparation(payload: SchedulePreparationClearIn, user=Depends(get_current_user)):
    """Remove green prep badge / preparation marker (ops and client-lead team)."""
    if not _can_manage_schedule_prep(user):
        raise HTTPException(status_code=403, detail="Forbidden")
    return await _clear_schedule_preparation_marker(
        therapist_id=payload.therapist_id,
        client_id=payload.client_id,
        session_date=payload.session_date,
        schedule_cell_id=payload.schedule_cell_id,
        time_slot=payload.time_slot,
        suppress_badge=payload.suppress_badge,
        delete_prep_history=payload.delete_prep_history,
        delete_sessions=payload.delete_sessions,
        suppressed_by=user["id"],
    )


@api.post("/schedule/preparations")
async def mark_schedule_preparation(payload: SchedulePreparationIn, user=Depends(get_current_user)):
    """Mark preparation complete for a scheduled session slot."""
    _require_same_day_session(user, payload.session_date)
    if user.get("role") == "therapist" and not _has_full_client_access(user):
        uid = await _resolve_user_therapist_id(user) or user.get("id")
        if payload.therapist_id != uid:
            raise HTTPException(status_code=403, detail="Forbidden")
        if not await _therapist_assigned_to_client(uid, payload.client_id):
            raise HTTPException(status_code=403, detail="Forbidden")
    anchor_cell = await _validate_prep_client_matches_cell(
        schedule_cell_id=payload.schedule_cell_id,
        client_id=payload.client_id,
        therapist_id=payload.therapist_id,
        cell_child_name=payload.cell_child_name,
    )
    if not anchor_cell and payload.schedule_cell_id:
        anchor_cell = await db.schedule_cells.find_one(
            {"id": payload.schedule_cell_id}, {"_id": 0},
        )
    doc = await _upsert_schedule_preparation(
        therapist_id=payload.therapist_id,
        client_id=payload.client_id,
        session_date=payload.session_date,
        prepared_by=user["id"],
        time_slot=payload.time_slot,
        schedule_cell_id=payload.schedule_cell_id,
        week_start=payload.week_start,
        day=payload.day,
        notes=payload.notes,
        internal_note=payload.internal_note,
    )
    return doc


@api.patch("/schedule/preparations/note")
async def update_schedule_preparation_note(payload: SchedulePreparationNoteIn, user=Depends(get_current_user)):
    """Save internal prep notes for a schedule slot (not shown on the public grid)."""
    if user.get("role") == "therapist" and not _has_full_client_access(user):
        uid = await _resolve_user_therapist_id(user) or user.get("id")
        if payload.therapist_id != uid:
            raise HTTPException(status_code=403, detail="Forbidden")
        if not await _therapist_assigned_to_client(uid, payload.client_id):
            raise HTTPException(status_code=403, detail="Forbidden")
    session_date = (payload.session_date or "")[:10]
    slot = (payload.time_slot or "").strip()
    internal_note = (payload.internal_note or "").strip()
    q = _prep_history_key(payload.therapist_id, payload.client_id, session_date, slot)
    existing = await db.schedule_preparations.find_one(q, {"_id": 0})
    if existing:
        await db.schedule_preparations.update_one(
            {"id": existing["id"]},
            {"$set": {"internal_note": internal_note, "updated_at": now_iso()}},
        )
    else:
        client = await db.clients.find_one(
            _active_client_filter({"id": payload.client_id}), {"_id": 0, "name": 1},
        )
        await _upsert_schedule_preparation(
            therapist_id=payload.therapist_id,
            client_id=payload.client_id,
            session_date=session_date,
            prepared_by=user["id"],
            time_slot=slot,
            schedule_cell_id=payload.schedule_cell_id,
            client_name=(client or {}).get("name"),
            internal_note=internal_note,
        )
    hist_q = _prep_history_key(payload.therapist_id, payload.client_id, session_date, slot)
    hist = await db.prep_history.find_one(hist_q, {"_id": 0, "id": 1})
    if hist:
        await db.prep_history.update_one(
            {"id": hist["id"]},
            {"$set": {"internal_note": internal_note, "updated_at": now_iso()}},
        )
    return {"ok": True, "internal_note": internal_note}


@api.get("/clients/{cid}/prep-history")
async def list_client_prep_history(cid: str, user=Depends(get_current_user)):
    """All preparation log entries for a client — includes records without an invoice sheet."""
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0, "id": 1})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if user.get("role") == "therapist" and not _has_full_client_access(user):
        uid = await _resolve_user_therapist_id(user) or user["id"]
        if not await _therapist_assigned_to_client(uid, cid):
            raise HTTPException(status_code=403, detail="Forbidden")
    await _sync_schedule_preparations_to_prep_history(cid)
    items = await db.prep_history.find({"client_id": cid}, {"_id": 0}).sort(
        [("session_date", -1), ("prepared_at", -1)]
    ).to_list(500)
    # Hide future-dated prep rows by default (these confuse schedule badges + history).
    today = now_iso()[:10]
    items = [it for it in items if (_session_date_iso(it.get("session_date")) or "") <= today]
    therapists = {
        t["id"]: t
        async for t in db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "key": 1})
    }
    for item in items:
        if not item.get("therapist_name"):
            item["therapist_name"] = therapist_schedule_display_name(
                therapists.get(item.get("therapist_id"))
            )
    return items


@api.delete("/prep-history/{hid}")
async def delete_prep_history(hid: str, user=Depends(get_current_user)):
    """Remove a preparation log entry and clear its schedule badge."""
    if not _can_manage_schedule_prep(user):
        raise HTTPException(status_code=403, detail="Forbidden")
    rec = await db.prep_history.find_one({"id": hid}, {"_id": 0})
    if not rec:
        return {"ok": True}
    await db.prep_history.delete_one({"id": hid})
    await _clear_schedule_preparation_marker(
        therapist_id=rec.get("therapist_id"),
        client_id=rec.get("client_id"),
        session_date=rec.get("session_date"),
        schedule_cell_id=rec.get("schedule_cell_id"),
        time_slot=rec.get("time_slot"),
        suppress_badge=True,
        delete_prep_history=False,
        delete_sessions=True,
        suppressed_by=user["id"],
    )
    return {"ok": True}


@api.patch("/prep-history/{hid}")
async def link_prep_history_invoice(hid: str, payload: PrepHistoryInvoiceLinkIn, user=Depends(get_current_user)):
    """Attach or update invoice sheet reference on a preparation log entry."""
    rec = await db.prep_history.find_one({"id": hid}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Not found")
    if user.get("role") == "therapist" and not _has_full_client_access(user):
        uid = await _resolve_user_therapist_id(user) or user["id"]
        if not await _therapist_assigned_to_client(uid, rec["client_id"]):
            raise HTTPException(status_code=403, detail="Forbidden")
    patch: dict = {}
    if payload.invoice_id is not None:
        if payload.invoice_id:
            inv = await db.invoices.find_one({"id": payload.invoice_id, "client_id": rec["client_id"]}, {"_id": 0, "id": 1})
            if not inv:
                raise HTTPException(status_code=400, detail="Invoice not found for this client")
        patch["invoice_id"] = payload.invoice_id or None
    if payload.notes is not None:
        patch["notes"] = payload.notes
    if patch:
        await db.prep_history.update_one({"id": hid}, {"$set": patch})
    return await db.prep_history.find_one({"id": hid}, {"_id": 0})


async def _notification_user_ids(user: dict) -> List[str]:
    """IDs that may own in-app notifications for this login (admin user + linked therapist)."""
    ids = [user["id"]]
    tid = await _resolve_user_therapist_id(user)
    if tid and tid not in ids:
        ids.append(tid)
    return ids


async def _notify(user_id: str, ntype: str, title: str, message: str, **extra):
    doc = {
        "id": str(uuid.uuid4()), "user_id": user_id, "type": ntype,
        "title": title, "message": message, "read": False,
        "acknowledged": False, "created_at": now_iso(),
        **extra,
    }
    await db.notifications.insert_one(doc)
    return doc

async def _notify_admins(ntype: str, title: str, message: str):
    """Send notification to portal admins and HR ops users."""
    admins = await db.users.find({"role": "admin"}, {"_id": 0, "id": 1, "email": 1, "is_hr_ops": 1}).to_list(50)
    for a in admins:
        stub = {"role": "admin", "email": a.get("email"), "is_hr_ops": a.get("is_hr_ops")}
        if _is_portal_admin(stub) or _is_hr_ops(stub):
            await _notify(a["id"], ntype, title, message)


async def _notify_hr_ops(ntype: str, title: str, message: str, **extra):
    admins = await db.users.find({"role": "admin"}, {"_id": 0, "id": 1, "email": 1, "is_hr_ops": 1}).to_list(50)
    for a in admins:
        stub = {"role": "admin", "email": a.get("email"), "is_hr_ops": a.get("is_hr_ops")}
        if _is_hr_ops(stub):
            await _notify(a["id"], ntype, title, message, **extra)


def _can_view_all_purchases(user: dict) -> bool:
    return (
        _is_portal_admin(user)
        or _is_hr_ops(user)
        or _is_client_lead(user)
        or _is_walaa_ops(user)
        or _is_jenan(user)
    )


async def _notify_ops_leads(ntype: str, title: str, message: str, **extra):
    """Walaa, Maha, Fahda, Jenan (+ Walaa ops account)."""
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "email": 1, "key": 1, "name": 1}).to_list(300)
    notified: set = set()
    for t in therapists:
        stub = {"email": t.get("email"), "key": t.get("key"), "name": t.get("name"), "role": "therapist"}
        if (_is_client_lead(stub) or _is_walaa_ops(stub)) and t["id"] not in notified:
            await _notify(t["id"], ntype, title, message, **extra)
            notified.add(t["id"])


async def _notify_purchase_submitted(purchaser_name: str, item: str, category: str):
    title = "New staff purchase logged"
    message = f"{purchaser_name}: {item} ({category}) — pending review"
    extra = {"link": "/purchases"}
    await _notify_ops_leads("purchase_new", title, message, **extra)
    await _notify_hr_ops("purchase_new", title, message, **extra)
    body = f"{message}\n"
    portal = _portal_base_url()
    if portal:
        body += f"\nReview in portal: {portal}/purchases\n"
    body += "\n— Boost Growth Portal"
    await _send_urgent_email(await _jenan_recipient_email(), title, body)


def _urgent_email_subject(subject: str) -> str:
    s = (subject or "").strip()
    if s.startswith("[عاجل]") or s.startswith("[Urgent]"):
        return s
    return f"[عاجل] [Urgent] {s}"


def _portal_base_url() -> str:
    for key in ("PORTAL_URL", "FRONTEND_URL", "PUBLIC_URL", "RAILWAY_PUBLIC_DOMAIN"):
        val = (os.environ.get(key) or "").strip().rstrip("/")
        if not val:
            continue
        if key == "RAILWAY_PUBLIC_DOMAIN" and not val.startswith("http"):
            return f"https://{val}"
        return val
    return ""


async def _jenan_recipient_email() -> str:
    """Canonical inbox for Jenan — always jsalmuhaisin@ (map login aliases)."""
    tid = await _jenan_therapist_id()
    if tid:
        t = await db.therapists.find_one({"id": tid}, {"_id": 0, "email": 1})
        if t and t.get("email"):
            raw = t["email"].strip().lower()
            return THERAPIST_LOGIN_EMAIL_ALIASES.get(raw, JENAN_EMAIL)
    return JENAN_EMAIL


async def _hr_ops_recipient_emails() -> List[str]:
    """HR inbox recipients — seeded HR account plus any is_hr_ops admin users."""
    emails: set = set()
    admins = await db.users.find({"role": "admin"}, {"_id": 0, "email": 1, "is_hr_ops": 1}).to_list(50)
    for a in admins:
        stub = {"role": "admin", "email": a.get("email"), "is_hr_ops": a.get("is_hr_ops")}
        if _is_hr_ops(stub) and a.get("email"):
            emails.add(a["email"].lower().strip())
    if not emails:
        emails.add(HR_OPS_EMAIL.lower())
    return sorted(emails)


async def _send_urgent_email(to: str, subject: str, body: str) -> dict:
    return await _send_email_stub(to, _urgent_email_subject(subject), body)


async def _email_hr_ops_urgent(subject: str, body: str) -> List[dict]:
    results = []
    for addr in await _hr_ops_recipient_emails():
        results.append(await _send_urgent_email(addr, subject, body))
    return results


MANAGER_HR_NOTIFY_STATUSES = frozenset({"pending_hr", "pending_manager", "rejected"})


async def _notify_hr_manager_decision(
    *,
    ntype: str,
    therapist_name: str,
    summary: str,
    decision_status: str,
    admin_note: Optional[str],
):
    """HR in-app + email when the direct manager saves any review outcome."""
    decision_labels = {
        "pending_hr": "Approve & forward to HR",
        "pending_manager": "Pending manager review",
        "rejected": "Rejected by manager",
    }
    decision_label = decision_labels.get(decision_status, decision_status)
    hr_title = f"Manager review — {decision_label}"
    hr_msg = f"{therapist_name or 'Staff'}: {summary}"
    await _notify_hr_ops(ntype, hr_title, hr_msg)
    hr_body = f"{hr_msg}\n\nManager decision: {decision_label}"
    if admin_note:
        hr_body += f"\n\nManager note: {admin_note}"
    portal = _portal_base_url()
    if portal:
        hr_body += f"\n\nReview in portal: {portal}/requests"
    hr_body += "\n\n— Boost Growth Portal"
    await _email_hr_ops_urgent(hr_title, hr_body)


async def _notify_request_submitted(title: str, message: str, *, email_subject: Optional[str] = None):
    """Jenan and HR both get in-app + urgent email when a therapist submits a request."""
    jenan_id = await _jenan_therapist_id()
    if jenan_id:
        await _notify(jenan_id, "request_new", title, message)
    await _notify_hr_ops("request_new", title, message)
    body = f"{message}\n"
    portal = _portal_base_url()
    if portal:
        body += f"\nReview in portal: {portal}/requests\n"
    body += "\n— Boost Growth Portal"
    subj = email_subject or title
    await _send_urgent_email(await _jenan_recipient_email(), subj, body)
    await _email_hr_ops_urgent(subj, body)


def _display_leave_type(leave_type: Optional[str]) -> str:
    """Human-friendly leave type label for emails."""
    t = (leave_type or "").strip().lower()
    if t in ("annual", "annual leave"):
        return "Annual"
    if t in ("unpaid", "absence"):
        return "Unpaid"
    if t in ("sick", "sick leave"):
        return "Sick"
    if t in ("permission",):
        return "Permission"
    return (leave_type or "Leave").strip() or "Leave"


async def _notify_leave_submitted(
    *,
    therapist_name: str,
    leave_type: Optional[str],
    start_date: str,
    end_date: str,
    days: float,
    notes: Optional[str] = None,
):
    """Jenan gets in-app + email on new leave (HR notified only after manager forwards)."""
    leave_label = _display_leave_type(leave_type)
    title = f"New leave request from {therapist_name or 'Therapist'}"
    summary = f"{leave_label} — {start_date} → {end_date} ({days:g} day(s))"
    jenan_id = await _jenan_therapist_id()
    if jenan_id:
        await _notify(jenan_id, "leave_request", title, summary)
    else:
        await _notify_admins("leave_request", title, summary)

    body = (
        "A therapist has submitted a new leave request and it is pending your review.\n\n"
        f"Therapist: {therapist_name or '—'}\n"
        f"Leave type: {leave_label}\n"
        f"Date range: {start_date} → {end_date}\n"
        f"Total days: {days:g}\n"
    )
    if (notes or "").strip():
        body += f"\nNotes:\n{notes.strip()}\n"
    portal = _portal_base_url()
    if portal:
        body += f"\nReview in portal: {portal}/manager\n"
    body += "\n— Boost Growth Portal"
    await _send_urgent_email(await _jenan_recipient_email(), title, body)


async def _resend_leave_notification(leave: dict, therapist: Optional[dict], *, also_in_app: bool = True) -> dict:
    """Re-send Jenan urgent email (and optional in-app) for an existing leave row."""
    payload = {
        "therapist_name": therapist_schedule_display_name(therapist) if therapist else "Therapist",
        "leave_type": leave.get("leave_type"),
        "start_date": leave.get("start_date") or "",
        "end_date": leave.get("end_date") or "",
        "days": float(leave.get("days") or 0),
        "notes": leave.get("notes"),
    }
    leave_label = _display_leave_type(payload["leave_type"])
    title = f"New leave request from {payload['therapist_name'] or 'Therapist'}"
    summary = f"{leave_label} — {payload['start_date']} → {payload['end_date']} ({payload['days']:g} day(s))"
    jenan_id = await _jenan_therapist_id()
    if also_in_app and jenan_id:
        await _notify(jenan_id, "leave_request", title, summary)
    body = (
        "A therapist has submitted a new leave request and it is pending your review.\n\n"
        f"Therapist: {payload['therapist_name'] or '—'}\n"
        f"Leave type: {leave_label}\n"
        f"Date range: {payload['start_date']} → {payload['end_date']}\n"
        f"Total days: {payload['days']:g}\n"
    )
    if (payload.get("notes") or "").strip():
        body += f"\nNotes:\n{payload['notes'].strip()}\n"
    portal = _portal_base_url()
    if portal:
        body += f"\nReview in portal: {portal}/manager\n"
    body += "\n— Boost Growth Portal"
    email_result = await _send_urgent_email(await _jenan_recipient_email(), title, body)
    return {
        "leave_id": leave.get("id"),
        "therapist_id": leave.get("therapist_id"),
        "therapist_name": payload["therapist_name"],
        "status": leave.get("status"),
        "email_to": await _jenan_recipient_email(),
        "email_status": email_result.get("status"),
        "email_error": email_result.get("error"),
    }


async def _walaa_notify_user_ids() -> List[str]:
    t = await db.therapists.find_one(
        {"email": {"$regex": r"^walaa@boostgrowthsa\.com$", "$options": "i"}},
        {"_id": 0, "id": 1},
    )
    if t:
        return [t["id"]]
    for t in await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "key": 1}).to_list(200):
        if (t.get("key") or "").lower() == "mswalaa":
            return [t["id"]]
        name = (t.get("name") or "").lower().replace("ms.", "").replace("ms ", "").strip()
        if name.startswith("walaa"):
            return [t["id"]]
    return []


SCHEDULE_DAYS_AR = ("الأحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس")


async def _mark_parent_cancel_pending(cell_id: str) -> None:
    await db.schedule_cells.update_one(
        {"id": cell_id},
        {
            "$set": {"parent_notify_pending": True, "parent_cancel_marked_at": now_iso()},
            "$unset": {"parent_notify_sent_at": ""},
        },
    )


async def _clear_parent_cancel_pending(cell_id: str) -> None:
    await db.schedule_cells.update_one(
        {"id": cell_id},
        {
            "$set": {"parent_notify_pending": False},
            "$unset": {
                "parent_cancel_marked_at": "",
                "parent_notify_sent_at": "",
                "parent_notify_sent_by": "",
                "parent_notify_message": "",
            },
        },
    )


async def _notify_parent_cancel_pending(cell: dict, actor: str = "") -> None:
    child = (cell.get("child_name") or "—").strip()
    slot = cell.get("time_slot") or ""
    msg = f"Parent WhatsApp needed: {child} — therapist cancellation at {slot}"
    if actor:
        msg = f"{actor}: {msg}"
    title = "Parent cancellation — WhatsApp pending"
    extra = {"schedule_cell_id": cell.get("id")} if cell.get("id") else {}
    notified: set = set()
    admins = await db.users.find({"role": "admin"}, {"_id": 0, "id": 1, "email": 1, "is_hr_ops": 1}).to_list(50)
    for a in admins:
        stub = {"role": "admin", "email": a.get("email"), "is_hr_ops": a.get("is_hr_ops")}
        if _is_portal_admin(stub) or _is_hr_ops(stub):
            uid = a["id"]
            if uid not in notified:
                await _notify(uid, "parent_cancel_pending", title, msg, **extra)
                notified.add(uid)
    for wid in await _walaa_notify_user_ids():
        if wid not in notified:
            await _notify(wid, "parent_cancel_pending", title, msg, **extra)
            notified.add(wid)


async def _jenan_therapist_id() -> Optional[str]:
    t = await db.therapists.find_one(
        {"email": {"$regex": r"^jsalmuhaisin@boostgrowthsa\.com$", "$options": "i"}},
        {"_id": 0, "id": 1},
    )
    if t:
        return t["id"]
    for t in await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "key": 1}).to_list(200):
        if (t.get("key") or "").lower() == "msjenan":
            return t["id"]
        name = (t.get("name") or "").lower().replace("ms.", "").replace("ms ", "").strip()
        if name.startswith("jenan"):
            return t["id"]
    return None


SCHEDULE_CHILD_NAME_ALIASES = {
    "abdularahman": "abdulrahman",
    "aljouhrah": "aljoharah",
    "ameerah": "ameirah",
    "mohmmed": "mohammed",
    "alaqeel": "alaqel",
}

# Schedule grid abbreviations (e.g. "Abdulaziz A" on Shatha's row)
SCHEDULE_SHORT_LABEL_FILES = {
    "abdulaziz a": "024",
    "abdulaziz w": "040",
    "khalid": "072",
    "khalid ibrahim": "072",
    "mohammed alaqeel": "027",
    "mohammed alaqel": "027",
    "mohmmed alaqel": "027",
}


def _apply_schedule_child_name_aliases(name: str) -> str:
    """Fix common Excel/schedule typos before client lookup."""
    raw = (name or "").strip()
    if not raw:
        return raw
    parts = raw.split()
    first = parts[0].lower()
    if first in SCHEDULE_CHILD_NAME_ALIASES:
        parts[0] = SCHEDULE_CHILD_NAME_ALIASES[first].title() if parts[0][0].isupper() else SCHEDULE_CHILD_NAME_ALIASES[first]
        return " ".join(parts)
    return raw


async def _find_client_by_schedule_child_name(child_name: str) -> Optional[dict]:
    """Match client by schedule cell child_name (name, file_no, first-name, aliases)."""
    name = (child_name or "").strip()
    if not name:
        return None
    parts = _split_schedule_child_names(name)
    if len(parts) > 1:
        matched: List[dict] = []
        for part in parts:
            hit = await _find_client_by_schedule_child_name(part)
            if hit:
                matched.append(hit)
        unique = {c["id"]: c for c in matched}
        if len(unique) == 1:
            return next(iter(unique.values()))
        if len(unique) > 1:
            return None
    lookup_name = parts[0] if parts else name
    fields = {
        "_id": 0, "id": 1, "name": 1, "file_no": 1,
        "main_therapist_id": 1, "co_therapist_ids": 1,
        "schedule_color": 1, "color": 1,
    }
    short_key = _normalize_intake_name(lookup_name)
    if short_key in SCHEDULE_SHORT_LABEL_FILES:
        by_file = await _find_client_by_file_no(SCHEDULE_SHORT_LABEL_FILES[short_key])
        if by_file:
            client = await db.clients.find_one(_active_client_filter({"id": by_file["id"]}), fields)
            if client:
                return client

    async def _lookup(label: str) -> Optional[dict]:
        label = (label or "").strip()
        if not label:
            return None
        client = await db.clients.find_one(_active_client_filter({"name": label}), fields)
        if client:
            return client
        client = await db.clients.find_one(
            _active_client_filter({"name": {"$regex": f"^{re.escape(label)}($|\\s)", "$options": "i"}}),
            fields,
        )
        if client:
            return client
        items = await db.clients.find(_active_client_filter(), fields).to_list(500)
        for c in items:
            cn = (c.get("name") or "").strip()
            if cn and (label == cn or label.startswith(cn + " ")):
                return c
        first = label.split()[0] if label.split() else label
        if len(first) >= 3:
            fl = first.lower()
            by_first = [
                c for c in items
                if (c.get("name") or "").strip().split()[0].lower() == fl
            ]
            if len(by_first) == 1:
                return by_first[0]
        return None

    m = re.match(r"^(\d{2,3})\b", lookup_name)
    if m:
        by_file = await _find_client_by_file_no(m.group(1))
        if by_file:
            return await db.clients.find_one(_active_client_filter({"id": by_file["id"]}), fields)
    m = re.search(r"\((\d{2,3})\)", lookup_name)
    if m:
        by_file = await _find_client_by_file_no(m.group(1))
        if by_file:
            return await db.clients.find_one(_active_client_filter({"id": by_file["id"]}), fields)

    client = await _lookup(lookup_name)
    if client:
        return client
    aliased = _apply_schedule_child_name_aliases(lookup_name)
    if aliased != lookup_name:
        client = await _lookup(aliased)
        if client:
            return client
    return None


async def _ensure_co_therapist_from_schedule(therapist_id: str, child_name: Optional[str]) -> None:
    """When a therapist is scheduled for a child, add them as co-therapist so they see Client Info."""
    if not therapist_id or not (child_name or "").strip():
        return
    client = await _find_client_by_schedule_child_name(child_name)
    if not client:
        return
    if client.get("main_therapist_id") == therapist_id:
        return
    co_ids = list(client.get("co_therapist_ids") or [])
    if therapist_id in co_ids:
        return
    co_ids.append(therapist_id)
    await db.clients.update_one({"id": client["id"]}, {"$set": {"co_therapist_ids": co_ids}})


@api.post("/schedule")
async def create_schedule_cell(payload: ScheduleCellIn, _=Depends(schedule_edit_or_admin)):
    cid = str(uuid.uuid4())
    doc = _strip_session_cell_color({"id": cid, **payload.model_dump(), "created_at": now_iso()})
    await db.schedule_cells.insert_one(doc)
    doc.pop("_id", None)
    if doc.get("child_name"):
        await _ensure_co_therapist_from_schedule(doc.get("therapist_id"), doc.get("child_name"))
    if doc.get("therapist_id"):
        await _notify(doc["therapist_id"], "schedule", "New session added",
                      f"{doc.get('service_code')} | {doc.get('child_name') or ''} at {doc.get('time_slot')}")
    return doc

def _normalize_schedule_cell_state(state: Optional[str]) -> str:
    """Treat blank/null as normal so cancellations can be cleared via API."""
    s = (state or "").strip()
    if not s or s == "normal":
        return "normal"
    return s


@api.put("/schedule/{cid}")
async def update_schedule_cell(cid: str, payload: ScheduleCellIn, user=Depends(schedule_edit_or_admin)):
    prev = await db.schedule_cells.find_one({"id": cid}, {"_id": 0})
    prev_state = (prev or {}).get("state")
    update = _strip_session_cell_color(payload.model_dump())
    update["state"] = _normalize_schedule_cell_state(update.get("state"))
    await db.schedule_cells.update_one({"id": cid}, {"$set": update})
    cell = await db.schedule_cells.find_one({"id": cid}, {"_id": 0})
    if cell and cell.get("child_name"):
        await _ensure_co_therapist_from_schedule(cell.get("therapist_id"), cell.get("child_name"))
    actor = _actor_display(user)
    new_state = (cell or {}).get("state") or "normal"
    if prev_state in ("cancel_therapist", "cancel_child") and new_state not in ("cancel_therapist", "cancel_child"):
        await _clear_parent_cancel_pending(cid)
        cell = await db.schedule_cells.find_one({"id": cid}, {"_id": 0})
    if cell and cell.get("state") == "cancel_therapist":
        await _mark_parent_cancel_pending(cid)
        await _notify_parent_cancel_pending(cell, actor)
        await _log_therapist_cancel_prep_history(cell, user.get("id") or "")
    if cell and cell.get("therapist_id"):
        title = "Schedule update"
        detail = f"{cell.get('service_code')} | {cell.get('child_name') or ''} at {cell.get('time_slot')}"
        if cell.get("state") == "cancel_therapist":
            title = "Session marked as Therapist Cancellation"
            await _notify_admins("cancel_alert", "Therapist cancellation",
                                 f"{actor} marked {cell.get('child_name') or '—'} session on day {cell.get('day')} at {cell.get('time_slot')} as Therapist Cancel")
        elif cell.get("state") == "cancel_child":
            title = "Session marked as Client Cancellation"
            await _notify_admins("cancel_alert", "Client cancellation",
                                 f"{actor} marked {cell.get('child_name') or '—'} session on day {cell.get('day')} at {cell.get('time_slot')} as Client Cancel")
        await _notify(cell["therapist_id"], "schedule", title,
                      f"{actor} updated the schedule: {detail}",
                      actor_id=user.get("id"), actor_name=actor)
    return cell

@api.post("/schedule/{cid}/parent-whatsapp-sent")
async def mark_parent_whatsapp_sent(cid: str, payload: ParentWhatsAppSentIn, user=Depends(get_current_user)):
    if not _can_parent_cancellation_ops(user):
        raise HTTPException(status_code=403, detail="Parent cancellation ops access required")
    cell = await db.schedule_cells.find_one({"id": cid}, {"_id": 0})
    if not cell:
        raise HTTPException(status_code=404, detail="Schedule cell not found")
    await db.schedule_cells.update_one(
        {"id": cid},
        {"$set": {
            "parent_notify_pending": False,
            "parent_notify_sent_at": now_iso(),
            "parent_notify_sent_by": user.get("id"),
            "parent_notify_message": payload.message or "",
        }},
    )
    return await db.schedule_cells.find_one({"id": cid}, {"_id": 0})

@api.post("/schedule/{cid}/duplicate")
async def duplicate_cell(cid: str, _=Depends(schedule_edit_or_admin)):
    cell = await db.schedule_cells.find_one({"id": cid}, {"_id": 0})
    if not cell:
        raise HTTPException(status_code=404, detail="Not found")
    new_cell = {**cell, "id": str(uuid.uuid4()), "created_at": now_iso()}
    await db.schedule_cells.insert_one(new_cell)
    new_cell.pop("_id", None)
    return new_cell

@api.delete("/schedule/{cid}")
async def delete_schedule_cell(cid: str, _=Depends(schedule_edit_or_admin)):
    await db.schedule_cells.delete_one({"id": cid})
    return {"ok": True}

@api.post("/schedule/{cid}/notify")
async def notify_schedule(cid: str, body: ScheduleNotifyIn, user=Depends(schedule_edit_or_admin)):
    cell = await db.schedule_cells.find_one({"id": cid}, {"_id": 0})
    if not cell:
        raise HTTPException(status_code=404, detail="Schedule cell not found")
    actor = _actor_display(user)
    msg = body.message or f"Notice about session: {cell.get('child_name') or ''}"
    title = f"Notice from {actor}"
    recipients = body.recipient_ids or ([cell["therapist_id"]] if cell.get("therapist_id") else [])
    if not recipients:
        raise HTTPException(status_code=400, detail="No recipients selected")
    sent = []
    for rid in recipients:
        if body.send_in_app:
            n = await _notify(
                rid, "schedule_alert", title, msg,
                schedule_cell_id=cid, requires_ack=True,
                actor_id=user.get("id"), actor_name=actor,
            )
            sent.append({"user_id": rid, "notification_id": n["id"]})
        if body.send_email:
            therapist = await db.therapists.find_one({"id": rid}, {"_id": 0})
            if therapist and therapist.get("email"):
                subj = f"[Boost Growth] Notice from {actor}"
                email_body = (
                    f"Hello {therapist.get('name', '')},\n\n"
                    f"{actor} sent you a schedule notice:\n\n"
                    f"{msg}\n\n"
                    f"Session: {cell.get('service_code') or '—'} | {cell.get('child_name') or '—'}\n"
                    f"Day: {cell.get('day')} | Time: {cell.get('time_slot') or '—'}\n\n"
                    f"— Boost Growth Portal"
                )
                await _send_email_stub(therapist["email"], subj, email_body)
    return {"ok": True, "sent": sent}

@api.get("/schedule/{cid}/notification-receipts")
async def schedule_notification_receipts(cid: str, _=Depends(schedule_edit_or_admin)):
    items = await db.notifications.find(
        {"schedule_cell_id": cid}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    therapists = {t["id"]: t async for t in db.therapists.find({}, {"_id": 0, "id": 1, "name": 1})}
    out = []
    for n in items:
        tid = n.get("user_id")
        out.append({
            **n,
            "therapist_name": therapist_schedule_display_name(therapists.get(tid)) if tid in therapists else None,
        })
    return out

# ------------------- Clients & Sessions -------------------
@api.get("/clients")
async def list_clients(user=Depends(get_current_user)):
    if _has_full_client_access(user):
        return await db.clients.find(_active_client_filter(), {"_id": 0}).sort("file_no", 1).to_list(500)
    # therapist: see only assigned (main or co)
    items = await db.clients.find(_active_client_filter(), {"_id": 0}).sort("file_no", 1).to_list(500)
    uid = user["id"]
    return [c for c in items if c.get("main_therapist_id") == uid or uid in (c.get("co_therapist_ids") or [])]


@api.get("/clients/supervision-caseload")
async def supervision_caseload(user=Depends(get_current_user)):
    """Caseload split by clinical supervisor (Ms. Fahda / Ms. Maha) for ops leads."""
    if not _has_full_client_access(user):
        raise HTTPException(status_code=403, detail="Access denied")
    clients = await db.clients.find(_active_client_filter(), {"_id": 0}).sort("file_no", 1).to_list(500)
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
    t_by_id = {t["id"]: t.get("name") for t in therapists}
    buckets: dict = {"fahda": [], "maha": [], "other": []}

    def _service_label(c: dict) -> Optional[str]:
        st = c.get("service_type")
        if st:
            return _normalize_service_type(st) or str(st).strip()
        locs = c.get("locations") or []
        types = sorted({l.get("service_type") for l in locs if l.get("service_type")})
        if types:
            return " / ".join(types)
        return None

    for c in clients:
        bucket = _supervisor_bucket(c.get("supervisor"))
        main_name = t_by_id.get(c.get("main_therapist_id"))
        row = {
            "id": c.get("id"),
            "file_no": str(c.get("file_no") or "").zfill(3) if c.get("file_no") else None,
            "name": c.get("name"),
            "supervisor": c.get("supervisor"),
            "service": _service_label(c),
            "status": c.get("status") or "Active",
            "main_therapist": main_name,
        }
        if bucket in ("fahda", "maha"):
            buckets[bucket].append(row)
        else:
            buckets["other"].append(row)

    return {
        "fahda": buckets["fahda"],
        "maha": buckets["maha"],
        "other": buckets["other"],
        "counts": {
            "fahda": len(buckets["fahda"]),
            "maha": len(buckets["maha"]),
            "other": len(buckets["other"]),
            "active_fahda": sum(1 for r in buckets["fahda"] if (r.get("status") or "Active") != "Inactive"),
            "active_maha": sum(1 for r in buckets["maha"] if (r.get("status") or "Active") != "Inactive"),
        },
    }

async def _resolve_user_therapist_id(user: dict) -> Optional[str]:
    """Map logged-in user to therapist id (handles client-lead admin logins)."""
    uid = user.get("id")
    if user.get("role") == "therapist" and uid:
        return uid
    email = (user.get("email") or "").lower().strip()
    if email:
        t = await db.therapists.find_one({"email": {"$regex": f"^{re.escape(email)}$", "$options": "i"}}, {"_id": 0, "id": 1})
        if t:
            return t["id"]
    key = (user.get("key") or "").lower()
    if key:
        t = await db.therapists.find_one({"key": key}, {"_id": 0, "id": 1})
        if t:
            return t["id"]
    name = (user.get("name") or "").lower().replace("ms.", "").replace("ms ", "").strip()
    first = name.split()[0] if name else ""
    if first in FULL_CLIENT_NAME_TOKENS:
        for t in await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(200):
            tn = (t.get("name") or "").lower().replace("ms.", "").replace("ms ", "").strip()
            tfirst = tn.split()[0] if tn else ""
            if tfirst == first:
                return t["id"]
    return uid


@api.get("/clients/resolve-schedule-name")
async def resolve_client_by_schedule_name(child_name: str, user=Depends(get_current_user)):
    """Resolve a schedule cell child_name to a client the current user may log sessions for."""
    client = await _find_client_by_schedule_child_name(child_name)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if _has_full_client_access(user):
        return await db.clients.find_one(_active_client_filter({"id": client["id"]}), {"_id": 0})
    tid = await _resolve_user_therapist_id(user)
    if not tid:
        raise HTTPException(status_code=403, detail="Access denied")
    if client.get("main_therapist_id") == tid or tid in (client.get("co_therapist_ids") or []):
        return await db.clients.find_one(_active_client_filter({"id": client["id"]}), {"_id": 0})
    scheduled = await db.schedule_cells.find_one(
        {"therapist_id": tid, "child_name": {"$regex": f"^{re.escape((child_name or '').strip())}($|\\s)"}},
        {"_id": 0, "id": 1},
    )
    if scheduled:
        await _ensure_co_therapist_from_schedule(tid, child_name)
        return await db.clients.find_one(_active_client_filter({"id": client["id"]}), {"_id": 0})
    raise HTTPException(status_code=403, detail="Access denied")


@api.post("/clients")
async def create_client(payload: ClientIn, _=Depends(client_lead_or_admin)):
    cid = str(uuid.uuid4())
    data = payload.model_dump()
    data["locations"] = [l for l in (data.get("locations") or [])]
    data["status"] = _normalize_client_status(data.get("status"))
    doc = {"id": cid, **data, "created_at": now_iso()}
    await db.clients.insert_one(doc)
    doc.pop("_id", None)
    return doc

async def _user_can_edit_client_records(user: dict, client: dict) -> bool:
    if _is_portal_admin(user) or _is_client_lead(user) or _is_hr_ops(user):
        return True
    tid = await _resolve_user_therapist_id(user)
    if not tid:
        return False
    return client.get("main_therapist_id") == tid or tid in (client.get("co_therapist_ids") or [])


@api.put("/clients/{cid}")
async def update_client(cid: str, payload: ClientIn, user=Depends(get_current_user)):
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    is_lead = _is_portal_admin(user) or _is_client_lead(user) or _is_hr_ops(user)
    can_edit_records = await _user_can_edit_client_records(user, client)

    if not is_lead:
        if not can_edit_records:
            raise HTTPException(status_code=403, detail="Admin access required")
        patch: dict = {}
        if payload.drive_links is not None:
            patch["drive_links"] = payload.drive_links
        if payload.record_files is not None:
            clean = []
            for rf in payload.record_files or []:
                if not isinstance(rf, dict):
                    continue
                entry = {k: v for k, v in rf.items() if k != "file_data"}
                if entry.get("id"):
                    clean.append(entry)
            patch["record_files"] = clean
        if not patch:
            raise HTTPException(status_code=403, detail="Therapists can only update drive links and record files")
        await db.clients.update_one({"id": cid}, {"$set": patch})
        return await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})

    data = payload.model_dump()
    data["locations"] = [l for l in (data.get("locations") or [])]
    data["status"] = _normalize_client_status(data.get("status"))
    if data.get("record_files") is not None:
        data["record_files"] = [
            {k: v for k, v in (rf or {}).items() if k != "file_data"}
            for rf in (data.get("record_files") or [])
            if isinstance(rf, dict) and rf.get("id")
        ]
    await db.clients.update_one({"id": cid}, {"$set": data})
    return await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})

class ClientScheduleColorIn(BaseModel):
    color: Optional[str] = None

@api.put("/clients/{cid}/schedule-color")
async def update_client_schedule_color(cid: str, body: ClientScheduleColorIn, _=Depends(ops_or_admin)):
    """Set schedule_color on client and propagate to all schedule cells with matching child_name."""
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    color = body.color
    await db.clients.update_one({"id": cid}, {"$set": {"schedule_color": color}})
    return {"ok": True, "schedule_color": color, "client_id": cid}

@api.delete("/clients/{cid}")
async def delete_client(cid: str, _=Depends(client_lead_or_admin)):
    """Soft-delete: mark deleted=true; sessions/invoices are preserved."""
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0, "id": 1})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    await db.clients.update_one({"id": cid}, {"$set": {"deleted": True, "deleted_at": now_iso()}})
    return {"ok": True}

@api.get("/admin/clients/deleted")
async def list_deleted_clients(_=Depends(admin_only)):
    return await db.clients.find({"deleted": True}, {"_id": 0}).sort("deleted_at", -1).to_list(500)

@api.post("/admin/clients/{cid}/restore")
async def restore_client(cid: str, _=Depends(admin_only)):
    result = await db.clients.update_one(
        {"id": cid, "deleted": True},
        {"$set": {"deleted": False}, "$unset": {"deleted_at": ""}},
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Deleted client not found")
    return {"ok": True}

@api.delete("/admin/clients/{cid}/permanent")
async def permanent_delete_client(cid: str, _=Depends(admin_only)):
    client = await db.clients.find_one({"id": cid, "deleted": True}, {"_id": 0, "id": 1})
    if not client:
        raise HTTPException(status_code=404, detail="Deleted client not found")
    await db.sessions.delete_many({"client_id": cid})
    await db.invoices.delete_many({"client_id": cid})
    await db.progress_reports.delete_many({"client_id": cid})
    await db.clients.delete_one({"id": cid})
    return {"ok": True}

# ------------------- Progress Reports (per client) -------------------
PROGRESS_STATUSES = {"uploaded", "reviewed", "resolved"}

class ProgressReportIn(BaseModel):
    title: str
    url: Optional[str] = None
    status: Optional[str] = "uploaded"   # uploaded | reviewed | resolved
    notes: Optional[str] = None
    report_date: Optional[str] = None    # ISO date when the report was made

class ProgressStatusIn(BaseModel):
    status: str

class ProgressStepsIn(BaseModel):
    uploaded: Optional[bool] = None
    uploaded_by: Optional[str] = None
    uploaded_at: Optional[str] = None
    reviewed: Optional[bool] = None
    reviewed_by: Optional[str] = None
    reviewed_at: Optional[str] = None
    resolved: Optional[bool] = None
    resolved_by: Optional[str] = None
    resolved_at: Optional[str] = None

class ProgressReportLinkIn(BaseModel):
    url: Optional[str] = None

SUPERVISOR_CLIENT_FILES = {
    "msMaha": ["035", "037", "038", "040", "041", "042", "047", "052", "054", "060", "063", "065", "070"],
    "msFahda": ["009", "011", "018", "023", "024", "027", "030", "034", "061", "062", "068", "072", "079"],
}

_SUPERVISOR_BUCKET_LABELS = {"fahda": "Ms. Fahda", "maha": "Ms. Maha", "jenan": "Ms. Jenan"}


def _supervisor_bucket(supervisor: Optional[str]) -> Optional[str]:
    """Map free-text supervisor labels to canonical buckets: fahda, maha, jenan."""
    if not supervisor:
        return None
    s = str(supervisor).strip().lower()
    s = re.sub(r"^ms\.?\s*", "", s)
    if s.startswith("fahd"):
        return "fahda"
    if s.startswith("maha"):
        return "maha"
    if s.startswith("jenan"):
        return "jenan"
    return None


def _normalize_supervisor_value(val) -> Optional[str]:
    """Normalize Excel/portal supervisor spellings (Fahdah, Fahdh, Maha, …) to display names."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    bucket = _supervisor_bucket(s)
    if bucket:
        return _SUPERVISOR_BUCKET_LABELS[bucket]
    return s

async def _client_file_no(client_id: str) -> Optional[str]:
    c = await db.clients.find_one(_active_client_filter({"id": client_id}), {"_id": 0, "file_no": 1})
    if not c or not c.get("file_no"):
        return None
    return str(c["file_no"]).zfill(3)

def _is_supervisor_for_file(user: dict, file_no: str) -> bool:
    key = user.get("key") or ""
    fn = str(file_no or "").zfill(3)
    return fn in (SUPERVISOR_CLIENT_FILES.get(key) or [])

async def _can_edit_progress_step(user: dict, report_id: str, step: str) -> bool:
    if _has_full_client_access(user):
        return True
    if user.get("role") == "admin":
        return True
    doc = await db.progress_reports.find_one({"id": report_id}, {"_id": 0, "client_id": 1})
    if not doc:
        return False
    if step == "uploaded":
        fn = await _client_file_no(doc["client_id"])
        if fn and _is_supervisor_for_file(user, fn):
            return True
        if user.get("role") != "therapist":
            return False
        client = await db.clients.find_one(
            _active_client_filter({"id": doc["client_id"]}),
            {"_id": 0, "main_therapist_id": 1, "co_therapist_ids": 1},
        )
        if not client:
            return False
        uid = user["id"]
        return client.get("main_therapist_id") == uid or uid in (client.get("co_therapist_ids") or [])
    if step in ("reviewed", "resolved"):
        fn = await _client_file_no(doc["client_id"])
        return bool(fn and _is_supervisor_for_file(user, fn))
    return False

@api.get("/progress-reports/summary")
async def get_progress_reports_summary(user=Depends(get_current_user)):
    pipeline = [
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": "$client_id",
            "uploaded": {"$first": "$uploaded"},
            "reviewed": {"$first": "$reviewed"},
            "resolved": {"$first": "$resolved"},
            "count": {"$sum": 1},
        }},
    ]
    results = await db.progress_reports.aggregate(pipeline).to_list(500)
    return {
        r["_id"]: {
            "uploaded": bool(r.get("uploaded")),
            "reviewed": bool(r.get("reviewed")),
            "resolved": bool(r.get("resolved")),
            "count": r.get("count", 0),
        }
        for r in results
    }

@api.get("/clients/{cid}/progress-reports")
async def list_progress_reports(cid: str, user=Depends(get_current_user)):
    if not _has_full_client_access(user):
        client = await db.clients.find_one(
            _active_client_filter({"id": cid}),
            {"_id": 0, "main_therapist_id": 1, "co_therapist_ids": 1, "file_no": 1},
        )
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        uid = user["id"]
        assigned = client.get("main_therapist_id") == uid or uid in (client.get("co_therapist_ids") or [])
        fn = str(client.get("file_no") or "").strip()
        supervisor = bool(fn and _is_supervisor_for_file(user, fn))
        if not assigned and not supervisor:
            raise HTTPException(status_code=403, detail="Forbidden")
    return await db.progress_reports.find({"client_id": cid}, {"_id": 0, "file_data": 0}).sort("created_at", -1).to_list(200)

@api.post("/clients/{cid}/progress-reports")
async def create_progress_report(cid: str, payload: ProgressReportIn, user=Depends(get_current_user)):
    rid = str(uuid.uuid4())
    doc = {
        "id": rid,
        "client_id": cid,
        "title": payload.title.strip(),
        "url": (payload.url or "").strip() or None,
        "notes": payload.notes,
        "report_date": payload.report_date,
        "uploaded": False,
        "uploaded_by": None,
        "uploaded_at": None,
        "reviewed": False,
        "reviewed_by": None,
        "reviewed_at": None,
        "resolved": False,
        "resolved_by": None,
        "resolved_at": None,
        "created_by": user.get("name") or user.get("email"),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.progress_reports.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/progress-reports/{rid}")
async def update_progress_report(rid: str, payload: ProgressReportIn, _=Depends(admin_only)):
    status = (payload.status or "uploaded").lower()
    if status not in PROGRESS_STATUSES:
        raise HTTPException(status_code=400, detail=f"Status must be one of {sorted(PROGRESS_STATUSES)}")
    update = {
        "title": payload.title.strip(),
        "url": (payload.url or "").strip() or None,
        "status": status,
        "notes": payload.notes,
        "report_date": payload.report_date,
        "updated_at": now_iso(),
    }
    await db.progress_reports.update_one({"id": rid}, {"$set": update})
    return await db.progress_reports.find_one({"id": rid}, {"_id": 0})

@api.put("/progress-reports/{rid}/status")
async def set_progress_report_status(rid: str, payload: ProgressStatusIn, _=Depends(admin_only)):
    status = (payload.status or "").lower()
    if status not in PROGRESS_STATUSES:
        raise HTTPException(status_code=400, detail=f"Status must be one of {sorted(PROGRESS_STATUSES)}")
    await db.progress_reports.update_one({"id": rid}, {"$set": {"status": status, "updated_at": now_iso()}})
    return await db.progress_reports.find_one({"id": rid}, {"_id": 0})

@api.put("/progress-reports/{rid}/link")
async def update_progress_report_link(rid: str, payload: ProgressReportLinkIn, user=Depends(get_current_user)):
    """Save a Google Drive link to the Word/doc file for editing."""
    report = await db.progress_reports.find_one({"id": rid}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    if not await _can_access_progress_report(user, report):
        raise HTTPException(status_code=403, detail="Forbidden")
    url = (payload.url or "").strip() or None
    await db.progress_reports.update_one(
        {"id": rid}, {"$set": {"url": url, "updated_at": now_iso()}}
    )
    return await db.progress_reports.find_one({"id": rid}, {"_id": 0})


@api.put("/progress-reports/{rid}/steps")
async def update_progress_report_steps(rid: str, payload: ProgressStepsIn, user=Depends(get_current_user)):
    report = await db.progress_reports.find_one({"id": rid}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    update = {}
    for field in (
        "uploaded", "uploaded_by", "uploaded_at",
        "reviewed", "reviewed_by", "reviewed_at",
        "resolved", "resolved_by", "resolved_at",
    ):
        val = getattr(payload, field, None)
        if val is not None:
            update[field] = val
    for step in ("uploaded", "reviewed", "resolved"):
        if step in update and not await _can_edit_progress_step(user, rid, step):
            raise HTTPException(status_code=403, detail=f"Not allowed to update '{step}'")
    if update:
        update["updated_at"] = now_iso()
        await db.progress_reports.update_one({"id": rid}, {"$set": update})
        if any(k in update for k in ("uploaded", "reviewed", "resolved")):
            client = await db.clients.find_one(
                _active_client_filter({"id": report["client_id"]}),
                {"_id": 0, "name": 1, "main_therapist_id": 1},
            )
            if client and client.get("main_therapist_id"):
                therapist = await db.therapists.find_one(
                    {"id": client["main_therapist_id"]},
                    {"_id": 0, "email": 1, "name": 1},
                )
                if therapist and therapist.get("email"):
                    steps_changed = [k for k in ("uploaded", "reviewed", "resolved") if k in update]
                    await _send_email_stub(
                        therapist["email"],
                        f"[Boost Growth] Progress report updated — {client.get('name', '')}",
                        f"Hello {therapist.get('name', '')},\n\nProgress report steps updated: {', '.join(steps_changed)}.\n\n— Boost Growth Portal",
                    )
    doc = await db.progress_reports.find_one({"id": rid}, {"_id": 0})
    return doc or {}

@api.delete("/progress-reports/{rid}")
async def delete_progress_report(rid: str, _=Depends(admin_only)):
    report = await db.progress_reports.find_one({"id": rid}, {"_id": 0, "file_path": 1})
    if report and report.get("file_path"):
        fp = UPLOAD_DIR / report["file_path"]
        if fp.exists():
            fp.unlink()
    await db.progress_reports.delete_one({"id": rid})
    return {"ok": True}


async def _can_access_progress_report(user: dict, report: dict) -> bool:
    if _has_full_client_access(user):
        return True
    client = await db.clients.find_one(
        _active_client_filter({"id": report["client_id"]}),
        {"_id": 0, "main_therapist_id": 1, "co_therapist_ids": 1, "file_no": 1},
    )
    if not client:
        return False
    uid = user["id"]
    if client.get("main_therapist_id") == uid or uid in (client.get("co_therapist_ids") or []):
        return True
    fn = await _client_file_no(report["client_id"])
    return bool(fn and _is_supervisor_for_file(user, fn))


@api.post("/progress-reports/{rid}/file")
async def upload_progress_report_file(rid: str, file: UploadFile = File(...), user=Depends(get_current_user)):
    report = await db.progress_reports.find_one({"id": rid}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    if not await _can_access_progress_report(user, report):
        raise HTTPException(status_code=403, detail="Forbidden")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="File required")
    ext = Path(file.filename).suffix or ".pdf"
    stored = f"pr_{rid}{ext}"
    if report.get("file_path"):
        old = UPLOAD_DIR / report["file_path"]
        if old.exists() and old.name != stored:
            old.unlink()
    content = await file.read()
    file_data = _persist_upload(stored, content)
    await db.progress_reports.update_one({"id": rid}, {"$set": {
        "file_path": stored,
        "file_name": file.filename,
        "file_data": file_data,
        "file_uploaded_at": now_iso(),
    }})
    doc = await db.progress_reports.find_one({"id": rid}, {"_id": 0})
    return _strip_file_data(doc) or {}


@api.get("/progress-reports/{rid}/file")
async def download_progress_report_file(rid: str, user=Depends(get_current_user)):
    report = await db.progress_reports.find_one({"id": rid}, {"_id": 0})
    if not report or not report.get("file_path"):
        raise HTTPException(status_code=404, detail="No file")
    if not await _can_access_progress_report(user, report):
        raise HTTPException(status_code=403, detail="Forbidden")
    content = _load_upload(report.get("file_path"), report.get("file_data"))
    if not content:
        raise HTTPException(status_code=404, detail=FILE_UNAVAILABLE_DETAIL)
    return _bytes_file_response(content, report.get("file_name") or report["file_path"])


@api.delete("/progress-reports/{rid}/file")
async def delete_progress_report_file(rid: str, user=Depends(get_current_user)):
    report = await db.progress_reports.find_one({"id": rid}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    if user.get("role") != "admin" and not await _can_access_progress_report(user, report):
        raise HTTPException(status_code=403, detail="Forbidden")
    if report.get("file_path"):
        fp = UPLOAD_DIR / report["file_path"]
    if fp.exists():
        fp.unlink()
    await db.progress_reports.update_one({"id": rid}, {"$set": {
        "file_path": None, "file_name": None, "file_data": None, "file_uploaded_at": None,
    }})
    return {"ok": True}


APR2026_PROGRESS_REPORTS = [
    {"file_no": "009", "title": "009's Progress Report - Apr 2026",
     "url": "https://docs.google.com/document/d/14c29YPvhWaZirB5Qc-_47qP7Q_04-IOZEhpWk76WiU0/edit"},
    {"file_no": "040", "title": "40's Progress Report - Apr 2026",
     "url": "https://docs.google.com/document/d/1uPUgFPz944AqlHXFXT3oVOpar6JETzsQQ3a6rd3XTK8/edit"},
    {"file_no": "042", "title": "042's Progress Report - Apr 2026",
     "url": "https://docs.google.com/document/d/14tyu4xNlG4AmzALpjYwuWwKflwqk_buX-i2rsovxKRY/edit"},
    {"file_no": "047", "title": "047's Progress Report - Apr 2026",
     "url": "https://drive.google.com/file/d/1eD8w2NQ5WCRtZrODhX33RHQYGT2jbHvM/view"},
    {"file_no": "070", "title": "070's Progress Report - Apr 2026",
     "url": "https://drive.google.com/file/d/1tI5z5vrDDVaApSOAsp4HbkcawcDe42ll/view"},
    {"file_no": "072", "title": "072's Progress Report - Apr 2026",
     "url": "https://docs.google.com/document/d/19UY48orOHqV-ItptNFVxUgRyLgWeWgi8gIy--SmbMHA/edit"},
    {"file_no": "024", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "035", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "038", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "041", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "052", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "054", "title": "54's Progress Report - Apr 2026", "url": ""},
    {"file_no": "011", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "034", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "060", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "061", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "062", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "063", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "064", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "068", "title": "Progress Report - Apr 2026", "url": ""},
    {"file_no": "027", "title": "Progress Report - Apr 2026", "url": ""},
]


async def _find_client_by_file_no(file_no: str) -> Optional[dict]:
    raw = str(file_no or "").strip()
    padded = raw.zfill(3)
    for candidate in {raw, padded, raw.lstrip("0") or raw}:
        hit = await db.clients.find_one(_active_client_filter({"file_no": candidate}), {"_id": 0, "id": 1, "file_no": 1, "name": 1})
        if hit:
            return hit
    return None


@api.post("/admin/seed-progress-reports-apr2026")
async def seed_apr_progress_reports(_=Depends(admin_only)):
    """Idempotently seed April 2026 progress reports with Drive links where available."""
    inserted = 0
    skipped = 0
    missing = []
    report_date = "2026-04-30"
    for item in APR2026_PROGRESS_REPORTS:
        client = await _find_client_by_file_no(item["file_no"])
        if not client:
            missing.append(item["file_no"])
            continue
        cid = client["id"]
        existing = await db.progress_reports.find_one({
            "client_id": cid,
            "title": item["title"],
            "report_date": report_date,
        })
        if existing:
            skipped += 1
            continue
        url = (item.get("url") or "").strip() or None
        await db.progress_reports.insert_one({
            "id": str(uuid.uuid4()),
            "client_id": cid,
            "title": item["title"],
            "report_date": report_date,
            "url": url,
            "notes": "April 2026",
            "uploaded": False,
            "uploaded_by": None,
            "uploaded_at": None,
            "reviewed": False,
            "reviewed_by": None,
            "reviewed_at": None,
            "resolved": False,
            "resolved_by": None,
            "resolved_at": None,
            "file_name": None,
            "file_path": None,
            "file_uploaded_at": None,
            "created_by": "admin-seed",
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })
        inserted += 1
    return {
        "inserted": inserted,
        "skipped": skipped,
        "missing_clients": missing,
        "message": f"Seeded {inserted} progress reports ({skipped} already existed)",
    }


class DeleteClientSessionsIn(BaseModel):
    file_no: str


@api.get("/admin/client-lookup/{file_no}")
async def admin_lookup_client_by_file_no(file_no: str, _=Depends(ops_or_admin)):
    """Preview client name and session/invoice counts before bulk delete."""
    client = await _find_client_by_file_no(file_no)
    if not client:
        raise HTTPException(status_code=404, detail=f"Client file_no {file_no} not found")
    cid = client["id"]
    sessions_count = await db.sessions.count_documents({"client_id": cid})
    invoices_count = await db.invoices.count_documents({"client_id": cid})
    return {
        "client_id": cid,
        "file_no": client.get("file_no"),
        "name": client.get("name"),
        "sessions_count": sessions_count,
        "invoices_count": invoices_count,
    }


@api.post("/admin/delete-client-sessions-invoices")
async def admin_delete_client_sessions_invoices(body: DeleteClientSessionsIn, _=Depends(ops_or_admin)):
    client = await _find_client_by_file_no(body.file_no)
    if not client:
        raise HTTPException(status_code=404, detail=f"Client file_no {body.file_no} not found")
    cid = client["id"]
    s_result = await db.sessions.delete_many({"client_id": cid})
    i_result = await db.invoices.delete_many({"client_id": cid})
    return {
        "client_id": cid,
        "client_name": client.get("name"),
        "file_no": client.get("file_no"),
        "sessions_deleted": s_result.deleted_count,
        "invoices_deleted": i_result.deleted_count,
        "message": f"Deleted {s_result.deleted_count} sessions and {i_result.deleted_count} invoices for {client.get('name')}",
    }


class PurgeTherapistIn(BaseModel):
    name_pattern: str


@api.get("/admin/therapist-search")
async def admin_therapist_search(q: str = "", _=Depends(admin_only)):
    """Find therapists/users matching a name pattern (e.g. naja)."""
    regex = {"$regex": q.strip() or ".", "$options": "i"}
    therapists = await db.therapists.find({"name": regex}, {"_id": 0, "pin_hash": 0, "password_hash": 0}).sort("name", 1).to_list(50)
    users = await db.users.find(
        {"$or": [{"name": regex}, {"username": regex}, {"email": regex}]},
        {"_id": 0, "password_hash": 0, "pin_hash": 0},
    ).to_list(50)
    for t in therapists:
        t["schedule_cells"] = await db.schedule_cells.count_documents({"therapist_id": t["id"]})
    return {"therapists": therapists, "users": users, "query": q}


@api.post("/admin/purge-therapist")
async def admin_purge_therapist(body: PurgeTherapistIn, _=Depends(admin_only)):
    """Delete therapist(s), linked users, and schedule cells by name pattern."""
    pat = body.name_pattern.strip()
    if not pat:
        raise HTTPException(status_code=400, detail="name_pattern required")
    regex = {"$regex": pat, "$options": "i"}
    found = await db.therapists.find({"name": regex}, {"_id": 0}).to_list(50)
    if not found:
        found_users = await db.users.find({"$or": [{"name": regex}, {"username": regex}]}, {"_id": 0}).to_list(50)
        for u in found_users:
            await db.users.delete_one({"id": u["id"]})
        return {"therapists_deleted": 0, "users_deleted": len(found_users), "schedule_cells_deleted": 0, "names": []}
    names, sc_total, u_total = [], 0, 0
    for t in found:
        sc = await db.schedule_cells.delete_many({"therapist_id": t["id"]})
        uc = await db.users.delete_many({"$or": [{"therapist_id": t["id"]}, {"name": t.get("name")}]})
        await db.therapists.delete_one({"id": t["id"]})
        names.append(t.get("name"))
        sc_total += sc.deleted_count
        u_total += uc.deleted_count
    return {
        "therapists_deleted": len(found),
        "users_deleted": u_total,
        "schedule_cells_deleted": sc_total,
        "names": names,
        "message": f"Removed {len(found)} therapist(s): {', '.join(names)}",
    }


class ClearRequestsIn(BaseModel):
    confirm: str


@api.post("/admin/clear-requests")
async def admin_clear_all_requests(body: ClearRequestsIn, _=Depends(admin_only)):
    if body.confirm != "DELETE":
        raise HTTPException(status_code=400, detail='Type "DELETE" to confirm')
    r = await db.requests.delete_many({})
    l = await db.leaves.delete_many({})
    return {
        "requests_deleted": r.deleted_count,
        "leaves_deleted": l.deleted_count,
        "message": f"Deleted {r.deleted_count} requests and {l.deleted_count} leave requests",
    }


# ------------------- Invoices (per client; manual numbers) -------------------
@api.get("/clients/{cid}/invoices")
async def list_invoices(cid: str, service_type: Optional[str] = None, user=Depends(get_current_user)):
    items = await db.invoices.find({"client_id": cid}, {"_id": 0}).to_list(200)
    for inv in items:
        inv["is_closed"] = bool(inv.get("is_closed"))
    items = sorted(items, key=lambda i: (_invoice_num_key(i), (i.get("start_date") or i.get("created_at") or "")))
    if service_type:
        code = _normalize_service_type(service_type)
        if code:
            items = [i for i in items if _normalize_service_type(i.get("service_type")) == code]
    return items


@api.get("/invoices")
async def list_all_invoices(user=Depends(get_current_user)):
    """All invoices for clients the user can access (attendance card week windows)."""
    clients = await db.clients.find(
        _active_client_filter({"status": {"$ne": "Inactive"}}),
        {"_id": 0, "id": 1, "main_therapist_id": 1, "co_therapist_ids": 1},
    ).to_list(500)
    if not _has_full_client_access(user):
        uid = user["id"]
        clients = [
            c for c in clients
            if c.get("main_therapist_id") == uid or uid in (c.get("co_therapist_ids") or [])
        ]
    client_ids = {c["id"] for c in clients}
    if not client_ids:
        return []
    items = await db.invoices.find({"client_id": {"$in": list(client_ids)}}, {"_id": 0}).to_list(5000)
    return items

# ------------------- Billing / payment tracking -------------------
BILLING_REMINDER_EMAILS = frozenset({
    "wabuissa@boostgrowthsa.com",
    "walaa@boostgrowthsa.com",
    "admin@boostgrowthsa.com",
    "jsalmuhaisin@boostgrowthsa.com",
})


async def _billing_reminder_recipients() -> list:
    """Walaa + portal admin accounts (ADMIN_EMAIL env + users with role=admin)."""
    emails = set(BILLING_REMINDER_EMAILS)
    admin = (os.environ.get("ADMIN_EMAIL") or "").strip().lower()
    if admin:
        emails.add(admin)
    admins = await db.users.find({"role": "admin"}, {"_id": 0, "email": 1}).to_list(20)
    for u in admins:
        e = (u.get("email") or "").strip().lower()
        if e:
            emails.add(e)
    return sorted(emails)


def _normalize_payment_status(raw: Optional[str]) -> str:
    s = (raw or "pending").strip().lower()
    if s in ("complete", "paid", "done"):
        return "complete"
    if s == "partial":
        return "partial"
    return "pending"


def _effective_payment_status(inv: dict) -> str:
    status = _normalize_payment_status(inv.get("payment_status"))
    if status == "complete":
        return "complete"
    amount = float(inv.get("amount") or 0)
    paid = float(inv.get("amount_paid") or 0)
    if amount > 0 and paid > 0 and paid < amount:
        return "partial"
    if status == "partial":
        return "partial"
    return "pending"


def _days_between(from_iso: str, to_iso: str) -> int:
    try:
        a = datetime.fromisoformat(str(from_iso)[:10])
        b = datetime.fromisoformat(str(to_iso)[:10])
        return (b - a).days
    except Exception:
        return 0


def _billing_row(inv: dict, client: dict, today: str) -> dict:
    status = _effective_payment_status(inv)
    start = (inv.get("start_date") or inv.get("created_at") or "")[:10]
    amount = float(inv.get("amount") or 0)
    paid = float(inv.get("amount_paid") or 0)
    remaining = round(max(0, amount - paid), 2) if amount > 0 else None
    reminder = (inv.get("next_payment_reminder_at") or "")[:10] or None
    days_unpaid = _days_between(start, today) if status in ("pending", "partial") and start else 0
    days_until_reminder = _days_between(today, reminder) if reminder else None
    return {
        "invoice_id": inv.get("id"),
        "invoice_number": inv.get("invoice_number"),
        "client_id": client.get("id"),
        "client_name": client.get("name"),
        "file_no": client.get("file_no"),
        "service_type": inv.get("service_type"),
        "payment_status": status,
        "amount": amount or None,
        "amount_paid": paid or None,
        "amount_remaining": remaining,
        "start_date": start or None,
        "days_unpaid": days_unpaid,
        "next_payment_reminder_at": reminder,
        "days_until_reminder": days_until_reminder,
        "payment_notes": inv.get("payment_notes"),
        "is_closed": bool(inv.get("is_closed")),
        "package_size": inv.get("package_size"),
    }


async def _process_payment_reminders(force: bool = False) -> dict:
    """Email admins 1–2 days before next_payment_reminder_at on partial invoices."""
    today = now_iso()[:10]
    if not force:
        meta = await db.meta.find_one({"key": "billing_reminders_last_run"})
        if meta and meta.get("date") == today:
            recipients = await _billing_reminder_recipients()
            await _reload_email_settings_from_db()
            return {
                "sent": 0,
                "skipped": True,
                "recipients": recipients,
                "matched": 0,
                "email_results": [],
                "provider_configured": bool(
                    _mailgun_configured() or _brevo_configured() or _resend_configured() or _smtp_configured()
                ),
            }

    partial_invs = await db.invoices.find(
        {"is_closed": {"$ne": True}, "next_payment_reminder_at": {"$exists": True, "$ne": None, "$ne": ""}},
        {"_id": 0},
    ).to_list(500)
    sent = 0
    matched = 0
    recipients = await _billing_reminder_recipients()
    email_results_map: dict = {}
    for inv in partial_invs:
        if _effective_payment_status(inv) != "partial":
            continue
        reminder = (inv.get("next_payment_reminder_at") or "")[:10]
        if not reminder:
            continue
        days_until = _days_between(today, reminder)
        if force:
            # Manual send: today through 2 days ahead, or up to 7 days overdue
            if days_until > 2 or days_until < -7:
                continue
        elif days_until not in (1, 2):
            continue
        last_sent = (inv.get("last_payment_reminder_sent_at") or "")[:10]
        if last_sent == today:
            continue
        client = await db.clients.find_one({"id": inv.get("client_id")}, {"_id": 0, "name": 1, "file_no": 1})
        cname = (client or {}).get("name") or "Client"
        fno = (client or {}).get("file_no") or "—"
        inv_no = inv.get("invoice_number") or inv.get("id", "")[:8]
        amount = float(inv.get("amount") or 0)
        paid = float(inv.get("amount_paid") or 0)
        remaining = round(max(0, amount - paid), 2) if amount else None
        when = "tomorrow" if days_until == 1 else "in 2 days"
        subj = f"Payment reminder · {cname} · {inv_no}"
        body = (
            f"Reminder: follow up with {cname} (File #{fno}) about the next payment installment.\n\n"
            f"Invoice: {inv_no}\n"
            f"Reminder date: {reminder} ({when})\n"
        )
        if remaining is not None:
            body += f"Amount remaining: {remaining} SAR\n"
        if inv.get("payment_notes"):
            body += f"Notes: {inv.get('payment_notes')}\n"
        body += "\n— Boost Growth Staff Portal"
        matched += 1
        for email in recipients:
            r = await _send_email_stub(email, subj, body)
            email_results_map[email] = {
                "to": email,
                "status": r.get("status"),
                "error": r.get("error"),
            }
        await db.invoices.update_one(
            {"id": inv["id"]},
            {"$set": {"last_payment_reminder_sent_at": today}},
        )
        sent += 1

    await db.meta.update_one(
        {"key": "billing_reminders_last_run"},
        {"$set": {"date": today, "sent": sent, "at": now_iso()}},
        upsert=True,
    )
    await _reload_email_settings_from_db()
    return {
        "sent": sent,
        "skipped": False,
        "recipients": recipients,
        "matched": matched,
        "email_results": list(email_results_map.values()),
        "provider_configured": bool(
            _mailgun_configured() or _brevo_configured() or _resend_configured() or _smtp_configured()
        ),
    }


@api.get("/billing/dashboard")
async def billing_dashboard(user=Depends(billing_view_or_ops)):
    """Open invoices needing payment attention — unpaid, partial, reminders."""
    await _process_payment_reminders()
    today = now_iso()[:10]
    clients = await db.clients.find(
        _active_client_filter({"status": {"$ne": "Inactive"}}),
        {"_id": 0},
    ).to_list(500)
    client_map = {c["id"]: c for c in clients}
    client_ids = list(client_map.keys())
    if not client_ids:
        return {"summary": {"unpaid": 0, "partial": 0, "reminders_soon": 0}, "unpaid": [], "partial": [], "items": []}

    invoices = await db.invoices.find(
        {"client_id": {"$in": client_ids}, "is_closed": {"$ne": True}},
        {"_id": 0},
    ).to_list(5000)

    unpaid, partial, items = [], [], []
    reminders_soon = 0
    for inv in invoices:
        client = client_map.get(inv.get("client_id"))
        if not client:
            continue
        row = _billing_row(inv, client, today)
        if row["payment_status"] == "complete":
            continue
        items.append(row)
        if row["payment_status"] == "pending":
            unpaid.append(row)
        elif row["payment_status"] == "partial":
            partial.append(row)
            if row["days_until_reminder"] is not None and 0 <= row["days_until_reminder"] <= 2:
                reminders_soon += 1

    items.sort(key=lambda r: (-(r.get("days_unpaid") or 0), r.get("client_name") or ""))
    # One billing row per client — keep the invoice with the highest INV number
    by_client: Dict[str, dict] = {}
    for row in items:
        cid = row.get("client_id")
        if not cid:
            continue
        inv_num = row.get("invoice_number") or ""
        score = _invoice_num_key({"invoice_number": inv_num})
        prev = by_client.get(cid)
        if not prev or score >= _invoice_num_key({"invoice_number": prev.get("invoice_number") or ""}):
            by_client[cid] = row
    items = list(by_client.values())
    items.sort(key=lambda r: (-(r.get("days_unpaid") or 0), r.get("client_name") or ""))
    unpaid = [r for r in items if r["payment_status"] == "pending"]
    partial = [r for r in items if r["payment_status"] == "partial"]
    unpaid.sort(key=lambda r: -(r.get("days_unpaid") or 0))
    partial.sort(key=lambda r: (r.get("days_until_reminder") if r.get("days_until_reminder") is not None else 999, r.get("client_name") or ""))
    reminders_soon = sum(
        1 for r in partial
        if r.get("days_until_reminder") is not None and 0 <= r.get("days_until_reminder") <= 2
    )

    return {
        "summary": {
            "unpaid": len(unpaid),
            "partial": len(partial),
            "reminders_soon": reminders_soon,
        },
        "unpaid": unpaid,
        "partial": partial,
        "items": items,
    }


@api.put("/invoices/{iid}/payment")
async def update_invoice_payment(iid: str, body: InvoicePaymentIn, user=Depends(ops_or_admin)):
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    patch = {}
    if body.payment_status is not None:
        patch["payment_status"] = _normalize_payment_status(body.payment_status)
    if body.amount is not None:
        patch["amount"] = body.amount
    if body.installment_percent is not None:
        patch["installment_percent"] = body.installment_percent
        amount = float(patch.get("amount", inv.get("amount") or 0))
        if amount > 0 and body.installment_percent:
            patch["amount_paid"] = round(amount * float(body.installment_percent) / 100, 2)
    if body.amount_paid is not None and body.installment_percent is None:
        patch["amount_paid"] = body.amount_paid
    if body.next_payment_reminder_at is not None:
        patch["next_payment_reminder_at"] = body.next_payment_reminder_at or None
    if body.payment_notes is not None:
        patch["payment_notes"] = body.payment_notes
    if not patch:
        return inv
    # Auto partial when paid < total
    amount = float(patch.get("amount", inv.get("amount") or 0))
    paid = float(patch.get("amount_paid", inv.get("amount_paid") or 0))
    if patch.get("payment_status") != "complete" and amount > 0 and paid > 0:
        if paid >= amount:
            patch["payment_status"] = "complete"
        elif patch.get("payment_status") != "pending":
            patch["payment_status"] = "partial"
        elif paid < amount:
            patch["payment_status"] = "partial"
    await db.invoices.update_one({"id": iid}, {"$set": patch})
    updated = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if updated and not updated.get("is_closed"):
        client_patch = {"payment_status": _effective_payment_status(updated)}
        await db.clients.update_one({"id": updated["client_id"]}, {"$set": client_patch})
    return updated


@api.post("/billing/send-reminders")
async def billing_send_reminders(_=Depends(ops_or_admin)):
    """Manual trigger for payment reminder emails (also runs once daily via dashboard)."""
    result = await _process_payment_reminders(force=True)
    return result


def _month_bounds(month: str) -> tuple[str, str]:
    """(startISO, endISO) for YYYY-MM inclusive bounds."""
    import calendar as _cal
    try:
        y, m = [int(x) for x in (month or "").split("-")[:2]]
    except Exception:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")
    if m < 1 or m > 12:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")
    last_day = _cal.monthrange(y, m)[1]
    start = f"{y:04d}-{m:02d}-01"
    end = f"{y:04d}-{m:02d}-{last_day:02d}"
    return start, end


def _client_name_match(label: str, client_name: str) -> bool:
    """Best-effort match schedule cell label to canonical client name."""
    a = (label or "").strip().lower()
    b = (client_name or "").strip().lower()
    if not a or not b:
        return False
    if a == b or a.startswith(b + " "):
        return True
    af = a.split()[0] if a.split() else ""
    bf = b.split()[0] if b.split() else ""
    return len(bf) >= 3 and af == bf


async def _weekly_planned_hours_from_schedule(client: dict, today_iso: str) -> float:
    """Approx weekly planned hours from schedule_cells for current + next week."""
    from datetime import date, timedelta
    cname = (client or {}).get("name") or ""
    if not cname:
        return 0.0
    try:
        d = date.fromisoformat(today_iso[:10])
    except Exception:
        return 0.0
    weeks = [
        _normalize_week_start(d.isoformat()),
        _normalize_week_start((d + timedelta(days=7)).isoformat()),
    ]
    total = 0.0
    weeks_with_data = 0
    for ws in weeks:
        cells = await db.schedule_cells.find({"week_start": ws}, {"_id": 0}).to_list(5000)
        if not cells:
            continue
        week_hours = 0.0
        for cell in cells:
            if cell.get("state") in ("cancel_child", "cancel_therapist", "available"):
                continue
            label = _schedule_cell_child_label(cell)
            if not _client_name_match(label, cname):
                continue
            dur = float(cell.get("duration") or 0)
            if dur > 0:
                week_hours += dur
        if week_hours > 0:
            weeks_with_data += 1
            total += week_hours
    if weeks_with_data == 0:
        return 0.0
    return total / weeks_with_data


async def _weekly_actual_hours_from_sessions(client_id: str, today_iso: str, lookback_days: int = 28) -> float:
    """Avg completed-hours per week over lookback window."""
    from datetime import date, timedelta
    try:
        today = date.fromisoformat(today_iso[:10])
    except Exception:
        return 0.0
    start = (today - timedelta(days=lookback_days)).isoformat()
    sess = await db.sessions.find(
        {"client_id": client_id, "status": "Completed", "session_date": {"$gte": start}},
        {"_id": 0, "hours": 1},
    ).to_list(5000)
    total = sum(float(s.get("hours") or 0) for s in sess)
    weeks = max(1.0, lookback_days / 7.0)
    return total / weeks


async def _invoice_hours_used(inv: dict) -> float:
    """Billable hours for HS invoice (Completed + Cancelled)."""
    if not inv:
        return 0.0
    cid = inv.get("client_id")
    if not cid:
        return 0.0
    sess = await db.sessions.find({"client_id": cid}, {"_id": 0}).to_list(20000)
    used = 0.0
    for s in sess:
        if s.get("status") not in ("Completed", "Cancelled"):
            continue
        if not _session_linked_to_invoice(s, inv):
            continue
        used += float(s.get("hours") or 0)
    return used


@api.get("/billing/invoice-calendar")
async def billing_invoice_calendar(month: str = Query(..., description="YYYY-MM"), user=Depends(billing_view_or_ops)):
    """Forecast invoice end dates + include manual calendar entries."""
    today = now_iso()[:10]
    start, end = _month_bounds(month)
    clients = await db.clients.find(_billing_active_client_filter(), {"_id": 0, "id": 1, "name": 1, "file_no": 1, "package_hours": 1}).to_list(800)
    client_map = {c["id"]: c for c in clients if c.get("id")}
    invoices = await db.invoices.find({"is_closed": {"$ne": True}}, {"_id": 0}).to_list(5000)

    manual = await db.invoice_calendar_manual.find(
        {"date": {"$gte": start, "$lte": end}},
        {"_id": 0},
    ).sort("date", 1).to_list(2000)

    events: list = []
    for m in manual:
        events.append({**m, "kind": "manual"})

    for inv in invoices:
        cid = inv.get("client_id")
        client = client_map.get(cid)
        if not client:
            continue
        st = _normalize_service_type(inv.get("service_type"))
        if st != "HS":
            continue
        predicted = (inv.get("period_to") or "")[:10] or None
        source = "invoice.period_to" if predicted else "forecast"
        pkg = float(inv.get("package_size") or client.get("package_hours") or 24)
        used = await _invoice_hours_used(inv)
        remaining = max(0.0, pkg - used)
        weekly = 0.0
        if not predicted:
            planned = await _weekly_planned_hours_from_schedule(client, today)
            actual = await _weekly_actual_hours_from_sessions(cid, today)
            if planned > 0 and actual > 0:
                weekly = planned * 0.6 + actual * 0.4
            else:
                weekly = planned or actual or 0.0
            if weekly > 0 and remaining > 0:
                from datetime import date, timedelta
                weeks_needed = int((remaining + weekly - 1e-9) // weekly)  # floor
                if remaining > weeks_needed * weekly:
                    weeks_needed += 1
                base = date.fromisoformat(today)
                predicted = (base + timedelta(days=weeks_needed * 7)).isoformat()
            elif remaining <= 0:
                predicted = today
        if not predicted:
            continue
        if predicted < start or predicted > end:
            continue
        events.append({
            "id": f"forecast-{inv.get('id')}",
            "kind": "forecast",
            "date": predicted,
            "due_date": predicted,
            "source": source,
            "client_id": cid,
            "client_name": client.get("name"),
            "file_no": client.get("file_no"),
            "invoice_id": inv.get("id"),
            "invoice_number": inv.get("invoice_number"),
            "package_size": pkg,
            "hours_used": round(used, 2),
            "hours_remaining": round(remaining, 2),
            "weekly_hours": round(weekly, 2) if weekly else None,
        })

    events.sort(key=lambda e: ((e.get("date") or ""), (e.get("client_name") or ""), (e.get("title") or "")))
    return {"month": month, "start": start, "end": end, "events": events}


@api.get("/billing/invoice-calendar/manual")
async def list_invoice_calendar_manual(month: str = Query(..., description="YYYY-MM"), user=Depends(ops_or_admin)):
    start, end = _month_bounds(month)
    items = await db.invoice_calendar_manual.find(
        {"date": {"$gte": start, "$lte": end}},
        {"_id": 0},
    ).sort("date", 1).to_list(2000)
    return items


@api.post("/billing/invoice-calendar/manual")
async def create_invoice_calendar_manual(body: InvoiceCalendarManualIn, user=Depends(ops_or_admin)):
    date_key = (body.date or "")[:10]
    if not date_key or len(date_key) != 10:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
    title = (body.title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="title is required")
    doc = {
        "id": str(uuid.uuid4()),
        "title": title,
        "date": date_key,
        "client_id": body.client_id,
        "invoice_id": body.invoice_id,
        "notes": (body.notes or "").strip() or None,
        "created_at": now_iso(),
        "created_by": user.get("id"),
    }
    await db.invoice_calendar_manual.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/billing/invoice-calendar/manual/{mid}")
async def update_invoice_calendar_manual(mid: str, body: InvoiceCalendarManualIn, user=Depends(ops_or_admin)):
    existing = await db.invoice_calendar_manual.find_one({"id": mid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Manual entry not found")
    date_key = (body.date or existing.get("date") or "")[:10]
    title = (body.title or existing.get("title") or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="title is required")
    patch = {
        "title": title,
        "date": date_key,
        "client_id": body.client_id,
        "invoice_id": body.invoice_id,
        "notes": (body.notes or "").strip() or None,
        "updated_at": now_iso(),
        "updated_by": user.get("id"),
    }
    await db.invoice_calendar_manual.update_one({"id": mid}, {"$set": patch})
    updated = await db.invoice_calendar_manual.find_one({"id": mid}, {"_id": 0})
    return updated


@api.delete("/billing/invoice-calendar/manual/{mid}")
async def delete_invoice_calendar_manual(mid: str, user=Depends(ops_or_admin)):
    await db.invoice_calendar_manual.delete_one({"id": mid})
    return {"ok": True}

# ------------------- Package status (last open invoice) -------------------
def _invoice_num_key(inv: dict) -> int:
    m = _INV_NUM_RE.search((inv.get("invoice_number") or "").strip())
    return int(m.group(1)) if m else 0


def _sort_invoices_by_date(invoices: list) -> list:
    """Newest / highest INV number first (for open-invoice pickers)."""
    return sorted(
        invoices,
        key=lambda i: (_invoice_num_key(i), (i.get("start_date") or i.get("created_at") or "")),
        reverse=True,
    )


def _last_open_invoice(invoices: list, service_code: str) -> Optional[dict]:
    for inv in _sort_invoices_by_date(invoices):
        if inv.get("is_closed"):
            continue
        if _normalize_service_type(inv.get("service_type")) == service_code:
            return inv
    return None


def _sorted_invoices_for_client(client_id: str, invoices: list) -> list:
    client_invs = [i for i in (invoices or []) if i.get("client_id") == client_id]
    return sorted(
        client_invs,
        key=lambda i: (_invoice_num_key(i), (i.get("start_date") or i.get("created_at") or "")),
    )


def _invoice_window_bounds(inv: dict, sorted_client_invoices: list) -> tuple:
    start = (inv.get("start_date") or inv.get("created_at") or "0000-00-00")[:10]
    end = None
    for idx, i in enumerate(sorted_client_invoices):
        if i.get("id") == inv.get("id"):
            if idx + 1 < len(sorted_client_invoices):
                nxt = (sorted_client_invoices[idx + 1].get("start_date") or "")[:10]
                if nxt:
                    end = nxt
            break
    return start, end


def _session_linked_to_invoice(s: dict, inv: dict) -> bool:
    inv_id = inv.get("id")
    inv_num = (inv.get("invoice_number") or "").strip()
    if s.get("invoice_id") == inv_id:
        return True
    if inv_num and (s.get("source_invoice") or "").strip() == inv_num:
        return True
    return False


def _session_has_invoice_link(s: dict) -> bool:
    return bool(s.get("invoice_id") or (s.get("source_invoice") or "").strip())


def _normalize_session_date_iso(raw) -> str:
    """Canonical YYYY-MM-DD for sorting and window checks."""
    if not raw:
        return ""
    s = str(raw).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d = m.group(1), m.group(2).zfill(2), m.group(3).zfill(2)
        return f"{y}-{mo}-{d}"
    m = re.match(r"^(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})", s)
    if m:
        d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        return f"{y}-{mo}-{d}"
    return s[:10]


def _session_date_sort_key(s: dict):
    iso = _normalize_session_date_iso(s.get("session_date"))
    return iso or "9999-99-99"


def _session_in_invoice_date_window(s: dict, inv: dict, sorted_client_invoices: list) -> bool:
    d = _normalize_session_date_iso(s.get("session_date"))
    if not d:
        return False
    start, end = _invoice_window_bounds(inv, sorted_client_invoices)
    start = _normalize_session_date_iso(start) or start
    if d < start:
        return False
    if end:
        end = _normalize_session_date_iso(end) or end
        if d >= end:
            return False
    return True


def _orphan_belongs_to_invoice(s: dict, inv: dict, sorted_client_invoices: list) -> bool:
    if _session_has_invoice_link(s):
        return False
    if not _session_in_invoice_date_window(s, inv, sorted_client_invoices):
        return False
    d = _normalize_session_date_iso(s.get("session_date"))
    for other in sorted_client_invoices:
        if other.get("id") == inv.get("id"):
            continue
        start, end = _invoice_window_bounds(other, sorted_client_invoices)
        if not end:
            continue
        start = _normalize_session_date_iso(start) or start
        end = _normalize_session_date_iso(end) or end
        if d >= start and d < end:
            return False
    return True


def _sessions_for_invoice(inv: dict, sessions: list, client_invoices: list = None) -> list:
    inv_id = inv.get("id")
    inv_num = (inv.get("invoice_number") or "").strip()
    cid = inv.get("client_id")
    sorted_invs = _sorted_invoices_for_client(cid, client_invoices or [])
    out = []
    seen = set()
    for s in sessions:
        if s.get("client_id") != cid:
            continue
        linked_id = s.get("invoice_id")
        if linked_id and linked_id != inv_id:
            continue
        if _session_linked_to_invoice(s, inv):
            sid = s.get("id")
            if sid and sid not in seen:
                out.append(s)
                seen.add(sid)
    if sorted_invs:
        for s in sessions:
            if s.get("client_id") != cid:
                continue
            sid = s.get("id")
            if sid in seen:
                continue
            if _orphan_belongs_to_invoice(s, inv, sorted_invs):
                out.append(s)
                if sid:
                    seen.add(sid)
    out.sort(key=_session_date_sort_key)
    return out


def _client_service_codes(client: dict, invoices: list) -> list:
    """HS/SS codes from client profile and locations — never assume both unless profile says so."""
    codes: set = set()
    cst = _normalize_service_type(client.get("service_type"))
    raw = (client.get("service_type") or "").strip().upper()

    if cst == "HS":
        codes.add("HS")
    elif cst == "SS":
        codes.add("SS")
    elif cst == "AVC":
        return ["AVC"]
    elif raw and ("HS+SS" in raw or "HS/SS" in raw or "+" in raw):
        codes.update({"HS", "SS"})
    elif cst is None and raw:
        if "HOME" in raw:
            codes.add("HS")
        if "SCHOOL" in raw:
            codes.add("SS")

    for loc in client.get("locations") or []:
        lc = _normalize_service_type(loc.get("service"))
        if lc in ("HS", "SS"):
            codes.add(lc)

    if not codes:
        for inv in invoices:
            if not inv.get("is_closed"):
                st = _normalize_service_type(inv.get("service_type"))
                if st in ("HS", "SS"):
                    codes.add(st)

    if not codes:
        codes.add("HS")
    return sorted(codes)


def _is_school_day(d: datetime) -> bool:
    """Sun–Thu are school days (Python weekday: Mon=0 … Sun=6)."""
    return d.weekday() in (6, 0, 1, 2, 3)


def _collect_school_days(from_date: datetime, count: int) -> list:
    out = []
    d = from_date
    guard = 0
    while len(out) < count and guard < 400:
        if _is_school_day(d):
            out.append(d.strftime("%Y-%m-%d"))
        d += timedelta(days=1)
        guard += 1
    return out


def _school_week_windows(anchor_iso: str, total_weeks: int = 4) -> list:
    if not anchor_iso:
        return []
    try:
        anchor = datetime.fromisoformat(str(anchor_iso)[:10])
    except Exception:
        return []
    school_days = _collect_school_days(anchor, int(total_weeks) * 5)
    windows = []
    for k in range(int(total_weeks)):
        chunk = school_days[k * 5:(k + 1) * 5]
        windows.append({
            "week_number": k + 1,
            "dates": chunk,
            "start": chunk[0] if chunk else None,
            "end": chunk[-1] if chunk else None,
        })
    return windows


def _school_week_for_date(date_iso: str, anchor_iso: str, total_weeks: int = 4) -> Optional[int]:
    if not date_iso or not anchor_iso:
        return None
    d = str(date_iso)[:10]
    for w in _school_week_windows(anchor_iso, total_weeks):
        if d in w["dates"]:
            return w["week_number"]
    return None


def _day_name_from_date(date_iso: str) -> str:
    try:
        return datetime.fromisoformat(str(date_iso)[:10]).strftime("%a")
    except Exception:
        return ""


def _school_week_period_ended(end_iso: str) -> bool:
    """True once the last school day (Thu) of the week has passed."""
    if not end_iso:
        return False
    today = now_iso()[:10]
    return today > str(end_iso)[:10]


def _resolve_ss_week_status(
    end_iso: str,
    attended: int,
    school_days: int,
    session_count: int,
    manual_override: Optional[str] = None,
) -> str:
    if manual_override in ("open", "excluded"):
        return "Open"
    if manual_override == "completed":
        return "Completed"
    if session_count <= 0:
        return "Not started"
    if _school_week_period_ended(end_iso):
        return "Completed"
    if attended >= min(5, school_days):
        return "Completed"
    return "In Progress"


def _week_counts_as_done(week_status: str, session_count: int, manual_override: Optional[str]) -> bool:
    if manual_override in ("open", "excluded"):
        return False
    if manual_override == "completed":
        return True
    return week_status == "Completed" and session_count > 0


def _weeks_done_for_invoice(
    sessions: list,
    anchor_iso: str,
    total_weeks: int,
    week_overrides: Optional[dict] = None,
) -> int:
    """Count SS weeks credited: active weeks that ended, or manual overrides."""
    summary = _ss_week_summary_for_invoice(sessions, anchor_iso, total_weeks, week_overrides)
    return sum(1 for w in summary if w.get("countsAsDone"))


def _ss_week_summary_for_invoice(
    sessions: list,
    anchor_iso: str,
    total_weeks: int = 4,
    week_overrides: Optional[dict] = None,
) -> list:
    """Per-week status for attendance cards (matches frontend computeSsWeekSummary)."""
    overrides = week_overrides or {}
    windows = _school_week_windows(anchor_iso, total_weeks)
    completed = [s for s in (sessions or []) if s.get("status") == "Completed" and s.get("session_date")]
    out = []
    for w in windows:
        wnum = w["week_number"]
        dates = w.get("dates") or []
        school_days = len(dates) or 5
        attended = sum(1 for s in completed if str(s["session_date"])[:10] in dates)
        week_sessions = [s for s in (sessions or []) if str(s.get("session_date") or "")[:10] in dates]
        end_iso = w.get("end")
        manual = overrides.get(str(wnum)) or overrides.get(wnum)
        week_status = _resolve_ss_week_status(
            end_iso, attended, school_days, len(week_sessions), manual
        )
        label = "(upcoming)"
        if dates:
            label = f"{dates[0]} - {dates[-1]}"
        override_key = "open" if manual == "excluded" else manual
        out.append({
            "weekNumber": wnum,
            "weekStatus": week_status,
            "label": label,
            "attended": attended,
            "schoolDays": school_days,
            "endISO": end_iso,
            "countsAsDone": _week_counts_as_done(week_status, len(week_sessions), manual),
            "manual": override_key in ("open", "completed"),
            "overrideKey": override_key if override_key in ("open", "completed") else None,
        })
    return out


def _package_status_level_hs(remaining: float, total: float) -> str:
    if total <= 0 or remaining <= 0:
        return "expired"
    pct = (remaining / total) * 100
    if pct <= 15:
        return "critical"
    if pct <= 30:
        return "low"
    return "good"


def _package_status_level_ss_weeks(remaining_weeks: float) -> str:
    if remaining_weeks <= 0:
        return "expired"
    if remaining_weeks <= 1:
        return "critical"
    if remaining_weeks <= 2:
        return "low"
    return "good"


def _package_status_level_ss_sessions(remaining: float, total: float) -> str:
    if total <= 0 or remaining <= 0:
        return "expired"
    pct = (remaining / total) * 100
    if pct < 20:
        return "critical"
    if pct <= 40:
        return "low"
    return "good"


def _display_invoice_for_service(invoices: list, service_code: str) -> Optional[dict]:
    """Open invoice first; otherwise most recent for service (payment badge on attendance list)."""
    inv = _last_open_invoice(invoices, service_code)
    if inv:
        return inv
    typed = [i for i in invoices if _normalize_service_type(i.get("service_type")) == service_code]
    sorted_inv = _sort_invoices_by_date(typed)
    return sorted_inv[0] if sorted_inv else None


def _invoice_payment_fields(inv: Optional[dict]) -> dict:
    if not inv:
        return {"payment_status": "pending", "package_end_date": None}
    return {
        "payment_status": inv.get("payment_status") or "pending",
        "package_end_date": inv.get("period_to"),
    }


def _compute_package_status_row(client: dict, service_code: str, invoices: list, sessions: list) -> dict:
    display_inv = _display_invoice_for_service(invoices, service_code)
    pay_fields = _invoice_payment_fields(display_inv)
    inv = _last_open_invoice(invoices, service_code)
    base = {
        "client_id": client["id"],
        "client_name": client.get("name") or "",
        "file_no": client.get("file_no"),
        "service_type": service_code,
        "invoice_id": None,
        "invoice_number": None,
        "package_size": None,
        "used": 0,
        "remaining": 0,
        "remaining_pct": 0,
        "status": "none",
        "unit": "hours" if service_code == "HS" else "weeks",
        "label": "No open invoice",
        "current_week": None,
        "total_weeks": None,
        **pay_fields,
    }
    if not inv:
        return base

    inv_sessions = _sessions_for_invoice(inv, sessions, invoices)
    pkg = float(inv.get("package_size") or (24 if service_code == "HS" else 4))

    if service_code == "HS":
        used = sum(
            float(s.get("hours") or 0)
            for s in inv_sessions
            if s.get("status") in ("Completed", "Cancelled")
        )
        remaining = max(0, round(pkg - used, 2))
        pct = round((remaining / pkg) * 100, 1) if pkg > 0 else 0
        level = _package_status_level_hs(remaining, pkg)
        if inv.get("is_closed"):
            level = "expired"
        label = f"{int(pkg) if pkg == int(pkg) else pkg}h · {remaining}h left"
        return {
            **base,
            "invoice_id": inv["id"],
            "invoice_number": inv.get("invoice_number"),
            "package_size": pkg,
            "used": round(used, 2),
            "remaining": remaining,
            "remaining_pct": pct,
            "status": level,
            "unit": "hours",
            "label": label,
            **pay_fields,
        }

    # SS — always 4 school weeks per invoice
    total_weeks = 4
    anchor = inv.get("start_date") or client.get("cycle_start_date") or now_iso()[:10]
    week_overrides = inv.get("week_overrides") or {}
    weeks_done = _weeks_done_for_invoice(inv_sessions, anchor, total_weeks, week_overrides)
    remaining_w = max(0, total_weeks - weeks_done)
    current_w = min(total_weeks, weeks_done + 1) if weeks_done < total_weeks else total_weeks
    level = _package_status_level_ss_weeks(remaining_w)
    if inv.get("is_closed"):
        level = "expired"
    if remaining_w <= 1 and remaining_w > 0:
        label = "Last week!"
    else:
        label = f"Wk {current_w} of {total_weeks}"
    week_summary = _ss_week_summary_for_invoice(inv_sessions, anchor, total_weeks, week_overrides)
    return {
        **base,
        "invoice_id": inv["id"],
        "invoice_number": inv.get("invoice_number"),
        "package_size": total_weeks,
        "used": weeks_done,
        "remaining": remaining_w,
        "remaining_pct": round((remaining_w / total_weeks) * 100, 1) if total_weeks else 0,
        "status": level,
        "unit": "weeks",
        "label": label,
        "current_week": current_w,
        "total_weeks": total_weeks,
        "week_summary": week_summary,
        **pay_fields,
    }


def _package_status_for_client(client: dict, invoices: list, sessions: list) -> list:
    client_invs = [i for i in invoices if i.get("client_id") == client["id"]]
    client_sess = [s for s in sessions if s.get("client_id") == client["id"]]
    codes = _client_service_codes(client, client_invs)
    return [_compute_package_status_row(client, code, client_invs, client_sess) for code in codes]


@api.get("/clients/package-status")
async def list_clients_package_status(user=Depends(get_current_user)):
    clients = await db.clients.find(
        _active_client_filter({"status": {"$ne": "Inactive"}}), {"_id": 0}
    ).sort("name", 1).to_list(500)
    if not _has_full_client_access(user):
        uid = user["id"]
        clients = [
            c for c in clients
            if c.get("main_therapist_id") == uid or uid in (c.get("co_therapist_ids") or [])
        ]
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(5000)
    client_ids = {c["id"] for c in clients}
    sessions = await db.sessions.find(
        {"client_id": {"$in": list(client_ids)}}, {"_id": 0}
    ).to_list(20000) if client_ids else []
    rows = []
    for c in clients:
        rows.extend(_package_status_for_client(c, invoices, sessions))
    order = {"critical": 0, "expired": 1, "low": 2, "good": 3, "none": 4}
    rows.sort(key=lambda r: (order.get(r["status"], 9), r.get("client_name") or ""))
    return rows


@api.get("/clients/{cid}/package-status")
async def get_client_package_status(cid: str, user=Depends(get_current_user)):
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if not _has_full_client_access(user):
        uid = user["id"]
        if client.get("main_therapist_id") != uid and uid not in (client.get("co_therapist_ids") or []):
            fn = str(client.get("file_no") or "").strip()
            if not (fn and _is_supervisor_for_file(user, fn)):
                raise HTTPException(status_code=403, detail="Forbidden")
    invoices = await db.invoices.find({"client_id": cid}, {"_id": 0}).to_list(500)
    sessions = await db.sessions.find({"client_id": cid}, {"_id": 0}).to_list(5000)
    return _package_status_for_client(client, invoices, sessions)


@api.post("/clients/{cid}/invoices")
async def create_invoice(cid: str, payload: InvoiceIn, user=Depends(ops_or_admin)):
    inv_id = str(uuid.uuid4())
    st = _normalize_service_type(payload.service_type)
    pkg_size = payload.package_size
    if st == "SS" and (pkg_size is None or pkg_size > 12):
        pkg_size = 4
    doc = {
        "id": inv_id,
        "client_id": cid,
        "invoice_number": payload.invoice_number.strip(),
        "notes": payload.notes,
        "amount": payload.amount,
        "period_from": payload.period_from,
        "period_to": payload.period_to,
        "package_size": pkg_size,
        "payment_status": payload.payment_status or "pending",
        "start_date": payload.start_date or now_iso()[:10],
        "service_type": _normalize_service_type(payload.service_type),
        "is_closed": bool(payload.is_closed) if payload.is_closed is not None else False,
        "close_date": payload.close_date,
        "week_overrides": payload.week_overrides or {},
        "amount_paid": float(payload.amount_paid or 0),
        "next_payment_reminder_at": payload.next_payment_reminder_at,
        "payment_notes": payload.payment_notes,
        "created_by": user["id"],
        "created_at": now_iso(),
    }
    await db.invoices.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/invoices/{iid}/week-overrides")
async def update_invoice_week_overrides(iid: str, body: dict, _=Depends(ops_or_admin)):
    """Manual SS week marks (holiday skip / force complete)."""
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0, "id": 1})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    raw = body.get("week_overrides") if isinstance(body.get("week_overrides"), dict) else {}
    clean = {}
    for k, v in raw.items():
        if v == "excluded":
            clean[str(k)] = "open"
        elif v in ("open", "completed"):
            clean[str(k)] = v
    await db.invoices.update_one({"id": iid}, {"$set": {"week_overrides": clean}})
    return {"ok": True, "week_overrides": clean}

@api.put("/invoices/{iid}")
async def update_invoice(iid: str, payload: InvoiceIn, _=Depends(ops_or_admin)):
    update = {
        "invoice_number": payload.invoice_number.strip(),
        "notes": payload.notes,
        "amount": payload.amount,
        "period_from": payload.period_from,
        "period_to": payload.period_to,
        "package_size": payload.package_size,
        "payment_status": payload.payment_status,
        "start_date": payload.start_date,
        "service_type": _normalize_service_type(payload.service_type) if payload.service_type is not None else None,
        "is_closed": bool(payload.is_closed) if payload.is_closed is not None else False,
        "close_date": payload.close_date,
        "amount_paid": payload.amount_paid,
        "next_payment_reminder_at": payload.next_payment_reminder_at,
        "payment_notes": payload.payment_notes,
    }
    if payload.installment_percent is not None:
        update["installment_percent"] = payload.installment_percent
        amount = float(update.get("amount") if update.get("amount") is not None else 0)
        if amount == 0:
            inv_existing = await db.invoices.find_one({"id": iid}, {"_id": 0, "amount": 1})
            amount = float((inv_existing or {}).get("amount") or 0)
        if amount > 0 and payload.installment_percent:
            update["amount_paid"] = round(amount * float(payload.installment_percent) / 100, 2)
    if payload.week_overrides is not None:
        update["week_overrides"] = payload.week_overrides
    if payload.ss_week_count is not None:
        update["ss_week_count"] = max(4, int(payload.ss_week_count))
    update = {k: v for k, v in update.items() if v is not None or k in (
        "notes", "amount", "period_from", "period_to", "package_size", "service_type",
        "close_date", "amount_paid", "next_payment_reminder_at", "payment_notes",
    )}
    if payload.payment_status is not None:
        update["payment_status"] = _normalize_payment_status(payload.payment_status)
    await db.invoices.update_one({"id": iid}, {"$set": update})
    updated = await db.invoices.find_one({"id": iid}, {"_id": 0})
    # Keep client card payment badge in sync with the active (open) invoice
    if updated and not updated.get("is_closed"):
        client_patch = {}
        if payload.payment_status is not None or payload.amount_paid is not None:
            client_patch["payment_status"] = _effective_payment_status(updated)
        if payload.period_to is not None:
            client_patch["package_end_date"] = payload.period_to
        if client_patch:
            await db.clients.update_one({"id": updated["client_id"]}, {"$set": client_patch})
    return updated

@api.delete("/invoices/{iid}")
async def delete_invoice(iid: str, _=Depends(ops_or_admin)):
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0, "id": 1, "invoice_number": 1})
    if inv:
        inv_num = (inv.get("invoice_number") or "").strip()
        q = {"invoice_id": iid}
        if inv_num:
            q = {"$or": [{"invoice_id": iid}, {"source_invoice": inv_num}]}
        await db.sessions.delete_many(q)
    await db.invoices.delete_one({"id": iid})
    return {"ok": True}

@api.post("/clients/{cid}/invoices/sync-from-excel")
async def sync_invoices_from_excel(cid: str, file: UploadFile = File(...), user=Depends(ops_or_admin)):
    """Detect invoice sheets dynamically by inspecting an uploaded client workbook (.xlsx).
    Imports BOTH invoices (by sheet name) and the session rows inside each sheet.
    Idempotent: matches invoices by invoice_number, sessions by (client_id, session_date, start_time).
    """
    import openpyxl, io
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")
    return await _ingest_workbook_for_client(cid, client, wb, user["id"], origin="excel-sync")


class SyncFromDriveIn(BaseModel):
    drive_url: str


@api.post("/clients/{cid}/invoices/sync-from-drive")
async def sync_invoices_from_drive(cid: str, payload: SyncFromDriveIn, user=Depends(ops_or_admin)):
    """Fetch a Google Sheets document by URL and import all invoices + sessions.

    The sheet MUST be shared as 'Anyone with the link can view'. We hit the
    public xlsx export endpoint (no OAuth needed): 
        https://docs.google.com/spreadsheets/d/{ID}/export?format=xlsx
    """
    from drive_sync import fetch_workbook_from_url, extract_sheet_id
    import re as _re

    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    url = (payload.drive_url or "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="drive_url is required")
    if not extract_sheet_id(url):
        raise HTTPException(status_code=400, detail="Could not extract Google Sheets ID from URL. Make sure it is a /spreadsheets/d/<id>/... link.")
    try:
        wb = fetch_workbook_from_url(url)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not fetch sheet from Drive (make sure 'Anyone with the link' has view access): {e}")
    await db.clients.update_one({"id": cid}, {"$set": {"attendance_sheet_url": url}})
    return await _ingest_workbook_for_client(cid, client, wb, user["id"], origin="drive-sync")


def _merge_drive_client_meta(link_meta: dict, client: dict) -> dict:
    """Build client patch from Drive folder crawl (case summary, intake phone, links)."""
    from drive_sync import (
        fetch_case_summary_content,
        fetch_intake_parent_phone,
        extract_parent_phone_from_text,
        resolve_parent_phone_from_links,
        resolve_client_birth_date,
    )

    patch = {
        "drive_folder_id": link_meta.get("folder_id"),
        "drive_url": link_meta.get("folder_url"),
        "drive_links": link_meta.get("links") or [],
    }
    if link_meta.get("case_summary_url"):
        patch["case_summary_url"] = link_meta["case_summary_url"]
    if link_meta.get("intake_file_url"):
        patch["intake_file_url"] = link_meta["intake_file_url"]
    cs_url = link_meta.get("case_summary_url") or client.get("case_summary_url")
    if cs_url:
        patch["case_summary_url"] = cs_url
        try:
            fetched = fetch_case_summary_content(cs_url)
            if fetched.get("sections"):
                patch["case_summary_sections"] = fetched
            elif link_meta.get("case_summary_url"):
                patch["case_summary_sections"] = fetched
        except Exception as exc:
            logger.warning(f"Case summary fetch failed: {exc}")
    intake_url = link_meta.get("intake_file_url") or client.get("intake_file_url")
    try:
        ph = resolve_parent_phone_from_links(link_meta, client)
        if ph:
            patch["parent_phone"] = ph
    except Exception as exc:
        logger.warning(f"Parent phone resolve failed: {exc}")
    if not patch.get("parent_phone") and intake_url:
        try:
            ph = fetch_intake_parent_phone(intake_url)
            if ph:
                patch["parent_phone"] = ph
        except Exception as exc:
            logger.warning(f"Intake phone fetch failed: {exc}")
    if not patch.get("parent_phone") and cs_url:
        try:
            cs_content = fetch_case_summary_content(cs_url)
            blob_parts: List[str] = []
            for sec in cs_content.get("sections") or []:
                blob_parts.extend(sec.get("paragraphs") or [])
                for tbl in sec.get("tables") or []:
                    for row in tbl:
                        if isinstance(row, list):
                            blob_parts.extend(str(c) for c in row)
            ph = extract_parent_phone_from_text(" ".join(blob_parts))
            if ph:
                patch["parent_phone"] = ph
        except Exception as exc:
            logger.warning(f"Case summary phone fallback failed: {exc}")
    try:
        birth_iso = resolve_client_birth_date(
            case_summary_url=patch.get("case_summary_url") or client.get("case_summary_url"),
            intake_file_url=patch.get("intake_file_url") or client.get("intake_file_url"),
            case_summary_sections=patch.get("case_summary_sections") or client.get("case_summary_sections"),
        )
        if birth_iso:
            patch["birth_date"] = birth_iso
    except Exception as exc:
        logger.warning(f"Birth date resolve failed: {exc}")
    return patch


@api.post("/clients/{cid}/sync-drive-links")
async def sync_client_drive_links(cid: str, user=Depends(client_lead_or_admin)):
    """Crawl the client's Active Clients Drive folder for document links (excludes attendance sheets)."""
    from drive_sync import (
        extract_folder_id,
        fetch_doc_text,
        list_active_client_folders,
        list_client_folder_links,
        parse_case_summary_text,
        ACTIVE_CLIENTS_FOLDER_ID,
    )

    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    folder_id = client.get("drive_folder_id")
    if not folder_id:
        file_no = str(client.get("file_no") or "").strip().zfill(3)
        for entry in list_active_client_folders(ACTIVE_CLIENTS_FOLDER_ID):
            if entry["file_no"] == file_no:
                folder_id = entry["folder_id"]
                break
    if not folder_id:
        folder_id = extract_folder_id(client.get("drive_url") or "")
    if not folder_id:
        raise HTTPException(status_code=400, detail="No Drive folder found for this client")

    meta = list_client_folder_links(folder_id)
    patch = _merge_drive_client_meta({**meta, "folder_id": folder_id}, client)
    sections = patch.get("case_summary_sections")

    await db.clients.update_one({"id": cid}, {"$set": patch})
    return {
        "ok": True,
        **meta,
        "case_summary_sections": sections,
        "parent_phone": patch.get("parent_phone"),
        "intake_file_url": patch.get("intake_file_url") or meta.get("intake_file_url"),
        "case_summary_url": patch.get("case_summary_url") or meta.get("case_summary_url"),
        "links_count": len(meta.get("links") or []),
        "message": (
            f"Synced {len(meta.get('links') or [])} links"
            + (f" · phone {patch['parent_phone']}" if patch.get("parent_phone") else " · no phone found in Drive files")
            + ("" if meta.get("intake_file_url") else " · no intake file detected")
        ),
    }


@api.get("/clients/{cid}/case-summary")
async def get_client_case_summary(cid: str, refresh: bool = False, user=Depends(get_current_user)):
    """Return structured case summary sections; optionally refresh from Google Doc."""
    from drive_sync import fetch_case_summary_content

    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    url = (client.get("case_summary_url") or "").strip()
    cached = client.get("case_summary_sections")

    if cached and cached.get("sections") and not refresh:
        return {"url": url or None, "sections": cached["sections"], "cached": True}

    if refresh and url:
        try:
            sections = fetch_case_summary_content(url)
            await db.clients.update_one({"id": cid}, {"$set": {"case_summary_sections": sections}})
            return {"url": url, "sections": sections.get("sections", []), "cached": False}
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not fetch case summary: {exc}")

    return {"url": url or None, "sections": (cached or {}).get("sections", []) if cached else [], "cached": bool(cached)}


class CaseSummaryUpdateIn(BaseModel):
    case_summary_url: Optional[str] = None
    case_summary_text: Optional[str] = None
    refresh: bool = True


async def _user_can_edit_case_summary(user: dict, client: dict) -> bool:
    if _has_full_client_access(user) or _is_portal_admin(user):
        return True
    tid = await _resolve_user_therapist_id(user)
    if not tid:
        return False
    return client.get("main_therapist_id") == tid or tid in (client.get("co_therapist_ids") or [])


@api.put("/clients/{cid}/case-summary")
async def update_client_case_summary(cid: str, body: CaseSummaryUpdateIn, user=Depends(get_current_user)):
    """Specialists may edit case summary content in-portal; supervisors/admins may edit any client."""
    from drive_sync import fetch_case_summary_content, parse_case_summary_text

    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if not await _user_can_edit_case_summary(user, client):
        raise HTTPException(status_code=403, detail="Case summary edit not allowed")

    # Direct in-portal edit (preferred — no Drive round-trip)
    if body.case_summary_text is not None:
        text = (body.case_summary_text or "").strip()
        fetched = parse_case_summary_text(text) if text else {"sections": [], "raw_preview": ""}
        patch = {
            "case_summary_sections": fetched,
            "case_summary_updated_at": now_iso(),
            "case_summary_updated_by": user.get("id"),
            "case_summary_source": "portal",
        }
        await db.clients.update_one({"id": cid}, {"$set": patch})
        url = (client.get("case_summary_url") or "").strip() or None
        return {"url": url, "sections": fetched.get("sections", []), "cached": False}

    patch: dict = {}
    if body.case_summary_url is not None:
        patch["case_summary_url"] = (body.case_summary_url or "").strip() or None
    if patch:
        await db.clients.update_one({"id": cid}, {"$set": patch})
        client = {**client, **patch}

    url = (client.get("case_summary_url") or "").strip()
    sections = client.get("case_summary_sections")
    if body.refresh and url:
        try:
            fetched = fetch_case_summary_content(url)
            await db.clients.update_one({"id": cid}, {"$set": {"case_summary_sections": fetched}})
            return {"url": url, "sections": fetched.get("sections", []), "cached": False}
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not fetch case summary: {exc}")

    if sections and sections.get("sections"):
        return {"url": url or None, "sections": sections["sections"], "cached": True}
    return {"url": url or None, "sections": [], "cached": False}


class CaseSummaryRemindIn(BaseModel):
    message: Optional[str] = None


@api.post("/clients/{cid}/case-summary/remind")
async def remind_case_summary_update(
    cid: str,
    body: Optional[CaseSummaryRemindIn] = None,
    user=Depends(get_current_user),
):
    """Email the main therapist to update the case summary (supervisors / ops)."""
    if not (_has_full_client_access(user) or _is_jenan(user) or _is_portal_admin(user) or _is_hr_ops(user)):
        raise HTTPException(status_code=403, detail="Reminder not allowed")
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    main_tid = client.get("main_therapist_id")
    if not main_tid:
        raise HTTPException(status_code=400, detail="No main therapist assigned")
    therapist = await db.therapists.find_one({"id": main_tid}, {"_id": 0, "name": 1, "email": 1})
    if not therapist or not (therapist.get("email") or "").strip():
        raise HTTPException(status_code=400, detail="Therapist has no email on file")
    to = therapist["email"].strip()
    cname = client.get("name") or "Client"
    fno = client.get("file_no") or "—"
    sender = user.get("name") or user.get("email") or "Supervisor"
    subj = f"Please update case summary · {cname}"
    custom = ((body.message if body else None) or "").strip()
    body_text = (
        f"Hello {therapist.get('name') or 'there'},\n\n"
        f"Please review and update the case summary for {cname} (File #{fno}) in the staff portal.\n\n"
    )
    if custom:
        body_text += f"Message from {sender}:\n{custom}\n\n"
    body_text += f"Requested by: {sender}\n\n— Boost Growth Staff Portal"
    result = await _send_email_stub(to, subj, body_text)
    return {"ok": True, "to": to, "email_status": result.get("status"), "error": result.get("error")}


async def _user_can_access_client_records(user: dict, client: dict) -> bool:
    if _has_full_client_access(user) or _is_portal_admin(user) or _is_hr_ops(user):
        return True
    return await _user_can_edit_client_records(user, client)


@api.post("/clients/{cid}/records/upload")
async def upload_client_record(
    cid: str,
    file: UploadFile = File(...),
    title: str = Form(""),
    user=Depends(get_current_user),
):
    """Upload a session file to the client's records (therapist with caseload access or ops)."""
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if not await _user_can_access_client_records(user, client):
        raise HTTPException(status_code=403, detail="Forbidden")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="File required")
    content = await file.read()
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 15MB)")
    fid = str(uuid.uuid4())
    ext = Path(file.filename).suffix.lower() or ""
    stored = f"client_record_{cid}_{fid}{ext}"
    file_data = _persist_upload(stored, content)
    entry = {
        "id": fid,
        "title": (title or "").strip() or file.filename,
        "file_name": file.filename,
        "stored": stored,
        "file_data": file_data,
        "uploaded_at": now_iso(),
        "uploaded_by": user.get("id"),
    }
    await db.client_record_files.update_one(
        {"client_id": cid, "id": fid},
        {"$set": {"client_id": cid, **entry}},
        upsert=True,
    )
    meta = {k: v for k, v in entry.items() if k != "file_data"}
    records = list(client.get("record_files") or [])
    records.append(meta)
    await db.clients.update_one({"id": cid}, {"$set": {"record_files": records}})
    return meta


@api.get("/clients/{cid}/records/{fid}")
async def download_client_record(cid: str, fid: str, user=Depends(get_current_user)):
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if not await _user_can_access_client_records(user, client):
        raise HTTPException(status_code=403, detail="Forbidden")
    meta = next((r for r in (client.get("record_files") or []) if r.get("id") == fid), None)
    if not meta:
        raise HTTPException(status_code=404, detail="Record file not found")
    stored_doc = await db.client_record_files.find_one({"client_id": cid, "id": fid}, {"_id": 0})
    content = _load_upload(
        meta.get("stored") or (stored_doc or {}).get("stored"),
        (stored_doc or {}).get("file_data"),
    )
    if not content:
        raise HTTPException(status_code=404, detail=FILE_UNAVAILABLE_DETAIL)
    return _bytes_file_response(content, meta.get("file_name") or meta.get("stored") or "file")


class SyncActiveClientsIn(BaseModel):
    folder_url: Optional[str] = None
    file_nos: Optional[List[str]] = None
    dry_run: bool = False
    ensure_missing_clients: bool = True


_MASTER_BY_FILE_NO = {r[0]: r for r in MASTER_CLIENTS}


async def _ensure_client_for_drive_folder(file_no: str, folder_title: str = "") -> Optional[dict]:
    """Create a portal client for a Drive folder when missing (uses MASTER_CLIENTS or folder title)."""
    existing = await _find_client_by_file_no(file_no)
    if existing:
        return await db.clients.find_one(_active_client_filter({"id": existing["id"]}), {"_id": 0})

    fn = str(file_no or "").strip().zfill(3)
    row = _MASTER_BY_FILE_NO.get(fn)
    name = row[1] if row else ""
    if not name and folder_title:
        parts = str(folder_title).split("|", 1)
        if len(parts) > 1:
            name = parts[1].strip()
    if not name:
        return None

    key_to_id: dict = {}
    for t in await db.therapists.find({}, {"_id": 0, "id": 1, "key": 1}).to_list(200):
        if t.get("key"):
            key_to_id[t["key"]] = t["id"]

    main_id = None
    co_ids: List[str] = []
    sup_name = None
    pkg = 24
    service = None
    address = None
    if row:
        main_id = key_to_id.get(row[2])
        co_ids = [key_to_id[k] for k in row[3] if k in key_to_id]
        sup_k = row[5]
        if sup_k in key_to_id:
            tdoc = await db.therapists.find_one({"id": key_to_id[sup_k]}, {"_id": 0, "name": 1})
            sup_name = tdoc.get("name") if tdoc else None
        pkg, service, address = row[4], row[6], row[7]

    cid = str(uuid.uuid4())
    doc = {
        "id": cid,
        "file_no": fn,
        "name": name,
        "package_hours": pkg,
        "service_type": service,
        "address": address,
        "main_therapist_id": main_id,
        "co_therapist_ids": co_ids,
        "supervisor": sup_name,
        "color": "#7A8A6A",
        "billing_mode": "hours",
        "payment_status": "pending",
        "status": "Active",
        "created_at": now_iso(),
    }
    await db.clients.insert_one(doc)
    logger.info("Created missing client #%s %s from Drive folder", fn, name)
    return doc


def _arabic_drive_child_line(r: dict) -> str:
    fn = r.get("file_no") or "?"
    name = (r.get("client_name") or "").strip()
    label = f"#{fn}" + (f" {name}" if name else "")
    st = r.get("status")
    if st == "synced":
        added = len(r.get("invoices_added") or [])
        updated = len(r.get("invoices_updated") or [])
        sess = int(r.get("sessions_added") or 0)
        warn = r.get("warning")
        base = f"{label}: ✓ {added} فاتورة جديدة"
        if updated:
            base += f" · {updated} محدّثة"
        base += f" · {sess} جلسة"
        if warn:
            base += f" ⚠ {warn[:48]}"
        return base
    if st == "meta_synced":
        return f"{label}: روابط Drive فقط (لا Attendance Sheet)"
    if st == "skipped":
        return f"{label}: تخطّي — {r.get('reason') or 'غير معروف'}"
    if st == "error":
        err = (r.get("error") or "خطأ")[:72]
        return f"{label}: ❌ {err}"
    if st == "dry_run":
        return f"{label}: معاينة — {'يوجد Sheet' if r.get('sheet_url') else 'بدون Sheet'}"
    if st == "created":
        return f"{label}: ✓ أُنشئ في البوابة ثم {r.get('follow_up') or 'بانتظار المزامنة'}"
    return f"{label}: {st or '؟'}"


def _build_drive_sync_arabic_report(results: List[dict], *, totals: dict) -> str:
    lines: List[str] = []
    lines.append(
        f"المجموع: {totals.get('synced', 0)} مزامنة · "
        f"{totals.get('meta_synced', 0)} روابط فقط · "
        f"{totals.get('skipped', 0)} تخطّى · "
        f"{totals.get('errors', 0)} أخطاء"
    )
    for r in results:
        lines.append(_arabic_drive_child_line(r))
    return "\n".join(lines)


async def _import_clients_from_google_sheet_url(
    sheet_url: str,
    *,
    replace_missing: bool = False,
) -> dict:
    """Download Master Sheet (Active Clients tab) and upsert client records."""
    import httpx

    export_url = _google_sheet_export_url(sheet_url)
    async with httpx.AsyncClient(follow_redirects=True, timeout=120.0) as client:
        resp = await client.get(export_url)
    if resp.status_code != 200:
        raise HTTPException(
            status_code=400,
            detail=f"Could not download clients sheet (HTTP {resp.status_code})",
        )
    rows = _read_clients_import_rows(resp.content, "active_clients.xlsx")
    return await _import_clients_from_rows(rows, replace_missing)


async def _import_schedule_from_google_sheet_url(
    sheet_url: str,
    week_start: str,
    *,
    clear_existing: bool = True,
    sheet_name: Optional[str] = None,
) -> dict:
    """Import one schedule week from a public Google Sheets workbook."""
    import httpx

    week_start = _normalize_week_start(week_start)
    export_url = _google_sheet_export_url(sheet_url)
    async with httpx.AsyncClient(follow_redirects=True, timeout=120.0) as client:
        resp = await client.get(export_url)
    if resp.status_code != 200:
        raise HTTPException(
            status_code=400,
            detail=f"Could not download schedule sheet (HTTP {resp.status_code})",
        )
    content = resp.content
    grid, merge_anchors, merge_skip, used_sheet, sheet_names, fill_states = _load_schedule_xlsx_bytes(
        content, sheet_name
    )
    if not sheet_name:
        picked = _pick_sheet_for_week(sheet_names, week_start)
        if picked and picked != used_sheet:
            grid, merge_anchors, merge_skip, used_sheet, _, fill_states = _load_schedule_xlsx_bytes(
                content, picked
            )
    week_start, week_warning = _resolve_import_week_start(week_start, used_sheet)
    therapists = await db.therapists.find(
        {}, {"_id": 0, "pin_hash": 0, "password_hash": 0}
    ).to_list(100)
    t_by_name = _build_schedule_therapist_name_map(therapists)
    inserted, skipped = await _import_schedule_grid(
        grid,
        week_start,
        t_by_name,
        clear_existing,
        merge_anchors=merge_anchors,
        merge_skip=merge_skip,
        cell_fill_states=fill_states,
    )
    await _relink_prep_markers_after_schedule_import(week_start)
    return {
        "cells_inserted": inserted,
        "week_start": week_start,
        "skipped_therapists": skipped[:20],
        "sheet_used": used_sheet,
        "merge_spans_detected": len(merge_anchors),
        "week_start_warning": week_warning,
        "prep_relinked": True,
    }


async def _bulk_sync_active_clients_from_drive(
    *,
    folder_url: Optional[str] = None,
    file_nos: Optional[List[str]] = None,
    dry_run: bool = False,
    ensure_missing_clients: bool = True,
    user_id: str = "drive-sync",
) -> dict:
    """Bulk-sync attendance workbooks from the Active Clients Drive folder."""
    from drive_sync import (
        ACTIVE_CLIENTS_FOLDER_ID,
        extract_folder_id,
        fetch_workbook_from_url,
        list_active_client_folders,
        list_client_folder_links,
        resolve_attendance_sheet_url,
    )

    parent_id = extract_folder_id(folder_url or "") or ACTIVE_CLIENTS_FOLDER_ID
    folders = list_active_client_folders(parent_id)
    if file_nos:
        wanted = {str(x).strip().zfill(3) for x in file_nos if str(x).strip()}
        folders = [f for f in folders if f["file_no"] in wanted]

    results: List[dict] = []
    for entry in folders:
        file_no = entry["file_no"]
        client = await _find_client_by_file_no(file_no)
        if not client and ensure_missing_clients and not dry_run:
            created = await _ensure_client_for_drive_folder(file_no, entry.get("title") or "")
            if created:
                client = created
                logger.info("Auto-created portal client #%s from Drive folder", file_no)
        if not client:
            results.append({"file_no": file_no, "status": "skipped", "reason": "client not in portal"})
            continue
        try:
            link_meta = list_client_folder_links(entry["folder_id"])
            meta_patch = _merge_drive_client_meta({**link_meta, "folder_id": entry["folder_id"]}, client)
            client_patch = {
                **{k: v for k, v in meta_patch.items() if k != "drive_folder_id"},
                "drive_url": entry.get("folder_url") or link_meta.get("folder_url"),
                "drive_folder_id": entry["folder_id"],
            }
            sheet_url = resolve_attendance_sheet_url(entry["folder_id"])
            if dry_run:
                results.append({
                    "file_no": file_no,
                    "client_name": client.get("name"),
                    "status": "dry_run",
                    "sheet_url": sheet_url,
                    "drive_links": len(link_meta.get("links") or []),
                    "case_summary_url": link_meta.get("case_summary_url"),
                    "parent_phone": meta_patch.get("parent_phone"),
                })
                continue
            if sheet_url:
                client_patch["attendance_sheet_url"] = sheet_url
            await db.clients.update_one({"id": client["id"]}, {"$set": client_patch})
            if not sheet_url:
                results.append({
                    "file_no": file_no,
                    "client_name": client.get("name"),
                    "status": "meta_synced",
                    "parent_phone": meta_patch.get("parent_phone"),
                    "reason": "drive links/phone updated; no attendance spreadsheet in folder",
                })
                continue
            wb = fetch_workbook_from_url(sheet_url)
            ingest = await _ingest_workbook_for_client(
                client["id"], client, wb, user_id, origin="drive-bulk-sync"
            )
            row = {
                "file_no": file_no,
                "client_name": client.get("name"),
                "status": "synced",
                "sheet_url": sheet_url,
                "parent_phone": meta_patch.get("parent_phone"),
                "drive_links": len(link_meta.get("links") or []),
                **ingest,
            }
            results.append(row)
            logger.info(
                "Drive sync #%s %s: +%s inv, %s sessions",
                file_no,
                client.get("name"),
                len(ingest.get("invoices_added") or []),
                ingest.get("sessions_added", 0),
            )
        except Exception as exc:
            results.append({
                "file_no": file_no,
                "client_name": client.get("name"),
                "status": "error",
                "error": str(exc),
            })

    synced = sum(1 for r in results if r.get("status") == "synced")
    meta_synced = sum(1 for r in results if r.get("status") == "meta_synced")
    skipped = sum(1 for r in results if r.get("status") == "skipped")
    errors = sum(1 for r in results if r.get("status") == "error")
    totals = {
        "synced": synced,
        "meta_synced": meta_synced,
        "skipped": skipped,
        "errors": errors,
    }
    sessions_total = sum(int(r.get("sessions_added") or 0) for r in results if r.get("status") == "synced")
    return {
        "ok": True,
        "parent_folder_id": parent_id,
        "total_folders": len(folders),
        **totals,
        "sessions_total": sessions_total,
        "message": (
            f"{synced} attendance synced · {meta_synced} drive-only (phones/links) · "
            f"{skipped} skipped · {errors} errors"
        ),
        "summary_ar": _build_drive_sync_arabic_report(results, totals=totals),
        "results": results,
    }


def _arabic_full_restore_summary(results: dict) -> str:
    parts: List[str] = []
    imp = results.get("clients") or {}
    if imp:
        parts.append(
            f"أطفال: {imp.get('created', 0)} جديد · {imp.get('updated', 0)} محدّث"
        )
    drv = results.get("drive") or {}
    if drv:
        parts.append(
            f"Drive: {drv.get('synced', 0)} مزامنة · "
            f"{drv.get('sessions_total', 0)} جلسة · {drv.get('errors', 0)} أخطاء"
        )
        if drv.get("summary_ar"):
            parts.append(drv["summary_ar"])
    sch = results.get("schedule") or {}
    if sch:
        parts.append(
            f"جدول {sch.get('week_start', '')}: {sch.get('cells_inserted', 0)} خلية"
        )
    rec = results.get("recover") or {}
    if rec.get("summary_ar"):
        parts.append(rec["summary_ar"])
    return " · ".join(parts) if parts else "اكتملت الاستعادة"


async def _run_full_restore_from_drive(
    *,
    folder_url: Optional[str] = None,
    week_start: Optional[str] = None,
    clients_sheet_url: Optional[str] = None,
    schedule_sheet_url: Optional[str] = None,
    skip_clients: bool = False,
    skip_drive: bool = False,
    skip_schedule: bool = False,
    skip_recover: bool = False,
    dry_run: bool = False,
    user_id: str = "full-restore",
) -> dict:
    """One-shot portal recovery: clients → Drive invoices/sessions → schedule → prep relink."""
    week = _normalize_week_start(week_start or TRIAL_WEEK_START)
    results: dict = {"ok": True, "week_start": week, "dry_run": dry_run}

    if not skip_clients:
        url = (clients_sheet_url or MASTER_CLIENTS_SHEET_URL).strip()
        if dry_run:
            results["clients"] = {"status": "dry_run", "sheet_url": url}
        else:
            results["clients"] = await _import_clients_from_google_sheet_url(url, replace_missing=False)

    if not skip_drive:
        results["drive"] = await _bulk_sync_active_clients_from_drive(
            folder_url=folder_url,
            dry_run=dry_run,
            user_id=user_id,
        )

    if not skip_schedule:
        url = (schedule_sheet_url or SCHEDULE_MASTER_SHEET_URL).strip()
        if dry_run:
            results["schedule"] = {"status": "dry_run", "sheet_url": url, "week_start": week}
        else:
            results["schedule"] = await _import_schedule_from_google_sheet_url(
                url, week, clear_existing=True
            )

    if not skip_recover and not dry_run:
        results["recover"] = await _run_auto_recover(store_backup=True)
        results["health_after"] = results["recover"].get("health_after")

    results["summary_ar"] = _arabic_full_restore_summary(results)
    return results


class FullRestoreFromDriveIn(BaseModel):
    folder_url: Optional[str] = None
    week_start: Optional[str] = None
    clients_sheet_url: Optional[str] = None
    schedule_sheet_url: Optional[str] = None
    skip_clients: bool = False
    skip_drive: bool = False
    skip_schedule: bool = False
    skip_recover: bool = False
    dry_run: bool = False


@api.post("/admin/full-restore-from-drive")
async def admin_full_restore_from_drive(
    body: FullRestoreFromDriveIn = FullRestoreFromDriveIn(),
    user=Depends(admin_only),
):
    """Full portal recovery from Google Drive + Master Sheet (one-click Admin)."""
    return await _run_full_restore_from_drive(
        folder_url=body.folder_url,
        week_start=body.week_start,
        clients_sheet_url=body.clients_sheet_url,
        schedule_sheet_url=body.schedule_sheet_url,
        skip_clients=body.skip_clients,
        skip_drive=body.skip_drive,
        skip_schedule=body.skip_schedule,
        skip_recover=body.skip_recover,
        dry_run=body.dry_run,
        user_id=user["id"],
    )


@api.post("/admin/sync-active-clients-from-drive")
async def sync_active_clients_from_drive(body: SyncActiveClientsIn, user=Depends(client_lead_or_admin)):
    """Bulk-sync attendance workbooks from the Active Clients Drive folder.

    Each subfolder is named like ``009 | Child Name``. Inside, we locate the
    Attendance Sheets subfolder and import the newest attendance spreadsheet
    (prefers tabs whose title contains the latest year, e.g. 2026).
    """
    return await _bulk_sync_active_clients_from_drive(
        folder_url=body.folder_url,
        file_nos=body.file_nos,
        dry_run=body.dry_run,
        ensure_missing_clients=body.ensure_missing_clients,
        user_id=user["id"],
    )


@api.get("/admin/drive-client-folders")
async def list_drive_client_folders(folder_url: Optional[str] = None, _=Depends(admin_only)):
    """List child folders in the Active Clients Drive directory (file_no + title)."""
    from drive_sync import ACTIVE_CLIENTS_FOLDER_ID, extract_folder_id, list_active_client_folders

    parent_id = extract_folder_id(folder_url or "") or ACTIVE_CLIENTS_FOLDER_ID
    folders = list_active_client_folders(parent_id)
    return {"parent_folder_id": parent_id, "total": len(folders), "folders": folders}


class SyncBirthDatesIn(BaseModel):
    file_nos: Optional[List[str]] = None
    dry_run: bool = False
    overwrite: bool = False


@api.post("/admin/sync-birth-dates-from-drive")
async def sync_birth_dates_from_drive(body: SyncBirthDatesIn, user=Depends(client_lead_or_admin)):
    """Bulk-sync child birth dates from case summaries and intake forms on Drive."""
    from drive_sync import resolve_client_birth_date

    query = _active_client_filter({"status": {"$ne": "Inactive"}})
    if body.file_nos:
        file_nos = [str(f).strip().zfill(3) for f in body.file_nos if str(f).strip()]
        query["file_no"] = {"$in": file_nos}
    clients = await db.clients.find(query, {"_id": 0}).sort("file_no", 1).to_list(500)

    updated = 0
    skipped = 0
    results: List[dict] = []
    for client in clients:
        name = client.get("name") or ""
        file_no = client.get("file_no") or ""
        existing = (client.get("birth_date") or "").strip()
        if existing and not body.overwrite:
            skipped += 1
            results.append({
                "file_no": file_no, "name": name, "status": "skipped",
                "birth_date": existing, "reason": "already set",
            })
            continue
        try:
            birth_iso = resolve_client_birth_date(
                case_summary_url=client.get("case_summary_url"),
                intake_file_url=client.get("intake_file_url"),
                case_summary_sections=client.get("case_summary_sections"),
            )
        except Exception as exc:
            results.append({
                "file_no": file_no, "name": name, "status": "error", "error": str(exc),
            })
            continue
        if not birth_iso:
            if body.overwrite and existing:
                if not body.dry_run:
                    await db.clients.update_one({"id": client["id"]}, {"$unset": {"birth_date": ""}})
                updated += 1
                results.append({
                    "file_no": file_no, "name": name,
                    "status": "cleared" if not body.dry_run else "would_clear",
                    "reason": "no birth date found in Drive files",
                })
            else:
                skipped += 1
                results.append({
                    "file_no": file_no, "name": name, "status": "skipped",
                    "reason": "no birth date found in Drive files",
                })
            continue
        if birth_iso == existing:
            skipped += 1
            results.append({
                "file_no": file_no, "name": name, "status": "skipped",
                "birth_date": birth_iso, "reason": "unchanged",
            })
            continue
        if not body.dry_run:
            await db.clients.update_one({"id": client["id"]}, {"$set": {"birth_date": birth_iso}})
        updated += 1
        results.append({
            "file_no": file_no, "name": name, "status": "updated" if not body.dry_run else "would_update",
            "birth_date": birth_iso,
        })

    return {
        "ok": True,
        "dry_run": body.dry_run,
        "total": len(clients),
        "updated": updated,
        "skipped": skipped,
        "message": (
            f"{'Would update' if body.dry_run else 'Updated'} {updated} client(s) · {skipped} skipped"
        ),
        "results": results,
    }


# ---------- Workbook parser shared by both sync endpoints ----------
import re as _re_top  # for module load

# ---------- Invoice service type (HS / SS) ----------
def _normalize_service_type(val: Optional[str]) -> Optional[str]:
    """Return canonical 'HS', 'SS', or 'AVC', or None if unknown."""
    if not val:
        return None
    s = str(val).strip().lower()
    if s in ("hs", "home session", "home", "home sessions"):
        return "HS"
    if s in ("ss", "school support", "school"):
        return "SS"
    if s in ("avc",):
        return "AVC"
    if "school support" in s:
        return "SS"
    if "home session" in s or "home sessions" in s:
        return "HS"
    if "ss" in s and "hs" not in s:
        return "SS"
    if "hs" in s and "ss" not in s:
        return "HS"
    if "avc" in s:
        return "AVC"
    return None


def _service_type_from_label(text: str) -> Optional[str]:
    """Parse 'Service: Home Sessions' / 'School Support' / 'HS' etc."""
    if not text:
        return None
    raw = str(text).strip()
    s = raw.lower()
    if s.startswith("service:"):
        s = s.split(":", 1)[1].strip()
    if s in ("hs", "ss", "avc"):
        return s.upper()
    if "school" in s:
        return "SS"
    if "home" in s:
        return "HS"
    if "avc" in s:
        return "AVC"
    return _normalize_service_type(raw)


def _session_blob_service_hint(session: dict) -> Optional[str]:
    """Detect HS| or SS| prefix in status / note / location."""
    for field in ("status", "note", "location"):
        raw = (session.get(field) or "").strip()
        if not raw:
            continue
        m = _re_top.match(r"^(HS|SS)\s*\|", raw, _re_top.IGNORECASE)
        if m:
            return m.group(1).upper()
        low = raw.lower()
        if low.startswith("hs ") or low.startswith("hs|"):
            return "HS"
        if low.startswith("ss ") or low.startswith("ss|"):
            return "SS"
    return None


async def _infer_invoice_service_type(inv_id: str, client_id: str, inv_num: str, client: Optional[dict] = None) -> Optional[str]:
    inv_num = (inv_num or "").strip()
    q: dict = {"client_id": client_id, "$or": [{"invoice_id": inv_id}]}
    if inv_num:
        q["$or"].append({"source_invoice": inv_num})
    sessions = await db.sessions.find(q, {"_id": 0, "status": 1, "note": 1, "location": 1}).to_list(1000)
    hs = ss = 0
    for s in sessions:
        hint = _session_blob_service_hint(s)
        if hint == "HS":
            hs += 1
        elif hint == "SS":
            ss += 1
    if ss > hs:
        return "SS"
    if hs > ss:
        return "HS"
    if client:
        cst = _normalize_service_type(client.get("service_type"))
        if cst in ("HS", "SS"):
            return cst
    return None


async def _migrate_invoice_service_types() -> int:
    """Re-infer HS/SS/AVC on invoices from sessions; client profile only when type is missing."""
    migrated = 0
    client_cache: dict = {}
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(5000)
    for inv in invoices:
        cid = inv.get("client_id") or ""
        if cid not in client_cache:
            client_cache[cid] = await db.clients.find_one(
                {"id": cid}, {"_id": 0, "service_type": 1}
            )
        client = client_cache.get(cid)

        inferred = await _infer_invoice_service_type(
            inv["id"], cid, inv.get("invoice_number", ""), client=None
        )
        if not inferred and not inv.get("service_type") and client:
            cst = _normalize_service_type(client.get("service_type"))
            if cst in ("HS", "SS", "AVC"):
                inferred = cst

        if not inferred:
            normalized = _normalize_service_type(inv.get("service_type"))
            if normalized and inv.get("service_type") != normalized:
                await db.invoices.update_one(
                    {"id": inv["id"]}, {"$set": {"service_type": normalized}}
                )
                migrated += 1
            continue

        if inv.get("service_type") != inferred:
            await db.invoices.update_one(
                {"id": inv["id"]}, {"$set": {"service_type": inferred}}
            )
            migrated += 1
    return migrated


THERAPIST_EMAIL_MIGRATIONS = [
    ("ralshatri@boostgrowthsa.com", "ralshatery@boostgrowthsa.com"),
    ("salhammamy@boostgrowthsa.com", "shalhammami@boostgrowthsa.com"),
    ("aalroman@boostgrowthsa.com", "a.alromman@boostgrowthsa.com"),
    ("salamri@boostgrowthsa.com", "shalamri@boostgrowthsa.com"),
    ("aalshreef@boostgrowthsa.com", "a.alshareef@boostgrowthsa.com"),
    ("naja@boostgrowthsa.com", "nalhamad@boostgrowthsa.com"),
]


async def _partial_payment_client_ids() -> tuple[list[str], list[dict]]:
    """Resolve client ids for PARTIAL_PAYMENT_CLIENT_FILE_NOS (e.g. Fahad Alyahya #011)."""
    ids: list[str] = []
    rows: list[dict] = []
    for fn in sorted(PARTIAL_PAYMENT_CLIENT_FILE_NOS):
        c = await db.clients.find_one(
            {"file_no": fn},
            {"_id": 0, "id": 1, "name": 1, "file_no": 1},
        )
        if c:
            ids.append(c["id"])
            rows.append(c)
    return ids, rows


async def _migrate_mark_all_payments_complete(force: bool = False) -> dict:
    """Mark every client + invoice paid except PARTIAL_PAYMENT_CLIENT_FILE_NOS (half-paid)."""
    flag = await db.meta.find_one({"key": "payment_status_bulk_complete_v2"})
    if flag and not force:
        return {"skipped": True, "reason": "already applied"}

    partial_ids, partial_clients = await _partial_payment_client_ids()

    inv_r = await db.invoices.update_many(
        {
            "client_id": {"$nin": partial_ids},
            "payment_status": {"$ne": "complete"},
        },
        {"$set": {"payment_status": "complete"}},
    )
    cl_r = await db.clients.update_many(
        {
            "id": {"$nin": partial_ids},
            "payment_status": {"$ne": "complete"},
        },
        {"$set": {"payment_status": "complete"}},
    )

    partial_inv_count = 0
    for cid in partial_ids:
        async for inv in db.invoices.find({"client_id": cid}, {"_id": 0, "id": 1, "amount": 1}):
            amount = float(inv.get("amount") or 0)
            patch: dict = {"payment_status": "partial"}
            if amount > 0:
                patch["amount_paid"] = round(amount / 2, 2)
            await db.invoices.update_one({"id": inv["id"]}, {"$set": patch})
            partial_inv_count += 1
        await db.clients.update_one({"id": cid}, {"$set": {"payment_status": "partial"}})

    await db.meta.update_one(
        {"key": "payment_status_bulk_complete_v2"},
        {
            "$set": {
                "done": True,
                "at": now_iso(),
                "invoices_complete": inv_r.modified_count,
                "clients_complete": cl_r.modified_count,
                "partial_clients": partial_clients,
                "partial_invoices": partial_inv_count,
            }
        },
        upsert=True,
    )
    return {
        "skipped": False,
        "invoices_updated": inv_r.modified_count,
        "clients_updated": cl_r.modified_count,
        "partial_clients": partial_clients,
        "partial_invoices_updated": partial_inv_count,
    }


async def _migrate_therapist_emails() -> int:
    """Fix therapist emails: rename old→new, then force-sync from MASTER_THERAPISTS by key/name."""
    updated = 0
    for old_email, new_email in THERAPIST_EMAIL_MIGRATIONS:
        old_l = old_email.lower()
        new_l = new_email.lower()
        r1 = await db.therapists.update_many(
            {"email": {"$regex": f"^{re.escape(old_l)}$", "$options": "i"}},
            {"$set": {"email": new_l}},
        )
        r2 = await db.users.update_many(
            {"email": {"$regex": f"^{re.escape(old_l)}$", "$options": "i"}},
            {"$set": {"email": new_l}},
        )
        updated += (r1.modified_count or 0) + (r2.modified_count or 0)

    existing_ts = await db.therapists.find(
        {}, {"_id": 0, "id": 1, "name": 1, "key": 1, "email": 1}
    ).to_list(500)

    for key, first, email, role, leave_balance, join_date in MASTER_THERAPISTS:
        new_l = email.lower()
        match = next((t for t in existing_ts if t.get("key") == key), None)
        if not match:
            match = next(
                (t for t in existing_ts if first.lower() in (t.get("name") or "").lower()),
                None,
            )
        if not match:
            continue
        tid = match["id"]
        old_l = (match.get("email") or "").lower()
        if old_l != new_l:
            await db.therapists.update_one({"id": tid}, {"$set": {"email": new_l}})
            updated += 1
        # users: match by therapist id or any prior email for this person
        for candidate in {old_l, new_l, *(o.lower() for o, n in THERAPIST_EMAIL_MIGRATIONS if n.lower() == new_l)}:
            if not candidate:
                continue
            r = await db.users.update_many({"email": candidate}, {"$set": {"email": new_l}})
            updated += r.modified_count or 0
        r_id = await db.users.update_one({"id": tid}, {"$set": {"email": new_l}})
        if r_id.modified_count:
            updated += r_id.modified_count

    jenan = await db.therapists.find_one({"email": "jsalmuhaisin@boostgrowthsa.com"}, {"_id": 0, "id": 1})
    if not jenan:
        hit = await db.therapists.find_one({"name": {"$regex": "Jenan", "$options": "i"}}, {"_id": 0, "id": 1})
        if hit:
            await db.therapists.update_one({"id": hit["id"]}, {"$set": {"email": "jsalmuhaisin@boostgrowthsa.com"}})
            updated += 1
    return updated


THERAPIST_DISPLAY_NAMES = {
    "msAbeer": "Ms. Abeer",
    "msAlhanouf": "Ms. Alhanouf",
    "msBodoor": "Ms. Bodour",
    "msFahda": "Ms. Fahda",
    "msFatimah": "Ms. Fatimah",
    "msHajer": "Ms. Hajar",
    "msJenan": "Ms. Jenan",
    "msMaha": "Ms. Maha",
    "msManal": "Ms. Manal",
    "msRahaf": "Ms. Rahaf",
    "msRazan": "Ms. Razan",
    "msShatha": "Ms. Shatha",
    "msShrooq": "Ms. Shroug",
    "msWaad": "Ms. Waad",
    "msNaja": "Ms. Najla",
    "msNajla": "Ms. Najla",
    "msAsma": "Ms. Asma",
    "msWalaa": "Ms. Walaa",
}

# Schedule column headers: first name + family (matches frontend scheduleConstants.js).
THERAPIST_FAMILY_NAMES = {
    "msMaha": "Althunayan",
    "msFahda": "Alghadeeb",
    "msRazan": "Alshatery",
    "msManal": "Aldosery",
    "msAsma": "Ahmed",
    "msHajer": "Alfulaij",
    "msRahaf": "Aljuhani",
    "msShatha": "Alhammami",
    "msAlhanouf": "Alromman",
    "msWaad": "Alhamed",
    "msNajla": "Alhamad",
    "msNaja": "Alhamad",
    "msBodoor": "Alkhlifah",
    "msFatimah": "Alkhater",
    "msShrooq": "Alamri",
    "msAbeer": "Alshareef",
    "msJenan": "Almuhaisin",
    "msWalaa": "Althunayan",
}

THERAPIST_FIRST_NAME_OVERRIDES = {
    "shrooq": "Shroug",
    "shroug": "Shroug",
    "bodoor": "Bodour",
    "hajer": "Hajar",
}

# Excel column order — 28 Jun 2026 sheet (matches frontend scheduleConstants.js).
THERAPIST_SCHEDULE_ORDER = [
    "msmaha", "msfahda", "msrazan", "msmanal", "mshajer", "msrahaf",
    "msshatha", "msalhanouf", "mswaad", "msfatimah", "msshoroq",
    "msabeer", "msnajla", "msasma", "msbodoor", "msjenan", "mswalaa",
]

_SCHEDULE_THERAPIST_SKIP = frozenset({
    "therapist's name", "therapists name", "days", "service",
    "school support", "home session", "outdoor session",
    "therapist cancelation", "client cancelation",
    "various-service (school & home)",
})


def therapist_schedule_display_name(t: Optional[dict]) -> str:
    """Full therapist label (first + family) — source of truth for schedule/portal display."""
    if not t:
        return ""
    raw = re.sub(r"^Ms\.?\s*", "", (t.get("name") or ""), flags=re.I).strip()
    first = (raw.split()[0] if raw.split() else raw) or raw
    first_lower = first.lower()
    if first_lower in THERAPIST_FIRST_NAME_OVERRIDES:
        first = THERAPIST_FIRST_NAME_OVERRIDES[first_lower]
    if first_lower == "najla":
        return "Najla Alhamad"
    key = t.get("key") or ""
    family = None
    for k, v in THERAPIST_FAMILY_NAMES.items():
        if k.lower() == key.lower():
            family = v
            break
    if family:
        parts = raw.split()
        if len(parts) >= 2 and parts[-1].lower() == family.lower():
            head = THERAPIST_FIRST_NAME_OVERRIDES.get(parts[0].lower(), parts[0])
            return " ".join([head] + parts[1:])
        return f"{first} {family}"
    return raw or (t.get("name") or "")


async def _migrate_personal_therapist_accounts() -> dict:
    """Ensure personal specialist logins resolve to the correct keyed therapist (e.g. Asma not Ahmed)."""
    actions: List[str] = []
    personal = {
        "asma@boostgrowthsa.com": ("msAsma", "Ms. Asma"),
    }
    for email, (key, canonical) in personal.items():
        removed = await db.users.delete_many({"email": email})
        if removed.deleted_count:
            actions.append(f"removed {removed.deleted_count} admin user row(s) for {email}")
        t = await _find_therapist_by_email(email)
        if not t:
            continue
        patch: dict = {}
        if (t.get("key") or "") != key:
            patch["key"] = key
        if (t.get("name") or "") != canonical:
            patch["name"] = canonical
        if not t.get("role"):
            patch["role"] = "therapist"
        if patch:
            await db.therapists.update_one({"id": t["id"]}, {"$set": patch})
            actions.append(f"updated {email} -> {canonical} ({key})")
    # Drop duplicate rows that kept a placeholder surname on the same email.
    for em, group in (
        (e, [x async for x in db.therapists.find(
            {"email": {"$regex": f"^{re.escape(e)}$", "$options": "i"}}, {"_id": 0}
        )])
        for e in personal
    ):
        if len(group) <= 1:
            continue
        scored = [(_therapist_record_score(t), t.get("created_at") or "", t) for t in group]
        scored.sort(key=lambda x: (-x[0], x[1]))
        for _, _, loser in scored[1:]:
            await db.schedule_cells.delete_many({"therapist_id": loser["id"]})
            await db.users.delete_many({"therapist_id": loser["id"]})
            await db.therapists.delete_one({"id": loser["id"]})
            actions.append(f"deduped {em}: removed {loser.get('name')} ({loser['id'][:8]})")
    return {"actions": actions}


async def _migrate_therapist_display_names() -> int:
    """Align therapist DB names with full first+family spelling; sync labels in related records."""
    updated = 0
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "key": 1}).to_list(500)
    for t in therapists:
        display = therapist_schedule_display_name(t)
        if display and (t.get("name") or "") != display:
            await db.therapists.update_one({"id": t["id"]}, {"$set": {"name": display}})
            await db.users.update_many({"therapist_id": t["id"]}, {"$set": {"name": display}})
            t["name"] = display
            updated += 1
    for t in therapists:
        tid = t.get("id")
        if not tid:
            continue
        display = therapist_schedule_display_name(t)
        if not display:
            continue
        for coll, field in (("requests", "therapist_name"), ("leaves", "therapist_name"), ("staff_purchases", "therapist_name")):
            res = await db[coll].update_many({"therapist_id": tid, field: {"$ne": display}}, {"$set": {field: display}})
            updated += res.modified_count
    return updated


@api.post("/admin/migrate-therapist-display-names")
async def admin_migrate_therapist_display_names(_=Depends(ops_or_admin)):
    n = await _migrate_therapist_display_names()
    return {"ok": True, "records_updated": n}


@api.post("/admin/migrate-therapist-emails")
async def admin_migrate_therapist_emails(_=Depends(admin_only)):
    """Manually run therapist email migration (also runs on server startup)."""
    n = await _migrate_therapist_emails()
    return {"ok": True, "records_updated": n}


_INV_SHEET_RE = _re_top.compile(r"^(copy of\s+)?inv[\s\-_]*\d+", _re_top.IGNORECASE)
_INV_NUM_RE = _re_top.compile(r"inv[\s\-_]*(\d+)", _re_top.IGNORECASE)
_HEADER_TOKENS = {"day", "days", "date", "status", "time", "hrs", "hours", "# of hrs", "therapist", "note", "notes"}
_SKIP_TAB_HINTS = (
    "summary", "info", "readme", "template", "cover", "index", "master",
    "dashboard", "lookup", "archive", "settings", "data", "pivot",
)


def _invoice_number_from_name(name: str) -> Optional[str]:
    sn = (name or "").strip()
    m = _INV_NUM_RE.search(sn)
    if not m:
        return None
    return f"INV{m.group(1)}"


def _invoice_sort_key(name: str) -> int:
    m = _INV_NUM_RE.search(name or "")
    return int(m.group(1)) if m else 0


def _is_copy_sheet_name(name: str) -> bool:
    return (name or "").lower().startswith("copy of ")


def _sheet_has_session_table(ws) -> bool:
    """True if worksheet looks like an invoice session table (Day/Date + Time/Hrs)."""
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        cells = [str(c).strip().lower() if c is not None else "" for c in (row or [])]
        joined = " ".join(cells)
        if "date" in cells and ("time" in cells or "hrs" in joined or "# of hrs" in joined):
            return True
    return False


def _discover_invoice_sheets(wb, client_file_no: str = None) -> list:
    """Invoice tabs only: INV-prefixed names or sheets whose header embeds an invoice number."""
    candidates: List[str] = []
    fn_raw = (client_file_no or "").strip()
    fn_padded = fn_raw.zfill(3) if fn_raw else ""
    fn_stripped = fn_raw.lstrip("0") or fn_raw

    for name in wb.sheetnames:
        sn = name.strip()
        sn_low = sn.lower()
        if any(h in sn_low for h in _SKIP_TAB_HINTS):
            continue
        if _INV_SHEET_RE.match(sn):
            candidates.append(name)
            continue
        sn_compact = _re_top.sub(r"[\s\-_]+", "", sn)
        if fn_padded and (fn_padded in sn_compact or (fn_stripped and fn_stripped in sn_compact)):
            try:
                ws = wb[name]
                # Some historical attendance files name tabs like "061 HS" without an invoice number.
                # If the tab clearly contains a session table and is tagged with this client's file_no,
                # treat it as an invoice sheet — downstream ingest will generate a stable invoice id.
                if _sheet_has_session_table(ws):
                    candidates.append(name)
                    continue
            except Exception:
                pass
        try:
            ws = wb[name]
            if not _sheet_has_session_table(ws):
                continue
            header_info = _parse_invoice_header(ws, sn)
            inv_from_header = header_info.get("invoice_number") or ""
            if _invoice_number_from_name(inv_from_header) or _invoice_number_from_name(sn):
                candidates.append(name)
        except Exception:
            continue

    # Deduplicate by invoice number — prefer non-"Copy of" tabs
    by_num: Dict[str, str] = {}
    for name in candidates:
        inv = _invoice_number_from_name(name)
        if not inv:
            try:
                inv = _invoice_number_from_name(
                    (_parse_invoice_header(wb[name], name).get("invoice_number") or "")
                )
            except Exception:
                inv = None
        if not inv:
            # Keep non-INV tabs (e.g. "061 HS") when they were accepted as candidates.
            # They will be ingested with a generated invoice_number later.
            inv = f"TAB:{name}"
        prev = by_num.get(inv)
        if not prev:
            by_num[inv] = name
        elif _is_copy_sheet_name(name) and not _is_copy_sheet_name(prev):
            continue
        elif not _is_copy_sheet_name(name) and _is_copy_sheet_name(prev):
            by_num[inv] = name

    return sorted(by_num.values(), key=_invoice_sort_key)


async def _reconcile_invoices_after_sync(cid: str, synced_nums: set) -> dict:
    """Drop duplicate/orphan invoices after Drive import; keep highest-session copy per INV number."""
    if not synced_nums:
        return {"removed": 0, "merged": 0}
    all_invs = await db.invoices.find({"client_id": cid}, {"_id": 0}).to_list(500)
    by_num: Dict[str, list] = {}
    for inv in all_invs:
        num = _invoice_number_from_name(inv.get("invoice_number") or "")
        if num:
            by_num.setdefault(num, []).append(inv)
    removed = merged = 0
    keep_by_num: Dict[str, str] = {}
    for num, group in by_num.items():
        if num not in synced_nums and len(group) == 1:
            raw = group[0].get("invoice_number") or ""
            if _re_top.match(r"^INV_[a-f0-9]{8}_\d+$", raw, _re_top.I):
                await db.invoices.delete_one({"id": group[0]["id"]})
                await db.sessions.delete_many({"invoice_id": group[0]["id"]})
                removed += 1
            continue
        if len(group) < 2:
            keep_by_num[num] = group[0]["id"]
            continue
        scored = []
        for inv in group:
            sess_n = await db.sessions.count_documents({"invoice_id": inv["id"]})
            scored.append((sess_n, inv.get("created_at") or "", inv))
        scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
        keep = scored[0][2]
        keep_by_num[num] = keep["id"]
        for _, _, dup in scored[1:]:
            await db.sessions.update_many(
                {"invoice_id": dup["id"]},
                {"$set": {"invoice_id": keep["id"], "source_invoice": num}},
            )
            await db.invoices.delete_one({"id": dup["id"]})
            removed += 1
            merged += 1
    for inv in all_invs:
        raw = (inv.get("invoice_number") or "").strip()
        if not _re_top.match(r"^INV_[a-f0-9]{8}_\d+$", raw, _re_top.I):
            continue
        norm = _invoice_number_from_name(inv.get("source_sheet") or "")
        keep_id = keep_by_num.get(norm or "")
        if norm and norm in synced_nums and keep_id and keep_id != inv["id"]:
            await db.sessions.update_many(
                {"invoice_id": inv["id"]},
                {"$set": {"invoice_id": keep_id, "source_invoice": norm}},
            )
            await db.invoices.delete_one({"id": inv["id"]})
            removed += 1
    return {"removed": removed, "merged": merged}


async def _cleanup_orphan_invoices() -> dict:
    """Remove auto-generated INV_{uuid}_N invoices when a real INV#### exists."""
    removed = 0
    invs = await db.invoices.find({}, {"_id": 0, "id": 1, "client_id": 1, "invoice_number": 1}).to_list(5000)
    by_client: Dict[str, List[dict]] = {}
    for inv in invs:
        by_client.setdefault(inv.get("client_id") or "", []).append(inv)
    for cid, group in by_client.items():
        if not cid:
            continue
        reals = [i for i in group if _invoice_num_key(i) > 0]
        autos = [i for i in group if _re_top.match(r"^INV_[a-f0-9]{8}_\d+$", i.get("invoice_number") or "", _re_top.I)]
        if not reals or not autos:
            continue
        keep = max(reals, key=_invoice_num_key)
        for inv in autos:
            await db.sessions.update_many(
                {"invoice_id": inv["id"]},
                {"$set": {"invoice_id": keep["id"], "source_invoice": keep["invoice_number"]}},
            )
            await db.invoices.delete_one({"id": inv["id"]})
            removed += 1
    return {"removed": removed}


def _parse_invoice_header(ws, sheet_name: str = "") -> dict:
    """Read the header section (rows 1–10) for invoice metadata and service type."""
    sn = (sheet_name or ws.title or "").strip()
    info = {
        "invoice_number": sn,
        "is_closed": False,
        "close_date": None,
        "package_size": None,
        "service_type": None,
    }
    rows: list = []
    for r in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in (r or [])])

    # Invoice number — tab name (INV0465) or embedded in rows 1–10
    if _INV_SHEET_RE.match(sn):
        info["invoice_number"] = _re_top.sub(r"[\s\-_]+", "", sn, flags=_re_top.IGNORECASE).upper()
    else:
        for row in rows[:10]:
            joined = " ".join(row)
            m = _re_top.search(r"(inv[\s\-_]*\d+)", joined, _re_top.IGNORECASE)
            if m:
                info["invoice_number"] = _re_top.sub(
                    r"[\s\-_]+", "", m.group(1), flags=_re_top.IGNORECASE
                ).upper()
                break

    flat = " | ".join(" ".join(r) for r in rows).lower()

    # Open / closed status (Excel uses "Closed", "CLOSED", or "Close")
    status_cell = ""
    for row in rows[:3]:
        for cell in row[2:4]:
            if cell:
                status_cell = cell.strip().lower()
                break
        if status_cell:
            break
    if status_cell.startswith("clos") or "closed" in flat:
        info["is_closed"] = True
        m = _re_top.search(r"clos(?:ed)?[^0-9]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", flat)
        if m:
            try:
                info["close_date"] = _normalize_date(m.group(1))
            except Exception:
                pass

    # Row 5 (or any header row): "Service: Home Sessions" in column A or any cell
    for row in rows:
        for ci, cell in enumerate(row):
            raw = (cell or "").strip()
            if not raw:
                continue
            if raw.lower().startswith("service:"):
                st = _service_type_from_label(raw)
                if st:
                    info["service_type"] = st
                    break
        if info.get("service_type"):
            break

    # Tab / row hints: "INV0465 | SS", standalone HS/SS cells, keywords in header block
    if not info.get("service_type"):
        sn_low = sn.lower()
        if _re_top.search(r"\bss\b|school", sn_low):
            info["service_type"] = "SS"
        elif _re_top.search(r"\bhs\b|home", sn_low):
            info["service_type"] = "HS"
    if not info.get("service_type"):
        for row in rows[:6]:
            for cell in row:
                cu = (cell or "").strip().upper()
                if cu in ("SS", "HS", "AVC"):
                    info["service_type"] = cu
                    break
                st = _service_type_from_label(cell)
                if st:
                    info["service_type"] = st
                    break
            if info.get("service_type"):
                break
    if not info.get("service_type"):
        if _re_top.search(r"school\s*support|\|\s*ss\b|\bss\s*\|", flat):
            info["service_type"] = "SS"
        elif _re_top.search(r"home\s*session|\|\s*hs\b|\bhs\s*\|", flat):
            info["service_type"] = "HS"
        elif _re_top.search(r"\bavc\b", flat):
            info["service_type"] = "AVC"

    if info.get("service_type"):
        info["service_type"] = _normalize_service_type(info["service_type"])

    # SS tabs: "4 Weeks", "School Support", "school Sessions"
    if not info.get("service_type"):
        if _re_top.search(r"school\s+support|school\s+session", flat):
            info["service_type"] = "SS"
        elif _re_top.search(r"4\s*week", flat):
            info["service_type"] = "SS"

    # Package size — "# Paid SESH.: 24 Hours" or "4 Weeks"
    m_weeks = _re_top.search(r"paid\s+sesh[^0-9]*(\d+)\s*week", flat)
    if m_weeks:
        info["service_type"] = info.get("service_type") or "SS"
        info["package_size"] = 4
    else:
        m = _re_top.search(r"paid\s+sesh[^0-9]*([\d.]+)", flat)
        if m:
            try:
                info["package_size"] = float(m.group(1))
            except Exception:
                pass
    if info.get("service_type") == "SS" and (info.get("package_size") is None or info.get("package_size") > 12):
        info["package_size"] = 4
    return info


def _normalize_date(s: str) -> Optional[str]:
    s = (s or "").strip()
    if not s:
        return None
    # Try ISO first
    try:
        d = datetime.fromisoformat(s[:10])
        return d.strftime("%Y-%m-%d")
    except Exception:
        pass
    # Try D/M/YYYY or D-M-YYYY (Boost Growth display format — day first)
    for sep in ("/", "-"):
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                try:
                    d, mo, y = int(parts[0]), int(parts[1]), int(parts[2])
                    if y < 100:
                        y += 2000
                    return datetime(y, mo, d).strftime("%Y-%m-%d")
                except Exception:
                    continue
    return None


def _swap_month_day_iso(iso: str) -> Optional[str]:
    """Swap Y-M-D month/day when both are <= 12 (MM/DD vs DD/MM mistake)."""
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", (iso or "")[:10])
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if mo == d or mo > 12 or d > 12:
        return None
    try:
        return datetime(y, d, mo).strftime("%Y-%m-%d")
    except ValueError:
        return None


def _session_likely_swapped_month_day(session_iso: str, peer_isos: list) -> Optional[str]:
    """Return corrected ISO when this session's month looks like a day/month swap."""
    swapped = _swap_month_day_iso(session_iso)
    if not swapped or swapped == session_iso:
        return None
    peers = [p[:10] for p in peer_isos if p and p[:10] != session_iso[:10]]
    if len(peers) < 2:
        return None

    def month_key(iso):
        return iso[:7]

    from collections import Counter
    peer_months = Counter(month_key(p) for p in peers)
    cur_month = month_key(session_iso)
    swap_month = month_key(swapped)
    cur_peer_count = peer_months.get(cur_month, 0)
    swap_peer_count = peer_months.get(swap_month, 0)

    if swap_peer_count < 2:
        return None
    if cur_peer_count > 1:
        return None
    if swap_peer_count <= cur_peer_count:
        return None

    try:
        swap_dt = datetime.fromisoformat(swapped)
        peer_dts = sorted(datetime.fromisoformat(p) for p in peers)
        span_start = peer_dts[0]
        span_end = peer_dts[-1]
        margin = timedelta(days=45)
        if swap_dt < span_start - margin or swap_dt > span_end + margin:
            return None
    except Exception:
        return None
    return swapped


def _parse_time_range(s: str) -> tuple:
    """Return (start_HH:MM, end_HH:MM, hours_diff) or ('','',0)."""
    s = (s or "").strip()
    if not s:
        return "", "", 0.0
    m = _re_top.search(r"(\d{1,2})(?::(\d{2}))?\s*[-–]\s*(\d{1,2})(?::(\d{2}))?", s)
    if not m:
        return "", "", 0.0
    h1, m1, h2, m2 = m.group(1), m.group(2) or "00", m.group(3), m.group(4) or "00"
    start = f"{int(h1):02d}:{int(m1):02d}"
    end = f"{int(h2):02d}:{int(m2):02d}"
    diff = ((int(h2) * 60 + int(m2)) - (int(h1) * 60 + int(m1))) / 60.0
    if diff < 0:
        diff += 24
    return start, end, round(diff, 2)


async def _ingest_workbook_for_client(cid: str, client: dict, wb, user_id: str, origin: str = "import") -> dict:
    """Iterate invoice tabs in the workbook, create invoices + sessions idempotently."""
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(200)
    # name token (Manal / Hajer / etc.) -> id
    name_to_id = {}
    for t in therapists:
        name = (t.get("name") or "").replace("Ms.", "").replace("ms.", "").strip()
        if not name:
            continue
        first = name.split()[0].lower()
        name_to_id[first] = t["id"]

    def _resolve_therapist_ids(cell: str) -> list:
        s = (cell or "").lower()
        # Split by hyphen, slash, or comma
        parts = [p.strip() for p in _re_top.split(r"[-/,]", s) if p.strip()]
        out = []
        for p in parts:
            tok = p.split()[0]
            if tok in name_to_id and name_to_id[tok] not in out:
                out.append(name_to_id[tok])
            elif len(tok) >= 3:
                for k, tid in name_to_id.items():
                    if tid in out:
                        continue
                    if k.startswith(tok[:3]) or tok.startswith(k[:3]):
                        out.append(tid)
                        break
        return out

    matched_sheets = _discover_invoice_sheets(wb, client.get("file_no"))
    debug_sheets = []
    all_tabs = list(wb.sheetnames)
    existing_inv = {i["invoice_number"]: i for i in await db.invoices.find(
        {"client_id": cid}, {"_id": 0}
    ).to_list(500)}
    pkg_default = client.get("package_hours") or 24
    existing_sessions = await db.sessions.find(
        {"client_id": cid}, {"_id": 0, "session_date": 1, "start_time": 1, "sync_key": 1}
    ).to_list(5000)
    existing_key = {(s.get("session_date"), s.get("start_time") or "") for s in existing_sessions}
    existing_sync = {s["sync_key"] for s in existing_sessions if s.get("sync_key")}

    invoices_added, invoices_updated, sessions_added, sessions_skipped = [], [], 0, 0

    for tab_idx, name in enumerate(matched_sheets):
        ws = wb[name]
        clean = name.strip()
        header_info = _parse_invoice_header(ws, clean)
        inv_num = header_info.get("invoice_number") or clean
        normalized = _invoice_number_from_name(inv_num) or _invoice_number_from_name(clean)
        if normalized:
            inv_num = normalized
        elif not _INV_SHEET_RE.match(inv_num):
            inv_num = f"INV_{cid[:8]}_{tab_idx + 1}"
        header_info["invoice_number"] = inv_num
        sheet_hs = sheet_ss = 0
        debug_sheets.append({"sheet": clean, "invoice_number": inv_num})
        # Upsert invoice — match by invoice_number or legacy tab name
        inv_pkg = header_info.get("package_size") or pkg_default
        existing = existing_inv.get(inv_num) or existing_inv.get(clean)
        if existing:
            update = {
                "close_date": header_info.get("close_date"),
                "package_size": inv_pkg,
                "invoice_number": inv_num,
            }
            if header_info["is_closed"]:
                update["is_closed"] = True
                if header_info.get("close_date"):
                    update["close_date"] = header_info["close_date"]
            detected_st = _normalize_service_type(header_info.get("service_type"))
            if detected_st:
                update["service_type"] = detected_st
            await db.invoices.update_one({"id": existing["id"]}, {"$set": update})
            invoices_updated.append(inv_num)
            inv_doc = {**existing, **update}
            existing_inv[inv_num] = inv_doc
        else:
            inv_doc = {
                "id": str(uuid.uuid4()), "client_id": cid,
                "invoice_number": inv_num, "notes": None, "amount": None,
                "period_from": None, "period_to": header_info.get("close_date"),
                "package_size": inv_pkg,
                "payment_status": "complete" if header_info["is_closed"] else "pending",
                "start_date": now_iso()[:10],
                "service_type": _normalize_service_type(header_info.get("service_type")),
                "is_closed": header_info["is_closed"],
                "close_date": header_info.get("close_date"),
                "source": origin, "source_sheet": clean,
                "created_by": user_id, "created_at": now_iso(),
            }
            await db.invoices.insert_one(inv_doc)
            invoices_added.append(inv_num)
            existing_inv[inv_num] = inv_doc

        # Find header row containing the session columns
        header_row_idx = None
        col_map = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            joined = " ".join(cells)
            if "date" in cells and ("status" in cells or "# of hrs" in joined or "hrs" in joined):
                header_row_idx = r_idx
                for ci, h in enumerate(cells):
                    if h in ("day", "days"):
                        col_map["day"] = ci
                    elif h == "date":
                        col_map["date"] = ci
                    elif h == "status":
                        col_map["status"] = ci
                    elif h == "time":
                        col_map["time"] = ci
                    elif h in ("# of hrs", "hrs", "hours"):
                        col_map["hours"] = ci
                    elif h == "therapist":
                        col_map["therapist"] = ci
                    elif h in ("note", "notes"):
                        col_map["note"] = ci
                break
        if header_row_idx is None or "date" not in col_map:
            continue

        # Earliest session_date for this invoice -> use as start_date
        earliest = None
        # Iterate session rows after the header
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            if row is None or all((c is None or (isinstance(c, str) and not c.strip())) for c in row):
                continue
            raw_date = row[col_map["date"]] if col_map["date"] < len(row) else None
            # Stop when we hit a totals/footer row
            joined = " ".join(str(c).lower() for c in row if c is not None)
            if "total" in joined and "session" in joined:
                break
            date_iso = None
            if isinstance(raw_date, datetime):
                date_iso = raw_date.strftime("%Y-%m-%d")
            elif raw_date:
                date_iso = _normalize_date(str(raw_date))
            if not date_iso:
                continue
            status_val = (str(row[col_map["status"]]).strip() if "status" in col_map and col_map["status"] < len(row) and row[col_map["status"]] else "").strip()
            # Normalize statuses — strip HS| / SS| service prefix if present
            status_l = status_val.lower()
            svc_hint = _session_blob_service_hint({"status": status_val, "note": "", "location": ""})
            if svc_hint == "HS":
                sheet_hs += 1
            elif svc_hint == "SS":
                sheet_ss += 1
            if _re_top.match(r"^(HS|SS)\s*\|", status_val, _re_top.IGNORECASE):
                status_val = _re_top.sub(r"^(HS|SS)\s*\|\s*", "", status_val, flags=_re_top.IGNORECASE).strip()
                status_l = status_val.lower()
            if status_l in ("completed", "complete", "delivered"):
                status_norm = "Completed"
            elif "no service" in status_l or status_l == "ns":
                status_norm = "No Service"
            elif "cancel" in status_l:
                status_norm = "Cancelled"
            elif "no show" in status_l or "no-show" in status_l:
                status_norm = "No Show"
            else:
                status_norm = status_val.title() if status_val else "Completed"
            time_str = str(row[col_map["time"]]).strip() if "time" in col_map and col_map["time"] < len(row) and row[col_map["time"]] else ""
            start_t, end_t, calc_h = _parse_time_range(time_str)
            hours_val = row[col_map["hours"]] if "hours" in col_map and col_map["hours"] < len(row) else None
            try:
                hours_f = float(hours_val) if hours_val not in (None, "", "—") else calc_h
            except Exception:
                hours_f = calc_h
            ther_cell = str(row[col_map["therapist"]]).strip() if "therapist" in col_map and col_map["therapist"] < len(row) and row[col_map["therapist"]] else ""
            ther_ids = _resolve_therapist_ids(ther_cell)
            note_val = str(row[col_map["note"]]).strip() if "note" in col_map and col_map["note"] < len(row) and row[col_map["note"]] else ""
            nh = _session_blob_service_hint({"note": note_val, "status": "", "location": ""})
            if nh == "HS":
                sheet_hs += 1
            elif nh == "SS":
                sheet_ss += 1

            sync_key = f"{inv_num}|{date_iso}|{row_idx}"
            key = (date_iso, start_t)
            if sync_key in existing_sync or key in existing_key:
                sessions_skipped += 1
                continue
            inv_st = _normalize_service_type(inv_doc.get("service_type") or header_info.get("service_type"))
            sess_doc = {
                "id": str(uuid.uuid4()),
                "client_id": cid,
                "session_date": date_iso,
                "day_name": _day_name_from_date(date_iso),
                "start_time": start_t or None,
                "end_time": end_t or None,
                "hours": hours_f if inv_st != "SS" else None,
                "status": status_norm,
                "therapist_ids": ther_ids,
                "note": note_val or None,
                "service_type": inv_st,
                "week_number": None,
                "source": origin,
                "source_invoice": inv_num,
                "invoice_id": inv_doc.get("id"),
                "sync_key": sync_key,
                "created_at": now_iso(),
            }
            await db.sessions.insert_one(sess_doc)
            existing_key.add(key)
            existing_sync.add(sync_key)
            sessions_added += 1
            if earliest is None or date_iso < earliest:
                earliest = date_iso

        # Stamp invoice start_date with earliest imported session date
        if earliest:
            await db.invoices.update_one({"id": inv_doc["id"]}, {"$set": {"start_date": earliest}})

        # Final service_type: Row 5 header → session majority → client profile
        st = _normalize_service_type(header_info.get("service_type"))
        if not st and sheet_hs > sheet_ss:
            st = "HS"
        elif not st and sheet_ss > sheet_hs:
            st = "SS"
        elif not st:
            client_st = _normalize_service_type(client.get("service_type"))
            if client_st in ("SS", "HS", "AVC"):
                st = client_st
        if st:
            await db.invoices.update_one({"id": inv_doc["id"]}, {"$set": {"service_type": st}})
            inv_doc["service_type"] = st
            # SS package is always 4 weeks
            if st == "SS":
                await db.invoices.update_one({"id": inv_doc["id"]}, {"$set": {"package_size": 4}})
                inv_doc["package_size"] = 4

        # Stamp service_type, day_name, week_number on all sessions for this invoice
        anchor = earliest or inv_doc.get("start_date")
        inv_sessions = await db.sessions.find(
            {"client_id": cid, "$or": [{"invoice_id": inv_doc["id"]}, {"source_invoice": inv_num}]},
            {"_id": 0, "id": 1, "session_date": 1},
        ).to_list(2000)
        for s in inv_sessions:
            sd = str(s.get("session_date") or "")[:10]
            patch = {"day_name": _day_name_from_date(sd), "service_type": st}
            if st == "SS":
                patch["week_number"] = _school_week_for_date(sd, anchor, 4)
                patch["hours"] = None
            elif st == "HS":
                patch["week_number"] = None
            await db.sessions.update_one({"id": s["id"]}, {"$set": patch})

    synced_nums = set()
    for n in invoices_added + invoices_updated:
        norm = _invoice_number_from_name(n)
        if norm:
            synced_nums.add(norm)
    reconcile = await _reconcile_invoices_after_sync(cid, synced_nums)

    return {
        "matched_sheets": matched_sheets,
        "workbook_tabs": all_tabs,
        "sheet_details": debug_sheets,
        "invoices_added": invoices_added,
        "invoices_updated": invoices_updated,
        "sessions_added": sessions_added,
        "sessions_skipped_existing": sessions_skipped,
        "invoices_reconciled": reconcile,
        "warning": (
            None if matched_sheets
            else f"No invoice sheets found. Tabs in file: {', '.join(all_tabs)}"
        ),
    }

# ------------------- Package reset (manual; admin only) -------------------
@api.post("/clients/{cid}/reset-package")
async def reset_package(cid: str, user=Depends(ops_or_admin)):
    """Reset used-hours counter to 0 by stamping `package_reset_at`.
    Existing sessions are kept; the frontend filters out sessions before this timestamp
    when computing used hours for the current cycle. Safe and reversible.
    """
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    ts = now_iso()
    await db.clients.update_one({"id": cid}, {"$set": {
        "package_reset_at": ts,
        "cycle_start_date": ts[:10],
    }})
    return {"ok": True, "package_reset_at": ts}

# ------------------- Sessions (Attendance log) -------------------
def _sessions_with_day_names(sessions: list) -> list:
    for s in sessions:
        if s.get("session_date"):
            s["day_name"] = _day_name_from_date(s["session_date"])
    return sessions


async def _sessions_for_invoice_query(client_id: str, invoice_id: str) -> list:
    """Sessions for ONE invoice — by invoice_id, source_invoice, or date window (orphans only)."""
    inv = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not inv:
        return []
    inv_num = (inv.get("invoice_number") or "").strip()
    q_or = [{"invoice_id": invoice_id}]
    if inv_num:
        q_or.append({"source_invoice": inv_num})
    direct = await db.sessions.find(
        {"client_id": client_id, "$or": q_or},
        {"_id": 0},
    ).sort("session_date", 1).to_list(2000)

    all_invs = await db.invoices.find({"client_id": client_id}, {"_id": 0}).to_list(200)
    sorted_invs = _sorted_invoices_for_client(client_id, all_invs)
    start, end = _invoice_window_bounds(inv, sorted_invs)

    orphan_q = {
        "client_id": client_id,
        "session_date": {"$gte": start},
        "$and": [
            {"$or": [{"invoice_id": None}, {"invoice_id": ""}, {"invoice_id": {"$exists": False}}]},
            {"$or": [{"source_invoice": None}, {"source_invoice": ""}, {"source_invoice": {"$exists": False}}]},
        ],
    }
    if end:
        orphan_q["session_date"]["$lt"] = end
    orphans = await db.sessions.find(orphan_q, {"_id": 0}).sort("session_date", 1).to_list(2000)

    seen = {s["id"] for s in direct if s.get("id")}
    merged = list(direct)
    for s in orphans:
        sid = s.get("id")
        if sid and sid in seen:
            continue
        if _session_in_invoice_date_window(s, inv, sorted_invs):
            merged.append(s)
            if sid:
                seen.add(sid)
                # Self-heal: link orphan session to this invoice for future queries
                patch = {"invoice_id": invoice_id}
                if inv_num:
                    patch["source_invoice"] = inv_num
                await db.sessions.update_one({"id": sid}, {"$set": patch})

    merged.sort(key=_session_date_sort_key)
    return merged


async def _therapist_assigned_to_client(therapist_id: str, client_id: str) -> bool:
    """True when therapist is main or co on the client caseload."""
    if not therapist_id or not client_id:
        return False
    client = await db.clients.find_one(
        _active_client_filter({"id": client_id}),
        {"_id": 0, "main_therapist_id": 1, "co_therapist_ids": 1},
    )
    if not client:
        return False
    return client.get("main_therapist_id") == therapist_id or therapist_id in (client.get("co_therapist_ids") or [])


@api.get("/sessions")
async def list_sessions(client_id: Optional[str] = None, invoice_id: Optional[str] = None, user=Depends(get_current_user)):
    if invoice_id and client_id:
        items = await _sessions_for_invoice_query(client_id, invoice_id)
    else:
        q = {}
        if client_id:
            q["client_id"] = client_id
        if invoice_id:
            inv = await db.invoices.find_one({"id": invoice_id}, {"_id": 0, "invoice_number": 1})
            if inv:
                inv_num = (inv.get("invoice_number") or "").strip()
                q["$or"] = [{"invoice_id": invoice_id}, {"source_invoice": inv_num}]
            else:
                q["invoice_id"] = invoice_id
        items = await db.sessions.find(q, {"_id": 0}).sort("session_date", -1).to_list(2000)
    if user.get("role") == "therapist" and not _has_full_client_access(user):
        uid = await _resolve_user_therapist_id(user) or user.get("id")
        # Caseload therapists see full client history for preparation review
        if not (client_id and await _therapist_assigned_to_client(uid, client_id)):
            items = [s for s in items if uid in (s.get("therapist_ids") or [])]
    return _sessions_with_day_names(items)

async def _user_in_session_therapists(user: dict, therapist_ids: list) -> bool:
    """True if logged-in user (or mapped therapist row) is on the session."""
    uid = await _resolve_user_therapist_id(user) or user.get("id")
    allowed = set(therapist_ids or [])
    return bool(uid and uid in allowed) or bool(user.get("id") and user["id"] in allowed)


async def _sync_prep_history_for_session(sess: dict, prepared_by: str, notes: Optional[str] = None) -> None:
    """Keep prep_history notes/therapists aligned with a logged session."""
    sid = sess.get("id")
    cid = sess.get("client_id")
    sd = (sess.get("session_date") or "")[:10]
    if not sid or not cid or not sd:
        return
    client = await db.clients.find_one({"id": cid}, {"_id": 0, "name": 1})
    tids = [t for t in (sess.get("therapist_ids") or []) if t]
    note_val = notes if notes is not None else sess.get("note")
    patch = {
        "notes": (note_val or "").strip(),
        "session_id": sid,
        "prepared_at": now_iso(),
        "source": "session",
    }
    await db.prep_history.update_many({"session_id": sid}, {"$set": patch})
    client_name = (client or {}).get("name")
    for prep_tid in tids:
        th = await db.therapists.find_one({"id": prep_tid}, {"_id": 0, "name": 1, "key": 1})
        await _upsert_prep_history(
            therapist_id=prep_tid,
            client_id=cid,
            session_date=sd,
            prepared_by=prepared_by,
            time_slot=sess.get("start_time") or "",
            client_name=client_name,
            notes=(note_val or "").strip() or None,
            invoice_id=sess.get("invoice_id"),
            session_id=sid,
            source="session",
        )


def _merge_session_therapist_ids(payload_ids: list, user: dict, resolved_uid: Optional[str]) -> list:
    """Ensure logging therapist stays on the session; keep co-therapists from payload."""
    ids = [t for t in (payload_ids or []) if t]
    if user.get("role") == "therapist":
        uid = resolved_uid or user.get("id")
        if uid and uid not in ids:
            ids.insert(0, uid)
        elif user.get("id") and user["id"] not in ids:
            ids.insert(0, user["id"])
    return ids


def _service_code_for_new_session(client: dict, payload: SessionIn) -> str:
    if payload.service_type:
        st = _normalize_service_type(payload.service_type)
        if st in ("HS", "SS"):
            return st
    if payload.location and client:
        for loc in client.get("locations") or []:
            if loc.get("address") == payload.location:
                st = _normalize_service_type(loc.get("service"))
                if st in ("HS", "SS"):
                    return st
    cst = _normalize_service_type(client.get("service_type") if client else None)
    if cst in ("HS", "SS"):
        return cst
    return "HS"


def _attach_open_invoice_to_session(doc: dict, client: dict, invoices: list) -> dict:
    """Link new session to the client's open invoice for the service type."""
    if doc.get("invoice_id"):
        return doc
    payload_st = doc.get("service_type")
    st_code = _normalize_service_type(payload_st) if payload_st else None
    if st_code not in ("HS", "SS"):
        fake = SessionIn(
            client_id=doc["client_id"],
            session_date=doc.get("session_date") or now_iso()[:10],
            location=doc.get("location"),
            service_type=payload_st,
        )
        st_code = _service_code_for_new_session(client, fake)
    open_inv = _last_open_invoice(invoices, st_code)
    if not open_inv:
        for code in ("HS", "SS"):
            open_inv = _last_open_invoice(invoices, code)
            if open_inv:
                break
    if open_inv:
        doc["invoice_id"] = open_inv["id"]
        inv_num = (open_inv.get("invoice_number") or "").strip()
        if inv_num:
            doc["source_invoice"] = inv_num
    return doc


@api.post("/sessions")
async def create_session(payload: SessionIn, user=Depends(get_current_user)):
    if (payload.status or "").strip() == "No Service":
        raise HTTPException(status_code=400, detail="No Service is no longer available")
    payload = _apply_session_time_edits(payload)
    _require_same_day_session(user, payload.session_date)
    _require_session_log_fields(user, payload)
    sid = str(uuid.uuid4())
    resolved_uid = await _resolve_user_therapist_id(user) if user.get("role") == "therapist" else None
    therapist_ids = _merge_session_therapist_ids(payload.therapist_ids, user, resolved_uid)
    client = await db.clients.find_one(_active_client_filter({"id": payload.client_id}), {"_id": 0})
    invs = await db.invoices.find({"client_id": payload.client_id}, {"_id": 0}).to_list(200)
    doc = {"id": sid, **payload.model_dump(), "therapist_ids": therapist_ids,
           "created_by": user["id"], "created_by_role": user["role"],
           "created_at": now_iso()}
    doc = _attach_open_invoice_to_session(doc, client or {}, invs)
    await db.sessions.insert_one(doc)
    doc.pop("_id", None)
    if payload.status in ("Completed", "No Show", "Cancelled"):
        try:
            await _ensure_session_schedule_prep_markers(
                doc, user["id"], notes=payload.note,
            )
        except Exception:
            logger.exception("Schedule prep sync failed for new session %s", sid)
    # Admin alerts
    cname = client.get("name") if client else "—"
    if user.get("role") == "therapist":
        await _notify_admins("session_log", f"New session logged ({payload.status})",
                             f"{user.get('name')} logged {payload.status} for {cname} ({payload.hours}h)")
    if payload.status in ("Cancelled", "No Show"):
        await _notify_admins("cancel_alert", f"Session {payload.status}: {cname}",
                             f"On {payload.session_date} ({user.get('name')})")
    # Low-hours alert — scoped to last open HS invoice only
    if client:
        invs = await db.invoices.find({"client_id": payload.client_id}, {"_id": 0}).to_list(200)
        open_hs = _last_open_invoice(invs, "HS")
        if open_hs and payload.status == "Completed":
            inv_sessions = await db.sessions.find(
                {"client_id": payload.client_id}, {"_id": 0}
            ).to_list(5000)
            matched = _sessions_for_invoice(open_hs, inv_sessions, invs)
            used_h = sum(
                float(s.get("hours") or 0)
                for s in matched
                if s.get("status") in ("Completed", "Cancelled")
            )
            pkg_h = float(open_hs.get("package_size") or client.get("package_hours") or 24)
            rem = pkg_h - used_h
            inv_label = open_hs.get("invoice_number") or "invoice"
            if 0 < rem <= 4:
                await _notify_admins("low_hours", f"⚠️ {cname} has only {rem}h left ({inv_label})",
                                     f"Pkg {pkg_h}h, used {used_h}h on {inv_label}. Consider package renewal.")
            elif rem <= 0:
                await _notify_admins("low_hours", f"🔴 {cname} package exhausted ({inv_label})",
                                     f"Used {used_h}h of {pkg_h}h on {inv_label}.")
    return doc

@api.put("/sessions/{sid}")
async def update_session(sid: str, payload: SessionIn, user=Depends(get_current_user)):
    sess = await db.sessions.find_one({"id": sid})
    if not sess:
        raise HTTPException(status_code=404, detail="Not found")
    if not _has_full_client_access(user) and not await _user_in_session_therapists(user, sess.get("therapist_ids") or []):
        raise HTTPException(status_code=403, detail="Forbidden")
    if not _session_editable_by_user(user, sess):
        raise HTTPException(
            status_code=403,
            detail="Sessions can only be edited within 1 hour of logging. Request an edit from admin if needed.",
        )
    if (payload.status or "").strip() == "No Service":
        raise HTTPException(status_code=400, detail="No Service is no longer available")
    payload = _apply_session_time_edits(payload)
    _require_same_day_session(user, payload.session_date)
    _require_session_log_fields(user, payload)
    resolved_uid = await _resolve_user_therapist_id(user) if user.get("role") == "therapist" else None
    therapist_ids = _merge_session_therapist_ids(payload.therapist_ids, user, resolved_uid)
    patch = payload.model_dump()
    patch["therapist_ids"] = therapist_ids
    patch["note"] = (payload.note or "").strip() or None
    await db.sessions.update_one({"id": sid}, {"$set": patch})
    updated = await db.sessions.find_one({"id": sid}, {"_id": 0})
    try:
        if updated.get("status") in ("Completed", "No Show", "Cancelled"):
            await _ensure_session_schedule_prep_markers(
                updated, user["id"], notes=patch.get("note"),
            )
        else:
            await _sync_prep_history_for_session(updated, user["id"], notes=patch.get("note"))
    except Exception:
        logger.exception("Prep history sync on session update failed")
    return updated

@api.delete("/sessions/{sid}")
async def delete_session(sid: str, user=Depends(get_current_user)):
    sess = await db.sessions.find_one({"id": sid})
    if not sess:
        return {"ok": True}
    if not _has_full_client_access(user) and not await _user_in_session_therapists(user, sess.get("therapist_ids") or []):
        raise HTTPException(status_code=403, detail="Forbidden")
    if not _session_editable_by_user(user, sess):
        raise HTTPException(
            status_code=403,
            detail="Sessions can only be edited within 1 hour of logging. Request an edit from admin if needed.",
        )
    try:
        await _cleanup_prep_for_deleted_session(sess)
    except Exception:
        logger.exception("cleanup prep for deleted session %s", sid)
    await db.sessions.delete_one({"id": sid})
    return {"ok": True}

# ------------------- Attendance Sheets (file upload, kept for backward compat) -------------------
@api.get("/clients/{cid}/sheets")
async def list_sheets(cid: str, user=Depends(get_current_user)):
    return await db.attendance_sheets.find({"client_id": cid}, {"_id": 0}).sort("page_number", 1).to_list(500)

@api.post("/clients/{cid}/sheets")
async def upload_sheet(cid: str,
                      title: str = Form(...),
                      session_date: str = Form(...),
                      therapist_id: Optional[str] = Form(None),
                      notes: Optional[str] = Form(None),
                      file: Optional[UploadFile] = File(None),
                      _=Depends(admin_only)):
    sid = str(uuid.uuid4())
    file_path = None
    file_name = None
    file_data = None
    if file:
        ext = Path(file.filename).suffix
        file_name = file.filename
        content = await file.read()
        file_path = f"{sid}{ext}"
        file_data = _persist_upload(file_path, content)
    last = await db.attendance_sheets.find_one({"client_id": cid}, sort=[("page_number", -1)])
    page_number = (last.get("page_number", 0) + 1) if last else 1
    doc = {"id": sid, "client_id": cid, "title": title, "session_date": session_date,
           "therapist_id": therapist_id, "notes": notes, "page_number": page_number,
           "file_name": file_name, "file_path": file_path, "file_data": file_data, "created_at": now_iso()}
    await db.attendance_sheets.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.delete("/sheets/{sid}")
async def delete_sheet(sid: str, _=Depends(ops_or_admin)):
    sheet = await db.attendance_sheets.find_one({"id": sid})
    if sheet and sheet.get("file_path"):
        fp = UPLOAD_DIR / sheet["file_path"]
        if fp.exists():
            fp.unlink()
    await db.attendance_sheets.delete_one({"id": sid})
    return {"ok": True}

@api.get("/sheets/{sid}/download")
async def download_sheet(sid: str, user=Depends(get_current_user)):
    sheet = await db.attendance_sheets.find_one({"id": sid}, {"_id": 0})
    if not sheet or not sheet.get("file_path"):
        raise HTTPException(status_code=404, detail="No file")
    content = _load_upload(sheet.get("file_path"), sheet.get("file_data"))
    if not content:
        raise HTTPException(status_code=404, detail=FILE_UNAVAILABLE_DETAIL)
    return _bytes_file_response(content, sheet.get("file_name") or sheet["file_path"])

# ------------------- Staff Purchases -------------------
async def _resolve_therapist_by_purchaser(purchaser: str) -> Optional[dict]:
    """Match spreadsheet purchaser name to a therapist record."""
    raw = (purchaser or "").strip()
    if not raw:
        return None
    low = raw.lower()
    first = low.split()[0] if low else ""
    first_aliases = {
        "walaa": "walaa", "maha": "maha", "jenan": "jenan",
        "fahda": "fahda", "fhdah": "fahda", "fahdah": "fahda",
        "fatima": "fatimah", "fatimah": "fatimah",
    }
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1}).to_list(500)
    if first in first_aliases:
        want = first_aliases[first]
        for t in therapists:
            tname = (t.get("name") or "").lower().replace("ms.", "").replace("ms ", "").strip()
            if want == "fahda" and ("fahd" in tname or "fhd" in tname):
                return t
            if want == "fatimah" and "fatim" in tname:
                return t
            if want in tname or tname.startswith(want):
                return t
    tokens = [t for t in re.split(r"[\s,]+", low) if t]
    best = None
    best_score = 0
    for t in therapists:
        tname = (t.get("name") or "").lower().replace("ms.", "").replace("ms ", "").strip()
        t_tokens = [x for x in re.split(r"[\s.]+", tname) if x]
        score = 0
        for tok in tokens:
            if tok in ("almuhaisin", "althunayan", "algadheeb", "algadeeb", "abueissa"):
                continue
            if any(tok in tt or tt.startswith(tok[:3]) for tt in t_tokens):
                score += 2
            if tok in tname:
                score += 3
        if score > best_score:
            best_score = score
            best = t
    return best if best_score >= 2 else None


def _normalize_purchase_status(s: str) -> str:
    v = (s or "pending").strip().lower()
    if v in PURCHASE_STATUSES:
        return v
    if v == "reimbursed":
        return "reimbursed"
    if v == "rejected":
        return "supervisor_rejected"
    if v == "approved":
        return "supervisor_approved"
    return "pending"


def _purchase_status_label(status: str) -> str:
    labels = {
        "pending": "Pending supervisor review",
        "supervisor_approved": "Approved by supervisor",
        "supervisor_rejected": "Rejected by supervisor",
        "pending_manager": "With manager Jenan",
        "manager_approved": "Approved by manager",
        "manager_rejected": "Rejected by manager",
        "approved": "Approved by supervisor",
        "rejected": "Rejected",
        "reimbursed": "Reimbursed",
    }
    return labels.get(status or "pending", status or "pending")


def _append_purchase_trail(existing: dict, user: dict, action: str, note: Optional[str] = None) -> list:
    trail = list(existing.get("approval_trail") or [])
    trail.append({
        "action": action,
        "by": user.get("id"),
        "by_name": _actor_display(user),
        "note": (note or "").strip() or None,
        "at": now_iso(),
    })
    return trail


_PURCHASE_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, "jul": 7,
    "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

PURCHASES_SHEET_ID = "10ZGq3ABQ1t-w32jrGZIu6Gxa2wevIJU2GLe9YWGdkIQ"
PURCHASES_SHEET_URL = f"https://docs.google.com/spreadsheets/d/{PURCHASES_SHEET_ID}/edit"


def _purchase_month_from_tab(tab_name: str) -> Optional[str]:
    name = (tab_name or "").strip()
    m = re.search(
        r"(January|February|March|April|May|June|July|August|September|October|November|December|"
        r"Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*[-_]?\s*(\d{4})?",
        name,
        re.I,
    )
    if not m:
        return None
    month = _PURCHASE_MONTH_MAP.get(m.group(1).lower()[:3])
    if not month:
        return None
    year = int(m.group(2)) if m.group(2) else datetime.now(timezone.utc).year
    return f"{year}-{month:02d}"


def _parse_loose_date(val) -> Optional[str]:
    if val is None:
        return None
    if hasattr(val, "isoformat"):
        try:
            return val.isoformat()[:10]
        except Exception:
            pass
    s = str(val).strip()
    if not s or s.lower() in ("none", "-"):
        return None
    m = re.match(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y}-{mo:02d}-{d:02d}"
    if re.match(r"\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return None


def _parse_purchase_total(val) -> tuple:
    if val is None:
        return None, ""
    if isinstance(val, (int, float)):
        return float(val), str(val)
    s = str(val).strip()
    nums = re.findall(r"[\d.]+", s.replace(",", ""))
    if nums:
        return float(nums[0]), s
    return None, s


def _is_emergent_website_item(item: str) -> bool:
    """Emergent / emergency website subscription — always filed under May in the official sheet."""
    s = (item or "").lower()
    if "emergent" in s and "website" in s:
        return True
    if "emergency" in s and "website" in s:
        return True
    return False


def _is_walaa_purchaser_name(name: str) -> bool:
    return bool(re.match(r"^walaa\b", (name or "").strip(), re.I))


def _walaa_emergent_website_month(purchase_date: str, item: str, purchaser_name: str) -> tuple:
    """Official sheet lists Walaa's Emergent Website under May — never June."""
    if not _is_emergent_website_item(item):
        return purchase_date[:10], purchase_date[:7]
    if not _is_walaa_purchaser_name(purchaser_name) and "walaa" not in (purchaser_name or "").lower():
        return purchase_date[:10], purchase_date[:7]
    year = purchase_date[:4] if len(purchase_date) >= 4 else str(datetime.now(timezone.utc).year)
    month_key = f"{year}-05"
    day = purchase_date[8:10] if len(purchase_date) >= 10 else "01"
    return f"{month_key}-{day}", month_key


def _read_purchases_xlsx(content: bytes, months: Optional[List[str]] = None) -> tuple:
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    records: List[dict] = []
    tabs_used: List[str] = []
    for tab in wb.sheetnames:
        purchase_month = _purchase_month_from_tab(tab)
        if not purchase_month:
            continue
        tabs_used.append(tab)
        if months and purchase_month not in months:
            continue
        ws = wb[tab]
        for r in range(3, (ws.max_row or 0) + 1):
            item = ws.cell(r, 2).value
            if item is None or not str(item).strip():
                continue
            item_str = str(item).strip()
            category = str(ws.cell(r, 3).value or "").strip()
            if not category:
                continue
            # Emergent/emergency website is May-only — skip if duplicated on June tab
            if purchase_month and purchase_month.endswith("-06") and _is_emergent_website_item(item_str):
                continue
            total, total_display = _parse_purchase_total(ws.cell(r, 7).value)
            if total is None and not str(ws.cell(r, 7).value or "").strip():
                continue
            purchaser = str(ws.cell(r, 8).value or "").strip()
            status_raw = str(ws.cell(r, 9).value or "").strip()
            reimb = _parse_loose_date(ws.cell(r, 10).value)
            row_no = ws.cell(r, 1).value
            try:
                row_no = int(float(row_no)) if row_no is not None else r - 2
            except (TypeError, ValueError):
                row_no = r - 2
            # Tab month is the accounting month; reimbursement may fall in an earlier month.
            if reimb and str(reimb)[:7] == purchase_month:
                purchase_date = str(reimb)[:10]
            else:
                purchase_date = f"{purchase_month}-01"
            records.append({
                "row_no": row_no,
                "item": str(item).strip(),
                "category": category,
                "description": str(ws.cell(r, 4).value or "").strip() or "-",
                "qty": str(ws.cell(r, 5).value or "1").strip(),
                "unit_price": str(ws.cell(r, 6).value or "").strip(),
                "total": total,
                "total_display": total_display or (str(total) if total is not None else ""),
                "purchaser": purchaser,
                "status": _normalize_purchase_status(status_raw),
                "reimbursement_date": reimb,
                "purchase_month": purchase_month,
                "purchase_date": purchase_date,
            })
    return records, tabs_used


async def _backfill_purchase_months() -> int:
    """Fill missing purchase_month from purchase_date / reimbursement_date (never overwrite tab month)."""
    fixed = 0
    items = await db.staff_purchases.find(
        {},
        {"_id": 0, "id": 1, "purchase_date": 1, "reimbursement_date": 1, "purchase_month": 1},
    ).to_list(5000)
    for doc in items:
        existing = (doc.get("purchase_month") or "").strip()
        if existing and len(existing) >= 7:
            continue
        src = doc.get("purchase_date") or doc.get("reimbursement_date")
        if not src or len(str(src)) < 7:
            continue
        pm_from_date = str(src)[:7]
        await db.staff_purchases.update_one(
            {"id": doc["id"]},
            {"$set": {"purchase_month": pm_from_date, "updated_at": now_iso()}},
        )
        fixed += 1
    return fixed


async def _repair_purchase_dates_from_month() -> int:
    """Align purchase_date with purchase_month when reimbursement fell in a different month."""
    fixed = 0
    items = await db.staff_purchases.find(
        {"purchase_month": {"$regex": r"^\d{4}-\d{2}$"}},
        {"_id": 0, "id": 1, "purchase_month": 1, "purchase_date": 1, "reimbursement_date": 1},
    ).to_list(5000)
    for doc in items:
        pm = (doc.get("purchase_month") or "").strip()[:7]
        pd = (doc.get("purchase_date") or "").strip()[:10]
        if not pm or len(pm) < 7:
            continue
        if pd and pd[:7] == pm:
            continue
        reimb = (doc.get("reimbursement_date") or "").strip()[:10]
        if reimb and reimb[:7] == pm:
            new_date = reimb
        else:
            new_date = f"{pm}-01"
        if pd == new_date:
            continue
        await db.staff_purchases.update_one(
            {"id": doc["id"]},
            {"$set": {"purchase_date": new_date, "updated_at": now_iso()}},
        )
        fixed += 1
    return fixed


def _purchase_month_key(doc: dict) -> str:
    pm = (doc.get("purchase_month") or "").strip()
    if len(pm) >= 7:
        return pm[:7]
    src = doc.get("purchase_date") or doc.get("reimbursement_date") or ""
    s = str(src)
    return s[:7] if len(s) >= 7 else ""


async def _ensure_purchases_sheet_synced() -> dict:
    """Import purchase months from the official Google Sheet when missing in the portal."""
    import httpx

    export_url = _google_sheet_export_url(PURCHASES_SHEET_URL)
    async with httpx.AsyncClient(follow_redirects=True, timeout=120.0) as client:
        resp = await client.get(export_url)
    if resp.status_code != 200:
        return {"skipped": f"sheet HTTP {resp.status_code}"}
    try:
        records, tabs_used = _read_purchases_xlsx(resp.content)
    except Exception as e:
        logger.warning("Purchases sheet parse failed during startup sync: %s", e)
        return {"skipped": "parse failed", "error": str(e)}
    if not records:
        return {"skipped": "no records parsed", "tabs_found": tabs_used}

    sheet_months = {r["purchase_month"] for r in records if r.get("purchase_month")}
    db_months: set = set()
    items = await db.staff_purchases.find(
        {},
        {"_id": 0, "purchase_month": 1, "purchase_date": 1, "reimbursement_date": 1},
    ).to_list(5000)
    for doc in items:
        pm = _purchase_month_key(doc)
        if pm:
            db_months.add(pm)

    missing = sorted(sheet_months - db_months)
    if not missing:
        sheet_count = await db.staff_purchases.count_documents({"sync_source": "google_sheet"})
        if sheet_count < 12:
            filtered = records
            missing = sorted(sheet_months)
        else:
            return {
                "skipped": "all sheet months present",
                "sheet_months": sorted(sheet_months),
                "db_months": sorted(db_months),
            }
    else:
        filtered = [r for r in records if r.get("purchase_month") in missing]

    result = await _upsert_purchases_from_sheet(filtered)
    await _fix_walaa_purchase_month_mismatch()
    pm_fixed = await _backfill_purchase_months()
    pd_fixed = await _repair_purchase_dates_from_month()
    result["missing_months_synced"] = missing
    result["tabs_found"] = tabs_used
    result["purchase_month_backfilled"] = pm_fixed
    result["purchase_date_repaired"] = pd_fixed
    return result


async def _fix_walaa_purchase_month_mismatch():
    """Walaa's Emergent/emergency website payment belongs in May (month 5), not June."""
    walaa = await db.therapists.find_one(
        {
            "$or": [
                {"email": {"$regex": r"walaa@boostgrowthsa\.com$", "$options": "i"}},
                {"email": {"$regex": r"wabuissa@boostgrowthsa\.com$", "$options": "i"}},
                {"key": "mswalaa"},
                {"name": {"$regex": r"^ms\.?\s*walaa\b", "$options": "i"}},
            ]
        },
        {"_id": 0, "id": 1},
    )
    walaa_id = walaa.get("id") if walaa else None
    owner_filter: List[dict] = [
        {"purchaser_name": {"$regex": r"^walaa\b", "$options": "i"}},
        {"therapist_name": {"$regex": r"walaa", "$options": "i"}},
    ]
    if walaa_id:
        owner_filter.insert(0, {"therapist_id": walaa_id})

    items = await db.staff_purchases.find(
        {
            "$and": [
                {"$or": owner_filter},
                {"item": {"$regex": r"emergent|emergency", "$options": "i"}},
            ]
        },
        {"_id": 0, "id": 1, "purchase_date": 1, "purchase_month": 1, "item": 1},
    ).to_list(20)
    items = [d for d in items if _is_emergent_website_item(d.get("item"))]

    has_may_row = any((d.get("purchase_month") or "").endswith("-05") for d in items)
    for doc in items:
        old_month = doc.get("purchase_month") or ""
        if old_month.endswith("-05"):
            continue
        if old_month.endswith("-06") and has_may_row:
            await db.staff_purchases.delete_one({"id": doc["id"]})
            logger.info(
                "Removed duplicate Walaa Emergent Website in %s (May row kept)",
                old_month,
            )
            continue
        if not old_month.endswith("-06"):
            continue
        year = old_month[:4] if len(old_month) >= 4 else str(datetime.now(timezone.utc).year)
        new_month = f"{year}-05"
        pd = (doc.get("purchase_date") or f"{old_month}-01")[:10]
        new_date = f"{new_month}{pd[7:]}" if len(pd) >= 10 else f"{new_month}-01"
        await db.staff_purchases.update_one(
            {"id": doc["id"]},
            {"$set": {"purchase_month": new_month, "purchase_date": new_date, "updated_at": now_iso()}},
        )
        logger.info(
            "Fixed Walaa Emergent Website month %s → %s (%s)",
            old_month,
            new_month,
            doc.get("item"),
        )


async def _upsert_purchases_from_sheet(records: List[dict]) -> dict:
    months = sorted({r["purchase_month"] for r in records if r.get("purchase_month")})
    if months:
        await db.staff_purchases.delete_many({
            "purchase_month": {"$in": months},
            "$or": [{"sync_source": "google_sheet"}, {"imported": True}],
        })
    # Drop mis-filed sheet rows (e.g. June tab rows moved to January by old backfill).
    for rec in records:
        t = await _resolve_therapist_by_purchaser(rec.get("purchaser") or "")
        if not t:
            continue
        match_q: dict = {
            "therapist_id": t["id"],
            "item": rec.get("item"),
            "$or": [{"sync_source": "google_sheet"}, {"imported": True}],
        }
        if rec.get("row_no") is not None:
            match_q["row_no"] = rec.get("row_no")
        elif rec.get("total") is not None:
            match_q["total"] = rec.get("total")
        await db.staff_purchases.delete_many(match_q)
    # Remove known duplicates that should not appear in June
    june_dupes = await db.staff_purchases.find(
        {"purchase_month": {"$regex": "-06$"}, "item": {"$regex": "emergent|emergency", "$options": "i"}},
        {"_id": 0, "id": 1, "item": 1},
    ).to_list(50)
    for doc in june_dupes:
        if _is_emergent_website_item(doc.get("item")):
            await db.staff_purchases.delete_one({"id": doc["id"]})
    inserted = 0
    skipped = 0
    for rec in records:
        t = await _resolve_therapist_by_purchaser(rec.get("purchaser") or "")
        if not t:
            skipped += 1
            continue
        doc = {
            "id": str(uuid.uuid4()),
            "row_no": rec.get("row_no"),
            "therapist_id": t["id"],
            "therapist_name": t.get("name"),
            "purchaser_name": rec.get("purchaser"),
            "item": rec.get("item"),
            "category": rec.get("category"),
            "description": rec.get("description") or "",
            "qty": rec.get("qty") or "1",
            "unit_price": rec.get("unit_price") or "",
            "total": rec.get("total"),
            "total_display": rec.get("total_display") or "",
            "status": rec.get("status") or "pending",
            "reimbursement_date": rec.get("reimbursement_date"),
            "purchase_date": rec.get("purchase_date"),
            "purchase_month": rec.get("purchase_month"),
            "sync_source": "google_sheet",
            "updated_at": now_iso(),
            "created_at": now_iso(),
        }
        await db.staff_purchases.insert_one(doc)
        inserted += 1
    return {"inserted": inserted, "skipped": skipped, "months": months, "total_rows": len(records)}


async def _therapist_email(therapist_id: str) -> Optional[str]:
    t = await db.therapists.find_one({"id": therapist_id}, {"_id": 0, "email": 1})
    email = (t or {}).get("email") or ""
    return email.strip() or None


def _schedule_cell_date_iso(cell: dict) -> Optional[str]:
    ws = cell.get("week_start")
    day = cell.get("day")
    if ws is None or day is None:
        return None
    try:
        base = datetime.fromisoformat(str(ws)[:10])
        return (base + timedelta(days=int(day))).strftime("%Y-%m-%d")
    except Exception:
        return None


async def _process_unprepared_session_alerts(force: bool = False) -> dict:
    """After midnight: alert ops leads about yesterday's scheduled sessions not prepared/logged."""
    now = datetime.now(timezone.utc)
    today = now.date().isoformat()
    yesterday = (now.date() - timedelta(days=1)).isoformat()
    meta = await db.meta.find_one({"key": "unprepared_alerts_last_run"})
    if not force and meta and (meta.get("date") or "")[:10] == today:
        return {"skipped": "already_ran_today", "alerts": 0}
    target_date = yesterday
    cells = await db.schedule_cells.find(
        {
            "child_name": {"$exists": True, "$nin": ["", None]},
            "state": {"$nin": ["cancel_therapist"]},
            "service_code": {"$nin": ["LEAVE", "BREAK", "AVC", "AVAILABLE", ""]},
        },
        {"_id": 0},
    ).to_list(8000)
    sessions = await db.sessions.find(
        {"status": "Completed", "session_date": target_date},
        {"_id": 0, "session_date": 1, "client_id": 1, "therapist_ids": 1},
    ).to_list(25000)
    preps = await db.prep_history.find(
        {"session_date": target_date},
        {"_id": 0, "client_id": 1, "therapist_id": 1, "session_date": 1},
    ).to_list(25000)
    logged = set()
    for s in sessions:
        sd = (s.get("session_date") or "")[:10]
        cid = s.get("client_id")
        for tid in s.get("therapist_ids") or []:
            logged.add((sd, tid, cid))
    for p in preps:
        logged.add((
            (p.get("session_date") or "")[:10],
            p.get("therapist_id"),
            p.get("client_id"),
        ))
    clients = await db.clients.find(_active_client_filter(), {"_id": 0, "id": 1, "name": 1}).to_list(600)
    name_to_id = {_normalize_intake_name(c.get("name") or ""): c["id"] for c in clients}
    id_to_name = {c["id"]: c.get("name") or "" for c in clients}
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "key": 1}).to_list(300)
    t_names = {t["id"]: therapist_schedule_display_name(t) for t in therapists}
    alerts = 0
    for cell in cells:
        slot_date = _schedule_cell_date_iso(cell)
        if slot_date != target_date:
            continue
        child = (cell.get("child_name") or "").strip()
        if not child:
            continue
        cid = name_to_id.get(_normalize_intake_name(child))
        tid = cell.get("therapist_id")
        if not cid or not tid:
            continue
        if (slot_date, tid, cid) in logged:
            continue
        dedupe = f"unprepared:{cell.get('id')}:{slot_date}"
        if await db.notifications.find_one({"type": "unprepared_session", "link": dedupe}):
            continue
        slot = cell.get("time_slot") or ""
        tname = t_names.get(tid) or "Therapist"
        cname = id_to_name.get(cid) or child
        msg = f"{tname} did not prepare/log {cname} — {slot_date} {slot}".strip()
        await _notify_ops_leads(
            "unprepared_session",
            "Unprepared session (end of day)",
            msg,
            link=dedupe,
            therapist_id=tid,
            client_id=cid,
            session_date=slot_date,
            schedule_cell_id=cell.get("id"),
        )
        await _notify_admins("unprepared_session", "Unprepared session (end of day)", msg)
        alerts += 1
    await db.meta.update_one(
        {"key": "unprepared_alerts_last_run"},
        {"$set": {"date": today, "at": now_iso(), "alerts": alerts, "target_date": target_date}},
        upsert=True,
    )
    return {"alerts": alerts, "date": today, "target_date": target_date}


async def _get_purchase_reminder_settings() -> dict:
    doc = await db.purchase_reminder_settings.find_one({"id": "default"}, {"_id": 0})
    if not doc:
        doc = {
            "id": "default",
            "day_of_month": 25,
            "enabled": True,
            "therapist_ids": [],
            "last_sent_month": None,
        }
    return doc


async def _send_purchase_reminders(force: bool = False) -> dict:
    settings = await _get_purchase_reminder_settings()
    if not settings.get("enabled") and not force:
        return {"sent": 0, "skipped": "disabled"}
    today = datetime.now(timezone.utc)
    month_key = today.strftime("%Y-%m")
    if not force and settings.get("last_sent_month") == month_key:
        return {"sent": 0, "skipped": "already_sent_this_month"}
    day = int(settings.get("day_of_month") or 25)
    if not force and today.day < day:
        return {"sent": 0, "skipped": "before_reminder_day"}
    tids = settings.get("therapist_ids") or []
    if not tids:
        return {"sent": 0, "skipped": "no_therapists_selected"}
    sent = 0
    email_results = []
    title = "Monthly purchase log reminder"
    message = (
        f"Please log your purchases for {today.strftime('%B %Y')} before month-end. "
        "Open the portal → Purchases to add items you bought for work."
    )
    await _reload_email_settings_from_db()
    provider_ok = bool(
        _mailgun_configured() or _brevo_configured() or _resend_configured() or _smtp_configured()
    )
    for tid in tids:
        await _notify(tid, "purchase_reminder", title, message)
        sent += 1
        email = await _therapist_email(tid)
        if email:
            er = await _send_email_stub(email, title, message)
            email_results.append({
                "to": email,
                "status": er.get("status"),
                "error": er.get("error"),
                "hint": er.get("hint"),
            })
    await db.purchase_reminder_settings.update_one(
        {"id": "default"},
        {"$set": {"last_sent_month": month_key, "updated_at": now_iso()}},
        upsert=True,
    )
    return {
        "sent": sent,
        "month": month_key,
        "email_results": email_results,
        "provider_configured": provider_ok,
    }


@api.post("/import/purchases-google")
async def import_purchases_google(body: dict = None, user=Depends(get_current_user)):
    """Sync staff purchases from the official Google Sheet (Jan–Jul tabs)."""
    if not (_is_jenan(user) or _is_walaa_ops(user) or _is_portal_admin(user)):
        raise HTTPException(status_code=403, detail="Not allowed")
    import httpx
    body = body or {}
    sheet_url = (body.get("url") or body.get("sheet_url") or PURCHASES_SHEET_URL).strip()
    months = body.get("months")
    export_url = _google_sheet_export_url(sheet_url)
    async with httpx.AsyncClient(follow_redirects=True, timeout=120.0) as client:
        resp = await client.get(export_url)
    if resp.status_code != 200:
        raise HTTPException(
            status_code=400,
            detail=f"Could not download purchases sheet (HTTP {resp.status_code}).",
        )
    try:
        records, tabs_used = _read_purchases_xlsx(resp.content, months=months)
    except Exception as e:
        logger.exception("Purchases Google Sheet parse failed")
        raise HTTPException(status_code=400, detail=f"Could not parse purchases sheet: {e}")
    result = await _upsert_purchases_from_sheet(records)
    await _fix_walaa_purchase_month_mismatch()
    await _repair_purchase_dates_from_month()
    result["tabs_found"] = tabs_used
    result["message"] = (
        f"Imported {result['inserted']} purchases ({result.get('skipped', 0)} skipped — unknown purchaser)"
        + (f" · tabs: {', '.join(tabs_used)}" if tabs_used else "")
    )
    result["sheet_url"] = PURCHASES_SHEET_URL
    return result


@api.get("/purchases/categories")
async def purchase_categories(_=Depends(get_current_user)):
    return PURCHASE_CATEGORIES


@api.get("/purchases/reminder-settings")
async def get_purchase_reminder_settings(_=Depends(ops_or_admin)):
    return await _get_purchase_reminder_settings()


@api.put("/purchases/reminder-settings")
async def update_purchase_reminder_settings(body: PurchaseReminderSettingsIn, _=Depends(ops_or_admin)):
    day = max(1, min(28, int(body.day_of_month or 25)))
    doc = {
        "id": "default",
        "day_of_month": day,
        "enabled": bool(body.enabled),
        "therapist_ids": [t for t in (body.therapist_ids or []) if t],
        "updated_at": now_iso(),
    }
    existing = await db.purchase_reminder_settings.find_one({"id": "default"}, {"_id": 0, "last_sent_month": 1})
    if existing:
        doc["last_sent_month"] = existing.get("last_sent_month")
    await db.purchase_reminder_settings.update_one({"id": "default"}, {"$set": doc}, upsert=True)
    return doc


@api.post("/purchases/send-reminders")
async def send_purchase_reminders(_=Depends(ops_or_admin)):
    return await _send_purchase_reminders(force=True)


@api.get("/purchases")
async def list_purchases(
    therapist_id: Optional[str] = None,
    status: Optional[str] = None,
    month: Optional[str] = None,
    user=Depends(get_current_user),
):
    q: dict = {}
    if _can_view_all_purchases(user):
        if therapist_id:
            q["therapist_id"] = therapist_id
    else:
        tid = await _resolve_user_therapist_id(user) or user.get("id")
        q["therapist_id"] = tid
    if status:
        q["status"] = _normalize_purchase_status(status)
    if month:
        month = month.strip()
        month_clauses = [{"purchase_month": month}]
        if len(month) >= 7 and month[4] == "-":
            mm = month[5:7]
            month_clauses.append({"purchase_month": {"$regex": f"-{re.escape(mm)}$"}})
            month_clauses.append({
                "$and": [
                    {"$or": [
                        {"purchase_month": {"$exists": False}},
                        {"purchase_month": None},
                        {"purchase_month": ""},
                    ]},
                    {"$or": [
                        {"purchase_date": {"$regex": f"^{re.escape(month)}"}},
                        {"reimbursement_date": {"$regex": f"^{re.escape(month)}"}},
                    ]},
                ]
            })
        q["$or"] = month_clauses
    items = await db.staff_purchases.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "key": 1}).to_list(500)
    t_by_id = {t["id"]: t for t in therapists}
    for it in items:
        pm = _purchase_month_key(it)
        if pm:
            it["purchase_month"] = pm
        t = t_by_id.get(it.get("therapist_id"))
        if t:
            it["therapist_name"] = therapist_schedule_display_name(t)
    return items


@api.post("/purchases")
async def create_purchase(payload: PurchaseIn, user=Depends(get_current_user)):
    own_tid = await _resolve_user_therapist_id(user) or user.get("id")
    requested_tid = (payload.therapist_id or "").strip() or None
    if requested_tid and _can_view_all_purchases(user):
        th = await db.therapists.find_one({"id": requested_tid}, {"_id": 0, "id": 1, "name": 1, "key": 1})
        if not th:
            raise HTTPException(status_code=400, detail="Therapist not found")
        tid = th["id"]
        display = therapist_schedule_display_name(th)
        purchaser_name = display
        therapist_name = display
    else:
        tid = own_tid
        if not tid:
            raise HTTPException(status_code=403, detail="Therapist profile required")
        th = await db.therapists.find_one({"id": tid}, {"_id": 0, "id": 1, "name": 1, "key": 1})
        display = therapist_schedule_display_name(th or user)
        purchaser_name = display
        therapist_name = display
    item = (payload.item or "").strip()
    category = (payload.category or "").strip()
    if not item or not category:
        raise HTTPException(status_code=400, detail="item and category required")
    purchase_date = (payload.purchase_date or now_iso()[:10])[:10]
    purchase_date, purchase_month = _walaa_emergent_website_month(purchase_date, item, purchaser_name or "")
    line_items = []
    if payload.line_items:
        for li in payload.line_items:
            li_item = (li.item or "").strip()
            if not li_item:
                continue
            line_items.append({
                "item": li_item,
                "qty": (li.qty or "1").strip(),
                "unit_price": (li.unit_price or "").strip(),
                "total": float(li.total) if li.total is not None else None,
            })
    if line_items and not item:
        item = " · ".join(x["item"] for x in line_items[:3])
        if len(line_items) > 3:
            item += f" (+{len(line_items) - 3} more)"
    line_total = sum(float(x["total"] or 0) for x in line_items if x.get("total") is not None)
    doc_total = float(payload.total) if payload.total is not None else (line_total or None)
    doc = {
        "id": str(uuid.uuid4()),
        "therapist_id": tid,
        "therapist_name": therapist_name,
        "purchaser_name": purchaser_name,
        "item": item,
        "category": category,
        "description": (payload.description or "").strip(),
        "qty": (payload.qty or "1").strip(),
        "unit_price": (payload.unit_price or "").strip(),
        "total": doc_total,
        "total_display": str(doc_total) if doc_total is not None else "",
        "line_items": line_items or None,
        "status": "pending",
        "approval_trail": [],
        "reimbursement_date": None,
        "purchase_date": purchase_date,
        "purchase_month": purchase_month,
        "notes": (payload.notes or "").strip() or None,
        "invoice_file_path": None,
        "invoice_file_name": None,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.staff_purchases.insert_one(doc)
    doc.pop("_id", None)
    await _notify_purchase_submitted(purchaser_name, item, category)
    return doc


@api.put("/purchases/{pid}")
async def update_purchase(pid: str, payload: PurchaseUpdate, user=Depends(get_current_user)):
    existing = await db.staff_purchases.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Purchase not found")
    tid = await _resolve_user_therapist_id(user) or user.get("id")
    is_owner = existing.get("therapist_id") == tid
    is_supervisor = _can_supervisor_review_purchases(user)
    is_manager = _can_manager_finalize_purchases(user)
    if not is_owner and not is_supervisor and not is_manager:
        raise HTTPException(status_code=403, detail="Forbidden")
    if is_owner and not is_supervisor and not is_manager and existing.get("status") != "pending":
        raise HTTPException(status_code=403, detail="Only pending entries can be edited")
    patch = {k: v for k, v in payload.model_dump().items() if v is not None and k not in ("forward_to_manager", "supervisor_note")}
    prev_status = existing.get("status") or "pending"
    new_status = patch.get("status")
    note = payload.supervisor_note
    trail = list(existing.get("approval_trail") or [])
    if new_status is not None:
        new_status = _normalize_purchase_status(new_status)
        if is_manager and not is_supervisor:
            if new_status in ("manager_approved", "approved", "reimbursed"):
                new_status = "manager_approved" if new_status != "reimbursed" else "reimbursed"
            elif new_status in ("manager_rejected", "rejected"):
                new_status = "manager_rejected"
            elif prev_status not in ("pending_manager", "manager_approved", "supervisor_approved", "approved"):
                raise HTTPException(status_code=403, detail="Manager can only act on forwarded purchases")
        elif is_supervisor and not is_manager:
            if new_status in ("supervisor_approved", "approved"):
                new_status = "supervisor_approved"
            elif new_status in ("supervisor_rejected", "rejected"):
                new_status = "supervisor_rejected"
            elif new_status == "pending":
                new_status = "pending"
            else:
                raise HTTPException(status_code=400, detail="Supervisor: use approved, rejected, or pending")
        elif not (is_supervisor or is_manager):
            raise HTTPException(status_code=403, detail="Status updates require supervisor or manager access")
        patch["status"] = new_status
        trail = _append_purchase_trail(existing, user, new_status, note)
        patch["approval_trail"] = trail
        if note:
            patch["supervisor_note"] = note
    if payload.forward_to_manager and is_supervisor:
        if prev_status not in ("supervisor_approved", "approved", "pending_manager") and patch.get("status") not in ("supervisor_approved",):
            if patch.get("status") != "supervisor_approved":
                patch["status"] = "supervisor_approved"
        patch["status"] = "pending_manager"
        trail = _append_purchase_trail({**existing, "approval_trail": trail}, user, "forwarded_to_manager", note)
        patch["approval_trail"] = trail
        jenan_id = await _jenan_therapist_id()
        if jenan_id:
            await _notify(
                jenan_id,
                "purchase_forwarded",
                "Purchase forwarded for final approval",
                f"{existing.get('purchaser_name') or 'Therapist'}: {existing.get('item')} — awaiting your review",
                link="/purchases",
            )
    if "purchase_date" in patch:
        item_name = patch.get("item") or existing.get("item") or ""
        purchaser = existing.get("purchaser_name") or existing.get("therapist_name") or ""
        pd, pm = _walaa_emergent_website_month(patch["purchase_date"][:10], item_name, purchaser)
        patch["purchase_date"] = pd
        patch["purchase_month"] = pm
    patch["updated_at"] = now_iso()
    await db.staff_purchases.update_one({"id": pid}, {"$set": patch})
    updated = await db.staff_purchases.find_one({"id": pid}, {"_id": 0})
    final_status = updated.get("status") or prev_status
    if final_status != prev_status and existing.get("therapist_id"):
        trail_text = " → ".join(
            _purchase_status_label(t.get("action")) for t in (updated.get("approval_trail") or [])[-3:]
        )
        await _notify(
            existing["therapist_id"],
            "purchase_update",
            "Purchase status updated",
            f"\"{existing.get('item')}\" — {_purchase_status_label(final_status)}"
            + (f" ({trail_text})" if trail_text else ""),
            link="/requests",
        )
    if is_manager and patch.get("status") == "reimbursed" and prev_status != "reimbursed":
        await _notify(
            existing["therapist_id"],
            "purchase_reimbursed",
            "Purchase reimbursed",
            f"Your purchase \"{existing.get('item')}\" has been marked reimbursed.",
        )
    if updated:
        updated["status_label"] = _purchase_status_label(updated.get("status"))
    return updated


@api.post("/purchases/{pid}/invoice")
async def upload_purchase_invoice(pid: str, file: UploadFile = File(...), user=Depends(get_current_user)):
    existing = await db.staff_purchases.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Purchase not found")
    tid = await _resolve_user_therapist_id(user) or user.get("id")
    if existing.get("therapist_id") != tid and not _can_view_all_purchases(user):
        raise HTTPException(status_code=403, detail="Forbidden")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="File required")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")
    ext = Path(file.filename).suffix.lower() or ".pdf"
    if ext not in (".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif"):
        raise HTTPException(status_code=400, detail="PDF or image only")
    stored = f"purchase_{pid}{ext}"
    file_data = _persist_upload(stored, content)
    await db.staff_purchases.update_one(
        {"id": pid},
        {"$set": {
            "invoice_file_path": stored,
            "invoice_file_name": file.filename,
            "invoice_file_data": file_data,
            "updated_at": now_iso(),
        }},
    )
    updated = await db.staff_purchases.find_one({"id": pid}, {"_id": 0, "invoice_file_data": 0})
    if updated:
        updated["invoice_url"] = f"/api/purchases/{pid}/invoice"
    return updated


@api.get("/purchases/{pid}/invoice")
async def get_purchase_invoice(pid: str, user=Depends(get_current_user)):
    existing = await db.staff_purchases.find_one({"id": pid}, {"_id": 0})
    if not existing or not existing.get("invoice_file_data"):
        raise HTTPException(status_code=404, detail="Invoice not found")
    tid = await _resolve_user_therapist_id(user) or user.get("id")
    if existing.get("therapist_id") != tid and not _can_view_all_purchases(user):
        raise HTTPException(status_code=403, detail="Forbidden")
    return _file_response_from_data(
        existing.get("invoice_file_data"),
        existing.get("invoice_file_name") or "invoice",
    )


@api.delete("/purchases/{pid}")
async def delete_purchase(pid: str, _=Depends(ops_or_admin)):
    r = await db.staff_purchases.delete_one({"id": pid})
    if not r.deleted_count:
        raise HTTPException(status_code=404, detail="Purchase not found")
    return {"ok": True}


# ------------------- Requests -------------------
def _enrich_request_attachment(req: dict) -> dict:
    if _request_has_attachment(req):
        req["attachment_url"] = f"/api/requests/{req['id']}/attachment"
    else:
        req.setdefault("attachment_url", None)
    return _strip_file_data(req)


@api.get("/requests")
async def list_requests(scope: Optional[str] = None, user=Depends(get_current_user)):
    """List requests. Default: caller's own. scope=staff: manager/HR staff queue (not own rows for Jenan)."""
    scope_norm = (scope or "").strip().lower()
    if scope_norm == "staff":
        if not _can_staff_request_scope(user):
            raise HTTPException(status_code=403, detail="Staff request access required")
        q: dict = {}
        if _is_jenan(user) and not _is_portal_admin(user) and not _is_hr_ops(user):
            q = {"therapist_id": {"$ne": user["id"]}}
    else:
        q = {"therapist_id": user["id"]}
    items = await db.requests.find(q, {"_id": 0, "attachment_file_data": 0}).sort("created_at", -1).to_list(500)
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "key": 1}).to_list(500)
    t_by_id = {t["id"]: t for t in therapists}
    out = []
    for it in items:
        t = t_by_id.get(it.get("therapist_id"))
        if t:
            it["therapist_name"] = therapist_schedule_display_name(t)
        out.append(_enrich_request_attachment(it))
    return out

@api.post("/requests")
async def create_request(payload: RequestIn, user=Depends(get_current_user)):
    if user.get("role") != "therapist":
        raise HTTPException(status_code=403, detail="Therapist only")
    rt = (payload.request_type or "general").strip()
    if rt == "other" and not (payload.description or "").strip():
        raise HTTPException(status_code=400, detail="Description is required for Other requests")
    title = (payload.title or "").strip()
    if rt == "companies" and not title:
        title = "Companies request"
    elif rt == "other" and not title:
        title = "Other request"
    rid = str(uuid.uuid4())
    initial_status = "pending_manager"
    if payload.requires_attachment:
        initial_status = "pending_attachment"
    th = await db.therapists.find_one({"id": user["id"]}, {"_id": 0, "id": 1, "name": 1, "key": 1})
    display = therapist_schedule_display_name(th or user)
    body = payload.model_dump()
    if title:
        body["title"] = title
    doc = {"id": rid, "therapist_id": user["id"], "therapist_name": display,
           **body, "status": initial_status, "admin_note": None,
           "created_at": now_iso(), "updated_at": now_iso(),
           "timeline": [{"event": "submitted", "at": now_iso(), "by": user.get("name")}]}
    await db.requests.insert_one(doc)
    doc.pop("_id", None)
    msg = f"{display}: {title or body.get('title', '')} (priority: {payload.priority})"
    await _notify_request_submitted(
        f"New {payload.request_type} request",
        msg,
        email_subject=f"New staff request: {title or body.get('title', '')}",
    )
    return _enrich_request_attachment(doc)


@api.post("/requests/{rid}/attachment")
async def add_request_attachment(
    rid: str,
    file: UploadFile = File(...),
    report_date: Optional[str] = Form(None),
    user=Depends(get_current_user),
):
    """Attach a file to an existing general request (submitted by the same therapist)."""
    if user.get("role") != "therapist":
        raise HTTPException(status_code=403, detail="Therapist only")
    req = await db.requests.find_one({"id": rid}, {"_id": 0})
    if not req or req.get("therapist_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Request not found")
    if req.get("request_type") == "attachment":
        raise HTTPException(status_code=400, detail="Use the dedicated attachment request type")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="File required")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")
    ext = Path(file.filename).suffix.lower() or ".pdf"
    if ext not in (".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif", ".doc", ".docx"):
        raise HTTPException(status_code=400, detail="PDF, image, or Word document only")
    stored = f"req_{rid}{ext}"
    file_data = _persist_upload(stored, content)
    report_date = (report_date or "").strip()[:10] or None
    patch = {
        "attachment_file_path": stored,
        "attachment_file_name": file.filename,
        "attachment_file_data": file_data,
        "updated_at": now_iso(),
    }
    if report_date:
        patch["report_date"] = report_date
    if req.get("status") == "pending_attachment" or (req.get("requires_attachment") and not req.get("attachment_file_path")):
        patch["status"] = "pending_manager"
        timeline = list(req.get("timeline") or [])
        timeline.append({"event": "attachment_uploaded", "at": now_iso(), "by": user.get("name")})
        patch["timeline"] = timeline
    await db.requests.update_one({"id": rid}, {"$set": patch})
    updated = {**req, **patch}
    return _enrich_request_attachment(updated)


@api.post("/requests/upload-attachment")
async def upload_request_attachment(
    report_date: str = Form(...),
    title: Optional[str] = Form(None),
    file: UploadFile = File(...),
    user=Depends(get_current_user),
):
    if user.get("role") != "therapist":
        raise HTTPException(status_code=403, detail="Therapist only")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="File required")
    report_date = (report_date or "").strip()[:10]
    if not report_date:
        raise HTTPException(status_code=400, detail="Report date required")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")
    ext = Path(file.filename).suffix.lower() or ".pdf"
    if ext not in (".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif", ".doc", ".docx"):
        raise HTTPException(status_code=400, detail="PDF, image, or Word document only")
    rid = str(uuid.uuid4())
    stored = f"req_{rid}{ext}"
    file_data = _persist_upload(stored, content)
    doc = {
        "id": rid,
        "therapist_id": user["id"],
        "therapist_name": user.get("name"),
        "title": (title or "").strip() or file.filename or "Report attachment",
        "description": None,
        "request_type": "attachment",
        "report_date": report_date,
        "attachment_file_path": stored,
        "attachment_file_name": file.filename,
        "attachment_file_data": file_data,
        "priority": "normal",
        "status": "pending_manager",
        "admin_note": None,
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "timeline": [{"event": "submitted", "at": now_iso(), "by": user.get("name")}],
    }
    await db.requests.insert_one(doc)
    doc.pop("_id", None)
    msg = f"{user.get('name')}: {doc['title']} (report date: {report_date})"
    await _notify_request_submitted(
        "New report attachment",
        msg,
        email_subject=f"New report attachment: {doc['title']}",
    )
    return _enrich_request_attachment(doc)


LEAVE_BALANCE_SHEET_URL = os.environ.get(
    "LEAVE_BALANCE_SHEET_URL",
    "https://docs.google.com/spreadsheets/d/10Y2lmPEPtzWKZeP2SGIGbbdzTDFmNt_D3oQhG6ko3WQ/edit",
)
_leave_grid_cache: Dict[str, tuple] = {}
LEAVE_GRID_CACHE_SECS = 300
_SKIP_SHEET_EMPLOYEE_NAMES = frozenset(
    {"done", "approved", "ongoing", "cancelled", "canceled", "employee", "employee "}
)


def _sheet_cell_date(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s or None


def _norm_person_name(name: str) -> str:
    n = re.sub(r"[^a-z0-9]", "", (name or "").lower().replace("ms.", "").replace("ms ", ""))
    # Google Sheet spelling variants → portal display names
    aliases = {
        "manaldossery": "manaldosery",
        "manaldosari": "manaldosery",
        "shathaalhammami": "shathaalhammami",
    }
    return aliases.get(n, n)


def _parse_leave_balance_sheet(wb, year: int) -> List[dict]:
    sheet_name = f"For {year}"
    if sheet_name not in wb.sheetnames:
        year_sheets = [s for s in wb.sheetnames if re.match(r"^For \d{4}$", s)]
        sheet_name = year_sheets[-1] if year_sheets else wb.sheetnames[0]
    ws = wb[sheet_name]
    agg: Dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        cells = list(row)
        for block in range(3):
            base = block * 10
            if base >= len(cells):
                break
            chunk = cells[base:base + 10]
            while len(chunk) < 10:
                chunk.append(None)
            emp_id, name, _start, _end, days, vtype, status, remaining, join, _notes = chunk[:10]
            if not name:
                continue
            name_s = str(name).strip()
            if name_s.lower() in _SKIP_SHEET_EMPLOYEE_NAMES:
                continue
            if not vtype:
                continue
            stat = str(status or "").strip().lower()
            if stat in ("cancelled", "canceled"):
                continue
            try:
                day_val = float(days or 0)
            except (TypeError, ValueError):
                day_val = 0.0
            key = _norm_person_name(name_s)
            if not key:
                continue
            if key not in agg:
                agg[key] = {
                    "employee_id": str(emp_id).strip() if emp_id is not None else None,
                    "name": name_s,
                    "annual_days": 0.0,
                    "sick_days": 0.0,
                    "permission_count": 0,
                    "permission_days": 0.0,
                    "unpaid_days": 0.0,
                    "remaining_raw": None,
                    "join_date": None,
                }
            rec = agg[key]
            vt = str(vtype).strip()
            if vt == "Annual":
                rec["annual_days"] += day_val
            elif vt in ("Sickleave", "Sick"):
                rec["sick_days"] += day_val
            elif vt == "Permission":
                rec["permission_count"] += 1
                rec["permission_days"] += day_val
            elif vt == "Unpaid":
                rec["unpaid_days"] += day_val
            if remaining is not None and str(remaining).strip():
                rec["remaining_raw"] = str(remaining).strip()
            if join and not rec["join_date"]:
                rec["join_date"] = _sheet_cell_date(join)
    out = []
    for rec in agg.values():
        out.append({
            **rec,
            "annual_days": round(rec["annual_days"], 1),
            "sick_days": round(rec["sick_days"], 1),
            "permission_days": round(rec["permission_days"], 1),
            "unpaid_days": round(rec["unpaid_days"], 1),
        })
    out.sort(key=lambda r: (r.get("name") or "").lower())
    return out


def _match_sheet_row_to_therapist(row: dict, therapists: List[dict]) -> Optional[dict]:
    row_norm = _norm_person_name(row.get("name"))
    if not row_norm:
        return None
    best = None
    best_score = 0
    for t in therapists:
        t_norm = _norm_person_name(t.get("name"))
        if not t_norm:
            continue
        if row_norm == t_norm:
            return t
        if row_norm in t_norm or t_norm in row_norm:
            score = min(len(row_norm), len(t_norm))
            if score > best_score:
                best_score = score
                best = t
        row_parts = set(re.split(r"\s+", (row.get("name") or "").lower()))
        t_parts = set(re.split(r"\s+", (t.get("name") or "").lower().replace("ms.", "")))
        overlap = len(row_parts & t_parts)
        if overlap >= 2 and overlap > best_score:
            best_score = overlap
            best = t
    return best


async def _leave_balance_grid_rows(year: int) -> List[dict]:
    cache_key = str(year)
    cached = _leave_grid_cache.get(cache_key)
    now_ts = datetime.now(timezone.utc).timestamp()
    if cached and (now_ts - cached[0]) < LEAVE_GRID_CACHE_SECS:
        sheet_rows = cached[1]
    else:
        from drive_sync import fetch_workbook_from_url
        wb = await asyncio.to_thread(fetch_workbook_from_url, LEAVE_BALANCE_SHEET_URL)
        sheet_rows = _parse_leave_balance_sheet(wb, year)
        _leave_grid_cache[cache_key] = (now_ts, sheet_rows)
    therapists = await db.therapists.find(
        {},
        {"_id": 0, "id": 1, "name": 1, "color": 1, "email": 1, "key": 1},
    ).to_list(200)
    matched_ids = set()
    out = []
    for row in sheet_rows:
        t = _match_sheet_row_to_therapist(row, therapists)
        tid = t.get("id") if t else None
        if tid:
            matched_ids.add(tid)
        out.append({
            "therapist_id": tid,
            "name": row.get("name"),
            "employee_id": row.get("employee_id"),
            "color": (t or {}).get("color"),
            "email": (t or {}).get("email"),
            "annual_days": row.get("annual_days", 0),
            "sick_days": row.get("sick_days", 0),
            "permission_count": row.get("permission_count", 0),
            "permission_days": row.get("permission_days", 0),
            "unpaid_days": row.get("unpaid_days", 0),
            "remaining": row.get("remaining_raw"),
            "join_date": row.get("join_date"),
            "sheet_matched": bool(t),
        })
    for t in therapists:
        if t["id"] in matched_ids:
            continue
        name_l = (t.get("name") or "").lower()
        if any(skip in name_l for skip in ("jenan", "walaa", "maha", "fahda", "asma", "bodoor", "bodour")):
            continue
        out.append({
            "therapist_id": t["id"],
            "name": t.get("name"),
            "employee_id": None,
            "color": t.get("color"),
            "email": t.get("email"),
            "annual_days": 0,
            "sick_days": 0,
            "permission_count": 0,
            "permission_days": 0,
            "unpaid_days": 0,
            "remaining": None,
            "join_date": None,
            "sheet_matched": False,
        })
    out.sort(key=lambda r: (r.get("name") or "").lower())
    return out


@api.get("/hr/leave-balance-grid")
async def hr_leave_balance_grid(
    year: Optional[int] = None,
    refresh: bool = False,
    _=Depends(hr_manager_access),
):
    """All therapists' leave counts from the shared vacations Google Sheet (single fetch, cached)."""
    yr = year or datetime.now(timezone.utc).year
    if refresh:
        _leave_grid_cache.pop(str(yr), None)
    rows = await _leave_balance_grid_rows(yr)
    return {"year": yr, "rows": rows, "cached_seconds": LEAVE_GRID_CACHE_SECS}


async def _sync_leave_balances_from_sheet(year: Optional[int] = None) -> dict:
    """Sync remaining leave balances from the vacations Google Sheet into therapist records."""
    yr = year or datetime.now(timezone.utc).year
    rows = await _leave_balance_grid_rows(yr)
    updated = 0
    skipped = 0
    for row in rows:
        tid = row.get("therapist_id")
        if not tid:
            skipped += 1
            continue
        remaining_raw = row.get("remaining")
        if remaining_raw is None or str(remaining_raw).strip() == "":
            skipped += 1
            continue
        try:
            remaining = float(str(remaining_raw).replace(",", "").strip())
        except (TypeError, ValueError):
            skipped += 1
            continue
        patch = {"leave_balance": remaining, "leave_balance_synced_at": now_iso(), "leave_balance_sync_year": yr}
        annual_used = row.get("annual_days")
        if annual_used is not None:
            try:
                patch["annual_balance"] = round(float(remaining) + float(annual_used), 1)
            except (TypeError, ValueError):
                pass
        master_jd = _master_join_date_for_therapist(await db.therapists.find_one({"id": tid}, {"_id": 0, "key": 1}) or {})
        if master_jd:
            patch["join_date"] = master_jd
        elif row.get("join_date"):
            patch["join_date"] = row["join_date"]
        await db.therapists.update_one({"id": tid}, {"$set": patch})
        updated += 1
    return {"year": yr, "updated": updated, "skipped": skipped, "total_rows": len(rows)}


@api.post("/hr/leave-balance-sync")
async def hr_leave_balance_sync(
    year: Optional[int] = None,
    refresh: bool = True,
    _=Depends(hr_manager_access),
):
    """Pull leave balances from the shared Google Sheet into the database."""
    yr = year or datetime.now(timezone.utc).year
    if refresh:
        _leave_grid_cache.pop(str(yr), None)
    return await _sync_leave_balances_from_sheet(yr)


@api.get("/hr/therapist/{tid}/profile")
async def hr_therapist_profile(tid: str, _=Depends(hr_manager_access)):
    """Therapist summary for manager review — requests, leave, contract, training uploads."""
    therapist = await db.therapists.find_one(
        {"id": tid},
        {"_id": 0, "pin_hash": 0, "password_hash": 0},
    )
    if not therapist:
        raise HTTPException(status_code=404, detail="Therapist not found")
    balance_row = await _balance_row_for_therapist(therapist)
    therapist = await db.therapists.find_one(
        {"id": tid},
        {"_id": 0, "pin_hash": 0, "password_hash": 0},
    )
    all_reqs = await db.requests.find({"therapist_id": tid}, {"_id": 0}).to_list(500)
    open_statuses = {"pending", "pending_manager", "pending_hr", "in_progress"}
    answered_statuses = {"approved", "done", "rejected"}
    req_open = sum(1 for r in all_reqs if r.get("status") in open_statuses)
    req_answered = sum(1 for r in all_reqs if r.get("status") in answered_statuses)
    trainings = sorted(
        [
            {
                "id": r.get("id"),
                "title": r.get("title"),
                "report_date": r.get("report_date"),
                "created_at": r.get("created_at"),
                "status": r.get("status"),
                "attachment_file_name": r.get("attachment_file_name"),
            }
            for r in all_reqs
            if r.get("request_type") == "attachment"
        ],
        key=lambda x: x.get("report_date") or x.get("created_at") or "",
        reverse=True,
    )
    alerts: List[dict] = []
    contract_start = therapist.get("contract_start") or therapist.get("join_date")
    annual_end = (
        therapist.get("annual_contract_end")
        or therapist.get("contract_period_end")
    )
    if annual_end:
        try:
            end_dt = datetime.fromisoformat(str(annual_end)[:10])
            days_left = (end_dt.date() - datetime.now().date()).days
            if 0 <= days_left <= 60:
                alerts.append({
                    "type": "annual_contract_expiry",
                    "message": f"Annual contract expires {annual_end[:10]} ({days_left} days left)",
                    "severity": "urgent" if days_left <= 30 else "warning",
                })
        except Exception:
            pass
    probation_end = therapist.get("probation_end")
    if not probation_end and contract_start:
        try:
            start_dt = datetime.fromisoformat(str(contract_start)[:10])
            probation_end = (start_dt.date() + timedelta(days=90)).isoformat()
        except Exception:
            probation_end = None
    trial_days_left: Optional[int] = None
    if probation_end:
        try:
            prob_dt = datetime.fromisoformat(str(probation_end)[:10])
            trial_days_left = (prob_dt.date() - datetime.now().date()).days
            if 0 <= trial_days_left <= 30:
                alerts.append({
                    "type": "probation_end",
                    "message": (
                        f"Trial period ends {probation_end[:10]} ({trial_days_left} days left)"
                    ),
                    "severity": "urgent" if trial_days_left <= 14 else "warning",
                })
        except Exception:
            trial_days_left = None
    bal = balance_row.get("remaining")
    if bal is not None and float(bal) < 5:
        alerts.append({
            "type": "low_leave",
            "message": f"Leave balance low: {bal} days remaining",
            "severity": "warning",
        })
    ym = datetime.now().strftime("%Y-%m")
    sessions = await db.sessions.find(
        {"therapist_ids": tid, "status": "Completed"},
        {"_id": 0, "hours": 1, "session_date": 1, "therapist_ids": 1},
    ).to_list(5000)
    hours_month = sum(
        float(s.get("hours") or 0)
        for s in sessions
        if (s.get("session_date") or "")[:7] == ym and tid in (s.get("therapist_ids") or [])
    )
    hours_total = sum(
        float(s.get("hours") or 0)
        for s in sessions
        if tid in (s.get("therapist_ids") or [])
    )
    clients = await db.clients.find(_active_client_filter(), {"_id": 0, "main_therapist_id": 1, "co_therapist_ids": 1}).to_list(500)
    assigned = sum(
        1 for c in clients
        if c.get("main_therapist_id") == tid or tid in (c.get("co_therapist_ids") or [])
    )
    return {
        "therapist": therapist,
        "requests": {"total": len(all_reqs), "open": req_open, "answered": req_answered},
        "leave_balance": balance_row.get("remaining"),
        "annual_balance": balance_row.get("allocated"),
        "contract_period_start": balance_row.get("contract_period_start"),
        "contract_period_end": balance_row.get("contract_period_end"),
        "contract_start": contract_start,
        "annual_contract_end": annual_end,
        "probation_end": probation_end,
        "trial_days_left": trial_days_left,
        "join_date": balance_row.get("join_date"),
        "trainings": trainings,
        "alerts": alerts,
        "hours_this_month": round(hours_month, 1),
        "hours_total": round(hours_total, 1),
        "assigned_clients": assigned,
        "monthly_evaluations": [
            {k: v for k, v in ev.items() if k != "file_data"}
            for ev in (therapist.get("monthly_evaluations") or [])
        ],
        "annual_evaluations": [
            {k: v for k, v in ev.items() if k != "file_data"}
            for ev in (therapist.get("annual_evaluations") or [])
        ],
        "manager_meetings": therapist.get("manager_meetings") or [],
    }


def _default_probation_end(contract_start: Optional[str]) -> Optional[str]:
    if not contract_start:
        return None
    try:
        start_dt = datetime.fromisoformat(str(contract_start)[:10])
        return (start_dt.date() + timedelta(days=90)).isoformat()
    except Exception:
        return None


def _parse_iso_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)[:10]).date()
    except Exception:
        return None


def _add_months(start: date, months: int) -> date:
    y = start.year + (start.month - 1 + months) // 12
    m = (start.month - 1 + months) % 12 + 1
    last_day = calendar.monthrange(y, m)[1]
    return date(y, m, min(start.day, last_day))


_HR_CALENDAR_SKIP_NAMES = ("jenan", "walaa", "maha", "fahda", "asma", "bodoor", "bodour")
EVAL_DUE_NOTICE_DAYS = 7


def _is_hr_calendar_therapist(therapist: dict) -> bool:
    name_l = (therapist.get("name") or "").lower()
    return not any(skip in name_l for skip in _HR_CALENDAR_SKIP_NAMES)


def _therapist_evaluation_schedule(therapist: dict, *, lookback_days: int = 730, horizon_days: int = 1825) -> List[dict]:
    """Trial-period evals every 3 months from contract start; annual evals every 12 months."""
    contract_start = therapist.get("contract_start") or therapist.get("join_date")
    start = _parse_iso_date(contract_start)
    if not start:
        return []
    today = datetime.now().date()
    window_start = today - timedelta(days=lookback_days)
    window_end = today + timedelta(days=horizon_days)
    probation_end = therapist.get("probation_end") or _default_probation_end(contract_start)
    prob_date = _parse_iso_date(probation_end)
    tid = therapist.get("id")
    tname = therapist.get("name") or "Therapist"
    entries: List[dict] = []

    months = 3
    while months <= 36:
        eval_date = _add_months(start, months)
        if prob_date and eval_date > prob_date:
            break
        if eval_date > window_end:
            break
        if eval_date >= window_start:
            entries.append({
                "therapist_id": tid,
                "therapist_name": tname,
                "eval_type": "trial",
                "eval_label": "Trial period evaluation",
                "date": eval_date.isoformat(),
                "contract_start": start.isoformat(),
            })
        months += 3

    for years in range(1, 11):
        eval_date = _add_months(start, years * 12)
        if eval_date > window_end:
            break
        if eval_date >= window_start:
            entries.append({
                "therapist_id": tid,
                "therapist_name": tname,
                "eval_type": "annual",
                "eval_label": "Annual evaluation",
                "date": eval_date.isoformat(),
                "contract_start": start.isoformat(),
            })

    entries.sort(key=lambda e: e["date"])
    return entries


async def _process_evaluation_due_alerts(force: bool = False) -> dict:
    """Notify Jenan 1 week before trial or annual evaluation due dates."""
    jenan_id = await _jenan_therapist_id()
    if not jenan_id:
        return {"sent": 0, "skipped": "no_jenan"}
    today = datetime.now().date()
    target = today + timedelta(days=EVAL_DUE_NOTICE_DAYS)
    target_iso = target.isoformat()
    therapists = await db.therapists.find(
        {},
        {"_id": 0, "id": 1, "name": 1, "contract_start": 1, "join_date": 1, "probation_end": 1},
    ).to_list(400)
    sent = 0
    for therapist in therapists:
        if not _is_hr_calendar_therapist(therapist):
            continue
        for entry in _therapist_evaluation_schedule(therapist):
            if entry["date"] != target_iso:
                continue
            dedupe = f"eval_due:{entry['therapist_id']}:{entry['eval_type']}:{entry['date']}"
            if await db.notifications.find_one({"type": "evaluation_due", "eval_dedupe": dedupe}):
                continue
            tname = entry["therapist_name"]
            label = entry["eval_label"]
            title = f"Evaluation due in 1 week — {tname}"
            msg = f"{label} for {tname} is due on {entry['date']} (7 days from today)."
            await _notify(
                jenan_id,
                "evaluation_due",
                title,
                msg,
                eval_dedupe=dedupe,
                therapist_id=entry["therapist_id"],
                link="/manager?tab=calendar",
                eval_date=entry["date"],
                eval_type=entry["eval_type"],
            )
            body = f"{msg}\n\nOpen calendar: {_portal_base_url()}/manager?tab=calendar\n\n— Boost Growth Portal"
            await _send_urgent_email(await _jenan_recipient_email(), title, body)
            sent += 1
    return {"sent": sent, "target_date": target_iso, "forced": force}


@api.put("/hr/therapist/{tid}/profile")
async def hr_therapist_profile_update(tid: str, payload: TherapistHrProfileUpdate, user=Depends(hr_manager_access)):
    therapist = await db.therapists.find_one({"id": tid}, {"_id": 0})
    if not therapist:
        raise HTTPException(status_code=404, detail="Therapist not found")
    patch: dict = {}
    trial_end = payload.probation_end or payload.trial_end
    if trial_end is not None:
        patch["probation_end"] = str(trial_end)[:10] if trial_end else None
    if payload.annual_contract_end is not None:
        patch["annual_contract_end"] = str(payload.annual_contract_end)[:10] if payload.annual_contract_end else None
    if payload.meeting_date:
        meetings = list(therapist.get("manager_meetings") or [])
        meetings.append({
            "id": str(uuid.uuid4()),
            "date": str(payload.meeting_date)[:10],
            "notes": (payload.meeting_notes or "").strip() or None,
            "created_at": now_iso(),
            "created_by": user.get("name") or _actor_display(user),
        })
        patch["manager_meetings"] = meetings
    if patch:
        await db.therapists.update_one({"id": tid}, {"$set": patch})
    return await hr_therapist_profile(tid, user)


@api.post("/hr/therapist/{tid}/evaluations")
async def hr_upload_therapist_evaluation(
    tid: str,
    file: UploadFile = File(...),
    eval_type: str = Form("monthly"),
    period: Optional[str] = Form(None),
    user=Depends(hr_manager_access),
):
    therapist = await db.therapists.find_one({"id": tid}, {"_id": 0, "id": 1})
    if not therapist:
        raise HTTPException(status_code=404, detail="Therapist not found")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="File required")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")
    ext = Path(file.filename).suffix.lower() or ".pdf"
    if ext not in (".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif", ".doc", ".docx"):
        raise HTTPException(status_code=400, detail="Unsupported file type")
    etype = (eval_type or "monthly").strip().lower()
    if etype not in ("monthly", "annual"):
        raise HTTPException(status_code=400, detail="eval_type must be monthly or annual")
    eval_id = str(uuid.uuid4())
    stored = f"eval_{tid}_{eval_id}{ext}"
    file_data = _persist_upload(stored, content)
    entry = {
        "id": eval_id,
        "file_path": stored,
        "file_name": file.filename,
        "file_data": file_data,
        "uploaded_at": now_iso(),
        "uploaded_by": user.get("name") or _actor_display(user),
    }
    if etype == "monthly":
        entry["month"] = (period or datetime.now().strftime("%Y-%m"))[:7]
        field = "monthly_evaluations"
    else:
        try:
            entry["year"] = int((period or str(datetime.now().year))[:4])
        except ValueError:
            entry["year"] = datetime.now().year
        field = "annual_evaluations"
    await db.therapists.update_one(
        {"id": tid},
        {"$push": {field: entry}},
    )
    safe = {k: v for k, v in entry.items() if k != "file_data"}
    return {"ok": True, "type": etype, "evaluation": safe}


@api.get("/hr/therapist/{tid}/evaluations/{eval_id}/file")
async def hr_download_therapist_evaluation(tid: str, eval_id: str, user=Depends(get_current_user)):
    therapist = await db.therapists.find_one({"id": tid}, {"_id": 0})
    if not therapist:
        raise HTTPException(status_code=404, detail="Therapist not found")
    if not (_is_portal_admin(user) or _is_hr_ops(user) or _is_jenan(user)):
        if therapist.get("id") != user["id"]:
            raise HTTPException(status_code=403, detail="Forbidden")
    entry = None
    for ev in (therapist.get("monthly_evaluations") or []) + (therapist.get("annual_evaluations") or []):
        if ev.get("id") == eval_id:
            entry = ev
            break
    if not entry:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    content = _load_upload(entry.get("file_path"), entry.get("file_data"))
    if not content:
        raise HTTPException(status_code=404, detail=FILE_UNAVAILABLE_DETAIL)
    return _bytes_file_response(content, entry.get("file_name") or entry.get("file_path") or "evaluation")


@api.post("/hr/therapist/{tid}/contract-reminder")
async def hr_contract_reminder(tid: str, user=Depends(hr_manager_access)):
    """Notify Jenan to prepare trial / annual contracts for this therapist."""
    therapist = await db.therapists.find_one({"id": tid}, {"_id": 0, "id": 1, "name": 1, "join_date": 1, "contract_start": 1, "probation_end": 1, "annual_contract_end": 1, "contract_period_end": 1})
    if not therapist:
        raise HTTPException(status_code=404, detail="Therapist not found")
    tname = therapist.get("name") or "Therapist"
    contract_start = therapist.get("contract_start") or therapist.get("join_date")
    trial_end = therapist.get("probation_end") or _default_probation_end(contract_start)
    annual_end = therapist.get("annual_contract_end") or therapist.get("contract_period_end")
    parts = [f"Contract prep reminder for {tname}:"]
    if trial_end:
        parts.append(f"  • Trial period ends: {str(trial_end)[:10]}")
    if annual_end:
        parts.append(f"  • Annual contract ends: {str(annual_end)[:10]}")
    if contract_start:
        parts.append(f"  • Contract start: {str(contract_start)[:10]}")
    msg = "\n".join(parts)
    title = f"Contract reminder — {tname}"
    jenan_id = await _jenan_therapist_id()
    if jenan_id:
        await _notify(jenan_id, "contract_reminder", title, msg, therapist_id=tid)
    await _notify_hr_ops("contract_reminder", title, msg)
    body = f"{msg}\n\nRequested by: {user.get('name') or 'Manager'}\n"
    portal = _portal_base_url()
    if portal:
        body += f"\nOpen profile: {portal}/manager?tab=profiles\n"
    body += "\n— Boost Growth Portal"
    await _send_urgent_email(await _jenan_recipient_email(), title, body)
    return {"ok": True, "message": "Jenan notified"}


@api.get("/hr/evaluation-calendar")
async def hr_evaluation_calendar(
    year: Optional[int] = Query(None),
    _=Depends(hr_manager_access),
):
    """Upcoming and past evaluation dates for all therapists (Manager Hub calendar)."""
    therapists = await db.therapists.find(
        {},
        {
            "_id": 0, "id": 1, "name": 1, "key": 1, "email": 1,
            "contract_start": 1, "join_date": 1, "probation_end": 1,
            "annual_contract_end": 1,
        },
    ).to_list(400)
    entries: List[dict] = []
    for therapist in therapists:
        if not _is_hr_calendar_therapist(therapist):
            continue
        entries.extend(_therapist_evaluation_schedule(therapist))
    if year is not None:
        y = str(year)
        entries = [e for e in entries if e["date"].startswith(y)]
    entries.sort(key=lambda e: (e["date"], e["therapist_name"].lower()))
    today_iso = datetime.now().date().isoformat()
    upcoming = sum(1 for e in entries if e["date"] >= today_iso)
    return {
        "year": year or datetime.now().year,
        "entries": entries,
        "summary": {
            "total": len(entries),
            "upcoming": upcoming,
            "past": len(entries) - upcoming,
            "trial": sum(1 for e in entries if e["eval_type"] == "trial"),
            "annual": sum(1 for e in entries if e["eval_type"] == "annual"),
        },
    }


@api.get("/my-performance")
async def therapist_performance(user=Depends(get_current_user)):
    """Therapist view: manager meetings and evaluation attachments."""
    tid = user["id"]
    therapist = await db.therapists.find_one({"id": tid}, {"_id": 0, "pin_hash": 0, "password_hash": 0})
    if not therapist:
        raise HTTPException(status_code=403, detail="Therapist profile required")
    therapist = await _ensure_contract_balance(therapist)
    contract_start = therapist.get("contract_start") or therapist.get("join_date")
    return {
        "therapist_id": tid,
        "therapist_name": therapist.get("name"),
        "contract_start": contract_start,
        "probation_end": therapist.get("probation_end") or _default_probation_end(contract_start),
        "annual_contract_end": therapist.get("annual_contract_end") or therapist.get("contract_period_end"),
        "manager_meetings": sorted(
            therapist.get("manager_meetings") or [],
            key=lambda m: m.get("date") or "",
            reverse=True,
        ),
        "monthly_evaluations": sorted(
            [{k: v for k, v in ev.items() if k != "file_data"} for ev in (therapist.get("monthly_evaluations") or [])],
            key=lambda e: e.get("month") or e.get("uploaded_at") or "",
            reverse=True,
        ),
        "annual_evaluations": sorted(
            [{k: v for k, v in ev.items() if k != "file_data"} for ev in (therapist.get("annual_evaluations") or [])],
            key=lambda e: str(e.get("year") or e.get("uploaded_at") or ""),
            reverse=True,
        ),
    }


@api.get("/tracking/inbox")
async def tracking_inbox(user=Depends(get_current_user)):
    if not (_is_portal_admin(user) or _is_hr_ops(user) or _is_jenan(user) or _is_walaa_ops(user)):
        raise HTTPException(status_code=403, detail="Inbox access required")
    leaves_pending_manager = await db.leaves.count_documents({"status": {"$in": list(PENDING_MANAGER_STATUSES)}})
    leaves_pending_hr = await db.leaves.count_documents({"status": "pending_hr"})
    requests_pending_manager = await db.requests.count_documents(
        {"status": {"$in": list(PENDING_MANAGER_REQUEST_STATUSES)}}
    )
    requests_pending_hr = await db.requests.count_documents({"status": "pending_hr"})
    requests_pending = requests_pending_manager + requests_pending_hr
    purchases_pending = await db.staff_purchases.count_documents({"status": "pending"})
    billing_reminders_soon = 0
    try:
        dash = await billing_dashboard(user)
        billing_reminders_soon = int((dash.get("summary") or {}).get("reminders_soon") or 0)
    except Exception:
        pass
    parent_cancellations_pending = 0
    if _can_parent_cancellation_ops(user):
        parent_cancellations_pending = await db.schedule_cells.count_documents({
            "state": "cancel_therapist",
            "parent_notify_pending": True,
        })
    unprepared_alerts = 0
    try:
        prep = await _process_unprepared_session_alerts(force=False)
        unprepared_alerts = int(prep.get("alerts") or 0)
    except Exception:
        logger.exception("Unprepared session alert check failed")
    return {
        "leaves_pending_manager": leaves_pending_manager,
        "leaves_pending_hr": leaves_pending_hr,
        "requests_pending_manager": requests_pending_manager,
        "requests_pending_hr": requests_pending_hr,
        "requests_pending": requests_pending,
        "purchases_pending": purchases_pending,
        "billing_reminders_soon": billing_reminders_soon,
        "parent_cancellations_pending": parent_cancellations_pending,
        "unprepared_sessions_today": unprepared_alerts,
    }


@api.get("/tracking/parent-cancellations")
async def list_parent_cancellations(user=Depends(get_current_user)):
    if not _can_parent_cancellation_ops(user):
        raise HTTPException(status_code=403, detail="Parent cancellation ops access required")
    cells = await db.schedule_cells.find(
        {"state": "cancel_therapist", "parent_notify_pending": True},
        {"_id": 0},
    ).sort("parent_cancel_marked_at", -1).to_list(200)
    therapist_ids = list({c.get("therapist_id") for c in cells if c.get("therapist_id")})
    therapist_names: dict = {}
    if therapist_ids:
        for t in await db.therapists.find(
            {"id": {"$in": therapist_ids}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(200):
            therapist_names[t["id"]] = t.get("name") or ""
    results = []
    for cell in cells:
        client = await _find_client_by_schedule_child_name(cell.get("child_name") or "")
        day_idx = cell.get("day")
        week_start = cell.get("week_start") or ""
        day_label = ""
        day_ar = ""
        if day_idx is not None:
            try:
                di = int(day_idx)
                if 0 <= di < len(SCHEDULE_DAYS_AR):
                    day_ar = SCHEDULE_DAYS_AR[di]
            except (TypeError, ValueError):
                pass
        if week_start and day_idx is not None:
            try:
                d = datetime.fromisoformat(str(week_start)[:10]) + timedelta(days=int(day_idx))
                day_label = d.strftime("%d %b %Y")
            except Exception:
                day_label = str(week_start)
        results.append({
            **cell,
            "parent_phone": (client or {}).get("parent_phone"),
            "parent_name": (client or {}).get("parent_name"),
            "client_id": (client or {}).get("id"),
            "therapist_name": therapist_names.get(cell.get("therapist_id"), ""),
            "day_label": day_label,
            "day_ar": day_ar,
        })
    return results


@api.get("/requests/{rid}/attachment")
async def download_request_attachment(rid: str, user=Depends(get_current_user)):
    req = await db.requests.find_one({"id": rid}, {"_id": 0})
    if not req or not _request_has_attachment(req):
        raise HTTPException(status_code=404, detail="No attachment")
    if not (_is_portal_admin(user) or _is_hr_ops(user) or _is_jenan(user)):
        if req.get("therapist_id") != user["id"]:
            raise HTTPException(status_code=403, detail="Forbidden")
    content = _load_upload(req.get("attachment_file_path"), req.get("attachment_file_data"))
    if not content:
        raise HTTPException(status_code=404, detail=FILE_UNAVAILABLE_DETAIL)
    fname = req.get("attachment_file_name") or req.get("attachment_file_path") or "attachment"
    return _bytes_file_response(content, fname)


@api.put("/requests/{rid}/status")
async def update_request_status(rid: str, payload: RequestStatusUpdate, user=Depends(get_current_user)):
    req = await db.requests.find_one({"id": rid})
    if not req:
        raise HTTPException(status_code=404, detail="Not found")
    prev_status = req.get("status")
    effective_prev = _normalize_request_status(prev_status)
    new_status = _coerce_manager_approve_to_hr(payload.status, payload.notify_hr)
    is_pa = _is_portal_admin(user)
    is_hr = _is_hr_ops(user)
    is_jenan_mgr = _is_jenan(user) and not is_pa
    notify_hr = payload.notify_hr if payload.notify_hr is not None else (not is_jenan_mgr)
    notify_therapist = payload.notify_therapist if payload.notify_therapist is not None else (not is_jenan_mgr)

    if is_pa:
        pass
    elif is_jenan_mgr:
        if effective_prev not in MANAGER_ACTIVE_REQUEST_STATUSES:
            raise HTTPException(status_code=403, detail="Manager can only act on pending manager requests")
        if _request_awaiting_attachment(req):
            raise HTTPException(status_code=400, detail="Attachment required before manager review")
        if effective_prev == "in_progress":
            if new_status not in ("pending_hr", "rejected", "in_progress"):
                raise HTTPException(status_code=400, detail="Manager must approve (forward to HR) or reject")
        elif new_status not in ("pending_hr", "rejected", "in_progress", "pending_manager", "approved"):
            raise HTTPException(status_code=400, detail="Manager must choose pending, approve, or reject")
    elif is_hr:
        if effective_prev != "pending_hr":
            raise HTTPException(status_code=403, detail="HR can only act on HR-pending requests")
        if new_status not in ("approved", "rejected", "in_progress", "done"):
            raise HTTPException(status_code=400, detail="HR must approve, reject, mark in progress, or complete")
    else:
        raise HTTPException(status_code=403, detail="Staff request management access required")

    timeline = req.get("timeline", [])
    timeline.append({"event": new_status, "at": now_iso(), "by": user.get("name") or _actor_display(user),
                     "note": payload.admin_note})
    await db.requests.update_one({"id": rid}, {"$set": {
        "status": new_status, "admin_note": payload.admin_note,
        "updated_at": now_iso(), "timeline": timeline,
    }})
    if is_jenan_mgr and effective_prev in MANAGER_ACTIVE_REQUEST_STATUSES and notify_hr and new_status in MANAGER_HR_NOTIFY_STATUSES:
        await _notify_hr_manager_decision(
            ntype="request",
            therapist_name=req.get("therapist_name") or "Staff",
            summary=req.get("title") or "Request",
            decision_status=new_status,
            admin_note=payload.admin_note,
        )
    status_map = {
        "pending": "Pending", "pending_manager": "Pending manager review",
        "pending_hr": "Pending HR review", "in_progress": "In Progress",
        "approved": "Approved", "rejected": "Rejected", "done": "Completed",
    }
    if notify_therapist:
        await _notify(req["therapist_id"], "request", "Request update",
                      f"Your request '{req['title']}' is now: {status_map.get(new_status, new_status)}")
    if new_status in ("approved", "done") and is_hr and notify_therapist:
        email = await _therapist_email(req.get("therapist_id"))
        if email:
            ts = now_iso()[:16].replace("T", " ")
            body = (
                f"Hello,\n\nYour request \"{req.get('title')}\" was {status_map.get(new_status, new_status)} "
                f"on {ts}.\n"
            )
            if payload.admin_note:
                body += f"\nNote from HR: {payload.admin_note}\n"
            body += "\n— Boost Growth Portal"
            await _send_email_stub(
                email,
                f"Request {status_map.get(new_status, new_status)} — {req.get('title')}",
                body,
            )
    return await db.requests.find_one({"id": rid}, {"_id": 0})

@api.delete("/requests/{rid}")
async def delete_request(rid: str, user=Depends(get_current_user)):
    req = await db.requests.find_one({"id": rid})
    if not req:
        return {"ok": True}
    if not _can_delete_staff_submission(user, req.get("therapist_id")):
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.requests.delete_one({"id": rid})
    return {"ok": True}

# ------------------- Notifications -------------------
@api.get("/notifications")
async def list_notifications(user=Depends(get_current_user)):
    if _has_full_client_access(user) or _is_walaa_ops(user) or _is_portal_admin(user) or _is_hr_ops(user):
        try:
            await _process_unprepared_session_alerts(force=False)
        except Exception:
            logger.exception("unprepared session alert check on notifications load")
    uids = await _notification_user_ids(user)
    return await db.notifications.find({"user_id": {"$in": uids}}, {"_id": 0}).sort("created_at", -1).limit(100).to_list(100)

@api.post("/notifications/{nid}/read")
async def mark_read(nid: str, user=Depends(get_current_user)):
    uids = await _notification_user_ids(user)
    await db.notifications.update_one({"id": nid, "user_id": {"$in": uids}}, {"$set": {"read": True}})
    return {"ok": True}

@api.post("/notifications/{nid}/acknowledge")
async def acknowledge_notification(nid: str, user=Depends(get_current_user)):
    uids = await _notification_user_ids(user)
    await db.notifications.update_one(
        {"id": nid, "user_id": {"$in": uids}},
        {"$set": {"read": True, "acknowledged": True, "acknowledged_at": now_iso()}},
    )
    return {"ok": True}

@api.post("/notifications/{nid}/notify-therapist")
async def notify_therapist_from_alert(nid: str, user=Depends(get_current_user)):
    """Admin/Walaa: send unprepared-session reminder to the assigned therapist."""
    if not (_is_portal_admin(user) or _is_hr_ops(user) or _is_walaa_ops(user) or _has_full_client_access(user)):
        raise HTTPException(status_code=403, detail="Admin access required")
    rec = await db.notifications.find_one({"id": nid}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Notification not found")
    tid = rec.get("therapist_id")
    if not tid:
        raise HTTPException(status_code=400, detail="No therapist linked to this alert")
    title = "Reminder: session not prepared"
    message = rec.get("message") or "Please prepare or log your scheduled session."
    await _notify(tid, "unprepared_reminder", title, message, link="/attendance")
    await db.notifications.update_one(
        {"id": nid},
        {"$set": {"therapist_notified_at": now_iso(), "therapist_notified_by": user.get("id")}},
    )
    return {"ok": True, "therapist_id": tid}

@api.post("/notifications/read-all")
async def mark_all_read(user=Depends(get_current_user)):
    uids = await _notification_user_ids(user)
    await db.notifications.update_many({"user_id": {"$in": uids}}, {"$set": {"read": True}})
    return {"ok": True}

PARENT_PHONES_JSON = ROOT_DIR / "data" / "parent_phones.json"


async def _apply_parent_phones(rows: List[dict], *, source: str = "import") -> dict:
    """Update parent_phone on clients matched by file_no."""
    updated: List[dict] = []
    skipped: List[dict] = []
    missing: List[dict] = []
    for row in rows:
        file_no = str(row.get("file_no") or "").strip()
        phone = str(row.get("parent_phone") or "").strip()
        if not file_no:
            continue
        if not phone:
            skipped.append({"file_no": file_no, "reason": "empty phone"})
            continue
        client = await _find_client_by_file_no(file_no)
        if not client:
            missing.append({"file_no": file_no, "name": row.get("client_name")})
            continue
        await db.clients.update_one(
            {"id": client["id"]},
            {"$set": {"parent_phone": phone, "parent_phone_source": source, "parent_phone_updated_at": now_iso()}},
        )
        updated.append({"file_no": client.get("file_no"), "name": client.get("name"), "parent_phone": phone})
    return {
        "ok": True,
        "updated": len(updated),
        "skipped": len(skipped),
        "missing": len(missing),
        "rows": updated,
        "missing_clients": missing,
        "message": f"Updated {len(updated)} parent phone(s)" + (f" · {len(missing)} file_no not found" if missing else ""),
    }


async def _apply_parent_phones_from_json_file() -> Optional[dict]:
    if not PARENT_PHONES_JSON.is_file():
        return None
    import json
    rows = json.loads(PARENT_PHONES_JSON.read_text(encoding="utf-8"))
    if not rows:
        return None
    return await _apply_parent_phones(rows, source="parent_phones.json")


@api.get("/admin/export-parent-phones")
async def export_parent_phones(_=Depends(client_lead_or_admin)):
    """CSV of all active clients' parent phone numbers for bulk editing."""
    import csv
    import io
    from fastapi.responses import Response

    clients = await db.clients.find(
        _active_client_filter(),
        {"_id": 0, "file_no": 1, "name": 1, "parent_phone": 1},
    ).sort("file_no", 1).to_list(500)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["file_no", "client_name", "parent_phone"])
    for c in clients:
        writer.writerow([
            c.get("file_no") or "",
            c.get("name") or "",
            c.get("parent_phone") or "",
        ])
    return Response(
        content="\ufeff" + buf.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="parent_phones.csv"'},
    )


@api.post("/admin/import-parent-phones")
async def import_parent_phones(
    file: UploadFile = File(...),
    _=Depends(client_lead_or_admin),
):
    """Upload CSV (file_no, client_name, parent_phone) to bulk-update parent phones."""
    import csv
    import io

    raw = await file.read()
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append({
            "file_no": row.get("file_no") or row.get("File No") or "",
            "client_name": row.get("client_name") or row.get("Client Name") or "",
            "parent_phone": row.get("parent_phone") or row.get("Parent Phone") or "",
        })
    if not rows:
        raise HTTPException(status_code=400, detail="CSV is empty or missing headers")
    return await _apply_parent_phones(rows, source="csv-upload")


@api.post("/admin/apply-parent-phones-seed")
async def apply_parent_phones_seed(_=Depends(client_lead_or_admin)):
    """Apply backend/data/parent_phones.json without redeploy."""
    result = await _apply_parent_phones_from_json_file()
    if not result:
        raise HTTPException(status_code=404, detail="parent_phones.json not found or empty")
    return result

# ------------------- Directory -------------------
@api.get("/directory")
async def list_directory(user=Depends(get_current_user)):
    return await db.directory.find({}, {"_id": 0}).to_list(500)

@api.post("/directory")
async def create_contact(payload: DirectoryContactIn, _=Depends(admin_only)):
    cid = str(uuid.uuid4())
    doc = {"id": cid, **payload.model_dump(), "created_at": now_iso()}
    await db.directory.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/directory/{cid}")
async def update_contact(cid: str, payload: DirectoryContactUpdate, _=Depends(admin_only)):
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No fields")
    await db.directory.update_one({"id": cid}, {"$set": update})
    return await db.directory.find_one({"id": cid}, {"_id": 0})

@api.delete("/directory/{cid}")
async def delete_contact(cid: str, _=Depends(admin_only)):
    await db.directory.delete_one({"id": cid})
    return {"ok": True}

# ------------------- Resources -------------------
@api.get("/resources")
async def list_resources(user=Depends(get_current_user)):
    items = await db.resources.find({}, {"_id": 0}).sort("sort_order", 1).to_list(500)
    if user.get("role") == "admin":
        return items
    # Therapists see only "therapist" and "all" visibility
    return [r for r in items if r.get("visibility") in ("therapist", "all")]

@api.post("/resources")
async def create_resource(payload: ResourceIn, _=Depends(admin_only)):
    rid = str(uuid.uuid4())
    doc = {"id": rid, **payload.model_dump(), "created_at": now_iso()}
    await db.resources.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/resources/{rid}")
async def update_resource(rid: str, payload: ResourceIn, _=Depends(admin_only)):
    await db.resources.update_one({"id": rid}, {"$set": payload.model_dump()})
    return await db.resources.find_one({"id": rid}, {"_id": 0})

@api.get("/clients/{cid}/billing-progress")
async def client_billing_progress(cid: str, user=Depends(get_current_user)):
    """Returns billing progress: hours-based or weeks-based.
    weeks-based: counts distinct weeks (Sun-Thu) where at least 1 Completed session exists,
                 since cycle_start_date; cycle ends when weeks_completed >= cycle_weeks.
    """
    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Not found")
    sessions = await db.sessions.find({"client_id": cid}, {"_id": 0}).to_list(2000)
    completed = [s for s in sessions if s.get("status") == "Completed"]
    mode = client.get("billing_mode") or "hours"
    if mode == "hours":
        used_h = sum(float(s.get("hours") or 0) for s in completed)
        pkg = float(client.get("package_hours") or 24)
        return {
            "mode": "hours", "used": round(used_h, 1), "package": pkg,
            "remaining": max(0.0, round(pkg - used_h, 1)),
            "percent": min(100, round((used_h / pkg) * 100)) if pkg else 0,
        }
    # weeks-based — 7-day windows from cycle_start_date anchor (same weekday each week)
    cycle_weeks = int(client.get("cycle_weeks") or 4)
    start_iso = client.get("cycle_start_date")
    if not start_iso:
        if completed:
            start_iso = min(s.get("session_date") or "9999" for s in completed)
        else:
            start_iso = datetime.now(timezone.utc).date().isoformat()
    try:
        start_d = datetime.fromisoformat(start_iso).date()
    except Exception:
        start_d = datetime.now(timezone.utc).date()
    weeks_done = 0
    week_breakdown = []
    for k in range(cycle_weeks):
        week_start = start_d + timedelta(days=7 * k)
        week_end = week_start + timedelta(days=7)
        ws_iso = week_start.isoformat()
        we_iso = week_end.isoformat()
        in_week = [s for s in completed if s.get("session_date") and ws_iso <= s.get("session_date") < we_iso]
        has = len(in_week) > 0
        if has:
            weeks_done += 1
        week_breakdown.append({
            "week_number": k + 1,
            "week_start": ws_iso,
            "week_end": we_iso,
            "sessions": len(in_week),
            "completed": has,
        })
    return {
        "mode": "weeks",
        "weeks_completed": weeks_done,
        "cycle_weeks": cycle_weeks,
        "cycle_start_date": start_d.isoformat(),
        "next_cycle_start": (start_d + timedelta(days=7 * cycle_weeks)).isoformat(),
        "remaining_weeks": max(0, cycle_weeks - weeks_done),
        "percent": round((weeks_done / cycle_weeks) * 100) if cycle_weeks else 0,
        "weeks": week_breakdown,
    }

@api.get("/clients/{cid}/sessions/export")
async def export_sessions_excel(cid: str, columns: Optional[str] = None, user=Depends(get_current_user)):
    """Export client's attendance sheet as Excel.
    If the client has invoices, each invoice becomes its own sheet/tab named with
    the invoice number (e.g. INV0451). Sessions in each tab are scoped to that
    invoice's window (>= invoice.start_date and < next invoice.start_date when
    sorted ascending). Otherwise, a single 'Attendance' sheet is produced.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    client = await db.clients.find_one(_active_client_filter({"id": cid}), {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if user.get("role") == "therapist" and user["id"] not in (client.get("co_therapist_ids") or []) + ([client.get("main_therapist_id")] if client.get("main_therapist_id") else []):
        raise HTTPException(status_code=403, detail="Forbidden")

    sessions = await db.sessions.find({"client_id": cid}, {"_id": 0}).sort("session_date", 1).to_list(2000)
    therapists = {t["id"]: t for t in await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    invoices = await db.invoices.find({"client_id": cid}, {"_id": 0}).sort("start_date", 1).to_list(200)

    head_fill = PatternFill("solid", fgColor="7A8A6A")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    sub_fill = PatternFill("solid", fgColor="EFE8D2")
    border = Border(left=Side(style="thin", color="B5B0A0"), right=Side(style="thin", color="B5B0A0"),
                    top=Side(style="thin", color="B5B0A0"), bottom=Side(style="thin", color="B5B0A0"))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    STATUS_FILLS = {"Completed": "D9EAD3", "Cancelled": "FCE0E8", "No Show": "FFF4C4", "No Service": "ECECEC"}

    def safe_sheet_name(name: str) -> str:
        # Excel limits: 31 chars, no []:*?/\\
        bad = '[]:*?/\\'
        for b in bad:
            name = name.replace(b, "-")
        return name[:31] or "Sheet"

    def write_invoice_sheet(ws, inv: Optional[dict], inv_sessions: list):
        """Write the Boost Growth-style header + session rows on `ws` for the given invoice."""
        pkg = float((inv or {}).get("package_size") or client.get("package_hours") or 24)
        used = sum(float(s.get("hours") or 0) for s in inv_sessions if s.get("status") == "Completed")
        rem = max(0.0, pkg - used)
        no_show = sum(1 for s in inv_sessions if s.get("status") == "No Show")
        no_service = sum(1 for s in inv_sessions if s.get("status") == "No Service")
        completed = sum(1 for s in inv_sessions if s.get("status") == "Completed")
        # Title
        ws.merge_cells("A1:G1")
        inv_label = (inv or {}).get("invoice_number") or "Attendance"
        status_label = "Closed" if (inv or {}).get("is_closed") else "Open"
        ws["A1"] = f"{inv_label} | {status_label}"
        ws["A1"].font = Font(bold=True, size=14, color="2C3625")
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 26
        # Patient info row 2
        ws["A2"] = "Patient's Name:"; ws["B2"] = client.get("name") or "—"
        ws["C2"] = "File NO.:"; ws["D2"] = client.get("file_no") or "—"
        ws["E2"] = "# Paid SESH.:"; ws["F2"] = f"{pkg}h"
        if (inv or {}).get("service_type"):
            ws["G2"] = (inv or {}).get("service_type")
        for c in "ABCDEFG":
            ws[f"{c}2"].fill = sub_fill
            ws[f"{c}2"].font = Font(bold=True, color="2C3625")
            ws[f"{c}2"].alignment = center
        COL_DEFS = [
            ("days", "Days"), ("date", "Date"), ("status", "Status"), ("time", "Time"),
            ("hours", "# of Hrs"), ("therapist", "Therapist"), ("note", "Note"),
            ("service", "Service type"), ("location", "Location"),
        ]
        if columns:
            wanted = {c.strip().lower() for c in columns.split(",") if c.strip()}
            active_cols = [c for c in COL_DEFS if c[0] in wanted]
        else:
            active_cols = COL_DEFS[:7]
        if not active_cols:
            active_cols = COL_DEFS[:7]
        ncols = len(active_cols)
        for i, (_, h) in enumerate(active_cols, 1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.fill = head_fill; cell.font = head_font; cell.alignment = center; cell.border = border
        ws.row_dimensions[4].height = 22
        row = 5
        for s in inv_sessions:
            sd = s.get("session_date") or ""
            try:
                dt = datetime.fromisoformat(sd)
                day_label = DAY_NAMES[dt.weekday()]
                date_label = f"{dt.day}/{dt.month}/{dt.year}"
            except Exception:
                day_label = "—"; date_label = sd
            therapist_names = " - ".join(
                ((therapists.get(tid) or {}).get("name", "?") or "?").replace("Ms. ", "")
                for tid in (s.get("therapist_ids") or [])
            )
            time_str = ""
            if s.get("start_time") and s.get("end_time"):
                time_str = f"{s['start_time']}-{s['end_time']}"
            values = {
                "days": day_label, "date": date_label, "status": s.get("status") or "—",
                "time": time_str, "hours": float(s.get("hours") or 0), "therapist": therapist_names,
                "note": s.get("note") or "", "service": s.get("service_type") or "—",
                "location": s.get("location") or "—",
            }
            for col_i, (key, _) in enumerate(active_cols, 1):
                val = values.get(key, "")
                cell = ws.cell(row=row, column=col_i, value=val)
                cell.alignment = center if key in ("days", "date", "status", "time", "hours") else Alignment(horizontal="left", vertical="center", wrap_text=True)
                if key == "status" and s.get("status") in STATUS_FILLS:
                    cell.fill = PatternFill("solid", fgColor=STATUS_FILLS[s["status"]])
                cell.border = border
            row += 1
        # Footer
        foot = row + 1
        ws.cell(row=foot, column=1, value="Total delivered Sessions:").font = Font(bold=True, color="2C3625")
        ws.cell(row=foot, column=2, value=completed)
        ws.cell(row=foot+1, column=1, value="Total NO Service (counted):").font = Font(bold=True, color="2C3625")
        ws.cell(row=foot+1, column=2, value=no_service)
        ws.cell(row=foot+2, column=1, value="Total No-Show:").font = Font(bold=True, color="2C3625")
        ws.cell(row=foot+2, column=2, value=no_show)
        ws.cell(row=foot+3, column=1, value="Total Counted Sessions:").font = Font(bold=True, color="2C3625")
        ws.cell(row=foot+3, column=2, value=completed + no_show)
        ws.cell(row=foot+4, column=1, value="Hours Remaining:").font = Font(bold=True, color="2C3625")
        ws.cell(row=foot+4, column=2, value=round(rem, 2))
        ws.cell(row=foot+5, column=1, value="Payment Status:").font = Font(bold=True, color="2C3625")
        ws.cell(row=foot+5, column=2, value="Paid" if (inv or {}).get("payment_status") == "complete" else "Payment Pending")
        width_map = {"days": 8, "date": 12, "status": 14, "time": 14, "hours": 10, "therapist": 24, "note": 32, "service": 12, "location": 16}
        for i, (key, _) in enumerate(active_cols, 1):
            ws.column_dimensions[chr(64 + i)].width = width_map.get(key, 12)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if invoices:
        # Build windows: each invoice covers [start_date, next_start) ; last invoice covers until end of time.
        sorted_inv = sorted(invoices, key=lambda i: (i.get("start_date") or "0000-00-00"))
        for idx, inv in enumerate(sorted_inv):
            start = inv.get("start_date") or "0000-00-00"
            end = sorted_inv[idx + 1]["start_date"] if (idx + 1 < len(sorted_inv) and sorted_inv[idx + 1].get("start_date")) else None
            inv_sessions = [s for s in sessions if (s.get("session_date") or "") >= start and (end is None or (s.get("session_date") or "") < end)]
            ws = wb.create_sheet(safe_sheet_name(inv.get("invoice_number") or f"INV-{idx+1}"))
            write_invoice_sheet(ws, inv, inv_sessions)
    else:
        ws = wb.create_sheet("Attendance")
        write_invoice_sheet(ws, None, sessions)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    from fastapi.responses import Response
    fname = f"attendance_{client.get('file_no') or 'client'}_{client.get('name','').replace(' ','_')}.xlsx"
    return Response(content=out.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})

# ------------------- Email Settings (admin) -------------------
class EmailSettingsIn(BaseModel):
    resend_api_key: Optional[str] = None
    brevo_api_key: Optional[str] = None
    mailgun_api_key: Optional[str] = None
    mailgun_domain: Optional[str] = None
    from_email: Optional[str] = None
    smtp_host: Optional[str] = None
    smtp_port: Optional[int] = None
    smtp_user: Optional[str] = None
    smtp_password: Optional[str] = None
    email_provider: Optional[str] = None  # auto | mailgun | brevo | resend | smtp

def _apply_email_settings(doc: dict) -> None:
    """Merge persisted email settings into process env (DB overrides; never clears Railway env)."""
    if not doc:
        return
    if doc.get("resend_api_key"):
        os.environ["RESEND_API_KEY"] = doc["resend_api_key"]
    if doc.get("brevo_api_key"):
        os.environ["BREVO_API_KEY"] = doc["brevo_api_key"]
    if doc.get("mailgun_api_key"):
        os.environ["MAILGUN_API_KEY"] = doc["mailgun_api_key"]
    if doc.get("mailgun_domain"):
        os.environ["MAILGUN_DOMAIN"] = doc["mailgun_domain"]
    if doc.get("from_email"):
        os.environ["EMAIL_FROM"] = doc["from_email"]
    if doc.get("smtp_host"):
        os.environ["SMTP_HOST"] = doc["smtp_host"]
    if doc.get("smtp_port"):
        os.environ["SMTP_PORT"] = str(doc["smtp_port"])
    if doc.get("smtp_user"):
        os.environ["SMTP_USER"] = doc["smtp_user"]
    if doc.get("smtp_password"):
        os.environ["SMTP_PASSWORD"] = str(doc["smtp_password"]).replace(" ", "")
    if doc.get("email_provider"):
        os.environ["EMAIL_PROVIDER"] = doc["email_provider"]

async def _reload_email_settings_from_db() -> None:
    doc = await db.settings.find_one({"key": "email"}, {"_id": 0})
    _apply_email_settings(doc or {})

def _parse_from_address() -> tuple:
    raw = _email_from_address()
    if "<" in raw and ">" in raw:
        name = raw.split("<")[0].strip().strip('"').strip() or "Boost Growth"
        email = raw.split("<")[-1].split(">")[0].strip()
        return name, email
    return "Boost Growth", raw.strip()

def _smtp_error_hint(err: str) -> str:
    e = (err or "").lower()
    if "101" in e or "network is unreachable" in e or "network unreachable" in e:
        return "Railway blocks Gmail SMTP. Use Mailgun (HTTPS) in Admin — it works on Railway."
    if "535" in e or "username and password not accepted" in e:
        return "Gmail rejected login. Check your App Password (16 characters, no spaces)."
    if "534" in e:
        return "Google Workspace may have SMTP disabled — contact your Google Workspace admin."
    if "550" in e or "relay" in e:
        return "Gmail will not send from this address — set From Email to match SMTP User."
    if "connection" in e or "timed out" in e:
        return "Could not connect to SMTP — Railway blocks port 587. Use Mailgun instead."
    return ""

def _email_from_address() -> str:
    return os.environ.get("EMAIL_FROM") or "Boost Growth <hr@boostgrowthsa.com>"

def _smtp_configured() -> bool:
    return bool(os.environ.get("SMTP_USER") and os.environ.get("SMTP_PASSWORD"))

def _resend_configured() -> bool:
    return bool(os.environ.get("RESEND_API_KEY"))

def _brevo_configured() -> bool:
    return bool(os.environ.get("BREVO_API_KEY"))

def _mailgun_configured() -> bool:
    return bool(os.environ.get("MAILGUN_API_KEY") and os.environ.get("MAILGUN_DOMAIN"))

async def _send_via_brevo(to: str, subject: str, body: str) -> str:
    api_key = os.environ.get("BREVO_API_KEY")
    if not api_key:
        raise ValueError("Brevo API key not configured")
    name, email = _parse_from_address()
    import httpx
    async with httpx.AsyncClient(timeout=30) as cli:
        r = await cli.post(
            "https://api.brevo.com/v3/smtp/email",
            headers={"api-key": api_key, "Content-Type": "application/json", "accept": "application/json"},
            json={
                "sender": {"name": name, "email": email},
                "to": [{"email": to}],
                "subject": subject,
                "htmlContent": f"<p>{body.replace(chr(10), '<br/>')}</p>",
                "textContent": body,
            },
        )
        if r.status_code in (200, 201, 202):
            return r.json().get("messageId") or "ok"
        raise ValueError(r.text[:500])

async def _send_via_resend(to: str, subject: str, body: str) -> str:
    api_key = os.environ.get("RESEND_API_KEY")
    if not api_key:
        raise ValueError("Resend API key not configured")
    import httpx
    async with httpx.AsyncClient(timeout=30) as cli:
        r = await cli.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"from": _email_from_address(), "to": [to], "subject": subject,
                  "html": f"<p>{body.replace(chr(10), '<br/>')}</p>"},
        )
        if r.status_code in (200, 202):
            return r.json().get("id") or "ok"
        raise ValueError(r.text[:500])

async def _send_via_mailgun(to: str, subject: str, body: str) -> str:
    api_key = os.environ.get("MAILGUN_API_KEY")
    domain = os.environ.get("MAILGUN_DOMAIN")
    if not api_key or not domain:
        raise ValueError("Mailgun API key and domain not configured")
    import httpx
    async with httpx.AsyncClient(timeout=30) as cli:
        r = await cli.post(
            f"https://api.mailgun.net/v3/{domain.strip()}/messages",
            auth=("api", api_key),
            data={
                "from": _email_from_address(),
                "to": to,
                "subject": subject,
                "text": body,
                "html": f"<p>{body.replace(chr(10), '<br/>')}</p>",
            },
        )
        if r.status_code in (200, 202):
            return r.json().get("id") or "ok"
        raise ValueError(r.text[:500])

def _send_via_smtp_sync(to: str, subject: str, body: str) -> None:
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    host = os.environ.get("SMTP_HOST") or "smtp.gmail.com"
    port = int(os.environ.get("SMTP_PORT") or "587")
    user = os.environ.get("SMTP_USER")
    password = os.environ.get("SMTP_PASSWORD")
    if not user or not password:
        raise ValueError("SMTP user/password not configured")
    password = str(password).replace(" ", "")
    from_addr = _email_from_address()
    # Gmail requires authenticated address to match sender
    if "<" in from_addr and ">" in from_addr:
        display_from = from_addr
        envelope_from = from_addr.split("<")[-1].split(">")[0].strip()
    else:
        display_from = user
        envelope_from = user
    if envelope_from.lower() != user.lower():
        display_from = user
        envelope_from = user
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = display_from
    msg["To"] = to
    msg.attach(MIMEText(body, "plain", "utf-8"))
    msg.attach(MIMEText(f"<p>{body.replace(chr(10), '<br/>')}</p>", "html", "utf-8"))
    with smtplib.SMTP(host, port, timeout=30) as server:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(user, password)
        server.sendmail(envelope_from, [to], msg.as_string())

@api.post("/admin/email-test-send")
async def email_test_send(payload: dict, _=Depends(admin_only)):
    """Test send to a target email. Returns clear error if provider rejects."""
    await _reload_email_settings_from_db()
    to = (payload.get("to") or "").strip()
    if not to:
        raise HTTPException(status_code=400, detail="Recipient email required")
    if not _mailgun_configured() and not _smtp_configured() and not _resend_configured() and not _brevo_configured():
        raise HTTPException(status_code=400, detail="No email provider configured. Save Mailgun settings in Admin.")
    result = await _send_email_stub(to,
        "Boost Growth — Test Email",
        "This is a test email from your Boost Growth Portal.\n\nIf you received this, email notifications are working correctly.\n\n— Boost Growth Portal")
    if result.get("status") == "failed" and result.get("error"):
        hint = _smtp_error_hint(result["error"])
        if hint:
            result["hint"] = hint
    return result

@api.get("/admin/email-settings")
async def get_email_settings(_=Depends(admin_only)):
    doc = await db.settings.find_one({"key": "email"}, {"_id": 0}) or {}
    has_resend = bool(doc.get("resend_api_key") or os.environ.get("RESEND_API_KEY"))
    has_brevo = bool(doc.get("brevo_api_key") or os.environ.get("BREVO_API_KEY"))
    has_mailgun = bool(
        (doc.get("mailgun_api_key") and doc.get("mailgun_domain"))
        or _mailgun_configured()
    )
    has_smtp = bool(
        (doc.get("smtp_user") and doc.get("smtp_password"))
        or _smtp_configured()
    )
    provider = doc.get("email_provider") or os.environ.get("EMAIL_PROVIDER") or "auto"
    active = "none"
    if provider == "mailgun" and has_mailgun:
        active = "mailgun"
    elif provider == "brevo" and has_brevo:
        active = "brevo"
    elif provider == "smtp" and has_smtp:
        active = "smtp"
    elif provider == "resend" and has_resend:
        active = "resend"
    elif has_mailgun:
        active = "mailgun"
    elif has_brevo:
        active = "brevo"
    elif has_resend:
        active = "resend"
    elif has_smtp:
        active = "smtp"
    return {
        "configured": has_mailgun or has_smtp or has_resend or has_brevo,
        "provider": provider,
        "active_provider": active,
        "smtp_configured": has_smtp,
        "resend_configured": has_resend,
        "brevo_configured": has_brevo,
        "mailgun_configured": has_mailgun,
        "mailgun_domain": doc.get("mailgun_domain") or os.environ.get("MAILGUN_DOMAIN") or "",
        "from_email": doc.get("from_email") or os.environ.get("EMAIL_FROM") or "Boost Growth <admin@boostgrowthsa.com>",
        "smtp_host": doc.get("smtp_host") or os.environ.get("SMTP_HOST") or "smtp.gmail.com",
        "smtp_port": doc.get("smtp_port") or int(os.environ.get("SMTP_PORT") or "587"),
        "smtp_user": doc.get("smtp_user") or os.environ.get("SMTP_USER") or "",
        "key_preview": (doc.get("resend_api_key") or "")[:8] + "..." if doc.get("resend_api_key") else None,
        "brevo_key_preview": (doc.get("brevo_api_key") or "")[:12] + "..." if doc.get("brevo_api_key") else None,
        "mailgun_key_preview": (doc.get("mailgun_api_key") or "")[:12] + "..." if doc.get("mailgun_api_key") else None,
    }

@api.post("/admin/email-settings")
async def save_email_settings(payload: EmailSettingsIn, _=Depends(admin_only)):
    update = {}
    if payload.resend_api_key and payload.resend_api_key.strip():
        key = payload.resend_api_key.strip()
        if not key.startswith("re_") or len(key) < 30:
            raise HTTPException(status_code=400,
                detail=f"Invalid Resend API key. Keys start with 're_' and are 30+ chars. You provided {len(key)} chars.")
        update["resend_api_key"] = key
    if payload.brevo_api_key and payload.brevo_api_key.strip():
        key = payload.brevo_api_key.strip()
        if key.startswith("xsmtpsib-"):
            raise HTTPException(status_code=400,
                detail="This is a Brevo SMTP key (xsmtpsib-). Use an API key (xkeysib-) from SMTP & API → API Keys.")
        if len(key) < 20:
            raise HTTPException(status_code=400, detail="Invalid Brevo API key (too short).")
        update["brevo_api_key"] = key
    if payload.mailgun_api_key and payload.mailgun_api_key.strip():
        key = payload.mailgun_api_key.strip()
        if len(key) < 20:
            raise HTTPException(status_code=400, detail="Invalid Mailgun API key (too short).")
        update["mailgun_api_key"] = key
    if payload.mailgun_domain and payload.mailgun_domain.strip():
        update["mailgun_domain"] = payload.mailgun_domain.strip().lower()
    if payload.from_email and payload.from_email.strip():
        update["from_email"] = payload.from_email.strip()
    if payload.smtp_host and payload.smtp_host.strip():
        update["smtp_host"] = payload.smtp_host.strip()
    if payload.smtp_port:
        update["smtp_port"] = int(payload.smtp_port)
    if payload.smtp_user is not None:
        update["smtp_user"] = payload.smtp_user.strip()
    if payload.smtp_password and payload.smtp_password.strip():
        update["smtp_password"] = payload.smtp_password.strip().replace(" ", "")
    if payload.email_provider and payload.email_provider.strip() in ("auto", "mailgun", "brevo", "smtp", "resend"):
        update["email_provider"] = payload.email_provider.strip()
    if not update:
        raise HTTPException(status_code=400, detail="No fields")
    update["updated_at"] = now_iso()
    unset = {}
    if update.get("email_provider") == "mailgun":
        unset = {"brevo_api_key": "", "resend_api_key": ""}
    elif update.get("email_provider") == "smtp":
        unset = {"brevo_api_key": "", "resend_api_key": ""}
    await db.settings.update_one(
        {"key": "email"},
        {"$set": update, "$setOnInsert": {"key": "email"}, **({"$unset": unset} if unset else {})},
        upsert=True,
    )
    doc = await db.settings.find_one({"key": "email"}, {"_id": 0}) or {}
    _apply_email_settings(doc)
    return {"ok": True, "configured": True}

@api.get("/admin/email-queue")
async def list_email_queue(_=Depends(admin_only)):
    return await db.email_queue.find({}, {"_id": 0}).sort("created_at", -1).limit(100).to_list(100)


@api.post("/admin/email-queue/retry")
async def admin_retry_email_queue(limit: int = Query(50, ge=1, le=200), _=Depends(admin_only)):
    """Re-send queued or failed emails from email_queue (newest failures first)."""
    pending = await db.email_queue.find(
        {"status": {"$in": ["queued", "queued_no_key", "failed"]}},
        {"_id": 0},
    ).sort("created_at", -1).limit(limit).to_list(limit)
    results = []
    for doc in pending:
        r = await _send_email_stub(doc.get("to") or "", doc.get("subject") or "", doc.get("body") or "")
        patch = {
            "last_retry_at": now_iso(),
            "last_retry_status": r.get("status"),
            "last_retry_error": r.get("error"),
        }
        if r.get("status") == "sent":
            patch["status"] = "sent"
            patch["provider"] = r.get("provider")
            patch["provider_id"] = r.get("provider_id")
        await db.email_queue.update_one({"id": doc["id"]}, {"$set": patch})
        results.append({
            "id": doc.get("id"),
            "to": doc.get("to"),
            "subject": doc.get("subject"),
            "previous_status": doc.get("status"),
            "retry_status": r.get("status"),
            "error": r.get("error"),
        })
    return {"retried": len(results), "results": results}

@api.delete("/resources/{rid}")
async def delete_resource(rid: str, _=Depends(admin_only)):
    await db.resources.delete_one({"id": rid})
    return {"ok": True}

# ------------------- Leaves / Vacations -------------------
DEFAULT_ANNUAL_BALANCE = 30  # baseline annual leave per contract year

LEAVE_DOC_TYPES = {"medical", "appointment", "other"}


def _master_join_date_by_key() -> dict:
    return {key: jd for key, _, _, _, _, jd in MASTER_THERAPISTS if jd}


def _master_join_date_for_therapist(therapist: dict) -> Optional[str]:
    key = (therapist or {}).get("key")
    if key:
        jd = _master_join_date_by_key().get(key)
        if jd:
            return jd
    return None


def _effective_join_date(therapist: dict) -> Optional[str]:
    """Master seed join_date wins over HR sheet (prevents bad sheet dates like Shatha Jun vs Apr)."""
    return _master_join_date_for_therapist(therapist) or (therapist or {}).get("join_date")


def _annual_leave_entitlement(therapist: dict) -> float:
    """Annual days allocated per contract year (not the HR-synced remaining balance)."""
    if therapist.get("annual_balance") is not None:
        return float(therapist["annual_balance"])
    return float(DEFAULT_ANNUAL_BALANCE)


def _contract_period_bounds(therapist: dict, ref=None):
    """Contract year from join_date anniversary (e.g. Apr → Apr), not calendar year."""
    from datetime import date as date_cls
    ref = ref or datetime.now(timezone.utc).date()
    join_raw = _effective_join_date(therapist) or f"{ref.year}-04-01"
    try:
        jd = date_cls.fromisoformat(str(join_raw)[:10])
    except ValueError:
        jd = date_cls(ref.year, 4, 1)
    sm, sd = jd.month, min(jd.day, 28)
    try:
        period_start = date_cls(ref.year, sm, sd)
    except ValueError:
        period_start = date_cls(ref.year, sm, 28)
    if ref < period_start:
        try:
            period_start = date_cls(ref.year - 1, sm, sd)
        except ValueError:
            period_start = date_cls(ref.year - 1, sm, 28)
    try:
        period_end = date_cls(period_start.year + 1, sm, sd) - timedelta(days=1)
    except ValueError:
        period_end = date_cls(period_start.year + 1, sm, 28) - timedelta(days=1)
    return period_start.isoformat(), period_end.isoformat()


async def _ensure_contract_balance(therapist: dict) -> dict:
    """Stamp contract-period bounds; preserve HR-synced leave_balance on join_date corrections."""
    if not therapist:
        return therapist
    effective_jd = _effective_join_date(therapist)
    if effective_jd and effective_jd != therapist.get("join_date"):
        await db.therapists.update_one(
            {"id": therapist["id"]},
            {"$set": {"join_date": effective_jd}},
        )
        therapist["join_date"] = effective_jd
    start, end = _contract_period_bounds(therapist)
    if therapist.get("contract_period_start") == start:
        return therapist
    patch = {"contract_period_start": start, "contract_period_end": end}
    prev_end = (therapist.get("contract_period_end") or "")[:10]
    ref = datetime.now(timezone.utc).date()
    rolled_new_year = False
    if prev_end:
        try:
            rolled_new_year = ref > datetime.fromisoformat(prev_end).date()
        except ValueError:
            rolled_new_year = False
    if rolled_new_year and therapist.get("leave_balance") is not None:
        patch["leave_balance"] = _annual_leave_entitlement(therapist)
    elif therapist.get("contract_period_start") is None and therapist.get("leave_balance") is None:
        patch["leave_balance"] = _annual_leave_entitlement(therapist)
    await db.therapists.update_one({"id": therapist["id"]}, {"$set": patch})
    therapist.update(patch)
    return therapist


async def _leave_usage_in_contract(therapist_id: str, start: str, end: str) -> dict:
    own = await db.leaves.find({
        "therapist_id": therapist_id,
        "start_date": {"$gte": start, "$lte": end},
    }, {"_id": 0}).to_list(500)
    used_annual = sum(
        float(l.get("days") or 0)
        for l in own
        if l.get("leave_type") == "Annual" and l.get("status") in ("approved", "done")
    )
    used_permission = sum(
        float(l.get("days") or 0)
        for l in own
        if l.get("leave_type") == "Permission"
        and l.get("status") in ("approved", "done")
        and l.get("is_paid", True)
    )
    pending = sum(
        float(l.get("days") or 0)
        for l in own
        if _normalize_leave_status(l.get("status")) in ("pending_manager", "pending_hr")
    )
    return {
        "leaves": own,
        "used_annual": used_annual,
        "used_permission": used_permission,
        "pending": pending,
    }


async def _balance_row_for_therapist(t: dict, year: Optional[int] = None) -> dict:
    """Unified leave balance — HR-synced remaining wins over raw entitlement math."""
    t = await _ensure_contract_balance(t)
    start, end = _contract_period_bounds(t)
    usage = await _leave_usage_in_contract(t["id"], start, end)
    own = usage["leaves"]
    used_annual = usage["used_annual"]
    used_permission = usage["used_permission"]
    pending = usage["pending"]
    used_unpaid = sum(
        float(l.get("days") or 0)
        for l in own
        if l.get("leave_type") == "Unpaid" and l.get("status") in ("approved", "done")
    )
    used_sick = sum(
        float(l.get("days") or 0)
        for l in own
        if l.get("leave_type") == "Sickleave" and l.get("status") in ("approved", "done")
    )
    permission_count = sum(1 for l in own if l.get("leave_type") == "Permission")
    other_requests_count = await db.requests.count_documents({
        "therapist_id": t["id"],
        "request_type": {"$nin": ["leave", "permission"]},
        "status": {"$in": ["pending_manager", "pending_hr", "in_progress"]},
    })
    allocated = _annual_leave_entitlement(t)
    computed_remaining = max(0.0, allocated - used_annual - used_permission)
    stored_remaining = t.get("leave_balance")
    if stored_remaining is not None:
        remaining = max(0.0, float(stored_remaining))
        allocated = max(allocated, round(used_annual + used_permission + pending + remaining, 1))
    else:
        remaining = computed_remaining
    return {
        "therapist_id": t["id"],
        "name": therapist_schedule_display_name(t),
        "color": t.get("color"),
        "email": t.get("email"),
        "join_date": _effective_join_date(t),
        "contract_period_start": start,
        "contract_period_end": end,
        "year": year or datetime.now(timezone.utc).year,
        "allocated": allocated,
        "used_annual": round(used_annual, 1),
        "used_permission": round(used_permission, 1),
        "permission_count": permission_count,
        "other_requests_count": other_requests_count,
        "used_unpaid": round(used_unpaid, 1),
        "used_sick": round(used_sick, 1),
        "pending": round(pending, 1),
        "remaining": round(remaining, 1),
        "computed_remaining": round(computed_remaining, 1),
        "leaves_count": len(own),
    }


def _leave_default_fields() -> dict:
    return {
        "document_url": None,
        "document_file_path": None,
        "document_file_name": None,
        "document_type": None,
        "document_verified": False,
        "schedule_impact": [],
        "timeline": [],
    }


def _normalize_leave_status(status: Optional[str]) -> str:
    s = (status or "pending").strip()
    return "pending_manager" if s == "pending" else s


def _normalize_request_status(status: Optional[str]) -> str:
    s = (status or "pending").strip()
    return "pending_manager" if s == "pending" else s


def _leave_requires_document(leave_type: Optional[str]) -> bool:
    return (leave_type or "") in LEAVE_DOC_REQUIRED_TYPES


def _leave_has_document(leave: dict) -> bool:
    return bool(leave.get("document_file_path") or leave.get("document_file_data"))


def _request_has_attachment(req: dict) -> bool:
    return bool(req.get("attachment_file_path") or req.get("attachment_file_data"))


def _request_awaiting_attachment(req: dict) -> bool:
    if req.get("status") == "pending_attachment":
        return True
    if req.get("requires_attachment") and not _request_has_attachment(req):
        return True
    return False


def _append_leave_timeline(leave: dict, event: str, actor: str) -> list:
    timeline = list(leave.get("timeline") or [])
    timeline.append({"event": event, "at": now_iso(), "by": actor})
    return timeline


def _schedule_day_index(dt: datetime) -> int:
    """0=Sun … 6=Sat (matches frontend Schedule)."""
    return (dt.weekday() + 1) % 7


def _week_start_sunday(iso: str) -> str:
    d = datetime.fromisoformat(str(iso)[:10])
    sun = d - timedelta(days=_schedule_day_index(d))
    return sun.strftime("%Y-%m-%d")


def _iter_dates_in_range(start_iso: str, end_iso: str):
    start = datetime.fromisoformat(str(start_iso)[:10])
    end = datetime.fromisoformat(str(end_iso)[:10])
    d = start
    while d <= end:
        yield d.strftime("%Y-%m-%d"), d
        d += timedelta(days=1)


async def _cancel_schedule_for_therapist(therapist_id: str, start_date: str, end_date: str) -> list:
    """Mark matching schedule cells as cancel_therapist; return impact list."""
    impacted = []
    seen_ids = set()
    for date_iso, d in _iter_dates_in_range(start_date, end_date):
        week_start = _week_start_sunday(date_iso)
        day_idx = _schedule_day_index(d)
        cells = await db.schedule_cells.find(
            {
                "therapist_id": therapist_id,
                "week_start": week_start,
                "day": day_idx,
                "service_code": {"$nin": ["LEAVE", "BREAK", ""]},
            },
            {"_id": 0},
        ).to_list(50)
        for cell in cells:
            if not (cell.get("child_name") or "").strip():
                continue
            cid = cell.get("id")
            if cid in seen_ids:
                continue
            seen_ids.add(cid)
            if cell.get("state") != "cancel_therapist":
                await db.schedule_cells.update_one({"id": cid}, {"$set": {"state": "cancel_therapist"}})
            impacted.append({
                "session_id": cid,
                "date": date_iso,
                "client_name": (cell.get("child_name") or "—").strip(),
                "time_slot": cell.get("time_slot"),
            })
    return impacted


def _enrich_leave_document_url(leave: dict) -> dict:
    if _leave_has_document(leave):
        leave["document_url"] = f"/api/leaves/{leave['id']}/document"
    else:
        leave.setdefault("document_url", None)
    leave.setdefault("document_verified", False)
    leave.setdefault("schedule_impact", leave.get("schedule_impact") or [])
    return _strip_file_data(leave)

@api.get("/leaves")
async def list_leaves(
    year: Optional[int] = None,
    scope: Optional[str] = None,
    user=Depends(get_current_user),
):
    q: dict = {}
    scope_norm = (scope or "").strip().lower()
    can_view_all = _can_view_all_leaves(user, scope_norm)
    if not can_view_all:
        therapist = await db.therapists.find_one({"id": user["id"]}, {"_id": 0})
        if therapist:
            therapist = await _ensure_contract_balance(therapist)
            start, end = _contract_period_bounds(therapist)
            q["start_date"] = {"$gte": start, "$lte": end}
        else:
            q["therapist_id"] = user["id"]
    else:
        yr = year or datetime.now(timezone.utc).year
        # HR/admin views: always include open requests even when start_date is outside the selected year.
        q = {"$or": [
            {"status": {"$in": list(OPEN_LEAVE_STATUSES)}},
            {"start_date": {"$gte": f"{yr}-01-01", "$lte": f"{yr}-12-31"}},
        ]}
    if not can_view_all:
        q["therapist_id"] = user["id"]
    items = await db.leaves.find(q, {"_id": 0, "document_file_data": 0}).sort("start_date", -1).to_list(2000)
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1, "color": 1, "key": 1, "join_date": 1, "contract_period_start": 1, "contract_period_end": 1, "annual_balance": 1, "leave_balance": 1}).to_list(100)
    t_by_id = {t["id"]: t for t in therapists}
    if scope_norm == "staff" and (_is_jenan(user) or _is_walaa_ops(user)):
        contract_bounds: dict = {}
        for t in therapists:
            t = await _ensure_contract_balance(t)
            contract_bounds[t["id"]] = _contract_period_bounds(t)
        filtered = []
        for it in items:
            tid = it.get("therapist_id")
            start, end = contract_bounds.get(tid, ("", "9999-12-31"))
            sd = (it.get("start_date") or "")[:10]
            in_contract = bool(start and sd and start <= sd <= end)
            is_open = _normalize_leave_status(it.get("status")) in (
                "pending_manager", "pending_hr", "pending_attachment", "in_progress"
            )
            if in_contract or is_open:
                it["in_current_contract"] = in_contract
                filtered.append(it)
        items = filtered
    for it in items:
        t = t_by_id.get(it.get("therapist_id"))
        if t:
            it["therapist_name"] = therapist_schedule_display_name(t)
            it["therapist_color"] = t.get("color")
            if can_view_all:
                it["therapist_email"] = t.get("email")
    return [_enrich_leave_document_url(it) for it in items]

@api.get("/leaves/balance")
async def leaves_balance(year: Optional[int] = None, scope: Optional[str] = None, user=Depends(get_current_user)):
    """Per-therapist balance for current contract year (anniversary from join_date)."""
    scope_norm = (scope or "").strip().lower()
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "color": 1, "email": 1, "annual_balance": 1, "leave_balance": 1, "join_date": 1, "contract_period_start": 1, "contract_period_end": 1, "key": 1, "leave_balance_sync_year": 1}).to_list(100)
    can_view_all = _is_portal_admin(user) or _is_hr_ops(user) or (scope_norm == "staff" and _is_jenan(user))
    if not can_view_all:
        therapists = [t for t in therapists if t["id"] == user["id"]]
    out = []
    for t in therapists:
        out.append(await _balance_row_for_therapist(t, year))
    return out

@api.post("/leaves")
async def create_leave(payload: LeaveIn, user=Depends(get_current_user)):
    if not _is_portal_admin(user) and not _is_hr_ops(user) and payload.therapist_id != user["id"]:
        raise HTTPException(status_code=403, detail="Therapist can only create own leaves")
    if (payload.leave_type or "") in ("Exam", "Emergency"):
        raise HTTPException(status_code=400, detail="This leave type is no longer accepted. Choose Annual, Sick, Unpaid, or Permission.")
    lid = str(uuid.uuid4())
    doc = {"id": lid, **payload.model_dump(), **_leave_default_fields(), "created_by": user["id"], "created_at": now_iso()}
    if user.get("role") != "admin":
        if _leave_requires_document(payload.leave_type) and not _leave_has_document(doc):
            doc["status"] = "pending_attachment"
        else:
            doc["status"] = "pending_manager"
        doc["timeline"] = [{"event": "submitted", "at": now_iso(), "by": _actor_display(user)}]
    await db.leaves.insert_one(doc)
    doc.pop("_id", None)
    if user.get("role") != "admin":
        time_part = ""
        if payload.leave_type == "Permission" and payload.start_time:
            time_part = f" {payload.start_time}"
            if payload.end_time:
                time_part += f"–{payload.end_time}"
        await _notify_leave_submitted(
            therapist_name=(user.get("name") or "Therapist").strip(),
            leave_type=payload.leave_type,
            start_date=payload.start_date,
            end_date=payload.end_date,
            days=float(payload.days or 0),
            notes=payload.notes,
        )
    return doc

@api.put("/leaves/{lid}")
async def update_leave(lid: str, payload: LeaveIn, user=Depends(get_current_user)):
    leave = await db.leaves.find_one({"id": lid})
    if not leave:
        raise HTTPException(status_code=404, detail="Not found")
    if not _is_portal_admin(user) and leave.get("therapist_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Forbidden")
    update = payload.model_dump()
    await db.leaves.update_one({"id": lid}, {"$set": update})
    return await db.leaves.find_one({"id": lid}, {"_id": 0})

@api.put("/leaves/{lid}/status")
async def update_leave_status(lid: str, payload: LeaveStatusUpdate, user=Depends(get_current_user)):
    leave = await db.leaves.find_one({"id": lid})
    if not leave:
        raise HTTPException(status_code=404, detail="Not found")
    prev_status = leave.get("status")
    effective_prev = _normalize_leave_status(prev_status)
    new_status = _coerce_manager_approve_to_hr(payload.status, payload.notify_hr)
    is_pa = _is_portal_admin(user)
    is_hr = _is_hr_ops(user)
    is_jenan_mgr = _is_jenan(user) and not is_pa
    notify_hr = payload.notify_hr if payload.notify_hr is not None else (not is_jenan_mgr)
    notify_therapist = payload.notify_therapist if payload.notify_therapist is not None else (not is_jenan_mgr)

    if is_pa:
        pass
    elif is_jenan_mgr:
        if effective_prev not in PENDING_MANAGER_STATUSES and effective_prev != "pending_attachment":
            raise HTTPException(status_code=403, detail="Manager can only act on pending manager requests")
        if effective_prev == "pending_attachment" or (_leave_requires_document(leave.get("leave_type")) and not _leave_has_document(leave)):
            raise HTTPException(status_code=400, detail="Document attachment required before review")
        if new_status not in ("pending_hr", "rejected", "pending_manager", "approved"):
            raise HTTPException(status_code=400, detail="Manager must choose pending, approve, or reject")
    elif is_hr:
        if effective_prev != "pending_hr":
            raise HTTPException(status_code=403, detail="HR can only act on HR-pending requests")
        if _leave_requires_document(leave.get("leave_type")) and not _leave_has_document(leave):
            raise HTTPException(status_code=400, detail="Document attachment required before approval")
        if new_status not in ("approved", "rejected"):
            raise HTTPException(status_code=400, detail="HR must approve or reject")
    else:
        raise HTTPException(status_code=403, detail="Leave management access required")

    is_paid = payload.is_paid if payload.is_paid is not None else leave.get("is_paid", True)
    deduct = payload.deduct_balance if payload.deduct_balance is not None else True
    if new_status == "approved" and leave.get("leave_type") == "Permission" and payload.is_paid is False:
        is_paid = False
        deduct = False
    if is_hr and not is_pa and new_status == "approved":
        lt = (leave.get("leave_type") or "").lower()
        if lt in ("unpaid", "absence") or not is_paid:
            deduct = False
    actor = _actor_display(user)
    timeline = _append_leave_timeline(leave, new_status, actor)
    await db.leaves.update_one({"id": lid}, {"$set": {
        "status": new_status, "admin_note": payload.admin_note,
        "decided_by": user.get("name") or actor, "decided_at": now_iso(),
        "is_paid": is_paid,
        "timeline": timeline,
    }})
    if is_jenan_mgr and effective_prev in MANAGER_FORWARD_HR_LEAVE_SOURCES and notify_hr and new_status in MANAGER_HR_NOTIFY_STATUSES:
        tname = leave.get("leave_type") or "Leave"
        summary = (
            f"{tname} {leave.get('days')}d "
            f"({leave.get('start_date')} → {leave.get('end_date')})"
        )
        await _notify_hr_manager_decision(
            ntype="leave_request",
            therapist_name=leave.get("therapist_name") or "Therapist",
            summary=summary,
            decision_status=new_status,
            admin_note=payload.admin_note,
        )
    # Deduct balance when newly approved (skip unpaid / absence)
    if new_status == "approved" and prev_status != "approved" and leave.get("therapist_id"):
        lt = (leave.get("leave_type") or "").lower()
        if deduct and is_paid and lt not in ("unpaid", "absence"):
            t = await db.therapists.find_one({"id": leave["therapist_id"]}, {"_id": 0, "leave_balance": 1})
            if t is not None and t.get("leave_balance") is not None:
                days = float(leave.get("days") or 0)
                new_bal = max(0.0, float(t["leave_balance"]) - days)
                await db.therapists.update_one(
                    {"id": leave["therapist_id"]},
                    {"$set": {"leave_balance": new_bal}},
                )
    # Notify therapist (in-app + email) when requested
    if leave.get("therapist_id") and notify_therapist:
        msg_map = {
            "approved": "Approved", "rejected": "Rejected", "done": "Completed",
            "cancelled": "Cancelled", "pending": "Pending", "pending_manager": "Pending manager review",
            "pending_hr": "Pending HR review",
        }
        label = msg_map.get(new_status, new_status)
        msg = (
            f"Your {leave.get('leave_type')} leave from {leave.get('start_date')} to "
            f"{leave.get('end_date')} ({leave.get('days')}d) is now {label}."
        )
        await _notify(leave["therapist_id"], "leave", f"Leave {label}", msg)
        if new_status in ("approved", "rejected"):
            tname = leave.get("leave_type") or "Leave"
            await _push_center_update(
                f"Leave {label.lower()}: {tname}",
                f"{leave.get('start_date')} → {leave.get('end_date')} ({leave.get('days')} day(s))",
            )
            therapist = await db.therapists.find_one({"id": leave["therapist_id"]}, {"_id": 0, "email": 1, "name": 1})
            if therapist and therapist.get("email"):
                await _send_email_stub(
                    therapist["email"],
                    f"[Boost Growth] Leave Request {label}",
                    f"Dear {therapist.get('name', '')},\n\nYour leave request from {leave.get('start_date')} to "
                    f"{leave.get('end_date')} has been {label.lower()}.\n\n— Boost Growth Portal",
                )
    return await db.leaves.find_one({"id": lid}, {"_id": 0})

@api.delete("/leaves/{lid}")
async def delete_leave(lid: str, user=Depends(get_current_user)):
    leave = await db.leaves.find_one({"id": lid})
    if not leave:
        return {"ok": True}
    if not _can_delete_staff_submission(user, leave.get("therapist_id")):
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.leaves.delete_one({"id": lid})
    return {"ok": True}


@api.post("/leaves/mark-absence")
async def mark_absence_without_request(payload: MarkAbsenceIn, admin=Depends(leave_manager)):
    """Admin: record absence/permission and optionally cancel schedule sessions."""
    t = await db.therapists.find_one({"id": payload.therapist_id}, {"_id": 0, "id": 1, "name": 1})
    if not t:
        raise HTTPException(status_code=404, detail="Therapist not found")
    days = max(1, (datetime.fromisoformat(payload.date_to[:10]) - datetime.fromisoformat(payload.date_from[:10])).days + 1)
    lid = str(uuid.uuid4())
    doc = {
        "id": lid,
        "therapist_id": payload.therapist_id,
        "start_date": payload.date_from[:10],
        "end_date": payload.date_to[:10],
        "days": float(days),
        "leave_type": payload.leave_type or "Absence",
        "status": "absent",
        "notes": payload.notes,
        "admin_note": "Marked absent by admin",
        **_leave_default_fields(),
        "created_by": admin["id"],
        "created_at": now_iso(),
        "decided_by": admin.get("name") or "Admin",
        "decided_at": now_iso(),
    }
    impact = []
    if payload.cancel_sessions:
        impact = await _cancel_schedule_for_therapist(payload.therapist_id, doc["start_date"], doc["end_date"])
        doc["schedule_impact"] = impact
    await db.leaves.insert_one(doc)
    doc.pop("_id", None)
    return {
        "leave": _enrich_leave_document_url(doc),
        "cancelled_sessions_count": len(impact),
        "sessions": impact,
        "message": f"Done. {len(impact)} session(s) cancelled for {t.get('name')}",
    }


@api.post("/leaves/{lid}/mark-absent")
async def mark_leave_absent(lid: str, payload: MarkAbsentIn, admin=Depends(leave_manager)):
    leave = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")
    impact = list(leave.get("schedule_impact") or [])
    if payload.cancel_sessions and leave.get("therapist_id"):
        new_impact = await _cancel_schedule_for_therapist(
            leave["therapist_id"], leave["start_date"], leave["end_date"]
        )
        existing_ids = {x.get("session_id") for x in impact}
        for row in new_impact:
            if row.get("session_id") not in existing_ids:
                impact.append(row)
    await db.leaves.update_one({"id": lid}, {"$set": {
        "status": "absent",
        "schedule_impact": impact,
        "decided_by": admin.get("name") or "Admin",
        "decided_at": now_iso(),
    }})
    updated = await db.leaves.find_one({"id": lid}, {"_id": 0})
    tname = leave.get("therapist_name") or "Therapist"
    therapists = await db.therapists.find_one({"id": leave.get("therapist_id")}, {"name": 1})
    if therapists:
        tname = therapists.get("name") or tname
    return {
        "leave": _enrich_leave_document_url(updated),
        "cancelled_sessions_count": len(impact),
        "sessions": impact,
        "message": f"Done. {len(impact)} session(s) cancelled on {leave.get('start_date')} → {leave.get('end_date')} for {tname}",
    }


@api.post("/leaves/{lid}/upload-document")
async def upload_leave_document(
    lid: str,
    file: UploadFile = File(...),
    document_type: Optional[str] = Form("other"),
    user=Depends(get_current_user),
):
    leave = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")
    if user.get("role") != "admin" and leave.get("therapist_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Forbidden")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="File required")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")
    ext = Path(file.filename).suffix.lower() or ".pdf"
    if ext not in (".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif"):
        raise HTTPException(status_code=400, detail="PDF or image only")
    stored = f"leave_{lid}{ext}"
    if leave.get("document_file_path"):
        old = UPLOAD_DIR / leave["document_file_path"]
        if old.exists() and old.name != stored:
            old.unlink()
    file_data = _persist_upload(stored, content)
    dtype = (document_type or "other").lower()
    if dtype not in LEAVE_DOC_TYPES:
        dtype = "other"
    await db.leaves.update_one({"id": lid}, {"$set": {
        "document_file_path": stored,
        "document_file_name": file.filename,
        "document_file_data": file_data,
        "document_type": dtype,
        "document_verified": False,
        "document_uploaded_at": now_iso(),
    }})
    if leave.get("status") == "pending_attachment":
        timeline = _append_leave_timeline(leave, "document_uploaded", _actor_display(user))
        await db.leaves.update_one({"id": lid}, {"$set": {
            "status": "pending_manager",
            "timeline": timeline,
            "updated_at": now_iso(),
        }})
    updated = await db.leaves.find_one({"id": lid}, {"_id": 0})
    return _enrich_leave_document_url(updated)


@api.get("/leaves/{lid}/document")
async def download_leave_document(lid: str, user=Depends(get_current_user)):
    leave = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not leave or not _leave_has_document(leave):
        raise HTTPException(status_code=404, detail="No document")
    if not (_is_portal_admin(user) or _is_hr_ops(user) or _is_jenan(user)):
        if leave.get("therapist_id") != user["id"]:
            raise HTTPException(status_code=403, detail="Forbidden")
    content = _load_upload(leave.get("document_file_path"), leave.get("document_file_data"))
    if not content:
        raise HTTPException(status_code=404, detail=FILE_UNAVAILABLE_DETAIL)
    fname = leave.get("document_file_name") or leave.get("document_file_path") or "document"
    return _bytes_file_response(content, fname)


@api.delete("/leaves/{lid}/document")
async def delete_leave_document(lid: str, user=Depends(get_current_user)):
    leave = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")
    if user.get("role") != "admin" and leave.get("therapist_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Forbidden")
    if leave.get("document_file_path"):
        fp = UPLOAD_DIR / leave["document_file_path"]
        if fp.exists():
            fp.unlink()
    await db.leaves.update_one({"id": lid}, {"$set": {
        "document_file_path": None,
        "document_file_name": None,
        "document_file_data": None,
        "document_type": None,
        "document_verified": False,
        "document_uploaded_at": None,
    }})
    return {"ok": True}


@api.put("/leaves/{lid}/verify-document")
async def verify_leave_document(lid: str, payload: LeaveDocumentVerifyIn, _=Depends(leave_manager)):
    leave = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")
    await db.leaves.update_one({"id": lid}, {"$set": {"document_verified": bool(payload.verified)}})
    updated = await db.leaves.find_one({"id": lid}, {"_id": 0})
    return _enrich_leave_document_url(updated)


@api.post("/admin/clear-leaves")
async def admin_clear_all_leaves(_=Depends(admin_only)):
    """Delete ALL leave records (test data cleanup)."""
    result = await db.leaves.delete_many({})
    return {"deleted": result.deleted_count, "message": f"Deleted {result.deleted_count} leave records"}


ONE_TIME_LEAVE_DELETE_TARGETS = {
    "hajar": {
        "meta_key": "one_time_leave_delete_hajar_v1",
        "email": "halfulaij@boostgrowthsa.com",
        "key": "msHajer",
        "therapist_id": "1eee3003-46e2-4051-a18a-ad76827f6d67",
        "label": "Ms. Hajar",
    },
    "razan": {
        "meta_key": "one_time_leave_delete_razan_v1",
        "email": "ralshatery@boostgrowthsa.com",
        "key": "msRazan",
        "therapist_id": "2832a061-2e3b-4a91-af66-8b5ec6ff00d4",
        "label": "Ms. Razan",
    },
}


async def _one_time_leave_delete_target(target: str) -> dict:
    cfg = ONE_TIME_LEAVE_DELETE_TARGETS.get((target or "").strip().lower())
    if not cfg:
        raise HTTPException(status_code=404, detail="Unknown target")
    return cfg


async def _resolve_one_time_leave_therapist(cfg: dict) -> Optional[dict]:
    therapist = await db.therapists.find_one(
        {"id": cfg["therapist_id"]},
        {"_id": 0, "pin_hash": 0, "password_hash": 0},
    )
    if not therapist:
        therapist = await db.therapists.find_one(
            {"$or": [{"email": cfg["email"]}, {"key": cfg["key"]}]},
            {"_id": 0, "pin_hash": 0, "password_hash": 0},
        )
    return therapist


async def _latest_leave_for_therapist(therapist_id: str) -> Optional[dict]:
    items = await db.leaves.find(
        {"therapist_id": therapist_id},
        {"_id": 0, "document_file_data": 0},
    ).sort([("created_at", -1), ("start_date", -1)]).to_list(1)
    return items[0] if items else None


async def _restore_leave_balance_if_needed(leave: dict) -> Optional[float]:
    if not leave or not leave.get("therapist_id"):
        return None
    status = _normalize_leave_status(leave.get("status"))
    if status not in ("approved", "done"):
        return None
    lt = (leave.get("leave_type") or "").lower()
    is_paid = leave.get("is_paid", True)
    if not is_paid or lt in ("unpaid", "absence"):
        return None
    t = await db.therapists.find_one({"id": leave["therapist_id"]}, {"_id": 0, "leave_balance": 1})
    if t is None or t.get("leave_balance") is None:
        return None
    days = float(leave.get("days") or 0)
    if days <= 0:
        return None
    new_bal = float(t["leave_balance"]) + days
    await db.therapists.update_one(
        {"id": leave["therapist_id"]},
        {"$set": {"leave_balance": new_bal}},
    )
    return new_bal


async def _one_time_leave_delete_status_block(target: str) -> dict:
    cfg = await _one_time_leave_delete_target(target)
    meta = await db.meta.find_one({"key": cfg["meta_key"]}, {"_id": 0})
    consumed = bool(meta)
    therapist = await _resolve_one_time_leave_therapist(cfg)
    latest = None
    if therapist:
        leave = await _latest_leave_for_therapist(therapist["id"])
        if leave:
            latest = _enrich_leave_document_url(leave)
            latest["therapist_name"] = therapist_schedule_display_name(therapist)
            latest["therapist_email"] = therapist.get("email")
    return {
        "target": target,
        "label": cfg["label"],
        "consumed": consumed,
        "available": not consumed,
        "therapist": {
            "id": therapist.get("id") if therapist else cfg.get("therapist_id"),
            "name": therapist_schedule_display_name(therapist) if therapist else cfg["label"],
            "email": (therapist or {}).get("email") or cfg["email"],
        } if therapist or cfg.get("therapist_id") else None,
        "latest_leave": latest,
        "meta": meta,
    }


class OneTimeLeaveDeleteConfirmIn(BaseModel):
    confirm: str


class ResendLeaveNotificationsIn(BaseModel):
    therapists: Optional[List[str]] = None
    include_pending_attachment: bool = True
    dry_run: bool = False
    also_notify_in_app: bool = True


class ResendPurchaseNotificationsIn(BaseModel):
    therapists: Optional[List[str]] = None
    statuses: Optional[List[str]] = None
    dry_run: bool = False


@api.get("/admin/leaves-audit")
async def admin_leaves_audit(q: str = Query(...), _=Depends(admin_only)):
    """List leave rows for therapists matching name/email/key fragment."""
    regex = {"$regex": (q or "").strip() or ".", "$options": "i"}
    therapists = await db.therapists.find(
        {"$or": [{"name": regex}, {"email": regex}, {"key": regex}]},
        {"_id": 0, "pin_hash": 0, "password_hash": 0},
    ).sort("name", 1).to_list(50)
    blocks = []
    for t in therapists:
        t = await _ensure_contract_balance(t)
        start, end = _contract_period_bounds(t)
        leaves = await db.leaves.find(
            {"therapist_id": t["id"]}, {"_id": 0, "document_file_data": 0},
        ).sort("start_date", -1).to_list(500)
        blocks.append({
            "therapist": {
                "id": t["id"],
                "name": therapist_schedule_display_name(t),
                "email": t.get("email"),
                "leave_balance": t.get("leave_balance"),
                "contract_period_start": start,
                "contract_period_end": end,
            },
            "leaves_total": len(leaves),
            "leaves_open": sum(
                1 for l in leaves if _normalize_leave_status(l.get("status")) in OPEN_LEAVE_STATUSES
            ),
            "leaves": [
                {
                    "id": l.get("id"),
                    "start_date": l.get("start_date"),
                    "end_date": l.get("end_date"),
                    "days": l.get("days"),
                    "leave_type": l.get("leave_type"),
                    "status": l.get("status"),
                    "created_at": l.get("created_at"),
                    "in_current_contract": start <= (l.get("start_date") or "") <= end,
                }
                for l in leaves
            ],
        })
    return {"therapists": blocks, "query": q}


class RecalculateLeaveBalancesIn(BaseModel):
    therapists: Optional[List[str]] = None
    sync_sheet: bool = True


@api.post("/admin/recalculate-leave-balances")
async def admin_recalculate_leave_balances(
    body: RecalculateLeaveBalancesIn = Body(default_factory=RecalculateLeaveBalancesIn),
    _=Depends(admin_only),
):
    """Re-sync HR sheet balances and refresh contract bounds (fixes join_date / remaining drift)."""
    sheet_result = None
    if body.sync_sheet:
        sheet_result = await _sync_leave_balances_from_sheet()
    if body.therapists:
        regex_clauses = []
        for pat in body.therapists:
            fragment = (pat or "").strip()
            if fragment:
                regex_clauses.append({"name": {"$regex": fragment, "$options": "i"}})
                regex_clauses.append({"email": {"$regex": fragment, "$options": "i"}})
                regex_clauses.append({"key": {"$regex": fragment, "$options": "i"}})
        q = {"$or": regex_clauses} if regex_clauses else {}
        therapists = await db.therapists.find(q, {"_id": 0, "pin_hash": 0, "password_hash": 0}).to_list(50)
    else:
        therapists = await db.therapists.find({}, {"_id": 0, "pin_hash": 0, "password_hash": 0}).to_list(200)
    rows = []
    for t in therapists:
        row = await _balance_row_for_therapist(t)
        rows.append({
            "therapist_id": row["therapist_id"],
            "name": row["name"],
            "join_date": row["join_date"],
            "contract_period_start": row["contract_period_start"],
            "contract_period_end": row["contract_period_end"],
            "remaining": row["remaining"],
            "computed_remaining": row.get("computed_remaining"),
            "used_annual": row["used_annual"],
            "pending": row["pending"],
            "allocated": row["allocated"],
        })
    return {"sheet_sync": sheet_result, "therapists": rows}


@api.post("/admin/resend-leave-notifications")
async def admin_resend_leave_notifications(
    body: ResendLeaveNotificationsIn = Body(default_factory=ResendLeaveNotificationsIn),
    _=Depends(admin_only),
):
    """Re-send Jenan urgent emails for open manager-pending leaves (optionally filter by therapist name)."""
    statuses = list(PENDING_MANAGER_STATUSES)
    if body.include_pending_attachment:
        statuses.append("pending_attachment")
    therapist_ids: Optional[set] = None
    if body.therapists:
        therapist_ids = set()
        for pat in body.therapists:
            fragment = (pat or "").strip()
            if not fragment:
                continue
            regex = {"$regex": fragment, "$options": "i"}
            matches = await db.therapists.find(
                {"$or": [{"name": regex}, {"email": regex}, {"key": regex}]},
                {"_id": 0, "id": 1},
            ).to_list(50)
            for m in matches:
                therapist_ids.add(m["id"])
    query: dict = {"status": {"$in": statuses}}
    if therapist_ids is not None:
        if not therapist_ids:
            return {"dry_run": body.dry_run, "matched": 0, "sent": [], "jenan_email": await _jenan_recipient_email()}
        query["therapist_id"] = {"$in": list(therapist_ids)}
    leaves = await db.leaves.find(query, {"_id": 0, "document_file_data": 0}).sort("created_at", -1).to_list(200)
    therapist_cache: dict = {}
    preview = []
    for leave in leaves:
        tid = leave.get("therapist_id")
        if tid and tid not in therapist_cache:
            therapist_cache[tid] = await db.therapists.find_one({"id": tid}, {"_id": 0, "id": 1, "name": 1, "key": 1, "email": 1})
        therapist = therapist_cache.get(tid)
        preview.append({
            "leave_id": leave.get("id"),
            "therapist_name": therapist_schedule_display_name(therapist) if therapist else "—",
            "status": leave.get("status"),
            "start_date": leave.get("start_date"),
            "end_date": leave.get("end_date"),
            "leave_type": leave.get("leave_type"),
        })
    if body.dry_run:
        return {
            "dry_run": True,
            "matched": len(leaves),
            "jenan_email": await _jenan_recipient_email(),
            "leaves": preview,
        }
    sent = []
    for leave in leaves:
        tid = leave.get("therapist_id")
        therapist = therapist_cache.get(tid) if tid else None
        sent.append(await _resend_leave_notification(leave, therapist, also_in_app=body.also_notify_in_app))
    return {
        "dry_run": False,
        "matched": len(leaves),
        "jenan_email": await _jenan_recipient_email(),
        "sent": sent,
    }


@api.post("/admin/resend-purchase-notifications")
async def admin_resend_purchase_notifications(
    body: ResendPurchaseNotificationsIn = Body(default_factory=ResendPurchaseNotificationsIn),
    _=Depends(admin_only),
):
    """Re-send Jenan urgent emails for open staff purchases (default: pending review)."""
    statuses = body.statuses or ["pending"]
    therapist_ids: Optional[set] = None
    if body.therapists:
        therapist_ids = set()
        for pat in body.therapists:
            fragment = (pat or "").strip()
            if not fragment:
                continue
            regex = {"$regex": fragment, "$options": "i"}
            matches = await db.therapists.find(
                {"$or": [{"name": regex}, {"email": regex}, {"key": regex}]},
                {"_id": 0, "id": 1},
            ).to_list(50)
            for m in matches:
                therapist_ids.add(m["id"])
    query: dict = {"status": {"$in": statuses}}
    if therapist_ids is not None:
        if not therapist_ids:
            return {"dry_run": body.dry_run, "matched": 0, "sent": [], "jenan_email": await _jenan_recipient_email()}
        query["therapist_id"] = {"$in": list(therapist_ids)}
    purchases = await db.staff_purchases.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    preview = [
        {
            "purchase_id": p.get("id"),
            "purchaser_name": p.get("purchaser_name") or p.get("therapist_name"),
            "item": p.get("item"),
            "category": p.get("category"),
            "status": p.get("status"),
            "purchase_date": p.get("purchase_date"),
        }
        for p in purchases
    ]
    if body.dry_run:
        return {
            "dry_run": True,
            "matched": len(purchases),
            "jenan_email": await _jenan_recipient_email(),
            "purchases": preview,
        }
    sent = []
    for p in purchases:
        name = p.get("purchaser_name") or p.get("therapist_name") or "Staff"
        item = p.get("item") or "—"
        category = p.get("category") or "—"
        title = "New staff purchase logged"
        message = f"{name}: {item} ({category}) — pending review"
        body_text = f"{message}\n"
        portal = _portal_base_url()
        if portal:
            body_text += f"\nReview in portal: {portal}/purchases\n"
        body_text += "\n— Boost Growth Portal"
        email_result = await _send_urgent_email(await _jenan_recipient_email(), title, body_text)
        sent.append({
            "purchase_id": p.get("id"),
            "purchaser_name": name,
            "email_to": await _jenan_recipient_email(),
            "email_status": email_result.get("status"),
            "email_error": email_result.get("error"),
        })
    return {
        "dry_run": False,
        "matched": len(purchases),
        "jenan_email": await _jenan_recipient_email(),
        "sent": sent,
    }


@api.get("/admin/one-time-leave-deletes")
async def admin_one_time_leave_deletes_status(_=Depends(admin_only)):
    hajar = await _one_time_leave_delete_status_block("hajar")
    razan = await _one_time_leave_delete_status_block("razan")
    return {"targets": {"hajar": hajar, "razan": razan}}


@api.post("/admin/one-time-leave-deletes/{target}/delete")
async def admin_one_time_leave_delete(target: str, body: OneTimeLeaveDeleteConfirmIn, user=Depends(admin_only)):
    if body.confirm != "DELETE":
        raise HTTPException(status_code=400, detail='Type "DELETE" to confirm')
    cfg = await _one_time_leave_delete_target(target)
    existing = await db.meta.find_one({"key": cfg["meta_key"]}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=403, detail=f"One-time delete for {cfg['label']} has already been used")
    therapist = await _resolve_one_time_leave_therapist(cfg)
    if not therapist:
        raise HTTPException(status_code=404, detail=f"Therapist not found for {cfg['label']}")
    leave = await _latest_leave_for_therapist(therapist["id"])
    if not leave:
        raise HTTPException(status_code=404, detail=f"No leave found for {cfg['label']}")
    restored_balance = await _restore_leave_balance_if_needed(leave)
    if leave.get("document_file_path"):
        fp = UPLOAD_DIR / leave["document_file_path"]
        if fp.exists():
            fp.unlink()
    result = await db.leaves.delete_one({"id": leave["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Leave not found")
    therapist_name = therapist_schedule_display_name(therapist)
    await db.meta.update_one(
        {"key": cfg["meta_key"]},
        {"$set": {
            "key": cfg["meta_key"],
            "target": target,
            "used_at": now_iso(),
            "used_by": _actor_display(user),
            "used_by_id": user.get("id"),
            "leave_id": leave.get("id"),
            "therapist_id": therapist.get("id"),
            "therapist_name": therapist_name,
            "therapist_email": therapist.get("email"),
            "leave_type": leave.get("leave_type"),
            "start_date": leave.get("start_date"),
            "end_date": leave.get("end_date"),
            "days": leave.get("days"),
            "status": leave.get("status"),
            "created_at": leave.get("created_at"),
            "balance_restored_to": restored_balance,
        }},
        upsert=True,
    )
    return {
        "ok": True,
        "target": target,
        "deleted_leave_id": leave.get("id"),
        "therapist_name": therapist_name,
        "balance_restored_to": restored_balance,
        "message": f"Deleted latest leave for {therapist_name}",
        "status": await _one_time_leave_delete_status_block(target),
    }


PROGRESS_REPORT_DRIVE_URLS = {
    "009": "https://docs.google.com/document/d/14c29YPvhWaZirB5Qc-_47qP7Q_04-IOZEhpWk76WiU0/edit",
    "024": "https://docs.google.com/document/d/1DS4n4WvIB2_lS-XaZD3gSYg8WcDa5k_RYLwAIOsX7Ig/edit",
    "038": "https://docs.google.com/document/d/1-cxoewBVcbyVXa-XuBziY4OFAD2bfSpYL2_SRK4p2Ik/edit",
    "040": "https://docs.google.com/document/d/1uPUgFPz944AqlHXFXT3oVOpar6JETzsQQ3a6rd3XTK8/edit",
    "042": "https://docs.google.com/document/d/14tyu4xNlG4AmzALpjYwuWwKflwqk_buX-i2rsovxKRY/edit",
    "047": "https://drive.google.com/file/d/1eD8w2NQ5WCRtZrODhX33RHQYGT2jbHvM/view",
    "063": "https://drive.google.com/file/d/1av1C994LOEuMY2ChsEl0t8QaS5fO3s7m/view",
    "070": "https://drive.google.com/file/d/1tI5z5vrDDVaApSOAsp4HbkcawcDe42ll/view",
    "072": "https://docs.google.com/document/d/19UY48orOHqV-ItptNFVxUgRyLgWeWgi8gIy--SmbMHA/edit",
    "034": "https://drive.google.com/file/d/1DRU9zPhF0fmS7RIOQFJ3FDaH7H7gkfTN/view",
}


@api.post("/admin/migrate-progress-report-urls")
async def admin_migrate_progress_report_urls(_=Depends(admin_only)):
    """One-time: set Drive URLs on existing Apr 2026 progress reports by client file_no."""
    updated = 0
    missing = []
    for file_no, url in PROGRESS_REPORT_DRIVE_URLS.items():
        client = await _find_client_by_file_no(file_no)
        if not client:
            missing.append(file_no)
            continue
        r = await db.progress_reports.update_many(
            {"client_id": client["id"], "title": {"$regex": "Apr 2026", "$options": "i"}},
            {"$set": {"url": url, "updated_at": now_iso()}},
        )
        updated += r.modified_count
    return {"updated": updated, "missing_clients": missing, "message": f"Updated {updated} progress report URLs"}


@api.post("/admin/repair-session-invoices")
async def admin_repair_session_invoices(_=Depends(admin_only)):
    """Backfill invoice_id on sessions; fix HS service_type; link orphans by date window."""
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(5000)
    inv_by_num = {}
    for inv in invoices:
        num = (inv.get("invoice_number") or "").strip()
        if num:
            inv_by_num[f"{inv['client_id']}|{num}"] = inv["id"]
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(20000)
    linked = typed = window_linked = 0
    clients = {c["id"]: c for c in await db.clients.find(_active_client_filter(), {"_id": 0, "id": 1, "service_type": 1}).to_list(500)}
    invs_by_client = {}
    for inv in invoices:
        invs_by_client.setdefault(inv["client_id"], []).append(inv)

    for s in sessions:
        patch = {}
        cid = s.get("client_id")
        if not s.get("invoice_id") and s.get("source_invoice"):
            key = f"{cid}|{(s.get('source_invoice') or '').strip()}"
            if key in inv_by_num:
                patch["invoice_id"] = inv_by_num[key]
                linked += 1
        client = clients.get(cid)
        if client and _normalize_service_type(client.get("service_type")) == "HS":
            if _normalize_service_type(s.get("service_type")) != "HS":
                patch["service_type"] = "HS"
                patch["week_number"] = None
                typed += 1
        if patch:
            await db.sessions.update_one({"id": s["id"]}, {"$set": patch})

    # Date-window backfill for sessions still without invoice link
    for cid, client_invs in invs_by_client.items():
        sorted_invs = _sorted_invoices_for_client(cid, client_invs)
        client_sessions = [s for s in sessions if s.get("client_id") == cid]
        for inv in sorted_invs:
            inv_num = (inv.get("invoice_number") or "").strip()
            for s in client_sessions:
                if _session_has_invoice_link(s):
                    continue
                if not _session_in_invoice_date_window(s, inv, sorted_invs):
                    continue
                patch = {"invoice_id": inv["id"]}
                if inv_num:
                    patch["source_invoice"] = inv_num
                await db.sessions.update_one({"id": s["id"]}, {"$set": patch})
                s["invoice_id"] = inv["id"]
                if inv_num:
                    s["source_invoice"] = inv_num
                window_linked += 1

    return {
        "invoice_ids_linked": linked,
        "service_types_fixed": typed,
        "window_linked": window_linked,
    }


class FixSwappedSessionDatesIn(BaseModel):
    dry_run: bool = True
    client_id: Optional[str] = None
    file_no: Optional[str] = None


@api.post("/admin/fix-swapped-session-dates")
async def admin_fix_swapped_session_dates(body: FixSwappedSessionDatesIn, _=Depends(client_lead_or_admin)):
    """Fix sessions where month/day were swapped (e.g. 06/11 meant June 11, stored as Nov 6)."""
    cq = _active_client_filter()
    if body.client_id:
        cq["id"] = body.client_id.strip()
    if body.file_no:
        cq["file_no"] = str(body.file_no).strip().zfill(3)
    clients = await db.clients.find(cq, {"_id": 0, "id": 1, "name": 1, "file_no": 1}).to_list(500)

    fixes: List[dict] = []
    applied = 0
    skipped = 0

    for client in clients:
        cid = client["id"]
        sessions = await db.sessions.find({"client_id": cid}, {"_id": 0}).to_list(5000)
        by_invoice: Dict[str, list] = {}
        for s in sessions:
            inv_key = s.get("invoice_id") or (s.get("source_invoice") or "").strip() or "__none__"
            by_invoice.setdefault(inv_key, []).append(s)

        for _inv_key, group in by_invoice.items():
            if len(group) < 3:
                continue
            peer_dates = [s.get("session_date") for s in group if s.get("session_date")]
            for s in group:
                iso = _normalize_session_date_iso(s.get("session_date"))
                if not iso:
                    continue
                corrected = _session_likely_swapped_month_day(iso, peer_dates)
                if not corrected:
                    continue
                start_t = s.get("start_time") or ""
                conflict = any(
                    _normalize_session_date_iso(o.get("session_date")) == corrected
                    and (o.get("start_time") or "") == start_t
                    and o.get("id") != s.get("id")
                    for o in group
                )
                if conflict:
                    skipped += 1
                    continue
                fixes.append({
                    "client_name": client.get("name"),
                    "file_no": client.get("file_no"),
                    "session_id": s.get("id"),
                    "invoice": s.get("source_invoice"),
                    "from_date": iso,
                    "to_date": corrected,
                    "status": s.get("status"),
                    "therapist_ids": s.get("therapist_ids") or [],
                })
                if not body.dry_run:
                    await db.sessions.update_one(
                        {"id": s["id"]},
                        {"$set": {
                            "session_date": corrected,
                            "day_name": _day_name_from_date(corrected),
                            "updated_at": now_iso(),
                        }},
                    )
                    applied += 1

    msg = f"Found {len(fixes)} session(s) with likely month/day swap"
    if body.dry_run:
        msg += " (preview — run again with dry_run=false to apply)"
    else:
        msg += f"; applied {applied}"
    if skipped:
        msg += f"; skipped {skipped} conflict(s)"

    return {
        "dry_run": body.dry_run,
        "candidates": len(fixes),
        "applied": applied,
        "skipped_conflicts": skipped,
        "fixes": fixes[:300],
        "message": msg,
    }


@api.post("/admin/mark-all-payments-complete")
async def admin_mark_all_payments_complete(_=Depends(admin_only)):
    """Mark all clients/invoices paid except partial-payment exceptions (Fahad #079 = half paid)."""
    result = await _migrate_mark_all_payments_complete(force=True)
    partial = result.get("partial_clients") or []
    names = ", ".join(f"#{c.get('file_no')} {c.get('name')}" for c in partial) or "none"
    result["message"] = (
        f"Marked {result.get('invoices_updated', 0)} invoice(s) and "
        f"{result.get('clients_updated', 0)} client(s) as paid. "
        f"Partial (half paid): {names}."
    )
    return result


# ------------------- Cancel-Notify (in-app + queued email) -------------------
async def _send_email_stub(to: str, subject: str, body: str) -> dict:
    """Send email via Brevo/Resend (HTTPS) or SMTP. Logs all attempts to email_queue."""
    await _reload_email_settings_from_db()
    provider_pref = os.environ.get("EMAIL_PROVIDER", "auto")

    def pick_provider():
        if provider_pref == "mailgun":
            return "mailgun" if _mailgun_configured() else None
        if provider_pref == "brevo":
            return "brevo" if _brevo_configured() else None
        if provider_pref == "resend":
            return "resend" if _resend_configured() else None
        if provider_pref == "smtp":
            return "smtp" if _smtp_configured() else None
        # auto — HTTPS first (works on Railway); SMTP last (blocked on Railway)
        if _mailgun_configured():
            return "mailgun"
        if _brevo_configured():
            return "brevo"
        if _resend_configured():
            return "resend"
        if _smtp_configured():
            return "smtp"
        return None

    chosen = pick_provider()
    queue_doc = {
        "id": str(uuid.uuid4()),
        "to": to, "subject": subject, "body": body,
        "status": "queued", "provider": chosen or "none",
        "created_at": now_iso(),
    }

    if not chosen:
        queue_doc["status"] = "queued_no_key"
        queue_doc["error"] = "No email provider configured. Add Mailgun in Admin."
        logger.info(f"Email queued (no provider): to={to} subject={subject}")
        await db.email_queue.insert_one(queue_doc)
        queue_doc.pop("_id", None)
        return queue_doc

    try:
        if chosen == "brevo":
            pid = await _send_via_brevo(to, subject, body)
            queue_doc["status"] = "sent"
            queue_doc["provider_id"] = pid
        elif chosen == "mailgun":
            pid = await _send_via_mailgun(to, subject, body)
            queue_doc["status"] = "sent"
            queue_doc["provider_id"] = pid
        elif chosen == "resend":
            pid = await _send_via_resend(to, subject, body)
            queue_doc["status"] = "sent"
            queue_doc["provider_id"] = pid
        else:
            await asyncio.to_thread(_send_via_smtp_sync, to, subject, body)
            queue_doc["status"] = "sent"
    except Exception as e:
        queue_doc["status"] = "failed"
        queue_doc["error"] = str(e)[:500]
        hint = _smtp_error_hint(str(e))
        if hint:
            queue_doc["hint"] = hint
        elif "sandbox" in str(e).lower() and "authorized" in str(e).lower():
            queue_doc["hint"] = "Mailgun sandbox: add the recipient email as an Authorized Recipient in Mailgun dashboard, or verify boostgrowth.org domain."
        elif "domain" in str(e).lower() and "not found" in str(e).lower():
            queue_doc["hint"] = "Check Mailgun Domain matches exactly (e.g. sandbox123.mailgun.org or mg.boostgrowth.org)."
        elif "unrecognised ip" in str(e).lower() or "unauthorized" in str(e).lower() or "authorised_ips" in str(e).lower():
            queue_doc["hint"] = "Brevo blocked the server IP. Open app.brevo.com/security/authorised_ips → authorize the server IP or disable IP blocking for the API."
        logger.warning(f"Email send failed ({chosen}) to {to}: {e}")

    await db.email_queue.insert_one(queue_doc)
    queue_doc.pop("_id", None)
    return queue_doc

@api.post("/schedule/cancel-notify")
async def schedule_cancel_notify(payload: CancelNotifyIn, user=Depends(schedule_edit_or_admin)):
    """Mark cell as cancelled (optional) + send in-app/email notifications to selected therapists."""
    cell = await db.schedule_cells.find_one({"id": payload.cell_id}, {"_id": 0})
    if not cell:
        raise HTTPException(status_code=404, detail="Schedule cell not found")
    actor = _actor_display(user)
    if payload.state:
        await db.schedule_cells.update_one({"id": payload.cell_id}, {"$set": {"state": payload.state}})
        if payload.state == "cancel_therapist":
            await _mark_parent_cancel_pending(payload.cell_id)
            cell = await db.schedule_cells.find_one({"id": payload.cell_id}, {"_id": 0})
            if cell:
                await _notify_parent_cancel_pending(cell, actor)
                await _log_therapist_cancel_prep_history(cell, user.get("id") or "")
    recipients = payload.recipient_ids or ([cell["therapist_id"]] if cell.get("therapist_id") else [])
    title = f"Notice from {actor}"
    if payload.state == "cancel_therapist":
        title = f"Session Cancelled — {actor}"
    elif payload.state == "cancel_child":
        title = f"Session Cancelled (Client) — {actor}"
    sent = []
    for rid in recipients:
        if payload.send_in_app:
            n = await _notify(
                rid, "schedule_cancel", title, payload.message,
                schedule_cell_id=payload.cell_id, requires_ack=True,
                actor_id=user.get("id"), actor_name=actor,
            )
            sent.append({"user_id": rid, "notification_id": n["id"]})
        therapist = await db.therapists.find_one({"id": rid}, {"_id": 0})
        send_mail = payload.send_email or payload.state == "cancel_therapist"
        if send_mail:
            recipient = payload.extra_email if len(recipients) == 1 and payload.extra_email else (therapist.get("email") if therapist else None)
            if recipient:
                client_name = (cell.get("child_name") or "—").strip()
                week_start = cell.get("week_start") or ""
                day_idx = cell.get("day")
                day_label = ""
                if week_start and day_idx is not None:
                    try:
                        d = datetime.fromisoformat(str(week_start)[:10]) + timedelta(days=int(day_idx))
                        day_label = d.strftime("%d %b %Y")
                    except Exception:
                        day_label = str(week_start)
                if payload.state == "cancel_therapist":
                    subj = f"Session Cancelled — {client_name} on {day_label or week_start} ({actor})"
                    body = (
                        f"Dear {therapist.get('name', '')},\n\n"
                        f"{actor} marked the session with {client_name} scheduled on {day_label or week_start} "
                        f"at {cell.get('time_slot') or '—'} as a therapist cancellation.\n\n"
                        f"{payload.message}\n\n— Boost Growth Portal"
                    )
                else:
                    subj = f"[Boost Growth] {title}"
                    body_lines = [
                        f"Hello {therapist.get('name') if therapist else ''},",
                        "",
                        f"{actor} sent a schedule notice:",
                        "",
                        payload.message,
                        "",
                        f"Cell: {cell.get('service_code')} | {client_name}",
                        f"Day: {day_label or cell.get('day')} | Time: {cell.get('time_slot')}",
                        "",
                        "— Boost Growth Portal",
                    ]
                    body = "\n".join(body_lines)
                await _send_email_stub(recipient, subj, body)
    return {"ok": True, "sent": sent}

# ------------------- Intake (admin only) -------------------
@api.get("/intake")
async def list_intake(_=Depends(client_lead_or_admin)):
    return await db.intake.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)


@api.get("/intake/meta")
async def get_intake_meta(_=Depends(client_lead_or_admin)):
    doc = await db.meta.find_one({"key": "intake_list_meta"}, {"_id": 0})
    return {
        "last_updated": (doc or {}).get("last_updated"),
        "updated_by": (doc or {}).get("updated_by"),
    }


@api.put("/intake/meta")
async def update_intake_meta(body: dict, user=Depends(get_current_user)):
    """Walaa, Maha, Fahda, Jenan may set the waiting-list last-updated date."""
    if not (_is_client_lead(user) or _is_jenan(user)):
        raise HTTPException(status_code=403, detail="Not allowed")
    last_updated = (body.get("last_updated") or "").strip() or None
    patch = {
        "key": "intake_list_meta",
        "last_updated": last_updated,
        "updated_by": user.get("name") or user.get("email"),
        "updated_at": now_iso(),
    }
    await db.meta.update_one({"key": "intake_list_meta"}, {"$set": patch}, upsert=True)
    return patch


@api.post("/intake")
async def create_intake(payload: IntakeIn, _=Depends(client_lead_or_admin)):
    iid = str(uuid.uuid4())
    data = payload.model_dump()
    name = (data.get("child_name") or "").strip()
    itype = data.get("intake_type") or "pre"
    cat = data.get("list_category") or ("school" if itype == "school" else "intake")
    doc = {
        "id": iid,
        **data,
        "child_name": name,
        "list_category": cat,
        "name_key": _intake_name_key(name, itype, cat),
        "sync_source": "manual",
        "created_at": now_iso(),
    }
    await db.intake.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/intake/{iid}")
async def update_intake(iid: str, payload: IntakeIn, _=Depends(client_lead_or_admin)):
    data = payload.model_dump()
    name = (data.get("child_name") or "").strip()
    itype = data.get("intake_type") or "pre"
    cat = data.get("list_category") or ("school" if itype == "school" else "intake")
    data["child_name"] = name
    data["list_category"] = cat
    data["name_key"] = _intake_name_key(name, itype, cat)
    if not data.get("sync_source"):
        data["sync_source"] = "manual"
    await db.intake.update_one({"id": iid}, {"$set": data})
    return await db.intake.find_one({"id": iid}, {"_id": 0})

@api.delete("/intake/{iid}")
async def delete_intake(iid: str, _=Depends(client_lead_or_admin)):
    await db.intake.delete_one({"id": iid})
    return {"ok": True}

@api.post("/admin/dedupe-intake")
async def admin_dedupe_intake(_=Depends(admin_only)):
    """Remove duplicate intake rows (same child name + pre/post type)."""
    removed = await _dedupe_intake_records()
    total = await db.intake.count_documents({})
    return {"ok": True, "removed": removed, "total": total, "message": f"Removed {removed} duplicate(s). {total} intake records remain."}


@api.post("/admin/dedupe-clients")
async def admin_dedupe_clients(_=Depends(admin_only)):
    """Soft-delete duplicate clients (same name or file_no); keep the record with data."""
    result = await _dedupe_duplicate_clients()
    total = await db.clients.count_documents(_active_client_filter())
    return {
        "ok": True,
        **result,
        "total_clients": total,
        "message": f"Removed {result['removed']} duplicate client(s). {total} active clients remain.",
    }

@api.post("/admin/seed-intake-master")
async def seed_intake_master(replace: bool = True, _=Depends(admin_only)):
    """Replace intake list from INTAKE_SEED (default) or upsert without deleting when replace=false."""
    if replace:
        await db.intake.delete_many({})
        created = 0
        for item in INTAKE_SEED:
            name = item.get("child_name", "").strip()
            itype = item.get("intake_type", "pre")
            if not name:
                continue
            await db.intake.insert_one({
                "id": str(uuid.uuid4()),
                "status": item.get("status") or "new",
                "priority": bool(item.get("priority")),
                "created_at": now_iso(),
                **item,
                "child_name": name,
                "intake_type": itype,
                "name_key": _intake_name_key(name, itype),
            })
            created += 1
        return {"created": created, "updated": 0, "total_seed": len(INTAKE_SEED), "replaced": True}

    created, updated = 0, 0
    for item in INTAKE_SEED:
        name = item.get("child_name", "").strip()
        itype = item.get("intake_type", "pre")
        if not name:
            continue
        match = await _find_intake_for_upsert(name, itype)
        doc = {**item, "child_name": name, "intake_type": itype, "name_key": _intake_name_key(name, itype)}
        if match:
            await db.intake.update_one({"id": match["id"]}, {"$set": doc})
            updated += 1
        else:
            await db.intake.insert_one({
                "id": str(uuid.uuid4()),
                "status": item.get("status") or "new",
                "priority": bool(item.get("priority")),
                "created_at": now_iso(),
                **doc,
            })
            created += 1
    return {"created": created, "updated": updated, "total_seed": len(INTAKE_SEED), "replaced": False}

# ------------------- Reports -------------------
@api.get("/reports/dashboard")
async def reports_dashboard(_=Depends(manager_reports_access)):
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(5000)
    clients = await db.clients.find(_active_client_filter(), {"_id": 0}).to_list(500)
    therapists = await db.therapists.find({}, {"_id": 0, "pin_hash": 0, "password_hash": 0}).to_list(50)
    requests = await db.requests.find({}, {"_id": 0}).to_list(500)
    cells = await db.schedule_cells.find({}, {"_id": 0}).to_list(5000)

    # Sessions per therapist
    per_t: dict = {}
    for t in therapists:
        per_t[t["id"]] = {"name": therapist_schedule_display_name(t), "color": t.get("color"),
                           "completed": 0, "cancelled": 0, "no_show": 0, "no_service": 0,
                           "hours": 0.0}
    for s in sessions:
        for tid in s.get("therapist_ids") or []:
            if tid in per_t:
                if s["status"] == "Completed":
                    per_t[tid]["completed"] += 1
                    per_t[tid]["hours"] += float(s.get("hours") or 0)
                elif s["status"] == "Cancelled":
                    per_t[tid]["cancelled"] += 1
                elif s["status"] == "No Show":
                    per_t[tid]["no_show"] += 1
                else:
                    per_t[tid]["no_service"] += 1

    # Per-client used hours + status
    per_c = []
    for c in clients:
        used = sum(float(s.get("hours") or 0) for s in sessions if s.get("client_id") == c["id"] and s.get("status") == "Completed")
        pkg = c.get("package_hours") or 24
        rem = max(0, pkg - used)
        if rem <= 0 or rem <= 2 or rem / pkg <= 0.2:
            status = "urgent"
        elif rem / pkg <= 0.35 or rem <= 4:
            status = "warning"
        else:
            status = "ok"
        per_c.append({"id": c["id"], "name": c["name"], "file_no": c.get("file_no"),
                      "color": c.get("color"), "pkg": pkg, "used": round(used, 1),
                      "rem": round(rem, 1), "status": status})

    # Cancellation breakdown from schedule cells (this week)
    sched_cancel_t = sum(1 for c in cells if c.get("state") == "cancel_therapist")
    sched_cancel_c = sum(1 for c in cells if c.get("state") == "cancel_child")

    return {
        "totals": {
            "therapists": len(therapists),
            "clients": len(clients),
            "sessions": len(sessions),
            "completed_sessions": sum(1 for s in sessions if s.get("status") == "Completed"),
            "total_hours": round(sum(float(s.get("hours") or 0) for s in sessions if s.get("status") == "Completed"), 1),
            "open_requests": sum(1 for r in requests if r.get("status") == "pending"),
            "urgent_clients": sum(1 for c in per_c if c["status"] == "urgent"),
            "warning_clients": sum(1 for c in per_c if c["status"] == "warning"),
            "schedule_cells": len(cells),
            "schedule_cancel_therapist": sched_cancel_t,
            "schedule_cancel_child": sched_cancel_c,
        },
        "per_therapist": list(per_t.values()),
        "per_client": sorted(per_c, key=lambda x: {"urgent":0,"warning":1,"ok":2}[x["status"]]),
    }

# ------------------- Imports -------------------
_INTAKE_HEADER_HINTS = frozenset({
    "child", "name", "phone", "service", "district", "diagnosis", "intake",
    "age", "parent", "area", "status", "notes", "language", "priority",
})


def _sanitize_cell(value):
    """Convert pandas/numpy scalars to BSON-safe Python types."""
    if value is None:
        return None
    import math
    if isinstance(value, float) and math.isnan(value):
        return None
    tn = type(value).__name__
    if tn in ("int64", "int32", "int16", "int8", "uint64", "uint32"):
        return int(value)
    if tn in ("float64", "float32"):
        v = float(value)
        return None if math.isnan(v) else v
    if tn in ("bool_", "bool8"):
        return bool(value)
    if hasattr(value, "isoformat"):
        try:
            return value.date().isoformat() if hasattr(value, "date") else value.isoformat()
        except Exception:
            return str(value)
    return value


def _clean_str(value) -> Optional[str]:
    v = _sanitize_cell(value)
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return None
    return s


def _normalize_intake_name(name: str) -> str:
    """Collapse spaces/punctuation so 'Yazeed Bu Sheet' matches 'yazeed bu sheet'."""
    if not name:
        return ""
    s = str(name).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^\w\s]", "", s, flags=re.UNICODE)
    return s.strip()


def _intake_name_key(name: str, intake_type: str, list_category: str = "intake") -> str:
    itype = (intake_type or "pre").lower()
    if itype == "school" or (list_category or "").lower() == "school":
        return f"{_normalize_intake_name(name)}|school|school"
    itype = "post" if "post" in itype else "pre"
    cat = (list_category or "intake").lower()
    return f"{_normalize_intake_name(name)}|{cat}|{itype}"


async def _find_intake_for_upsert(name: str, intake_type: str, list_category: str = "intake") -> Optional[dict]:
    key = _intake_name_key(name, intake_type, list_category)
    match = await db.intake.find_one({"name_key": key}, {"_id": 0, "id": 1, "created_at": 1})
    if match:
        return match
    itype = (intake_type or "pre").lower()
    cat = (list_category or "intake").lower()
    norm = _normalize_intake_name(name)
    if not norm:
        return None
    query = {"intake_type": itype}
    if cat == "school":
        query = {"$or": [{"list_category": "school"}, {"intake_type": "school"}]}
    for row in await db.intake.find(query, {"_id": 0, "id": 1, "child_name": 1, "created_at": 1}).to_list(500):
        if _normalize_intake_name(row.get("child_name", "")) == norm:
            return row
    return None


async def _dedupe_intake_records() -> int:
    """Remove duplicate intake rows; keep oldest record per normalized name + type."""
    rows = await db.intake.find({}, {"_id": 0, "id": 1, "child_name": 1, "intake_type": 1, "name_key": 1, "created_at": 1}).to_list(2000)
    groups: dict = {}
    for row in rows:
        key = row.get("name_key") or _intake_name_key(
            row.get("child_name", ""), row.get("intake_type", "pre"), row.get("list_category", "intake")
        )
        groups.setdefault(key, []).append(row)
    removed = 0
    for key, bucket in groups.items():
        if len(bucket) == 1:
            if not bucket[0].get("name_key"):
                await db.intake.update_one({"id": bucket[0]["id"]}, {"$set": {"name_key": key}})
            continue
        bucket.sort(key=lambda r: r.get("created_at") or "")
        keep_id = bucket[0]["id"]
        for dup in bucket[1:]:
            await db.intake.delete_one({"id": dup["id"]})
            removed += 1
        await db.intake.update_one({"id": keep_id}, {"$set": {"name_key": key}})
    return removed

def _normalize_table_column(name) -> str:
    """Excel headers like 'Child Name' / \"Child's Name\" / 'DOB/Age' → child_name / dob_age."""
    s = str(name).strip().lower()
    if s in ("#", "no.", "no"):
        return "row_num"
    s = s.replace("'", "").replace("'", "").replace("`", "")
    s = _re_top.sub(r"[\s/]+", "_", s)
    s = _re_top.sub(r"[^\w]+", "_", s)
    s = _re_top.sub(r"_+", "_", s).strip("_")
    if s in ("childs_name", "name_of_child", "student", "patient", "candidate"):
        return "child_name"
    if s in ("mobile", "mobile_no", "contact", "contact_number", "tel", "telephone", "phone_no"):
        return "phone"
    if s in ("pre_post", "prepost", "list_type"):
        return "intake_type"
    if s in ("note", "comment", "comments"):
        return "notes"
    if s in ("dis", "dis_", "location", "area", "region", "neighborhood", "location_district"):
        return "district"
    if s in ("school_start_date", "school_start", "start_date"):
        return "school_start_date"
    if s in ("school_name",):
        return "school_name"
    if s in ("time", "timing"):
        return "time_pref"
    if s in ("diagnosis_age", "diagnosis_age", "diag_age"):
        return "diagnosis_age"
    if s in ("birth_date", "date_of_birth", "dob", "child_dob", "birthdate"):
        return "birth_date"
    if s in ("dob_age", "age_dob", "age_year_of_birth"):
        return "dob_age"
    return s


def _detect_table_header_row(df_raw) -> int:
    """Find the row index that looks like a column header (not a title row)."""
    best_idx, best_score = 0, 0
    for idx in range(min(25, len(df_raw))):
        cells = [str(c).strip().lower() for c in df_raw.iloc[idx].tolist() if c is not None and str(c).strip()]
        if len(cells) < 2:
            continue
        score = 0
        for cell in cells:
            for hint in _INTAKE_HEADER_HINTS:
                if hint in cell:
                    score += 1
                    break
        if any("child" in c and "name" in c for c in cells):
            score += 3
        if any(c in ("name", "#", "no") for c in cells):
            score += 4
        if score > best_score:
            best_score, best_idx = score, idx
    return best_idx if best_score >= 2 else 0


def _pick_excel_sheet(xl, for_intake: bool = False):
    names = xl.sheet_names
    if not for_intake or len(names) == 1:
        return names[0]
    for name in names:
        low = name.lower()
        if any(k in low for k in ("pre", "post", "intake", "waiting", "list")):
            return name
    return names[0]


def _rows_from_dataframe(df) -> List[dict]:
    df.columns = [_normalize_table_column(c) for c in df.columns]
    df = df.where(df.notna(), None)
    rows = df.to_dict("records")
    out = []
    for r in rows:
        clean = {k: _sanitize_cell(v) for k, v in r.items()}
        if any(v is not None and str(v).strip().lower() not in ("", "nan", "none") for v in clean.values()):
            out.append(clean)
    return out


def _sheet_intake_type_hint(sheet_name: str) -> Optional[str]:
    low = (sheet_name or "").lower().strip()
    if "school" in low or "ss wait" in low or low.strip() in ("ss", "ss waiting", "school waiting", "school wait"):
        return "school"
    if "post" in low:
        return "post"
    if "pre" in low or "pending" in low:
        return "pre"
    return None


def _sheet_list_category(sheet_name: str, intake_type: str) -> str:
    low = (sheet_name or "").lower().strip()
    if "school" in low or intake_type == "school" or ("ss" in low and "waiting" in low):
        return "school"
    return "intake"


_INTAKE_JUNK_NAME_RE = _re_top.compile(
    r"(total\s*:|last\s+updated|symbols|colors|gold\s+star|gray\s+circle|normal\s+priority|"
    r"top\s+priority|timing\s+note|not\s+interested|waiting\s+list|pending\s+intake|post.intake|"
    r"pre.intake|child\s*name|^name$|^note$|^diagnosis$|^priority$|^service$|^phone$|^#|^pre$|^post$)",
    _re_top.IGNORECASE,
)
_HS_SS_SERVICE_RE = _re_top.compile(r"^(?:HS|SS)(?:\s*/\s*(?:HS|SS))?$", _re_top.IGNORECASE)


def _looks_like_person_name(name: str) -> bool:
    s = (name or "").strip()
    if len(s) < 2 or len(s) > 80:
        return False
    if _INTAKE_JUNK_NAME_RE.search(s):
        return False
    if s.replace(".", "").replace(" ", "").isdigit():
        return False
    if sum(c.isalpha() for c in s) < 2:
        return False
    return True


def _split_diagnosis_age(val) -> tuple:
    v = _clean_str(val)
    if not v:
        return None, None
    if _re_top.match(r"^\d{4}$", v):
        return None, v
    if _re_top.match(r"^\d+(\.\d+)?(\s*(year|yr|years|yo|y/o))?$", v, _re_top.IGNORECASE):
        return None, v
    return v, None


def _extract_service_from_row(r: dict) -> Optional[str]:
    svc = _clean_str(r.get("service") or r.get("service_type"))
    if svc:
        return svc
    for k in sorted(r.keys()):
        if "priority" not in k.lower():
            continue
        v = _clean_str(r.get(k))
        if v and _HS_SS_SERVICE_RE.match(v.strip()):
            return v
    return None


def _extract_priority_flag(r: dict) -> bool:
    pri_keys = sorted(k for k in r if "priority" in k.lower())
    for i, k in enumerate(pri_keys):
        v = r.get(k)
        if i == 0:
            sv = _clean_str(v)
            if sv and _HS_SS_SERVICE_RE.match(sv.strip()):
                continue
        if v is True:
            return True
        s = str(v or "")
        if "⭐" in s or "★" in s or s.strip().lower() in ("1", "true", "yes", "top", "star"):
            return True
    return False


def _extract_notes_and_language(r: dict) -> tuple:
    language = _clean_str(r.get("language"))
    note_cols = sorted(k for k in r if k == "notes" or k.startswith("notes_"))
    notes = None
    if len(note_cols) >= 2:
        notes = _clean_str(r.get(note_cols[0]))
        if not language:
            language = _clean_str(r.get(note_cols[1]))
    elif len(note_cols) == 1:
        v = _clean_str(r.get(note_cols[0]))
        if v and ("english" in v.lower() or "arabic" in v.lower()):
            language = v
        else:
            notes = v
    return notes, language


def _is_intake_data_row(r: dict) -> bool:
    name = _extract_intake_child_name(r)
    if not _looks_like_person_name(name):
        return False
    for v in r.values():
        if v is not None and "total:" in str(v).lower():
            return False
    row_num = r.get("row_num")
    if row_num is not None:
        try:
            n = int(float(str(row_num)))
            if n < 1 or n > 500:
                return False
        except (TypeError, ValueError):
            pass
    return True


def _is_school_intake_service(service: Optional[str]) -> bool:
    """SS / school-support children belong on the School Waiting list, not pre-intake."""
    s = _re_top.sub(r"\s+", "", (service or "").upper())
    if not s:
        return False
    if s in ("SS", "SCHOOL", "SCHOOLSUPPORT", "SCHOOLSERVICE"):
        return True
    if "SCHOOL" in s and "HS" not in s:
        return True
    return False


def _parse_intake_record(r: dict, sheet_name: str = "") -> Optional[dict]:
    """Map a Waiting List row to a DB-ready intake document."""
    name = _extract_intake_child_name(r)
    if not _looks_like_person_name(name):
        return None

    sheet_hint = _sheet_intake_type_hint(sheet_name)
    raw_type = _clean_str(r.get("intake_type") or r.get("type") or r.get("_sheet_intake_type")) or sheet_hint or "pre"
    intake_type = raw_type.lower()
    sheet_low = (sheet_name or "").lower().strip()

    # Queue is determined by SHEET TAB only — never move pre/post rows to school because of SS service.
    if sheet_hint == "school" or ("ss" in sheet_low and "waiting" in sheet_low):
        intake_type = "school"
    elif sheet_hint == "post" or ("post" in sheet_low and "intake" in sheet_low):
        intake_type = "post"
    elif sheet_hint == "pre" or "pending" in sheet_low:
        intake_type = "pre"
    elif intake_type == "school" and sheet_hint not in ("school",):
        intake_type = "pre"

    list_category = _sheet_list_category(sheet_name, intake_type)

    service = _extract_service_from_row(r)
    school_name = _clean_str(r.get("school_name"))
    school_start = _clean_str(r.get("school_start_date"))
    if list_category == "school" or intake_type == "school":
        if school_name and not service:
            service = service or "SS"
    elif school_name or school_start:
        # School metadata on a pre/post row stays as notes context only.
        pass

    diagnosis, age_from_da = _split_diagnosis_age(r.get("diagnosis_age"))
    notes, language = _extract_notes_and_language(r)
    birth_raw = _clean_str(r.get("birth_date") or r.get("date_of_birth") or r.get("dob"))
    birth_iso = _normalize_date(birth_raw) if birth_raw else None
    if not birth_iso:
        dob_age = _clean_str(r.get("dob_age"))
        if dob_age:
            birth_iso = _normalize_date(dob_age) or (
                f"{dob_age}-01-01" if _re_top.match(r"^\d{4}$", dob_age) else None
            )
    age_val = _clean_str(r.get("age")) or age_from_da
    if not age_val:
        da = _clean_str(r.get("dob_age"))
        if da and da != birth_raw and (not birth_iso or da not in (birth_iso, birth_iso[:4])):
            age_val = da

    doc = {
        "child_name": name,
        "intake_type": intake_type,
        "list_category": list_category,
        "parent_name": _clean_str(r.get("parent_name") or r.get("parent") or r.get("guardian")),
        "phone": _clean_str(r.get("phone") or r.get("parent_phone") or r.get("mobile")),
        "status": (_clean_str(r.get("status")) or "new").lower(),
        "notes": notes,
        "intake_date": _clean_str(r.get("intake_date") or r.get("date")),
        "birth_date": birth_iso,
        "age": age_val,
        "service": service,
        "district": _clean_str(
            r.get("district") or r.get("dis") or r.get("location") or r.get("area") or r.get("school_name")
        ),
        "diagnosis": diagnosis or _clean_str(r.get("diagnosis")),
        "school_start_date": _clean_str(r.get("school_start_date")),
        "time_pref": _clean_str(r.get("time_pref") or r.get("time") or r.get("time_preference")),
        "language": language,
        "priority": _extract_priority_flag(r),
    }
    return doc


def _dedupe_column_names(raw_headers) -> List[str]:
    """Excel duplicate headers (two 'Priority', two 'Note') → priority, priority_1, notes, notes_1."""
    seen: dict = {}
    out: List[str] = []
    for h in raw_headers:
        norm = _normalize_table_column(h)
        if norm in seen:
            seen[norm] += 1
            out.append(f"{norm}_{seen[norm]}")
        else:
            seen[norm] = 0
            out.append(norm)
    return out


def _intake_sheets_to_process(sheet_names: List[str]) -> List[str]:
    """When duplicate tab names exist (with/without trailing space), keep the best variant."""
    groups: dict = {}
    for s in sheet_names:
        key = re.sub(r"\s+", " ", (s or "").strip().lower())
        groups.setdefault(key, []).append(s)
    out: List[str] = []
    skip_keys = {"readme", "template", "settings", "dashboard"}
    for key, variants in groups.items():
        if key in skip_keys:
            continue
        if len(sheet_names) > 6 and not any(
            k in key for k in ("intake", "waiting", "pending", "post", "pre", "list", "school")
        ):
            continue
        out.append(max(variants, key=lambda v: (len(v), v)))
    return out


def _read_intake_rows(content: bytes, filename: str) -> tuple:
    """Parse each sheet once, filter junk rows, dedupe by name + intake type."""
    import pandas as pd
    import io

    parsed: List[dict] = []
    meta_parts: List[str] = []
    all_cols: List[str] = []

    def ingest_sheet(df_raw, sheet_name: str):
        hdr = _detect_table_header_row(df_raw)
        df = df_raw.iloc[hdr + 1:].copy()
        df.columns = _dedupe_column_names(df_raw.iloc[hdr].tolist())
        df = df.where(df.notna(), None)
        rows = []
        for _, series in df.iterrows():
            clean = {str(k): _sanitize_cell(v) for k, v in series.items()}
            if any(v is not None and str(v).strip().lower() not in ("", "nan", "none") for v in clean.values()):
                rows.append(clean)
        hint = _sheet_intake_type_hint(sheet_name)
        kept = 0
        for r in rows:
            if hint:
                r["_sheet_intake_type"] = hint
            if not _is_intake_data_row(r):
                continue
            doc = _parse_intake_record(r, sheet_name)
            if doc:
                parsed.append(doc)
                kept += 1
        if rows:
            all_cols.extend(list(rows[0].keys()))
        meta_parts.append(f"{sheet_name or 'csv'}:{kept}")

    if filename.lower().endswith(".csv"):
        df_raw = pd.read_csv(io.BytesIO(content), header=None)
        ingest_sheet(df_raw, "")
    else:
        xl = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
        for sheet in _intake_sheets_to_process(xl.sheet_names):
            try:
                df_raw = pd.read_excel(io.BytesIO(content), sheet_name=sheet, header=None, engine="openpyxl")
                if df_raw.empty:
                    continue
                ingest_sheet(df_raw, sheet)
            except Exception:
                logger.exception(f"Intake sheet skipped: {sheet}")
                continue

    # Dedupe: same child + intake type — later sheets win (newer tabs overwrite older)
    by_key: dict = {}
    for doc in parsed:
        cat = doc.get("list_category") or "intake"
        itype = doc.get("intake_type") or "pre"
        key = (_normalize_intake_name(doc["child_name"]), cat, itype)
        doc["name_key"] = _intake_name_key(doc["child_name"], itype, cat)
        by_key[key] = doc
    unique = list(by_key.values())

    meta = f"{len(unique)} records from {len(meta_parts)} sheet(s) — " + ", ".join(meta_parts)
    cols = list(dict.fromkeys(all_cols))[:20]
    return unique, cols, meta


def _read_table(file: UploadFile, for_intake: bool = False) -> List[dict]:
    """Read xlsx/csv into list of dicts with normalized lower-case keys."""
    import pandas as pd
    content = file.file.read()
    import io
    if for_intake:
        rows, _, _ = _read_intake_rows(content, file.filename or "")
        return rows
    if file.filename.lower().endswith(".csv"):
        df_raw = pd.read_csv(io.BytesIO(content), header=None)
        hdr = _detect_table_header_row(df_raw)
        df = pd.read_csv(io.BytesIO(content), header=hdr)
    else:
        xl = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
        sheet = _pick_excel_sheet(xl, for_intake=False)
        df_raw = pd.read_excel(io.BytesIO(content), sheet_name=sheet, header=None, engine="openpyxl")
        hdr = _detect_table_header_row(df_raw)
        df = pd.read_excel(io.BytesIO(content), sheet_name=sheet, header=hdr, engine="openpyxl")
    return _rows_from_dataframe(df)


def _extract_intake_child_name(r: dict) -> str:
    """Resolve child name from many possible Excel column layouts."""
    preferred_keys = (
        "child_name", "childs_name", "name", "child", "student_name",
        "client_name", "patient_name", "full_name", "candidate", "student",
    )
    header_like = {"child name", "child_name", "name", "child", "student name", "parent", "parent name", "phone", "service", "nan", "none"}

    for k in preferred_keys:
        v = r.get(k)
        if v is None:
            continue
        s = str(v).strip()
        if s and s.lower() not in header_like:
            return s

    for k, v in r.items():
        if v is None:
            continue
        kl = str(k).lower()
        if any(x in kl for x in ("parent", "guardian", "mother", "father", "note", "status", "phone", "service", "district", "diagnosis", "intake", "age", "date", "priority", "language", "time")):
            continue
        if "name" in kl or kl in ("child", "student", "patient"):
            s = str(v).strip()
            if s and s.lower() not in header_like and len(s) >= 2:
                return s

    # Fallback: first text column (including unnamed Excel columns)
    ordered_keys = sorted(r.keys(), key=lambda k: (0 if str(k).lower().startswith("unnamed") else 1, str(k)))
    for k in ordered_keys:
        v = r.get(k)
        if v is None:
            continue
        kl = str(k).lower()
        if kl in ("no", "num", "number", "#", "id", "sn", "sno", "index", "priority", "_sheet_intake_type"):
            continue
        if kl.startswith("_"):
            continue
        if any(x in kl for x in ("phone", "service", "district", "diagnosis", "status", "note", "age", "date", "language", "time")):
            continue
        s = str(v).strip()
        if len(s) >= 2 and s.lower() not in header_like and not s.replace(".", "", 1).isdigit():
            return s
    return ""

def _parse_active_clients_sheet(ws) -> List[dict]:
    """Parse Boost 'Active Clients' tab: file # (B), name (C), service (H), supervisor (J)."""
    by_file: dict = {}
    out: List[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        cells = list(row) + [None] * 12
        file_num, name, service, supervisor = cells[1], cells[2], cells[7], cells[9]
        if file_num and name:
            file_key = str(file_num).strip()
            if not file_key.replace("0", "").isdigit():
                continue
            if file_key in by_file:
                current = by_file[file_key]
            else:
                current = {
                    "file_no": file_key.zfill(3),
                    "name": str(name).strip(),
                    "services": [],
                    "supervisor": _normalize_supervisor_value(supervisor),
                }
                by_file[file_key] = current
                out.append(current)
            if service:
                svc = str(service).strip()
                if svc and svc not in current["services"]:
                    current["services"].append(svc)
            sup = _normalize_supervisor_value(supervisor)
            if sup:
                current["supervisor"] = sup
        elif by_file and service:
            last = out[-1] if out else None
            if last:
                svc = str(service).strip()
                if svc and svc not in last["services"]:
                    last["services"].append(svc)
    rows: List[dict] = []
    for c in out:
        svc_joined = " / ".join(c["services"]) if c["services"] else None
        rows.append({
            "file_no": c["file_no"],
            "name": c["name"],
            "service_type": svc_joined,
            "service": svc_joined,
            "supervisor": c.get("supervisor"),
        })
    return rows


def _read_clients_import_rows(content: bytes, filename: str) -> List[dict]:
    """Read client rows from CSV/Excel, including Boost Active Clients workbook layout."""
    import io
    import pandas as pd

    low = (filename or "").lower()
    if low.endswith((".xlsx", ".xls")):
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        active_ws = None
        for sn in wb.sheetnames:
            if sn.strip().lower().startswith("active client"):
                active_ws = wb[sn]
                break
        if active_ws is not None:
            parsed = _parse_active_clients_sheet(active_ws)
            wb.close()
            if parsed:
                return parsed
        wb.close()
        xl = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
        sheet = _pick_excel_sheet(xl, for_intake=False)
        df_raw = pd.read_excel(io.BytesIO(content), sheet_name=sheet, header=None, engine="openpyxl")
        hdr = _detect_table_header_row(df_raw)
        df = pd.read_excel(io.BytesIO(content), sheet_name=sheet, header=hdr, engine="openpyxl")
        return _rows_from_dataframe(df)
    if low.endswith(".csv"):
        df_raw = pd.read_csv(io.BytesIO(content), header=None)
        hdr = _detect_table_header_row(df_raw)
        df = pd.read_csv(io.BytesIO(content), header=hdr)
        return _rows_from_dataframe(df)
    raise HTTPException(400, "Upload .xlsx, .xls, or .csv")


async def _import_clients_from_rows(rows: List[dict], replace_missing: bool = False) -> dict:
    created, updated, skipped = 0, 0, 0
    file_nos_in_file: set = set()
    therapists = await db.therapists.find({}, {"_id": 0}).to_list(100)
    t_by_name = {t["name"].lower(): t["id"] for t in therapists}
    for r in rows:
        name = r.get("name") or r.get("child_name") or r.get("client_name") or r.get("full_name")
        if not name:
            skipped += 1
            continue
        file_no_raw = str(r.get("file_no") or r.get("file_number") or r.get("id") or r.get("file") or "").strip()
        if not file_no_raw or not file_no_raw.replace("0", "").isdigit():
            skipped += 1
            continue
        file_no = file_no_raw.zfill(3)
        file_nos_in_file.add(file_no)
        main_name = (r.get("main_therapist") or r.get("main") or "").strip().lower() if r.get("main_therapist") or r.get("main") else None
        main_id = t_by_name.get(main_name) if main_name else None
        svc_raw = r.get("service") or r.get("service_type")
        doc = {
            "name": str(name).strip(),
            "file_no": file_no,
            "package_hours": float(r.get("package_hours") or r.get("pkg") or 24),
            "supervisor": _normalize_supervisor_value(r.get("supervisor")),
            "service_type": str(svc_raw).strip() if svc_raw else None,
            "main_therapist_id": main_id,
            "co_therapist_ids": [],
            "color": r.get("color") or "#A2C4C9",
            "locations": [],
            "parent_name": r.get("parent_name") or r.get("parent"),
            "parent_phone": str(r.get("parent_phone") or r.get("phone") or "") or None,
            "age": str(r.get("age") or "") or None,
            "notes": r.get("notes"),
            "billing_mode": "hours",
        }
        match = await db.clients.find_one({"file_no": file_no}, {"_id": 0, "id": 1, "deleted": 1})
        if match and match.get("deleted"):
            skipped += 1
            continue
        if match:
            await db.clients.update_one({"id": match["id"]}, {"$set": doc})
            updated += 1
        else:
            await db.clients.insert_one({
                "id": str(uuid.uuid4()),
                "payment_status": "pending",
                "created_at": now_iso(),
                **doc,
            })
            created += 1
    removed_missing = 0
    if replace_missing and file_nos_in_file:
        r = await db.clients.update_many(
            _active_client_filter({"file_no": {"$nin": list(sorted(file_nos_in_file))}}),
            {"$set": {"deleted": True, "deleted_at": now_iso(), "dedupe_note": "missing from clients import file"}},
        )
        removed_missing = int(r.modified_count or 0)
    dedupe = await _dedupe_duplicate_clients()
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "removed_missing": removed_missing,
        "dedupe_removed": dedupe.get("removed") if isinstance(dedupe, dict) else None,
    }


@api.post("/import/clients")
async def import_clients(
    file: UploadFile = File(...),
    replace_missing: bool = Form(False),
    _=Depends(import_access),
):
    content = await file.read()
    rows = _read_clients_import_rows(content, file.filename or "")
    return await _import_clients_from_rows(rows, replace_missing)


@api.post("/admin/import-clients-and-sync")
async def admin_import_clients_and_sync(
    file: UploadFile = File(...),
    replace_missing: bool = Form(False),
    _=Depends(admin_only),
):
    """Upload Active Clients Excel/CSV, then auto-recover + relink prep badges."""
    content = await file.read()
    rows = _read_clients_import_rows(content, file.filename or "")
    import_result = await _import_clients_from_rows(rows, replace_missing)
    recover_result = await _run_auto_recover(store_backup=True)
    imp = import_result
    parts = [
        f"استيراد: {imp.get('created', 0)} جديد · {imp.get('updated', 0)} محدّث · {imp.get('skipped', 0)} تخطّى",
    ]
    if imp.get("removed_missing"):
        parts.append(f"{imp['removed_missing']} محذوف (غير موجود في الملف)")
    rec = recover_result.get("summary_ar") or ""
    if rec:
        parts.append(rec)
    parts.append(
        "ملاحظة: الفواتير والجلسات لا تُستورد من هذا الملف — استخدم Sync من Drive أو رفع Excel لكل طفل."
    )
    return {
        "ok": True,
        "import": import_result,
        "recover": recover_result,
        "health_after": recover_result.get("health_after"),
        "summary_ar": " · ".join(parts),
    }


WAITING_LIST_SHEET_ID = "14DLxQZOWSRnS_4kWeOsgKfpYMQiZ6hQiv2be_-J_hBg"
WAITING_LIST_SHEET_URL = f"https://docs.google.com/spreadsheets/d/{WAITING_LIST_SHEET_ID}/edit"


async def _purge_stale_google_intake(synced_keys: set) -> int:
    """Remove google-synced rows not in this import, per queue (pre / post / school)."""
    removed = 0
    queue_filters = [
        ("pre", {"intake_type": "pre", "list_category": {"$ne": "school"}}),
        ("post", {"intake_type": "post", "list_category": {"$ne": "school"}}),
        ("school", {"$or": [{"intake_type": "school"}, {"list_category": "school"}]}),
    ]
    for label, base_q in queue_filters:
        if label == "school":
            keys = [k for k in synced_keys if k.endswith("|school|school")]
        else:
            keys = [k for k in synced_keys if k.endswith(f"|intake|{label}")]
        if not keys:
            continue
        q = {**base_q, "name_key": {"$nin": keys}, "sync_source": {"$ne": "manual"}}
        r = await db.intake.delete_many(q)
        removed += r.deleted_count
    return removed


async def _reclassify_school_intake_records() -> int:
    """Deprecated — school queue is sheet-driven only. Kept for manual admin repair."""
    return 0


async def _upsert_intake_rows(rows: List[dict], detected_columns: List[str], parse_meta: str, replace_google_stale: bool = False) -> dict:
    created, updated, skipped = 0, 0, 0
    pre_count = sum(1 for d in rows if d.get("intake_type") == "pre")
    post_count = sum(1 for d in rows if d.get("intake_type") == "post")
    school_count = sum(1 for d in rows if d.get("intake_type") == "school" or d.get("list_category") == "school")
    synced_keys: set = set()
    for doc in rows:
        name = (doc.get("child_name") or "").strip()
        if not name:
            skipped += 1
            continue
        intake_type = doc.get("intake_type") or "pre"
        list_category = doc.get("list_category") or ("school" if intake_type == "school" else "intake")
        name_key = _intake_name_key(name, intake_type, list_category)
        synced_keys.add(name_key)
        match = await _find_intake_for_upsert(name, intake_type, list_category)
        db_doc = {k: v for k, v in doc.items() if v is not None}
        db_doc["child_name"] = name
        db_doc["list_category"] = list_category
        db_doc["name_key"] = name_key
        if replace_google_stale:
            db_doc["sync_source"] = "google_sheet"
        if match:
            await db.intake.update_one({"id": match["id"]}, {"$set": db_doc})
            await db.intake.delete_many({"name_key": name_key, "id": {"$ne": match["id"]}})
            updated += 1
        else:
            try:
                await db.intake.insert_one({
                    "id": str(uuid.uuid4()),
                    "created_at": now_iso(),
                    **db_doc,
                })
                created += 1
            except Exception as e:
                logger.warning(f"Intake insert skipped for {name}: {e}")
                skipped += 1
    if replace_google_stale and synced_keys:
        stale_removed = await _purge_stale_google_intake(synced_keys)
        stale = type("R", (), {"deleted_count": stale_removed})()
    else:
        stale = type("R", (), {"deleted_count": 0})()
    deduped = await _dedupe_intake_records()
    total_in_db = await db.intake.count_documents({})
    hint = None
    if not rows:
        hint = f"No intake rows found ({parse_meta}). Columns seen: {', '.join(detected_columns[:12])}"
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "removed_stale": stale.deleted_count,
        "deduped": deduped,
        "total_in_db": total_in_db,
        "rows_in_file": len(rows),
        "pre_count": pre_count,
        "post_count": post_count,
        "school_count": school_count,
        "detected_columns": detected_columns,
        "parse_meta": parse_meta,
        "message": f"{updated} updated, {created} added, {skipped} skipped, {deduped} duplicates removed · {total_in_db} total in list",
        "hint": hint,
    }


async def _upsert_school_waiting_rows(rows: List[dict], detected_columns: List[str], parse_meta: str, replace_google_stale: bool = False) -> dict:
    """Upsert school waiting list only — does not touch pre/post intake queues."""
    school_rows = [
        r for r in rows
        if r.get("intake_type") == "school" or r.get("list_category") == "school"
    ]
    created, updated, skipped = 0, 0, 0
    synced_keys: set = set()
    for doc in school_rows:
        name = (doc.get("child_name") or "").strip()
        if not name:
            skipped += 1
            continue
        intake_type = "school"
        list_category = "school"
        name_key = _intake_name_key(name, intake_type, list_category)
        synced_keys.add(name_key)
        match = await _find_intake_for_upsert(name, intake_type, list_category)
        db_doc = {k: v for k, v in doc.items() if v is not None}
        db_doc["child_name"] = name
        db_doc["intake_type"] = intake_type
        db_doc["list_category"] = list_category
        db_doc["name_key"] = name_key
        if replace_google_stale:
            db_doc["sync_source"] = "google_sheet"
        if match:
            await db.intake.update_one({"id": match["id"]}, {"$set": db_doc})
            await db.intake.delete_many({"name_key": name_key, "id": {"$ne": match["id"]}})
            updated += 1
        else:
            try:
                await db.intake.insert_one({
                    "id": str(uuid.uuid4()),
                    "created_at": now_iso(),
                    **db_doc,
                })
                created += 1
            except Exception as e:
                logger.warning(f"School intake insert skipped for {name}: {e}")
                skipped += 1
    stale_removed = 0
    if replace_google_stale and synced_keys:
        q = {
            "$or": [{"intake_type": "school"}, {"list_category": "school"}],
            "name_key": {"$nin": list(synced_keys)},
            "sync_source": {"$ne": "manual"},
        }
        r = await db.intake.delete_many(q)
        stale_removed = r.deleted_count
    deduped = await _dedupe_intake_records()
    school_count = await db.intake.count_documents({"$or": [{"intake_type": "school"}, {"list_category": "school"}]})
    hint = None
    if not school_rows:
        hint = f"No school waiting rows found ({parse_meta}). Columns seen: {', '.join(detected_columns[:12])}"
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "removed_stale": stale_removed,
        "deduped": deduped,
        "school_count": school_count,
        "rows_in_file": len(school_rows),
        "detected_columns": detected_columns,
        "parse_meta": parse_meta,
        "message": f"School waiting: {updated} updated, {created} added, {skipped} skipped · {school_count} in queue",
        "hint": hint,
    }


@api.post("/admin/fix-school-intake")
async def admin_fix_school_intake(_=Depends(import_access)):
    """Re-sync all waiting lists from Google Sheet tabs (pre / post / school)."""
    import httpx
    export_url = _google_sheet_export_url(WAITING_LIST_SHEET_URL)
    async with httpx.AsyncClient(follow_redirects=True, timeout=120.0) as client:
        resp = await client.get(export_url)
    if resp.status_code != 200:
        raise HTTPException(status_code=400, detail=f"Could not download sheet (HTTP {resp.status_code})")
    rows, detected_columns, parse_meta = _read_intake_rows(resp.content, "waiting_list.xlsx")
    result = await _upsert_intake_rows(rows, detected_columns, parse_meta, replace_google_stale=True)
    result["ok"] = True
    result["message"] = (
        f"Restored: {result.get('pre_count', 0)} pre · "
        f"{result.get('post_count', 0)} post · "
        f"{result.get('school_count', 0)} school"
    )
    return result


@api.post("/import/intake")
async def import_intake(file: UploadFile = File(...), _=Depends(import_access)):
    try:
        content = file.file.read()
        rows, detected_columns, parse_meta = _read_intake_rows(content, file.filename or "")
    except Exception as e:
        logger.exception("Intake file parse failed")
        raise HTTPException(status_code=400, detail=f"Could not read intake file: {e}")
    return await _upsert_intake_rows(rows, detected_columns, parse_meta)


@api.post("/import/intake-google")
async def import_intake_google(body: dict = None, _=Depends(import_access)):
    """Sync waiting list from the official Google Sheet (public export xlsx)."""
    import httpx
    body = body or {}
    sheet_url = (body.get("url") or body.get("sheet_url") or WAITING_LIST_SHEET_URL).strip()
    export_url = _google_sheet_export_url(sheet_url)
    async with httpx.AsyncClient(follow_redirects=True, timeout=120.0) as client:
        resp = await client.get(export_url)
    if resp.status_code != 200:
        raise HTTPException(
            status_code=400,
            detail=f"Could not download waiting list sheet (HTTP {resp.status_code}). Share link must be viewable.",
        )
    try:
        rows, detected_columns, parse_meta = _read_intake_rows(resp.content, "waiting_list.xlsx")
    except Exception as e:
        logger.exception("Intake Google Sheet parse failed")
        raise HTTPException(status_code=400, detail=f"Could not parse waiting list sheet: {e}")
    result = await _upsert_intake_rows(rows, detected_columns, parse_meta, replace_google_stale=True)
    result["sheet_id"] = WAITING_LIST_SHEET_ID
    result["sheet_url"] = WAITING_LIST_SHEET_URL
    return result


@api.post("/import/school-waiting-google")
async def import_school_waiting_google(body: dict = None, _=Depends(import_access)):
    """Sync school waiting list only from the SS waiting sheet tab — does not alter pre/post intake."""
    import httpx
    body = body or {}
    sheet_url = (body.get("url") or body.get("sheet_url") or WAITING_LIST_SHEET_URL).strip()
    export_url = _google_sheet_export_url(sheet_url)
    async with httpx.AsyncClient(follow_redirects=True, timeout=120.0) as client:
        resp = await client.get(export_url)
    if resp.status_code != 200:
        raise HTTPException(
            status_code=400,
            detail=f"Could not download school waiting sheet (HTTP {resp.status_code}). Share link must be viewable.",
        )
    try:
        rows, detected_columns, parse_meta = _read_intake_rows(resp.content, "school_waiting_list.xlsx")
    except Exception as e:
        logger.exception("School waiting Google Sheet parse failed")
        raise HTTPException(status_code=400, detail=f"Could not parse school waiting sheet: {e}")
    result = await _upsert_school_waiting_rows(rows, detected_columns, parse_meta, replace_google_stale=True)
    result["sheet_url"] = sheet_url
    return result


async def _ensure_school_waiting_records() -> dict:
    """Upsert the 6 official SS waiting-list clients — never touches pre/post intake."""
    created, updated = 0, 0
    for item in SCHOOL_WAITING_SEED:
        name = item.get("child_name", "").strip()
        if not name:
            continue
        match = await _find_intake_for_upsert(name, "school", "school")
        doc = {
            **item,
            "child_name": name,
            "intake_type": "school",
            "list_category": "school",
            "name_key": _intake_name_key(name, "school", "school"),
            "sync_source": "seed",
        }
        if match:
            await db.intake.update_one({"id": match["id"]}, {"$set": doc})
            updated += 1
        else:
            await db.intake.insert_one({
                "id": str(uuid.uuid4()),
                "status": item.get("status") or "new",
                "priority": bool(item.get("priority")),
                "created_at": now_iso(),
                **doc,
            })
            created += 1
    school_count = await db.intake.count_documents({"$or": [{"intake_type": "school"}, {"list_category": "school"}]})
    return {"created": created, "updated": updated, "school_count": school_count}


@api.post("/admin/seed-school-waiting")
async def seed_school_waiting(_=Depends(import_access)):
    """Seed / refresh the 6 SS waiting-list clients without altering pre/post intake."""
    result = await _ensure_school_waiting_records()
    result["ok"] = True
    result["message"] = f"School waiting: {result['created']} added, {result['updated']} updated · {result['school_count']} total"
    return result

# ------------------- Historical Schedule Loader -------------------
HISTORICAL_SCHEDULES = None  # lazy-loaded from JSON file

def _load_historical():
    global HISTORICAL_SCHEDULES
    if HISTORICAL_SCHEDULES is None:
        import json
        path = ROOT_DIR / "historical_schedules.json"
        if path.exists():
            HISTORICAL_SCHEDULES = json.loads(path.read_text())
        else:
            HISTORICAL_SCHEDULES = {}
    return HISTORICAL_SCHEDULES

@api.get("/import/historical-weeks")
async def list_historical_weeks(_=Depends(import_access)):
    data = _load_historical()
    return {"weeks": list(data.keys())}

@api.post("/import/historical-load")
async def import_historical(body: dict, _=Depends(import_access)):
    """Import all historical weeks into schedule_cells. body: {clear_existing?: bool}"""
    data = _load_historical()
    if not data:
        raise HTTPException(status_code=404, detail="No historical data file found")
    if body.get("clear_existing"):
        await db.schedule_cells.delete_many({})
    therapists = await db.therapists.find({}, {"_id": 0}).to_list(100)
    t_by_name = {t["name"]: t["id"] for t in therapists}
    DAYS_MAP = {"Sunday":0, "Monday":1, "Tuesday":2, "Wednesday":3, "Thursday":4}
    TIMES = ["8:00 AM - 9:00 AM","9:00 AM - 10:00 AM","10:00 AM - 11:00 AM",
             "11:00 AM - 12:00 PM","12:00 PM - 1:00 PM","1:00 PM - 2:00 PM",
             "2:00 PM - 3:00 PM","3:00 PM - 4:00 PM","4:00 PM - 5:00 PM",
             "5:00 PM - 6:00 PM"]
    inserted = 0
    weeks_loaded = 0
    for week_label, therapists_data in data.items():
        # parse week label like "26 Apr- 30 Apr" → use a fake ISO date for storage
        week_start_iso = f"hist:{week_label}"
        weeks_loaded += 1
        for entry in therapists_data:
            tname = entry.get("n")
            t_id = t_by_name.get(tname)
            if not t_id:
                continue
            for day_label, slots in entry.get("s", []):
                day_idx = DAYS_MAP.get(day_label)
                if day_idx is None:
                    continue
                for slot_idx, raw in enumerate(slots):
                    if not raw or not str(raw).strip():
                        continue
                    txt = str(raw).strip()
                    # parse service code
                    service = "SS"
                    child = None
                    note = None
                    custom = None
                    upper = txt.upper()
                    if upper.startswith("HS"): service = "HS"
                    elif upper.startswith("SS"): service = "SS"
                    elif upper.startswith("OS"): service = "OS"
                    elif "AVC" in upper: service = "AVC"
                    elif "SUPERVISION" in upper: service = "SUPERVISION"
                    elif "OBSERVATION" in upper: service = "OBSERVATION"
                    elif "MEETING" in upper: service = "MEETING"
                    elif "LEAVE" in upper: service = "LEAVE"
                    elif "BREAK" in upper: service = "BREAK"
                    # extract child name after | or W/
                    if "|" in txt:
                        child = txt.split("|", 1)[1].strip()
                    elif "W/" in txt:
                        child = txt.split("W/", 1)[1].strip()
                    elif "with" in txt.lower():
                        child = txt.lower().split("with", 1)[1].strip()
                    if child and "(" in child:
                        custom = child[child.find("(")+1:child.find(")")]
                        child = child[:child.find("(")].strip()
                    if slot_idx >= len(TIMES):
                        continue
                    if service in ("LEAVE", "BREAK", "AVC", "SUPERVISION", "OBSERVATION", "MEETING"):
                        note = txt
                    await db.schedule_cells.insert_one({
                        "id": str(uuid.uuid4()),
                        "therapist_id": t_id, "day": day_idx,
                        "time_slot": TIMES[slot_idx],
                        "service_code": service, "child_name": child,
                        "note": note, "custom_time": custom,
                        "state": "normal", "color": None, "duration": 1,
                        "week_start": week_start_iso, "created_at": now_iso(),
                    })
                    inserted += 1
    return {"weeks_loaded": weeks_loaded, "cells_inserted": inserted}

@api.post("/schedule/duplicate-week")
async def duplicate_week(body: dict, _=Depends(ops_or_admin)):
    """Copy all cells from source_week to target_week. body: {source_week, target_week, clear_target?}"""
    source = body.get("source_week"); target = body.get("target_week")
    if not source or not target:
        raise HTTPException(status_code=400, detail="source_week and target_week required")
    if body.get("clear_target"):
        await db.schedule_cells.delete_many({"week_start": target})
    cells = await db.schedule_cells.find({"week_start": source}, {"_id": 0}).to_list(5000)
    inserted = 0
    for c in cells:
        new_c = {**c, "id": str(uuid.uuid4()), "week_start": target,
                 "state": "normal", "color": None, "created_at": now_iso()}
        new_c = _strip_session_cell_color(new_c)
        await db.schedule_cells.insert_one(new_c)
        inserted += 1
    return {"copied": inserted}

@api.post("/import/list-sheets")
async def list_excel_sheets(file: UploadFile = File(...), _=Depends(import_access)):
    """Return the list of sheet names in an uploaded .xlsx file (helps user pick the right one)."""
    import openpyxl, io
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    return {"sheets": wb.sheetnames}


SCHEDULE_TIME_SLOTS = [
    "8:00 AM - 9:00 AM", "9:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
    "11:00 AM - 12:00 PM", "12:00 PM - 1:00 PM", "1:00 PM - 2:00 PM",
    "2:00 PM - 3:00 PM", "3:00 PM - 4:00 PM", "4:00 PM - 5:00 PM",
    "5:00 PM - 6:00 PM",
]
SCHEDULE_DAYS_MAP = {"sunday": 0, "monday": 1, "tuesday": 2, "wednesday": 3, "thursday": 4}


def _parse_hm_to_minutes(hm: str, ref_ampm: str = "AM") -> Optional[int]:
    """Parse '8:30', '1:30', etc. to minutes from midnight."""
    hm = (hm or "").strip().upper()
    if not hm:
        return None
    ampm = ref_ampm.upper()
    m = re.match(r"^(\d{1,2}):(\d{2})\s*(AM|PM)?$", hm, re.I)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if m.group(3):
        ampm = m.group(3).upper()
    elif h < 8 and ref_ampm == "PM":
        ampm = "PM"
    if ampm == "PM" and h != 12:
        h += 12
    if ampm == "AM" and h == 12:
        h = 0
    return h * 60 + mi


def _slot_bounds_minutes(time_slot: str):
    """Return (start_min, end_min) for a canonical time slot string."""
    parts = (time_slot or "").split(" - ")
    if len(parts) != 2:
        return None, None
    start_ref = "AM" if "AM" in parts[0].upper() else "PM"
    s = _parse_hm_to_minutes(parts[0].strip(), start_ref)
    e = _parse_hm_to_minutes(parts[1].strip(), "PM" if "PM" in parts[1].upper() else start_ref)
    return s, e


def _duration_from_custom(time_slot: str, custom: str, time_slots: list) -> float:
    """Session length in hours from a custom time range (supports 1.5h, 2.5h, …)."""
    if not custom or not str(custom).strip():
        return 1.0
    txt = str(custom).strip()
    m = re.search(r"([\d]{1,2}:[\d]{2})\s*[-–]\s*([\d]{1,2}:[\d]{2})", txt)
    if not m:
        return 1.0
    _, slot_end_ref = (time_slot.split(" - ") + ["AM"])[:2]
    ref = "PM" if "PM" in slot_end_ref.upper() else "AM"
    start_m = _parse_hm_to_minutes(m.group(1), ref)
    end_m = _parse_hm_to_minutes(m.group(2), "PM")
    if start_m is None or end_m is None:
        return 1.0
    if end_m <= start_m:
        end_m += 12 * 60
    total_min = end_m - start_m
    if total_min <= 0:
        return 1.0
    hours = total_min / 60.0
    return max(0.5, round(hours * 2) / 2)


def _time_range_from_text(txt: str) -> Optional[str]:
    """Pull the first HH:MM-HH:MM range from cell text (parentheses or inline)."""
    if not txt:
        return None
    m = re.search(r"([\d]{1,2}:[\d]{2})\s*[-–]\s*([\d]{1,2}:[\d]{2})", str(txt))
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2)}"


def _duration_slot_span(dur: float) -> int:
    """How many grid columns a duration covers (ceil of hours)."""
    d = float(dur) if dur else 1.0
    return max(1, int(d) if d == int(d) else int(d) + 1)


def _extract_horizontal_merges(ws):
    """Map Excel horizontal spans to 0-based grid indices: anchors -> colspan, skip -> covered cells."""
    anchors = {}
    skip = set()
    for merge_range in ws.merged_cells.ranges:
        min_r, max_r = merge_range.min_row, merge_range.max_row
        min_c, max_c = merge_range.min_col, merge_range.max_col
        if min_r != max_r or max_c <= min_c:
            continue
        r0 = min_r - 1
        c0 = min_c - 1
        span = max_c - min_c + 1
        anchors[(r0, c0)] = max(anchors.get((r0, c0), 1), span)
        for c in range(min_c + 1, max_c + 1):
            skip.add((r0, c - 1))

    # "Center Across Selection" looks merged in Excel but is not in merged_cells
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        c = 1
        while c <= max_col:
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is None or not str(val).strip():
                c += 1
                continue
            horiz = (cell.alignment.horizontal or "") if cell.alignment else ""
            if horiz in ("centerContinuous", "fill"):
                span = 1
                cc = c + 1
                while cc <= max_col:
                    nxt = ws.cell(row=r, column=cc)
                    if nxt.value not in (None, "") and str(nxt.value).strip():
                        break
                    span += 1
                    cc += 1
                if span > 1:
                    r0, c0 = r - 1, c - 1
                    anchors[(r0, c0)] = max(anchors.get((r0, c0), 1), span)
                    for cc in range(c + 1, c + span):
                        skip.add((r0, cc - 1))
                    c += span
                    continue
            c += 1
    return anchors, skip


def _is_empty_schedule_cell(val) -> bool:
    """True when Excel cell should stay blank on the schedule grid."""
    if val is None:
        return True
    txt = str(val).strip()
    if not txt:
        return True
    return txt in ("-", "—", "–", "N/A", "n/a")


def _normalize_schedule_therapist_label(name: str) -> str:
    """Strip Ms. prefix and parenthetical duplicates (e.g. 'Alhanouf (Alhanouf)')."""
    name = (name or "").strip()
    name = re.sub(r"^Ms\.?\s*", "", name, flags=re.I).strip()
    m = re.match(r"^(.+?)\s*\(\s*\1\s*\)\s*$", name, re.I)
    if m:
        name = m.group(1).strip()
    return name


def _build_schedule_therapist_name_map(therapists: list) -> dict:
    """Excel/header labels -> therapist id (first + family, aliases, display names)."""
    t_by_name: dict = {}

    def add(key: str, tid: str):
        k = (key or "").strip()
        if k:
            t_by_name[k] = tid
            t_by_name[k.lower()] = tid

    for t in therapists:
        tid = t["id"]
        raw = (t.get("name") or "").strip()
        add(raw, tid)
        short = _normalize_schedule_therapist_label(raw)
        add(short, tid)
        display = therapist_schedule_display_name(t)
        add(display, tid)
        parts = short.split()
        first = parts[0] if parts else short
        if first:
            add(first, tid)
        if first.lower() == "hajar":
            add("Hajer", tid)
        if first.lower() == "hajer":
            add("Hajar", tid)
        key = (t.get("key") or "").strip()
        family = None
        for k, v in THERAPIST_FAMILY_NAMES.items():
            if k.lower() == key.lower():
                family = v
                break
        if family and first:
            fo = THERAPIST_FIRST_NAME_OVERRIDES.get(first.lower(), first)
            add(f"{first} {family}", tid)
            if fo != first:
                add(f"{fo} {family}", tid)
            add(f"Ms. {first} {family}", tid)
            add(f"Ms. {fo} {family}", tid)
            if first.lower() in ("hajar", "hajer"):
                add(f"Hajer {family}", tid)
                add(f"Hajar {family}", tid)
    return t_by_name


def _parse_schedule_cell_text(txt: str):
    """Returns (service_code, child_name, custom_time, note) or None."""
    if _is_empty_schedule_cell(txt):
        return None
    txt = str(txt).strip()
    upper = txt.upper()
    custom = None
    note = None
    child = None
    service = "SS"
    if "AVC" in upper:
        service = "AVC"; note = txt
    elif "LEAVE" in upper:
        service = "LEAVE"; note = txt
    elif "BREAK" in upper:
        service = "BREAK"; note = txt
    elif "SUPERVISION" in upper:
        service = "SUPERVISION"; note = txt
    elif "OBSERVATION" in upper:
        service = "OBSERVATION"; note = txt
    elif "MEETING" in upper:
        service = "MEETING"; note = txt
    elif upper.startswith("HS"):
        service = "HS"
    elif upper.startswith("OS"):
        service = "OS"
    elif upper.startswith("SS"):
        service = "SS"
    if "|" in txt:
        child = txt.split("|", 1)[1].strip()
    elif "W/" in upper:
        idx = upper.index("W/")
        child = txt[idx + 2:].strip()
    elif " with " in txt.lower():
        child = txt.lower().split(" with ", 1)[1].strip().title()
    if service == "SUPERVISION" and not child:
        m = re.search(
            r"(?i)supervision\s*(?:[-–]\s*|\s+w/?\s*|\s+)(.+?)(?:\s*\(|$)",
            txt,
        )
        if m:
            child = m.group(1).strip()
    if child and "(" in child:
        m_open = child.find("(")
        m_close = child.find(")", m_open)
        if m_close > m_open:
            custom = child[m_open + 1:m_close].strip()
            child = child[:m_open].strip()
    if not custom:
        custom = _time_range_from_text(txt)
    if not child and not note:
        remainder = txt
        for prefix in ("HS", "SS", "OS"):
            if upper.startswith(prefix):
                remainder = re.sub(rf"^{prefix}\s*[\|\-:]*\s*", "", txt, flags=re.I).strip()
                break
        if remainder and remainder.upper() not in {
            "LEAVE", "BREAK", "AVC", "MEETING", "SUPERVISION", "OBSERVATION", "AVAILABLE",
        }:
            child = remainder
            if "(" in child:
                m_open = child.find("(")
                m_close = child.find(")", m_open)
                if m_close > m_open:
                    if not custom:
                        custom = child[m_open + 1:m_close].strip()
                    child = child[:m_open].strip()
    return service, child, custom, note


def _resolve_schedule_therapist(name: str, t_by_name: dict) -> Optional[str]:
    name = _normalize_schedule_therapist_label(name)
    if not name:
        return None
    if name in t_by_name:
        return t_by_name[name]
    nl = name.lower()
    if nl in t_by_name:
        return t_by_name[nl]
    first = name.split()[0] if name.split() else name
    if first in t_by_name:
        return t_by_name[first]
    fl = first.lower()
    if fl in t_by_name:
        return t_by_name[fl]
    # Unique match on "First Family" when Excel omits extra spacing/casing
    matches = {tid for key, tid in t_by_name.items() if key.lower() == nl}
    if len(matches) == 1:
        return next(iter(matches))
    prefix_matches = {
        tid for key, tid in t_by_name.items()
        if " " in key and (key.lower().startswith(nl) or nl.startswith(key.lower()))
    }
    if len(prefix_matches) == 1:
        return next(iter(prefix_matches))
    return None


def _normalize_schedule_grid(rows) -> List[List[str]]:
    grid = []
    for row in rows:
        if row is None:
            grid.append([])
            continue
        cells = []
        for c in row:
            if c is None:
                cells.append("")
            else:
                cells.append(str(c).strip())
        grid.append(cells)
    return grid


def _normalize_week_start(week_start: str) -> str:
    """Normalize any date to the Sunday (week start) ISO string."""
    from datetime import date, timedelta
    raw = (week_start or "").strip()[:10]
    d = date.fromisoformat(raw)
    days_since_sunday = (d.weekday() + 1) % 7
    sunday = d - timedelta(days=days_since_sunday)
    return sunday.isoformat()


def _schedule_state_from_excel_fill(cell) -> Optional[str]:
    """Detect cancel states from Excel cell background (pink = client, yellow = therapist)."""
    try:
        fill = cell.fill
        if not fill or getattr(fill, "fill_type", None) != "solid":
            return None
        color = ""
        for attr in ("fgColor", "start_color"):
            part = getattr(fill, attr, None)
            if part is None:
                continue
            rgb = getattr(part, "rgb", None) or getattr(part, "value", None)
            if rgb:
                s = str(rgb).upper()
                color = s[-6:] if len(s) >= 6 else s
                break
        if not color:
            return None
        if color in ("FCE0E8", "E8A4BD", "F4C2D0", "F8D7E3"):
            return "cancel_child"
        if color in ("FFF4C4", "E8C572", "FFE599", "FFF2CC"):
            return "cancel_therapist"
    except Exception:
        return None
    return None


def _extract_schedule_cell_fills(ws) -> dict:
    """Map 0-based (row, col) -> cancel state from Excel fills."""
    fills: dict = {}
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            state = _schedule_state_from_excel_fill(ws.cell(row=r, column=c))
            if state:
                fills[(r - 1, c - 1)] = state
    return fills


def _cell_import_meta_key(therapist_id: str, day: int, time_slot: str, child_name: Optional[str]) -> str:
    child = _normalize_intake_name(child_name or "")
    slot = (time_slot or "").strip()
    return f"{therapist_id}|{day}|{slot}|{child}"


async def _snapshot_week_cell_overrides(week_start: str) -> dict:
    """Preserve cancellation/cover metadata across Excel re-import."""
    overrides: dict = {}
    async for cell in db.schedule_cells.find({"week_start": week_start}, {"_id": 0}).batch_size(200):
        state = cell.get("state") or "normal"
        cover = (cell.get("cover_child_name") or "").strip() or None
        if state not in ("cancel_child", "cancel_therapist") and not cover:
            continue
        key = _cell_import_meta_key(
            cell.get("therapist_id") or "",
            cell.get("day", 0),
            cell.get("time_slot") or "",
            _schedule_cell_child_label(cell),
        )
        overrides[key] = {
            "state": state,
            "cover_child_name": cover,
            "parent_notify_pending": cell.get("parent_notify_pending"),
        }
    return overrides


async def _save_week_therapist_order(week_start: str, therapist_order: List[str]) -> None:
    """Persist Excel column order for schedule grid rendering."""
    if not therapist_order:
        return
    await db.schedule_weeks.update_one(
        {"week_start": week_start},
        {"$set": {"therapist_order": therapist_order, "updated_at": now_iso()}},
        upsert=True,
    )


async def _canonical_therapist_order_ids() -> List[str]:
    """Map THERAPIST_SCHEDULE_ORDER keys to therapist ids (Excel column order)."""
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "key": 1, "name": 1}).to_list(500)
    key_to_id = {(t.get("key") or "").lower(): t["id"] for t in therapists if t.get("key")}
    ordered = [key_to_id[k] for k in THERAPIST_SCHEDULE_ORDER if k in key_to_id]
    seen = set(ordered)
    seen_display: set = set()
    for tid in ordered:
        t = next((x for x in therapists if x["id"] == tid), None)
        if t:
            seen_display.add(therapist_schedule_display_name(t).lower())
    for t in therapists:
        if t["id"] in seen:
            continue
        disp = therapist_schedule_display_name(t).lower()
        if disp and disp in seen_display:
            continue
        ordered.append(t["id"])
        seen.add(t["id"])
        if disp:
            seen_display.add(disp)
    return ordered


async def _fix_schedule_therapist_order_duplicates() -> int:
    """Remove duplicate therapist rows from week therapist_order (same person twice)."""
    therapists = await db.therapists.find({}, {"_id": 0}).to_list(500)
    id_to_display = {
        t["id"]: therapist_schedule_display_name(t).lower()
        for t in therapists if t.get("id")
    }
    id_remap: dict = {}
    groups: dict = {}
    for t in therapists:
        tok = _therapist_identity_token(t)
        if tok:
            groups.setdefault(tok, []).append(t)
    for group in groups.values():
        if len(group) < 2:
            continue
        scored = sorted(group, key=lambda t: (-_therapist_record_score(t), t.get("created_at") or ""))
        winner = scored[0]
        for loser in scored[1:]:
            id_remap[loser["id"]] = winner["id"]

    updated = 0
    async for doc in db.schedule_weeks.find({}, {"_id": 0, "week_start": 1, "therapist_order": 1}):
        order = doc.get("therapist_order") or []
        if not order:
            continue
        new_order: List[str] = []
        seen_ids: set = set()
        seen_display: set = set()
        changed = False
        for tid in order:
            canonical = id_remap.get(tid, tid)
            disp = id_to_display.get(canonical, "")
            if canonical in seen_ids or (disp and disp in seen_display):
                changed = True
                continue
            if canonical != tid:
                changed = True
            new_order.append(canonical)
            seen_ids.add(canonical)
            if disp:
                seen_display.add(disp)
        if changed:
            await db.schedule_weeks.update_one(
                {"week_start": doc["week_start"]},
                {"$set": {"therapist_order": new_order, "updated_at": now_iso()}},
            )
            updated += 1
    return updated


async def _migrate_schedule_week_therapist_orders() -> int:
    """Fix therapist_order for weeks already imported (e.g. trial week 2026-06-28)."""
    canonical = await _canonical_therapist_order_ids()
    if not canonical:
        return 0
    updated = 0
    for week_start in ("2026-06-28",):
        n_cells = await db.schedule_cells.count_documents({"week_start": week_start})
        if not n_cells:
            continue
        doc = await db.schedule_weeks.find_one({"week_start": week_start}, {"therapist_order": 1})
        existing = (doc or {}).get("therapist_order") or []
        if existing != canonical:
            await db.schedule_weeks.update_one(
                {"week_start": week_start},
                {"$set": {"therapist_order": canonical, "updated_at": now_iso()}},
                upsert=True,
            )
            updated += 1
    return updated


async def _import_schedule_grid(
    grid: List[List[str]],
    week_start: str,
    t_by_name: dict,
    clear_existing: bool,
    merge_anchors: Optional[dict] = None,
    merge_skip: Optional[set] = None,
    cell_fill_states: Optional[dict] = None,
):
    merge_anchors = merge_anchors or {}
    merge_skip = merge_skip or set()
    cell_fill_states = cell_fill_states or {}
    preserved = await _snapshot_week_cell_overrides(week_start) if clear_existing else {}
    if clear_existing:
        await db.schedule_cells.delete_many({"week_start": week_start})
    inserted = 0
    skipped_unknown = []
    pending_cells: List[dict] = []
    therapist_order: List[str] = []
    time_slots = list(SCHEDULE_TIME_SLOTS)
    i = 0
    while i < len(grid):
        row = grid[i]
        joined = " ".join(c.lower() for c in row)
        if "therapist" in joined and "days" in joined and "8:00" in joined:
            time_col_start = 3
            for idx, c in enumerate(row):
                if c and "8:00" in c and ("AM" in c.upper() or "PM" in c.upper()):
                    time_col_start = idx
                    break
            header_times = []
            for c in row[time_col_start:]:
                if c and ("AM" in c.upper() or "PM" in c.upper()):
                    header_times.append(c)
            if header_times:
                time_slots = header_times
            current_t_id = None
            i += 1
            while i < len(grid):
                r = grid[i]
                if not any(r):
                    if i + 1 < len(grid) and not any(grid[i + 1]):
                        i += 1
                        break
                    i += 1
                    continue
                name_c = r[1] if len(r) > 1 else ""
                if name_c and name_c.lower() not in SCHEDULE_DAYS_MAP:
                    if name_c.lower() in _SCHEDULE_THERAPIST_SKIP:
                        i += 1
                        continue
                    tid = _resolve_schedule_therapist(name_c, t_by_name)
                    if tid:
                        current_t_id = tid
                        if tid not in therapist_order:
                            therapist_order.append(tid)
                    elif name_c and name_c not in skipped_unknown:
                        skipped_unknown.append(name_c)
                        current_t_id = None
                day_label = (r[2] if len(r) > 2 else "").lower()
                day_idx = SCHEDULE_DAYS_MAP.get(day_label)
                if day_idx is not None and current_t_id:
                    skip_until = -1
                    for slot_idx, ts in enumerate(time_slots):
                        if slot_idx <= skip_until:
                            continue
                        col_idx = time_col_start + slot_idx
                        if col_idx >= len(r):
                            break
                        if (i, col_idx) in merge_skip:
                            continue
                        val = r[col_idx].strip() if col_idx < len(r) else ""
                        parsed = _parse_schedule_cell_text(val)
                        canonical_ts = (
                            SCHEDULE_TIME_SLOTS[slot_idx]
                            if slot_idx < len(SCHEDULE_TIME_SLOTS)
                            else ts
                        )
                        if not parsed:
                            if _is_empty_schedule_cell(val):
                                await _clear_schedule_span(
                                    current_t_id, day_idx, canonical_ts, 1.0, week_start
                                )
                            continue
                        service, child, custom, note = parsed
                        merge_cols = float(merge_anchors.get((i, col_idx), 1))
                        custom_dur = _duration_from_custom(canonical_ts, custom, time_slots) if custom else 1.0
                        if custom and custom_dur > 1:
                            duration = custom_dur
                        elif merge_cols > 1:
                            duration = merge_cols
                        else:
                            duration = 1.0
                        span = _duration_slot_span(duration)
                        if span > 1:
                            skip_until = slot_idx + span - 1
                        if not clear_existing:
                            await _clear_schedule_span(
                                current_t_id, day_idx, canonical_ts, duration, week_start
                            )
                        meta_key = _cell_import_meta_key(current_t_id, day_idx, canonical_ts, child)
                        preserved_meta = preserved.get(meta_key) or {}
                        excel_state = cell_fill_states.get((i, col_idx))
                        cell_state = preserved_meta.get("state") or excel_state or "normal"
                        pending_cells.append({
                            "id": str(uuid.uuid4()),
                            "therapist_id": current_t_id,
                            "day": day_idx,
                            "time_slot": canonical_ts,
                            "service_code": service,
                            "child_name": child,
                            "note": note,
                            "custom_time": custom,
                            "state": cell_state,
                            "cover_child_name": preserved_meta.get("cover_child_name"),
                            "color": None,
                            "duration": duration,
                            "week_start": week_start,
                            "created_at": now_iso(),
                        })
                        if len(pending_cells) >= 100:
                            await db.schedule_cells.insert_many(pending_cells)
                            inserted += len(pending_cells)
                            pending_cells.clear()
                i += 1
            if pending_cells:
                await db.schedule_cells.insert_many(pending_cells)
                inserted += len(pending_cells)
                pending_cells.clear()
            continue
        i += 1
    await _save_week_therapist_order(week_start, therapist_order)
    return inserted, skipped_unknown


async def _relink_prep_markers_after_schedule_import(week_start: str) -> None:
    """Re-attach prep badges to new cell IDs after Excel re-import."""
    try:
        base = datetime.fromisoformat(str(week_start)[:10])
    except ValueError:
        return
    start = base.strftime("%Y-%m-%d")
    end = (base + timedelta(days=4)).strftime("%Y-%m-%d")
    await _sync_schedule_preparations_for_week(start, end)


async def _clear_schedule_span(
    therapist_id: str,
    day: int,
    time_slot: str,
    duration: float,
    week_start: str,
):
    """Remove existing cells overlapping this span (avoids duplicate 1h + 2h on re-import)."""
    slots = list(SCHEDULE_TIME_SLOTS)
    try:
        start_idx = slots.index(time_slot)
    except ValueError:
        await db.schedule_cells.delete_many({
            "therapist_id": therapist_id,
            "day": day,
            "week_start": week_start,
            "time_slot": time_slot,
        })
        return
    span = _duration_slot_span(duration)
    covered = [slots[start_idx + k] for k in range(span) if start_idx + k < len(slots)]
    if covered:
        await db.schedule_cells.delete_many({
            "therapist_id": therapist_id,
            "day": day,
            "week_start": week_start,
            "time_slot": {"$in": covered},
        })


def _google_sheet_export_url(sheet_url: str) -> str:
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", sheet_url or "")
    if not m:
        raise HTTPException(status_code=400, detail="Invalid Google Sheets URL")
    return f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx"


def _parse_week_start_from_sheet_name(name: str, ref_year: Optional[int] = None) -> Optional[str]:
    """Parse week Sunday from tab names like '7 Jun - 11 Jun' or '14 Jun - 18 Jun'."""
    from datetime import date, timedelta
    if not name:
        return None
    m = re.search(
        r"\b(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b",
        name,
        re.I,
    )
    if not m:
        return None
    day = int(m.group(1))
    month_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    month = month_map.get(m.group(2).lower())
    if not month:
        return None
    year = ref_year or date.today().year
    try:
        d = date(year, month, day)
    except ValueError:
        return None
    days_since_sunday = (d.weekday() + 1) % 7
    sunday = d - timedelta(days=days_since_sunday)
    return sunday.isoformat()


def _pick_sheet_for_week(sheet_names: List[str], week_start: str) -> Optional[str]:
    """Pick the tab whose name best matches week_start (e.g. 2026-06-07 → '7 Jun - 11 Jun')."""
    from datetime import date
    try:
        d = date.fromisoformat(week_start[:10])
    except ValueError:
        return sheet_names[0] if sheet_names else None
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    mon = month_names[d.month - 1]
    day = d.day
    pattern = re.compile(rf"\b{day}\s+{mon}\b", re.I)
    for name in sheet_names:
        if pattern.search(name):
            return name
    return sheet_names[0] if sheet_names else None


def _resolve_import_week_start(requested: str, sheet_name: Optional[str]) -> tuple:
    """Return (week_start, warning). Prefer dates parsed from the sheet tab name."""
    from datetime import date
    requested = _normalize_week_start(requested)
    ref_year = None
    try:
        ref_year = date.fromisoformat(requested[:10]).year
    except ValueError:
        pass
    parsed = _parse_week_start_from_sheet_name(sheet_name or "", ref_year)
    if parsed and parsed != requested:
        logger.info(
            f"Schedule import: sheet {sheet_name!r} implies week {parsed}, "
            f"overriding requested {requested}"
        )
        return parsed, (
            f"Sheet tab '{sheet_name}' is for week starting {parsed}; "
            f"used that instead of requested {requested}."
        )
    return requested, None


def _load_schedule_xlsx_bytes(content: bytes, sheet_name: Optional[str] = None):
    """Parse workbook bytes into grid + horizontal merge maps."""
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    if not sheet_name:
        sheet_name = wb.sheetnames[0] if wb.sheetnames else None
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    raw = []
    for r in range(1, max_row + 1):
        raw.append([ws.cell(row=r, column=c).value for c in range(1, max_col + 1)])
    grid = _normalize_schedule_grid(raw)
    merge_anchors, merge_skip = _extract_horizontal_merges(ws)
    fill_states = _extract_schedule_cell_fills(ws)
    return grid, merge_anchors, merge_skip, ws.title, wb.sheetnames, fill_states


async def _seed_schedule_week_2026_05_24(t_by_name: dict):
    """Seed May 24–28 2026 schedule if empty (idempotent)."""
    week_start = "2026-05-24"
    if await db.schedule_cells.count_documents({"week_start": week_start}) > 0:
        return 0
    seed_rows = [
        ("Ms. Maha", 0, "8:00 AM - 9:00 AM", "Leave"),
        ("Ms. Fahda", 0, "10:00 AM - 11:00 AM", "Supervision W/ Lulu (10-11:30)"),
        ("Ms. Fahda", 0, "11:00 AM - 12:00 PM", "HS | Saleh (11:30-1:30)"),
        ("Ms. Fahda", 0, "1:00 PM - 2:00 PM", "HS | Ibrahim"),
        ("Ms. Fahda", 1, "8:00 AM - 9:00 AM", "HS | Abdulaziz A"),
        ("Ms. Fahda", 1, "9:00 AM - 10:00 AM", "HS | Ibrahim"),
        ("Ms. Razan", 0, "8:00 AM - 9:00 AM", "Leave"),
        ("Ms. Manal", 0, "8:00 AM - 9:00 AM", "HS | Abdulaziz A"),
        ("Ms. Manal", 0, "9:00 AM - 10:00 AM", "HS | Omar"),
        ("Ms. Manal", 1, "8:00 AM - 9:00 AM", "HS | Salman"),
        ("Ms. Manal", 1, "9:00 AM - 10:00 AM", "HS | Saleh"),
        ("Ms. Hajer", 0, "8:00 AM - 9:00 AM", "Leave"),
        ("Ms. Rahaf", 0, "8:00 AM - 9:00 AM", "Leave"),
        ("Ms. Shatha", 0, "8:00 AM - 9:00 AM", "Leave"),
        ("Ms. Alhanouf", 0, "8:00 AM - 9:00 AM", "Leave"),
        ("Ms. Waad", 0, "8:00 AM - 9:00 AM", "Leave"),
        ("Ms. Bodoor", 0, "8:00 AM - 9:00 AM", "Leave"),
        ("Ms. Fatimah", 0, "8:00 AM - 9:00 AM", "HS | Lulu"),
        ("Ms. Fatimah", 0, "9:00 AM - 10:00 AM", "HS | Abdulaziz W"),
        ("Ms. Fatimah", 0, "10:00 AM - 11:00 AM", "AVC"),
        ("Ms. Fatimah", 1, "8:00 AM - 9:00 AM", "HS | Lulu"),
        ("Ms. Fatimah", 1, "9:00 AM - 10:00 AM", "HS | Abdulaziz W"),
        ("Ms. Shrooq", 0, "8:00 AM - 9:00 AM", "Leave"),
        ("Ms. Abeer", 0, "8:00 AM - 9:00 AM", "Leave"),
        ("Ms. Najla", 0, "8:00 AM - 9:00 AM", "HS | Abdulaziz A"),
        ("Ms. Najla", 0, "9:00 AM - 10:00 AM", "HS | Omar"),
        ("Ms. Najla", 1, "8:00 AM - 9:00 AM", "HS | Khalid"),
        ("Ms. Walaa", 0, "8:00 AM - 9:00 AM", "HS | Khalid"),
        ("Ms. Walaa", 1, "8:00 AM - 9:00 AM", "HS | Khalid"),
    ]
    inserted = 0
    for tname, day, slot, content in seed_rows:
        tid = _resolve_schedule_therapist(tname, t_by_name)
        if not tid:
            continue
        parsed = _parse_schedule_cell_text(content)
        if not parsed:
            continue
        service, child, custom, note = parsed
        await db.schedule_cells.insert_one({
            "id": str(uuid.uuid4()),
            "therapist_id": tid,
            "day": day,
            "time_slot": slot,
            "service_code": service,
            "child_name": child,
            "note": note,
            "custom_time": custom,
            "state": "normal",
            "color": None,
            "duration": 1,
            "week_start": week_start,
            "created_at": now_iso(),
        })
        inserted += 1
    if inserted:
        logger.info(f"Seeded {inserted} schedule cells for week {week_start}")
    return inserted


@api.post("/import/schedule-excel")
async def import_schedule_excel(file: UploadFile = File(...),
                                 week_start: str = Form(...),
                                 clear_existing: Optional[str] = Form(None),
                                 sheet_name: Optional[str] = Form(None),
                                 _=Depends(import_access)):
    """Parse Therapists' Schedule .xlsx or .csv and create cells for week_start."""
    import io
    content = await file.read()
    fname = (file.filename or "").lower()
    week_start = _normalize_week_start(week_start)
    logger.info(f"Schedule import week_start={week_start} (normalized to Sunday)")
    therapists = await db.therapists.find({}, {"_id": 0, "pin_hash": 0, "password_hash": 0}).to_list(100)
    t_by_name = _build_schedule_therapist_name_map(therapists)

    if fname.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig", errors="replace")
        grid = _normalize_schedule_grid(list(csv.reader(io.StringIO(text))))
        merge_anchors, merge_skip = {}, set()
        used_sheet = None
        fill_states = {}
    else:
        grid, merge_anchors, merge_skip, used_sheet, sheet_names, fill_states = _load_schedule_xlsx_bytes(content, sheet_name)
        if not sheet_name:
            picked = _pick_sheet_for_week(sheet_names, week_start)
            if picked and picked != used_sheet:
                grid, merge_anchors, merge_skip, used_sheet, _, fill_states = _load_schedule_xlsx_bytes(content, picked)
        logger.info(
            f"Schedule Excel sheet={used_sheet!r} merges: {len(merge_anchors)} anchors, {len(merge_skip)} covered"
        )

    week_start, week_warning = _resolve_import_week_start(week_start, used_sheet)
    inserted, skipped = await _import_schedule_grid(
        grid, week_start, t_by_name, clear_existing == "true",
        merge_anchors=merge_anchors, merge_skip=merge_skip,
        cell_fill_states=fill_states,
    )
    await _relink_prep_markers_after_schedule_import(week_start)
    return {
        "cells_inserted": inserted,
        "week_start": week_start,
        "skipped_therapists": skipped[:20],
        "sheet_used": used_sheet,
        "merge_spans_detected": len(merge_anchors),
        "week_start_warning": week_warning,
        "prep_relinked": True,
    }


@api.post("/import/schedule-google")
async def import_schedule_google(body: dict, _=Depends(import_access)):
    """Import a week directly from a public Google Sheets link (preserves merged cells)."""
    import httpx
    sheet_url = (body.get("url") or body.get("sheet_url") or "").strip()
    week_start = _normalize_week_start(body.get("week_start") or "")
    sheet_name = (body.get("sheet_name") or "").strip() or None
    clear_existing = body.get("clear_existing", True)
    if not sheet_url:
        raise HTTPException(status_code=400, detail="sheet_url required")
    export_url = _google_sheet_export_url(sheet_url)
    async with httpx.AsyncClient(follow_redirects=True, timeout=120.0) as client:
        resp = await client.get(export_url)
    if resp.status_code != 200:
        raise HTTPException(
            status_code=400,
            detail=f"Could not download Google Sheet (HTTP {resp.status_code}). Share link must be viewable.",
        )
    content = resp.content
    grid, merge_anchors, merge_skip, used_sheet, sheet_names, fill_states = _load_schedule_xlsx_bytes(content, sheet_name)
    if not sheet_name:
        picked = _pick_sheet_for_week(sheet_names, week_start)
        if picked and picked != used_sheet:
            grid, merge_anchors, merge_skip, used_sheet, _, fill_states = _load_schedule_xlsx_bytes(content, picked)
    logger.info(
        f"Google schedule import week={week_start} sheet={used_sheet!r} "
        f"merges={len(merge_anchors)} bytes={len(content)}"
    )
    week_start, week_warning = _resolve_import_week_start(week_start, used_sheet)
    therapists = await db.therapists.find({}, {"_id": 0, "pin_hash": 0, "password_hash": 0}).to_list(100)
    t_by_name = _build_schedule_therapist_name_map(therapists)
    inserted, skipped = await _import_schedule_grid(
        grid, week_start, t_by_name, bool(clear_existing),
        merge_anchors=merge_anchors, merge_skip=merge_skip,
        cell_fill_states=fill_states,
    )
    await _relink_prep_markers_after_schedule_import(week_start)
    return {
        "cells_inserted": inserted,
        "week_start": week_start,
        "skipped_therapists": skipped[:20],
        "sheet_used": used_sheet,
        "merge_spans_detected": len(merge_anchors),
        "available_sheets": sheet_names[:30],
        "week_start_warning": week_warning,
        "prep_relinked": True,
    }

@api.get("/")
async def root():
    return {"message": "Boost Growth Portal API", "status": "ok"}


@api.get("/health")
@api.head("/health")
async def health():
    return {"status": "ok"}


@api.get("/version")
async def app_version():
    """Deploy fingerprint — compare with GitHub main to confirm Railway picked up latest build."""
    version_file = ROOT_DIR / "BUILD_VERSION.txt"
    build_id = version_file.read_text(encoding="utf-8").strip() if version_file.is_file() else "unknown"
    js_hash = "unknown"
    index = ROOT_DIR / "static" / "index.html"
    if index.is_file():
        m = re.search(r"main\.([a-f0-9]+)\.js", index.read_text(encoding="utf-8"))
        if m:
            js_hash = m.group(1)
    return {"build": build_id, "js": js_hash, "status": "ok"}

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[
        "https://staff.boostgrowth.org",
        "http://staff.boostgrowth.org",
        "http://localhost:3000",
        "http://localhost:5173",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:5173",
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(GZipMiddleware, minimum_size=800)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ------------------- Seed Data (FROM BASE44 SOURCE) -------------------
THERAPIST_SEED = [
    {"name": "Ms. Maha", "color": "#7A8A6A", "email": "maha@boostgrowthsa.com"},
    {"name": "Ms. Fahda", "color": "#D4A64A", "email": "fahda@boostgrowthsa.com"},
    {"name": "Ms. Razan", "color": "#8FA481", "email": "razan@boostgrowthsa.com"},
    {"name": "Ms. Manal", "color": "#A4BCCB", "email": "manal@boostgrowthsa.com"},
    {"name": "Ms. Hajar", "color": "#C97B5C", "email": "halfulaij@boostgrowthsa.com"},
    {"name": "Ms. Rahaf", "color": "#9B7BAB", "email": "rahaf@boostgrowthsa.com"},
    {"name": "Ms. Shatha", "color": "#5C8B7E", "email": "shatha@boostgrowthsa.com"},
    {"name": "Ms. Alhanouf", "color": "#B89968", "email": "alhanouf@boostgrowthsa.com"},
    {"name": "Ms. Waad", "color": "#7B96B5", "email": "waad@boostgrowthsa.com"},
    {"name": "Ms. Fatimah", "color": "#6B9080", "email": "fatimah@boostgrowthsa.com"},
    {"name": "Ms. Shroug", "color": "#D49A60", "email": "shalamri@boostgrowthsa.com"},
    {"name": "Ms. Abeer", "color": "#8B7BA8", "email": "abeer@boostgrowthsa.com"},
    {"name": "Ms. Najla", "color": "#7BA890", "email": "najla@boostgrowthsa.com"},
    {"name": "Ms. Asma", "color": "#6A7F9B", "email": "asma@boostgrowthsa.com"},
    {"name": "Ms. Jenan", "color": "#7A8A6A", "email": "jsalmuhaisin@boostgrowthsa.com"},
]

CLIENT_SEED = [
    {"file_no":"009","name":"Saleh Ahusainy","main":"Ms. Waad","co":["Ms. Manal","Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#FFE599","locs":[{"service":"SS","address":"Alnakeel - Home Sweet Home"},{"service":"HS","address":"Alnakheel - 1st floor, apartment #7"},{"service":"HS","address":"Grandmother house"}]},
    {"file_no":"011","name":"Fahad Alyahya","main":"Ms. Alhanouf","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#A2C4C9","locs":[{"service":"HS","address":"Alyasmin - house no 3075"},{"service":"SS","address":"Talat School"}]},
    {"file_no":"018","name":"Layan AlSaud","main":"Ms. Jenan","co":[],"pkg":24,"sup":"Ms. Jenan","color":"#C9DAF8","locs":[{"service":"ABA","address":"Alaqiq"}]},
    {"file_no":"023","name":"Yahya Alqahtani","main":"Ms. Hajer","co":["Ms. Manal"],"pkg":24,"sup":"Ms. Fahda","color":"#D5A6BD","locs":[{"service":"HS","address":"Alaarid"}]},
    {"file_no":"024","name":"Abdulaziz Alrasheed","main":"Ms. Shatha","co":["Ms. Manal","Ms. Hajer"],"pkg":24,"sup":"Ms. Fahda","color":"#E6B8AF","locs":[{"service":"HS","address":"Alnada - Building #26, 3rd floor, apartment #23"}]},
    {"file_no":"027","name":"Mohmmed Alaqel","main":"Ms. Rahaf","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#FFF2CC","locs":[{"service":"HS","address":"AlMalqa - 331"}]},
    {"file_no":"030","name":"Husam Alturaigy","main":"Ms. Manal","co":["Ms. Shatha"],"pkg":24,"sup":"Ms. Fahda","color":"#B4A7D6","locs":[{"service":"SS","address":"Whales of the future daycare"},{"service":"HS","address":"Alwaha - Home #4B"}]},
    {"file_no":"034","name":"Aljouhrah Alduailij","main":"Ms. Asma","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#D9EAD3","locs":[{"service":"SS","address":"Alnakheel - Talat School"}]},
    {"file_no":"035","name":"Saad Alghamdi","main":"Ms. Shatha","co":["Ms. Hajer","Ms. Fatimah"],"pkg":24,"sup":"Ms. Maha","color":"#B6D7A8","locs":[{"service":"HS","address":"Al Aqiq - House in the corner"},{"service":"SS","address":"Al Motaqdimah Schools"}]},
    {"file_no":"037","name":"Suzan Alsultan","main":"Ms. Asma","co":[],"pkg":24,"sup":"Ms. Maha","color":"#FCE5CD","locs":[{"service":"HS","address":"King Fahad - Villa 1308"}]},
    {"file_no":"038","name":"Salman Alrasheed","main":"Ms. Manal","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Maha","color":"#F4CCCC","locs":[{"service":"SS","address":"Summer Camp - Stars of Knowledge School"},{"service":"HS","address":"Alnada - Building #26, 3rd floor, apartment #23"}]},
    {"file_no":"040","name":"Abdulaziz AlAbdulwahab","main":"Ms. Fatimah","co":["Ms. Fahda","Ms. Hajer"],"pkg":24,"sup":"Ms. Maha","color":"#6FA8DC","locs":[{"service":"HS","address":"Alraed - house no 8188"}]},
    {"file_no":"041","name":"Ameerah Alshehri","main":"Ms. Fahda","co":["Ms. Fatimah"],"pkg":24,"sup":"Ms. Maha","color":"#EA9999","locs":[{"service":"HS","address":"Roshen - Villa 277"}]},
    {"file_no":"042","name":"Sultan Aldamer","main":"Ms. Shroug","co":["Ms. Rahaf"],"pkg":24,"sup":"Ms. Maha","color":"#FFE599","locs":[{"service":"SS","address":"Bright Mind School"},{"service":"HS","address":"Alhada - No house number"}]},
    {"file_no":"047","name":"Alwaleed Alotaibi","main":"Ms. Hajer","co":["Ms. Alhanouf"],"pkg":24,"sup":"Ms. Maha","color":"#B4A7D6","locs":[{"service":"HS","address":"Alqairawan - house no 10"},{"service":"SS","address":"Al Motaqdimah Schools"}]},
    {"file_no":"052","name":"Sulaiman Alkhurashi","main":"Ms. Rahaf","co":["Ms. Maha"],"pkg":24,"sup":"Ms. Maha","color":"#F9CB9C","locs":[{"service":"HS","address":"Alsulaimanyah - house no 24"}]},
    {"file_no":"054","name":"Omar Alkhurashi","main":"Ms. Manal","co":["Ms. Maha"],"pkg":24,"sup":"Ms. Maha","color":"#D0E0E3","locs":[{"service":"HS","address":"Alsulaimanyah - house no 24"}]},
    {"file_no":"060","name":"Mohammed Albedayea","main":"Ms. Bodoor","co":["Ms. Shatha"],"pkg":24,"sup":"Ms. Maha","color":"#D9EAD3","locs":[{"service":"HS","address":"Alyasmin - Home no 14"},{"service":"SS","address":"Yas School"}]},
    {"file_no":"061","name":"Ibrahim Alnasir","main":"Ms. Rahaf","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#D9D2E9","locs":[{"service":"HS","address":"Alyasmin - Home no 39"}]},
    {"file_no":"062","name":"Lulu Almutair","main":"Ms. Razan","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#D5A6BD","locs":[{"service":"HS","address":"Almuroj - Home no 4"},{"service":"SS","address":"Alnakheel - Talat School"}]},
    {"file_no":"063","name":"Amani Ghaith","main":"Ms. Maha","co":[],"pkg":24,"sup":"Ms. Maha","color":"#FFF2CC","locs":[{"service":"HS","address":"Alnakheel"}]},
    {"file_no":"065","name":"Aser Alharbi","main":"Ms. Najla","co":["Ms. Maha"],"pkg":24,"sup":"Ms. Maha","color":"#F4CCCC","locs":[{"service":"HS","address":"Al Izdihar - First floor - House no 15"}]},
    {"file_no":"068","name":"Abdulrahman Alshawi","main":"Ms. Razan","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#C9DAF8","locs":[{"service":"HS","address":"AR Rayan - Home no 32"}]},
    {"file_no":"070","name":"Abdulelah Almuhana","main":"Ms. Abeer","co":["Ms. Maha"],"pkg":32,"sup":"Ms. Maha","color":"#CFE2F3","locs":[{"service":"HS","address":"Al-Manziliyah"}]},
    {"file_no":"072","name":"Khalid Bin Shuael","main":"Ms. Shatha","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#EAD1DC","locs":[{"service":"HS","address":"AlMursalat"}]},
    {"file_no":"076","name":"Sultan Abalkhail","main":"Ms. Shatha","co":[],"pkg":24,"sup":"Ms. Fahda","color":"#D0E0E3","locs":[{"service":"HS","address":"Al-Mursalat"},{"service":"SS","address":"Al-Mursalat"}]},
]

CLIENT_ATTENDANCE_SHEETS = {
    "038": "https://docs.google.com/spreadsheets/d/1O4xIX4pJzXfHw3b029kXOQZ8owI9F4RYSlBFuPJ3X7k/edit?usp=sharing",
}

OFFICIAL_CLIENT_FILE_NOS = frozenset(c["file_no"] for c in CLIENT_SEED)


class RestoreClientsIn(BaseModel):
    confirm: str


@api.post("/admin/restore-official-clients")
async def restore_official_clients(body: RestoreClientsIn, _=Depends(admin_only)):
    """Remove clients not in the official 25-client seed list and restore known profiles."""
    if (body.confirm or "").strip() != "RESTORE":
        raise HTTPException(status_code=400, detail="Type RESTORE to confirm")

    therapists_map = {t["name"]: t["id"] async for t in db.therapists.find({}, {"_id": 0, "name": 1, "id": 1})}
    deleted: List[str] = []
    for c in await db.clients.find({}, {"_id": 0, "id": 1, "file_no": 1, "name": 1}).to_list(500):
        fn = str(c.get("file_no") or "").strip()
        fn_norm = fn.zfill(3) if fn else None
        if fn_norm and fn_norm in OFFICIAL_CLIENT_FILE_NOS:
            continue
        cid = c["id"]
        await db.sessions.delete_many({"client_id": cid})
        await db.invoices.delete_many({"client_id": cid})
        await db.progress_reports.delete_many({"client_id": cid})
        await db.clients.delete_one({"id": cid})
        deleted.append(c.get("name") or fn or cid)

    created, updated = 0, 0
    for seed in CLIENT_SEED:
        match = await db.clients.find_one({"file_no": seed["file_no"]}, {"_id": 0, "id": 1, "deleted": 1})
        if match and match.get("deleted"):
            continue
        fields = {
            "file_no": seed["file_no"],
            "name": seed["name"],
            "package_hours": seed["pkg"],
            "supervisor": seed["sup"],
            "main_therapist_id": therapists_map.get(seed["main"]),
            "co_therapist_ids": [therapists_map[n] for n in seed["co"] if n in therapists_map],
            "color": seed["color"],
            "locations": seed["locs"],
            "billing_mode": "hours",
        }
        if seed["file_no"] in CLIENT_ATTENDANCE_SHEETS:
            fields["attendance_sheet_url"] = CLIENT_ATTENDANCE_SHEETS[seed["file_no"]]
        if seed["file_no"] in INACTIVE_CLIENT_FILE_NOS:
            fields["status"] = "Inactive"
        if match:
            await db.clients.update_one({"id": match["id"]}, {"$set": fields})
            updated += 1
        else:
            await db.clients.insert_one({
                "id": str(uuid.uuid4()),
                "payment_status": "pending",
                "created_at": now_iso(),
                **fields,
            })
            created += 1

    total = await db.clients.count_documents(_active_client_filter())
    return {
        "ok": True,
        "deleted_count": len(deleted),
        "deleted_names": deleted[:40],
        "created": created,
        "updated": updated,
        "total_clients": total,
        "message": (
            f"Removed {len(deleted)} unknown client(s). "
            f"Restored {len(CLIENT_SEED)} official clients ({created} new, {updated} updated). "
            f"Total now: {total}."
        ),
    }

# ------------------- Intake Seed (from Waiting_List_v4.xlsx) -------------------
INTAKE_SEED = [
    # Pre-Intake (16 — waiting for assessment)
    {"intake_type": "pre", "child_name": "Reema Idrees", "service": "HS", "district": "Irqah", "age": "2021", "diagnosis": "PWS", "priority": False},
    {"intake_type": "pre", "child_name": "Abdulaziz Alrajab", "service": "HS", "district": "Al Malqa", "age": "2023", "notes": "Online consultation", "priority": False},
    {"intake_type": "pre", "child_name": "Mansour", "service": "HS", "district": "Alyasmeen", "age": "2022", "diagnosis": "Speech delay", "priority": False},
    {"intake_type": "pre", "child_name": "Leen", "service": "HS", "district": "Al Raed", "age": "2010", "notes": "3hrs at school", "priority": False},
    {"intake_type": "pre", "child_name": "Ebrahim Alnami", "service": "SS", "district": "Alsulimania", "age": "2022", "diagnosis": "Premature 29 weeks", "time_pref": "Morning", "priority": False},
    {"intake_type": "pre", "child_name": "Naif Alblawi", "service": "SS", "district": "Qurtubah", "age": "2020", "diagnosis": "ADHD", "time_pref": "Evening", "priority": False},
    {"intake_type": "pre", "child_name": "Saad Alajaji", "service": "SS", "district": "AL-Suwaidi", "age": "2021", "time_pref": "Evening", "priority": False},
    {"intake_type": "pre", "child_name": "Reema Alotaibi", "service": "HS", "district": "AlArid", "diagnosis": "Speech delay", "time_pref": "Evening", "priority": True},
    {"intake_type": "pre", "child_name": "Waseem Aljohani", "service": "HS/SS", "district": "Alnarjis", "age": "2019", "diagnosis": "ADHD", "notes": "Dr.Turki", "priority": False},
    {"intake_type": "pre", "child_name": "Sultan Bandar", "service": "HS", "district": "Alyasmeen", "age": "2019", "diagnosis": "Speech delay/ADHD", "priority": False},
    {"intake_type": "pre", "child_name": "Feras AlFouzan", "service": "SS", "district": "AlFalah", "age": "2019", "diagnosis": "ASD level 1 nonverbal", "notes": "English, Ms.Jenan", "priority": False},
    {"intake_type": "pre", "child_name": "Saud Alshrafi", "service": "SS", "district": "Alyasmeen", "age": "2020", "diagnosis": "ADHD", "time_pref": "Morning", "priority": False},
    {"intake_type": "pre", "child_name": "Khalid Abunayyan", "service": "HS", "district": "Diriyah", "age": "2021", "diagnosis": "ADD", "notes": "English/Arabic", "priority": False},
    {"intake_type": "pre", "child_name": "Fahad Abdullatif", "service": "HS", "district": "Sidrah", "age": "2020", "diagnosis": "ADHD", "notes": "4:30 PM", "priority": False},
    {"intake_type": "pre", "child_name": "Mela Mohammed", "service": "SS", "district": "Tuwiq", "age": "2022", "diagnosis": "ADHD", "time_pref": "Morning", "priority": False},
    {"intake_type": "pre", "child_name": "Mansour Tonkar", "service": "SS", "district": "Al-Moroj", "age": "2019", "diagnosis": "ASD", "notes": "English", "priority": False},
    # Post-Intake (15 — assessed, waiting for slot)
    {"intake_type": "post", "child_name": "Mohammed Alnoweser", "service": "HS", "district": "King Fahad", "age": "3 yrs", "priority": False},
    {"intake_type": "post", "child_name": "Mohammed Alofi", "service": "HS", "phone": "554505400", "district": "AlAridh", "age": "6", "priority": False},
    {"intake_type": "post", "child_name": "Rakan Alaqel", "service": "HS", "phone": "538154083", "district": "Alnarjis", "age": "2019", "priority": False},
    {"intake_type": "post", "child_name": "Nawaf Alshweeb", "service": "HS", "district": "Um Alhamam", "age": "5.5 / ASD", "priority": False},
    {"intake_type": "post", "child_name": "Abdulkareem Kaki", "service": "HS", "priority": False},
    {"intake_type": "post", "child_name": "Abdulaziz Alzahrani", "service": "HS", "phone": "555341092", "district": "Almalqa", "age": "4", "priority": False},
    {"intake_type": "post", "child_name": "Yazeed Bu Sheet", "service": "SS", "phone": "555009662", "district": "Hitten", "diagnosis": "Autism", "priority": False},
    {"intake_type": "post", "child_name": "Misk Alsadoon", "service": "HS", "district": "Qurtubah", "notes": "with Ms.Fahda", "priority": False},
    {"intake_type": "post", "child_name": "Omar AlImazrou", "service": "HS", "phone": "534888855", "district": "AlArid", "age": "2023", "diagnosis": "Autism", "priority": True},
    {"intake_type": "post", "child_name": "Naif Alwhibi", "service": "SS/HS", "phone": "506128118", "district": "Ar Rabi", "age": "2020 / ASD", "priority": True},
    {"intake_type": "post", "child_name": "Ahmad Alshalfan", "service": "SS/HS", "phone": "505287407", "district": "Almalqa", "age": "2020", "diagnosis": "ADHD and GDD", "priority": True},
    {"intake_type": "post", "child_name": "Abdulelah Almuhana", "service": "HS", "phone": "966565544999", "district": "Al-Taawun", "age": "2021", "priority": True},
    {"intake_type": "post", "child_name": "Faisal Alzghaibi", "service": "HS", "phone": "966507479800", "district": "Alyasmeen", "age": "1445", "priority": False},
    {"intake_type": "post", "child_name": "Sultan Abalkhail", "service": "HS/SS", "district": "Al-Mursalat", "age": "2019", "priority": False},
    {"intake_type": "post", "child_name": "Leena Alshahrani", "service": "HS", "phone": "530511175", "district": "Alnarjis", "priority": False},
]

SCHOOL_WAITING_SEED = [
    {"intake_type": "school", "list_category": "school", "child_name": "Mohammed AlAqeel", "service": "SS", "school_start_date": "23/08/2026", "language": "Arabic", "status": "new", "priority": False},
    {"intake_type": "school", "list_category": "school", "child_name": "Ameerah Alshehri", "service": "SS", "school_start_date": "23/08/2026", "language": "Arabic", "status": "new", "priority": False},
    {"intake_type": "school", "list_category": "school", "child_name": "Abdulrahman Alshawi", "service": "SS", "language": "Arabic", "status": "new", "priority": False},
    {"intake_type": "school", "list_category": "school", "child_name": "khalid Abunyan", "service": "SS", "school_start_date": "12/07/2026", "language": "Arabic", "status": "new", "priority": False},
    {"intake_type": "school", "list_category": "school", "child_name": "Saif AlHoury", "service": "SS", "school_start_date": "05/07/2026", "language": "English", "status": "new", "priority": False},
    {"intake_type": "school", "list_category": "school", "child_name": "Aljouhrah Alduailij", "service": "SS", "school_start_date": "19/07/2025", "status": "new", "priority": False},
]

# ------------------- Directory Seed (Internal Team) -------------------
DIRECTORY_SEED = [
    {"name":"Genan Almuhaisen","role":"Direct Manager","phone":"","email":"genan@boostgrowthsa.com"},
    {"name":"Boost Growth (Main)","role":"Coordinator / General Inquiries","phone":"","email":"hello@boostgrowthsa.com"},
    {"name":"Ms. Walaa","role":"Operations","phone":"","email":"wabuissa@boostgrowthsa.com"},
    {"name":"Ms. Maha","role":"Supervisor","phone":"","email":"maha@boostgrowthsa.com"},
    {"name":"Ms. Fahdah","role":"Supervisor","phone":"","email":"fahda@boostgrowthsa.com"},
]

# ------------------- Resources Seed -------------------
RESOURCES_SEED = [
    {"title":"Therapist Drive","description":"Session materials · forms · training","url":"https://drive.google.com/drive/folders/1iMDwfucwzsEIl9WxwhJi_h6tg2vVtAFr","visibility":"therapist","icon":"Folders","bg":"#E5EBE1","color":"#3D4F35","sort_order":10},
    {"title":"Therapist Training Hub","description":"Protocols · lesson plans","url":"https://drive.google.com/drive/folders/1iMDwfucwzsEIl9WxwhJi_h6tg2vVtAFr","visibility":"therapist","icon":"Notebook","bg":"#EAF0F3","color":"#375568","sort_order":20},
    {"title":"Client Files","description":"Per-client folders","url":"https://drive.google.com/drive/folders/1iMDwfucwzsEIl9WxwhJi_h6tg2vVtAFr","visibility":"admin","icon":"Folders","bg":"#FAF0D1","color":"#6B5218","sort_order":30},
    {"title":"HR Files","description":"Employees · Contracts","url":"https://drive.google.com/drive/folders/1jWRO97gDHK_TfmZhTqCqm0SdBc6_b5bE","visibility":"admin","icon":"Files","bg":"#F1ECF7","color":"#4E3F70","sort_order":40},
    {"title":"Company Policies","description":"Internal policies & SOPs","url":"https://drive.google.com/drive/folders/11VQQ-o1QoDQV-ktygB1tlnRmqCs3mxAb","visibility":"all","icon":"Notebook","bg":"#F4E7D8","color":"#8B6918","sort_order":50},
]

CENTER_UPDATES_SEED = [
    {"title": "Summer schedule published", "body": "The June schedule is live — check your calendar for any location changes.", "date": "2025-06-01"},
    {"title": "New prep sheet workflow", "body": "Session preparation now tracks SS week progress automatically on the Attendance page.", "date": "2025-05-28"},
    {"title": "Eid closure reminder", "body": "Center closures are marked on your home calendar. Log sessions only on working days.", "date": "2025-05-20"},
]


OPS_THERAPIST_RECORDS = [
    {"key": "msWalaa", "name": "Ms. Walaa", "email": "wabuissa@boostgrowthsa.com", "role": "operations"},
    {"key": "msJenan", "name": "Ms. Jenan", "email": "jsalmuhaisin@boostgrowthsa.com", "role": "therapist"},
    {"key": "msMaha", "name": "Ms. Maha", "email": "msalthunayan@boostgrowthsa.com", "role": "therapist"},
    {"key": "msFahda", "name": "Ms. Fahda", "email": "falghadeeb@boostgrowthsa.com", "role": "therapist"},
]


async def _ensure_ops_therapist_records() -> int:
    """Ensure ops staff and client-lead supervisors exist in therapists for login, certificates, and training."""
    updated = 0
    for spec in OPS_THERAPIST_RECORDS:
        key = spec["key"]
        email = spec["email"].lower()
        existing = await db.therapists.find_one({"key": key}, {"_id": 0})
        if not existing:
            existing = await _find_therapist_by_email(email)
        if not existing:
            await db.therapists.insert_one({
                "id": str(uuid.uuid4()),
                "name": spec["name"],
                "email": email,
                "key": key,
                "role": spec.get("role", "operations"),
                "color": "#C4864A" if key == "msWalaa" else "#7A8A6A",
                "pin_hash": hash_password("0000"),
                "password_hash": hash_password(UNIFIED_LAUNCH_PASSWORD),
                "must_change_password": False,
                "created_at": now_iso(),
            })
            updated += 1
            continue
        patch: dict = {}
        if (existing.get("name") or "") != spec["name"]:
            patch["name"] = spec["name"]
        if (existing.get("email") or "").lower() != email:
            patch["email"] = email
        if existing.get("key") != key:
            patch["key"] = key
        if spec.get("role") and existing.get("role") != spec["role"]:
            patch["role"] = spec["role"]
        if not existing.get("password_hash"):
            patch["password_hash"] = hash_password(UNIFIED_LAUNCH_PASSWORD)
            patch["must_change_password"] = False
        if patch:
            await db.therapists.update_one({"id": existing["id"]}, {"$set": patch})
            updated += 1
    return updated


async def _run_startup():
    try:
        await db.users.create_index("email", unique=True)
        await db.therapists.create_index("id", unique=True)
        await db.schedule_cells.create_index([("week_start", 1), ("therapist_id", 1)])
        await db.schedule_cells.create_index("week_start")
        await db.notifications.create_index("user_id")
        await db.sessions.create_index([("client_id", 1), ("session_date", -1)])
        await db.sessions.create_index("invoice_id")
        await db.clients.create_index("id")
        await db.clients.create_index([("deleted", 1), ("file_no", 1)])
        await db.invoices.create_index([("client_id", 1), ("start_date", -1)])
        await db.center_test_attempts.create_index("created_at")
        await db.center_test_attempts.create_index("therapist_id")
        await db.therapist_certificates.create_index("therapist_id")

        admin_email = os.environ["ADMIN_EMAIL"].lower()
        admin_password = os.environ["ADMIN_PASSWORD"]
        admin_name = os.environ.get("ADMIN_NAME", "Admin")
        existing = await db.users.find_one({"email": admin_email})
        if not existing:
            await db.users.insert_one({"id": str(uuid.uuid4()), "email": admin_email,
                                       "password_hash": hash_password(admin_password),
                                       "name": admin_name, "role": "admin", "created_at": now_iso()})
            logger.info(f"Admin seeded: {admin_email}")
        # Never overwrite an existing admin password from ADMIN_PASSWORD on deploy —
        # specialists and ops leads change passwords in-app; resetting breaks login.

        hr_email = HR_OPS_EMAIL
        hr_existing = await db.users.find_one({"email": hr_email})
        if not hr_existing:
            await db.users.insert_one({
                "id": str(uuid.uuid4()), "email": hr_email,
                "password_hash": hash_password(HR_OPS_PASSWORD),
                "name": "HR", "role": "admin", "is_hr_ops": True, "created_at": now_iso(),
            })
            logger.info(f"HR ops user seeded: {hr_email}")
        else:
            await db.users.update_one({"email": hr_email}, {"$set": {"is_hr_ops": True}})

        # Seed therapists ONLY on first-time setup (count==0). NEVER overwrite existing data.
        th_count = await db.therapists.count_documents({})
        if th_count == 0:
            for s in THERAPIST_SEED:
                await db.therapists.insert_one({
                    "id": str(uuid.uuid4()), "name": s["name"], "color": s["color"],
                    "email": s.get("email"), "phone": None,
                    "pin_hash": hash_password("0000"),
                    "created_at": now_iso(),
                })
            logger.info(f"First-time seed: {len(THERAPIST_SEED)} therapists with PIN=0000")
        else:
            # Add any new seed therapists that don't exist yet (by name or email) — preserves existing UUIDs
            existing_names = {t["name"] async for t in db.therapists.find({}, {"_id": 0, "name": 1})}
            existing_emails = {
                (t.get("email") or "").lower()
                async for t in db.therapists.find({}, {"_id": 0, "email": 1})
                if t.get("email")
            }
            added = 0
            for s in THERAPIST_SEED:
                seed_email = (s.get("email") or "").lower()
                if s["name"] in existing_names or (seed_email and seed_email in existing_emails):
                    continue
                await db.therapists.insert_one({
                    "id": str(uuid.uuid4()), "name": s["name"], "color": s["color"],
                    "email": s.get("email"), "phone": None,
                    "pin_hash": hash_password("0000"),
                    "created_at": now_iso(),
                })
                added += 1
            if added:
                logger.info(f"Added {added} new therapist(s) without disturbing existing data")

        try:
            n_ops = await _ensure_ops_therapist_records()
            if n_ops:
                logger.info(f"Ops therapist records: {n_ops} created/updated")
        except Exception as e:
            logger.warning(f"Ops therapist records skipped: {e}")

        # Seed schedule week 2026-05-24 if empty
        t_map = {t["name"]: t["id"] async for t in db.therapists.find({}, {"_id": 0, "name": 1, "id": 1})}
        for t in await db.therapists.find({}, {"_id": 0, "name": 1, "id": 1}).to_list(100):
            short = t["name"].replace("Ms. ", "").strip()
            t_map[short] = t["id"]
            t_map[short.lower()] = t["id"]
        await _seed_schedule_week_2026_05_24(t_map)

        # Migrate invoice service_type (HS / SS) from sessions where missing
        try:
            n = await _migrate_invoice_service_types()
            if n:
                logger.info(f"Invoice service_type migration: updated {n} invoice(s)")
        except Exception as e:
            logger.warning(f"Invoice service_type migration skipped: {e}")

        try:
            n = await _migrate_therapist_emails()
            if n:
                logger.info(f"Therapist email migration: updated {n} record(s)")
        except Exception as e:
            logger.warning(f"Therapist email migration skipped: {e}")

        try:
            personal = await _migrate_personal_therapist_accounts()
            if personal.get("actions"):
                logger.info(f"Personal therapist accounts: {personal['actions']}")
        except Exception as e:
            logger.warning(f"Personal therapist account migration skipped: {e}")

        try:
            n = await _migrate_therapist_display_names()
            if n:
                logger.info(f"Therapist display-name migration: updated {n} record(s)")
        except Exception as e:
            logger.warning(f"Therapist display-name migration skipped: {e}")

        try:
            n = await _migrate_schedule_week_therapist_orders()
            if n:
                logger.info(f"Schedule week therapist_order migration: fixed {n} week(s)")
        except Exception as e:
            logger.warning(f"Schedule week therapist_order migration skipped: {e}")

        try:
            n = await _backfill_schedule_cell_colors_for_week("2026-06-28")
            if n:
                logger.info(f"Schedule cell color backfill for 2026-06-28: updated {n} cell(s)")
        except Exception as e:
            logger.warning(f"Schedule cell color backfill skipped: {e}")

        try:
            await _sync_schedule_preparations_for_week("2026-06-28", "2026-07-02")
            logger.info("Prep relink for week 2026-06-28 complete")
        except Exception as e:
            logger.warning(f"Prep relink for week 2026-06-28 skipped: {e}")

        try:
            pin = await _ensure_fahda_saleh_wed_prep_marker()
            if pin.get("ok"):
                logger.info(
                    "Fahda+Saleh Wed prep badge: cell=%s propagated=%s",
                    pin.get("cell_id"),
                    pin.get("propagated_cells"),
                )
        except Exception as e:
            logger.warning(f"Fahda+Saleh Wed prep badge fix skipped: {e}")

        try:
            pay = await _migrate_mark_all_payments_complete()
            if not pay.get("skipped"):
                logger.info(
                    f"Payment bulk complete: {pay.get('invoices_updated', 0)} invoice(s), "
                    f"{pay.get('clients_updated', 0)} client(s)"
                )
        except Exception as e:
            logger.warning(f"Payment bulk complete migration skipped: {e}")

        try:
            phones = await _apply_parent_phones_from_json_file()
            if phones and phones.get("updated"):
                logger.info(f"Parent phones seed: {phones.get('message')}")
        except Exception as e:
            logger.warning(f"Parent phones seed skipped: {e}")

        # Load persisted email settings from db.settings into env
        settings_doc = await db.settings.find_one({"key": "email"}, {"_id": 0})
        if settings_doc:
            _apply_email_settings(settings_doc)

        try:
            dedupe = await _dedupe_duplicate_clients()
            if dedupe.get("removed"):
                logger.info(f"Client dedupe: removed {dedupe['removed']} duplicate(s)")
        except Exception as e:
            logger.warning(f"Client dedupe skipped: {e}")

        try:
            th_dedupe = await _dedupe_duplicate_therapists()
            if th_dedupe.get("removed"):
                logger.info(f"Therapist dedupe: removed {th_dedupe['removed']} duplicate(s)")
        except Exception as e:
            logger.warning(f"Therapist dedupe skipped: {e}")

        try:
            id_dedupe = await _dedupe_therapists_by_identity()
            if id_dedupe.get("removed"):
                logger.info(f"Therapist identity dedupe: removed {id_dedupe['removed']} duplicate(s)")
        except Exception as e:
            logger.warning(f"Therapist identity dedupe skipped: {e}")

        try:
            order_fix = await _fix_schedule_therapist_order_duplicates()
            if order_fix:
                logger.info(f"Schedule therapist_order dedupe: fixed {order_fix} week(s)")
        except Exception as e:
            logger.warning(f"Schedule therapist_order dedupe skipped: {e}")

        try:
            prep_recovery = await _recover_misdated_week_prep(TRIAL_WEEK_START, TRIAL_WEEK_END)
            if any(prep_recovery.values()):
                logger.info(f"Prep recovery {TRIAL_WEEK_START}: {prep_recovery}")
        except Exception as e:
            logger.warning(f"Prep recovery for trial week skipped: {e}")

        try:
            backup = await _auto_backup_if_stale()
            if backup:
                logger.info(
                    f"Startup auto-backup stored ({backup.get('id', '')[:8]}…, "
                    f"clients={backup.get('totals', {}).get('clients')})"
                )
        except Exception as e:
            logger.warning(f"Startup auto-backup skipped: {e}")

        try:
            health = await _get_data_health_snapshot()
            logger.info(
                "Startup data counts: clients=%s invoices=%s sessions=%s prep=%s "
                "therapists=%s dup_groups=%s backups=%s",
                health.get("clients"),
                health.get("invoices"),
                health.get("sessions"),
                health.get("prep_history"),
                health.get("therapists"),
                health.get("duplicate_therapist_groups"),
                health.get("stored_backups"),
            )
        except Exception as e:
            logger.warning(f"Startup health log skipped: {e}")

        try:
            nat_clean = await _remove_therapists_without_nationality()
            if nat_clean.get("removed"):
                logger.info(f"Therapist nationality cleanup: removed {nat_clean['removed']}")
        except Exception as e:
            logger.warning(f"Therapist nationality cleanup skipped: {e}")

        try:
            if await _migrate_hr_password_once():
                logger.info("HR password migrated to Boost@2026 (one-time)")
        except Exception as e:
            logger.warning(f"HR password migration skipped: {e}")

        try:
            if await _ensure_walaa_ops_login_once():
                logger.info("Walaa ops login restored for wabuissa@ (one-time)")
        except Exception as e:
            logger.warning(f"Walaa ops login restore skipped: {e}")

        try:
            pw_n = await _migrate_bootstrap_therapist_passwords()
            if pw_n:
                logger.info(f"Therapist bootstrap passwords: updated {pw_n} record(s)")
        except Exception as e:
            logger.warning(f"Therapist bootstrap passwords skipped: {e}")

        try:
            inv_clean = await _cleanup_orphan_invoices()
            if inv_clean.get("removed"):
                logger.info(f"Orphan invoice cleanup: removed {inv_clean['removed']}")
        except Exception as e:
            logger.warning(f"Orphan invoice cleanup skipped: {e}")

        try:
            inactive_n = await _apply_inactive_client_status()
            if inactive_n:
                logger.info(f"Marked {inactive_n} client(s) inactive")
        except Exception as e:
            logger.warning(f"Inactive client migration skipped: {e}")

        # Seed clients ONLY on first-time setup (no active clients). Preserves user edits and soft-deletions.
        cl_count = await db.clients.count_documents(_active_client_filter())
        if cl_count == 0:
            therapists_map = {t["name"]: t["id"] async for t in db.therapists.find({}, {"_id": 0, "name": 1, "id": 1})}
            for c in CLIENT_SEED:
                if await db.clients.find_one({"file_no": c["file_no"]}, {"_id": 0, "id": 1}):
                    continue
                seed_doc = {
                    "id": str(uuid.uuid4()),
                    "file_no": c["file_no"], "name": c["name"],
                    "package_hours": c["pkg"], "supervisor": c["sup"],
                    "main_therapist_id": therapists_map.get(c["main"]),
                    "co_therapist_ids": [therapists_map[n] for n in c["co"] if n in therapists_map],
                    "color": c["color"], "locations": c["locs"],
                    "parent_name": None, "parent_phone": None, "age": None,
                    "notes": None, "created_at": now_iso(),
                }
                if c["file_no"] in INACTIVE_CLIENT_FILE_NOS:
                    seed_doc["status"] = "Inactive"
                await db.clients.insert_one(seed_doc)
            await db.meta.update_one({"key": "client_seed_version"},
                                     {"$set": {"version": 1, "updated_at": now_iso()}},
                                     upsert=True)
            logger.info(f"First-time seed: {len(CLIENT_SEED)} clients")

        # Seed Intake (only if empty — admin may manage manually)
        if await db.intake.count_documents({}) == 0:
            for item in INTAKE_SEED:
                name = item.get("child_name", "").strip()
                itype = item.get("intake_type", "pre")
                await db.intake.insert_one({
                    "id": str(uuid.uuid4()),
                    "status": "new",
                    "priority": bool(item.get("priority")),
                    "created_at": now_iso(),
                    **item,
                    "child_name": name,
                    "name_key": _intake_name_key(name, itype),
                })
            logger.info(f"Seeded {len(INTAKE_SEED)} intake records from waiting list")

        try:
            school_seed = await _ensure_school_waiting_records()
            if school_seed.get("created") or school_seed.get("updated"):
                logger.info(
                    "School waiting seed: %s added, %s updated (%s total)",
                    school_seed.get("created", 0),
                    school_seed.get("updated", 0),
                    school_seed.get("school_count", 0),
                )
        except Exception as e:
            logger.warning(f"School waiting seed skipped: {e}")

        # Seed Directory (only if empty)
        if await db.directory.count_documents({}) == 0:
            for item in DIRECTORY_SEED:
                await db.directory.insert_one({
                    "id": str(uuid.uuid4()), **item, "created_at": now_iso(),
                })
            logger.info(f"Seeded {len(DIRECTORY_SEED)} directory contacts")

        # Seed Resources (only if empty)
        if await db.resources.count_documents({}) == 0:
            for item in RESOURCES_SEED:
                await db.resources.insert_one({
                    "id": str(uuid.uuid4()),
                    "category": "drive",
                    **item,
                    "created_at": now_iso(),
                })
            logger.info(f"Seeded {len(RESOURCES_SEED)} resources")

        # Seed center updates (only if empty)
        if await db.center_updates.count_documents({}) == 0:
            for item in CENTER_UPDATES_SEED:
                await db.center_updates.insert_one({
                    "id": str(uuid.uuid4()),
                    **item,
                    "created_at": now_iso(),
                    "is_important": False,
                    "requires_ack": False,
                    "send_to_specialists": False,
                    "acknowledged_by": [],
                })
            logger.info(f"Seeded {len(CENTER_UPDATES_SEED)} center updates")

        # Seed Leaves (only if empty) — from leaves_seed.json (parsed Vacation 2026)
        if await db.leaves.count_documents({}) == 0:
            seed_path = ROOT_DIR / "leaves_seed.json"
            if seed_path.exists():
                import json
                seed = json.loads(seed_path.read_text())
                t_by_name = {t["name"]: t["id"] async for t in db.therapists.find({}, {"_id": 0, "name": 1, "id": 1})}
                inserted = 0
                for item in seed:
                    tid = t_by_name.get(item.get("therapist_name"))
                    if not tid:
                        continue
                    await db.leaves.insert_one({
                        "id": str(uuid.uuid4()),
                        "therapist_id": tid,
                        "start_date": item.get("start_date") or "",
                        "end_date": item.get("end_date") or "",
                        "days": item.get("days") or 0,
                        "leave_type": item.get("leave_type") or "Annual",
                        "status": item.get("status") or "done",
                        "notes": item.get("notes"),
                        "created_at": now_iso(),
                    })
                    inserted += 1
                logger.info(f"Seeded {inserted} leaves from Vacation 2026")

        # Seed staff purchases from spreadsheet (only if empty)
        if await db.staff_purchases.count_documents({}) == 0:
            seed_path = ROOT_DIR / "purchases_seed.json"
            if seed_path.exists():
                import json
                seed = json.loads(seed_path.read_text())
                inserted = 0
                for item in seed:
                    t = await _resolve_therapist_by_purchaser(item.get("purchaser") or "")
                    if not t:
                        logger.warning("Purchase seed: no therapist for %s", item.get("purchaser"))
                        continue
                    purchase_date = item.get("reimbursement_date") or "2026-01-01"
                    await db.staff_purchases.insert_one({
                        "id": str(uuid.uuid4()),
                        "row_no": item.get("row_no"),
                        "therapist_id": t["id"],
                        "therapist_name": t.get("name"),
                        "purchaser_name": item.get("purchaser"),
                        "item": item.get("item"),
                        "category": item.get("category"),
                        "description": item.get("description") or "",
                        "qty": item.get("qty") or "1",
                        "unit_price": item.get("unit_price") or "",
                        "total": item.get("total"),
                        "total_display": item.get("total_display") or str(item.get("total") or ""),
                        "status": _normalize_purchase_status(item.get("status") or "pending"),
                        "reimbursement_date": item.get("reimbursement_date"),
                        "purchase_date": purchase_date,
                        "purchase_month": purchase_date[:7],
                        "notes": None,
                        "created_at": now_iso(),
                        "updated_at": now_iso(),
                        "imported": True,
                    })
                    inserted += 1
                logger.info(f"Seeded {inserted} staff purchases from spreadsheet")

        if await db.purchase_reminder_settings.count_documents({}) == 0:
            await db.purchase_reminder_settings.insert_one({
                "id": "default",
                "day_of_month": 25,
                "enabled": True,
                "therapist_ids": [],
                "last_sent_month": None,
                "updated_at": now_iso(),
            })

        try:
            await _send_purchase_reminders(force=False)
        except Exception:
            logger.exception("Purchase reminder check failed")

        try:
            eval_alerts = await _process_evaluation_due_alerts(force=False)
            if eval_alerts.get("sent"):
                logger.info(
                    "Evaluation due alerts sent: %s (target %s)",
                    eval_alerts.get("sent"),
                    eval_alerts.get("target_date"),
                )
        except Exception:
            logger.exception("Evaluation due alert check failed")

        try:
            pm_fixed = await _backfill_purchase_months()
            if pm_fixed:
                logger.info(f"Backfilled purchase_month on {pm_fixed} staff purchase(s)")
        except Exception as e:
            logger.warning(f"Purchase month backfill skipped: {e}")

        try:
            pd_fixed = await _repair_purchase_dates_from_month()
            if pd_fixed:
                logger.info(f"Aligned purchase_date on {pd_fixed} staff purchase(s)")
        except Exception as e:
            logger.warning(f"Purchase date repair skipped: {e}")

        try:
            sync_result = await _ensure_purchases_sheet_synced()
            if sync_result.get("inserted"):
                logger.info(
                    "Synced missing purchase months from Google Sheet: %s (%s rows)",
                    sync_result.get("missing_months_synced"),
                    sync_result.get("inserted"),
                )
        except Exception as e:
            logger.warning(f"Purchase sheet startup sync skipped: {e}")

        try:
            lb_sync = await _sync_leave_balances_from_sheet()
            if lb_sync.get("updated"):
                logger.info(
                    "Synced leave balances from Google Sheet: %s therapist(s) for %s",
                    lb_sync.get("updated"),
                    lb_sync.get("year"),
                )
        except Exception as e:
            logger.warning(f"Leave balance sheet startup sync skipped: {e}")

        await _fix_walaa_purchase_month_mismatch()
    except Exception:
        logger.exception(
            "Background startup/seed failed — check MONGO_URL and MongoDB Atlas network access (0.0.0.0/0)"
        )


@app.on_event("startup")
async def startup():
    asyncio.create_task(_run_startup())
    logger.info("API ready; database init running in background")


# ------------------- Center training test + academic portfolio -------------------
import json as _json

_CENTER_TEST_PATH = ROOT_DIR / "center_test_questions.json"
_center_test_cache: Optional[dict] = None


def _load_center_test_data(test_id: Optional[str] = None) -> dict:
    global _center_test_cache
    if _center_test_cache is None:
        if not _CENTER_TEST_PATH.is_file():
            raise HTTPException(status_code=500, detail="Test questions file missing")
        _center_test_cache = _json.loads(_CENTER_TEST_PATH.read_text(encoding="utf-8"))
    data = _center_test_cache
    tid = (test_id or data.get("testId") or "default").strip()
    if tid and data.get("testId") and data.get("testId") != tid:
        raise HTTPException(status_code=404, detail="Assessment not found")
    return data


def _center_test_catalog_entry(data: dict) -> dict:
    return {
        "testId": data.get("testId", "default"),
        "courseName": data.get("courseName") or data.get("courseTopic", ""),
        "title": data.get("title", ""),
        "courseTopic": data.get("courseTopic", ""),
        "instructor": data.get("instructor", ""),
        "passThreshold": int(data.get("passThreshold", 70)),
        "questionCount": len(data.get("questions", [])),
    }


async def _optional_current_user(request: Request) -> Optional[dict]:
    try:
        return await get_current_user(request)
    except HTTPException:
        return None


async def _attempts_for_user(user: dict) -> List[dict]:
    tid = await _resolve_user_therapist_id(user)
    email = (user.get("email") or "").lower().strip()
    or_clauses: List[dict] = []
    if tid:
        or_clauses.append({"therapist_id": tid})
    if email:
        or_clauses.append({"therapist_email": email})
    if not or_clauses:
        return []
    rows = await db.center_test_attempts.find(
        {"$or": or_clauses}, {"_id": 0}
    ).sort([("created_at", -1)]).to_list(500)
    return rows


def _prepare_learning_attempts(attempts: List[dict]) -> List[dict]:
    """Number attempts per test; only passed attempts (80%+) include correct answer keys."""
    by_test: Dict[str, List[dict]] = {}
    for a in attempts:
        tid = a.get("test_id") or "default"
        by_test.setdefault(tid, []).append(a)

    prepared: List[dict] = []
    for _tid, group in by_test.items():
        chronological = sorted(group, key=lambda x: x.get("created_at") or "")
        for idx, a in enumerate(chronological, 1):
            row = dict(a)
            row["attempt_number"] = idx
            passed = bool(a.get("passed"))
            row["answers_unlocked"] = passed
            if passed:
                row["answers"] = list(a.get("answers") or [])
            else:
                row["answers"] = [
                    {
                        "question_id": ans.get("question_id"),
                        "question_text": ans.get("question_text"),
                        "selected": ans.get("selected"),
                        "selected_text": ans.get("selected_text"),
                        "is_correct": ans.get("is_correct"),
                    }
                    for ans in (a.get("answers") or [])
                ]
            prepared.append(row)
    prepared.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return prepared


class CenterTestSubmitIn(BaseModel):
    student_name: str
    answers: Dict[str, str]
    test_id: Optional[str] = None


@api.get("/center-test/catalog")
async def get_center_test_catalog():
    data = _load_center_test_data()
    return {"tests": [_center_test_catalog_entry(data)]}


@api.get("/center-test/questions")
async def get_center_test_questions(test_id: Optional[str] = None):
    data = _load_center_test_data(test_id)
    public_questions = []
    for q in data.get("questions", []):
        public_questions.append({
            "id": q["id"],
            "text": q["text"],
            "choices": q["choices"],
        })
    entry = _center_test_catalog_entry(data)
    return {
        **entry,
        "passThreshold": entry["passThreshold"],
        "questions": public_questions,
    }


@api.post("/center-test/attempts")
async def submit_center_test_attempt(payload: CenterTestSubmitIn, request: Request):
    name = (payload.student_name or "").strip()
    if len(name) < 3:
        raise HTTPException(status_code=400, detail="Please enter your full name")
    data = _load_center_test_data(payload.test_id)
    questions = data.get("questions", [])
    if not questions:
        raise HTTPException(status_code=500, detail="No questions configured")
    threshold = int(data.get("passThreshold", 70))
    answer_rows = []
    correct_count = 0
    for q in questions:
        qid = q["id"]
        selected = (payload.answers.get(qid) or "").strip().lower()
        correct = (q.get("correct") or "").strip().lower()
        is_correct = bool(selected) and selected == correct
        if is_correct:
            correct_count += 1
        selected_label = ""
        correct_label = ""
        for ch in q.get("choices", []):
            if ch["id"] == selected:
                selected_label = ch.get("text", "")
            if ch["id"] == correct:
                correct_label = ch.get("text", "")
        answer_rows.append({
            "question_id": qid,
            "question_text": q.get("text", ""),
            "selected": selected,
            "selected_text": selected_label,
            "correct": correct,
            "correct_text": correct_label,
            "is_correct": is_correct,
        })
    total = len(questions)
    percentage = round((correct_count / total) * 100) if total else 0
    passed = percentage >= threshold
    user = await _optional_current_user(request)
    therapist_id = None
    therapist_email = None
    if user:
        therapist_id = await _resolve_user_therapist_id(user) or (
            user.get("id") if user.get("role") == "therapist" else None
        )
        therapist_email = (user.get("email") or "").lower().strip() or None
    doc = {
        "id": str(uuid.uuid4()),
        "student_name": name,
        "test_id": data.get("testId", "default"),
        "course_name": data.get("courseName") or data.get("courseTopic", ""),
        "test_title": data.get("title", ""),
        "therapist_id": therapist_id,
        "therapist_email": therapist_email,
        "answers": answer_rows,
        "score": correct_count,
        "total": total,
        "percentage": percentage,
        "passed": passed,
        "pass_threshold": threshold,
        "created_at": now_iso(),
    }
    await db.center_test_attempts.insert_one(doc)
    doc.pop("_id", None)
    return doc


async def center_test_results_access(user: dict = Depends(get_current_user)) -> dict:
    """Training test results — ops leads, admin, HR, Jenan, Walaa, and can_view_reports."""
    if user.get("can_view_reports"):
        return user
    if (
        _is_portal_admin(user)
        or _is_walaa_ops(user)
        or _is_hr_ops(user)
        or _is_jenan(user)
        or _is_client_lead(user)
    ):
        return user
    raise HTTPException(status_code=403, detail="Training results access required")


def _can_upload_therapist_certificates(user: dict) -> bool:
    """Certificate upload — portal admin, HR, and Jenan only (not Walaa/Maha/Fahda)."""
    if _is_portal_admin(user) or _is_hr_ops(user) or _is_jenan(user):
        return True
    return False


async def certificate_upload_access(user: dict = Depends(get_current_user)) -> dict:
    if not _can_upload_therapist_certificates(user):
        raise HTTPException(status_code=403, detail="Certificate upload access required")
    return user


def _can_manage_center_test_attempts(user: dict) -> bool:
    """Delete training attempts — portal admin, Walaa ops, HR, and Jenan."""
    if _is_portal_admin(user) or _is_walaa_ops(user) or _is_hr_ops(user) or _is_jenan(user):
        return True
    return False


async def center_test_manage_access(user: dict = Depends(get_current_user)) -> dict:
    if not _can_manage_center_test_attempts(user):
        raise HTTPException(status_code=403, detail="Training test management access required")
    return user


@api.get("/center-test/attempts")
async def list_center_test_attempts(user=Depends(center_test_results_access)):
    rows = await db.center_test_attempts.find({}, {"_id": 0}).sort([("created_at", -1)]).to_list(2000)
    can_upload = _can_upload_therapist_certificates(user)
    therapists = []
    if can_upload:
        therapists = await db.therapists.find(
            {}, {"_id": 0, "id": 1, "name": 1, "email": 1}
        ).sort("name", 1).to_list(500)
    return {
        "attempts": rows,
        "can_delete_attempts": _can_manage_center_test_attempts(user),
        "can_upload_certificates": can_upload,
        "therapists": therapists,
    }


@api.delete("/center-test/attempts/{attempt_id}")
async def delete_center_test_attempt(attempt_id: str, _=Depends(center_test_manage_access)):
    res = await db.center_test_attempts.delete_one({"id": attempt_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Attempt not found")
    return {"ok": True, "deleted_id": attempt_id}


@api.get("/my-learning")
async def get_my_learning(user=Depends(get_current_user)):
    """Academic portfolio — available assessments, my attempts, my certificates."""
    tid = await _resolve_user_therapist_id(user) or (
        user.get("id") if user.get("role") == "therapist" else None
    )
    catalog = [_center_test_catalog_entry(_load_center_test_data())]
    raw_attempts = await _attempts_for_user(user)
    attempts = _prepare_learning_attempts(raw_attempts)
    passed_test_ids = sorted({
        a.get("test_id") or "default"
        for a in raw_attempts
        if a.get("passed")
    })
    cert_q: dict = {}
    if tid:
        cert_q["therapist_id"] = tid
    else:
        cert_q["therapist_id"] = "__none__"
    certs = await db.therapist_certificates.find(cert_q, {"_id": 0, "file_data": 0}).sort(
        [("issued_at", -1)]
    ).to_list(100)
    for c in certs:
        c["download_url"] = f"/api/therapist-certificates/{c['id']}/file"
    therapists = []
    can_upload = _can_upload_therapist_certificates(user)
    if can_upload:
        therapists = await db.therapists.find(
            {}, {"_id": 0, "id": 1, "name": 1, "email": 1}
        ).sort("name", 1).to_list(500)
    return {
        "catalog": catalog,
        "attempts": attempts,
        "passed_test_ids": passed_test_ids,
        "certificates": certs,
        "can_upload_certificates": can_upload,
        "therapists": therapists if can_upload else [],
        "user": {
            "id": user.get("id"),
            "name": user.get("name"),
            "email": user.get("email"),
            "therapist_id": tid,
        },
    }


async def _notify_therapist_certificate_ready(therapist: dict, course_name: str, cert_id: str) -> None:
    """In-app alert: certificate published to My Learning."""
    title = "Your certificate is ready"
    message = (
        f"Your certificate for {course_name} is now available in "
        "My Learning → My Certificates."
    )
    extra = {"link": "/my-learning", "certificate_id": cert_id, "course_name": course_name}
    notified: set = set()
    tid = therapist.get("id")
    if tid:
        await _notify(tid, "certificate_ready", title, message, **extra)
        notified.add(tid)
    email = (therapist.get("email") or "").lower().strip()
    if email:
        u = await db.users.find_one({"email": email}, {"_id": 0, "id": 1})
        if u and u["id"] not in notified:
            await _notify(u["id"], "certificate_ready", title, message, **extra)


@api.post("/therapist-certificates")
async def upload_therapist_certificate(
    therapist_id: str = Form(...),
    course_name: str = Form(...),
    title: str = Form(""),
    issued_at: Optional[str] = Form(None),
    notify_trainee: str = Form("true"),
    file: UploadFile = File(...),
    user=Depends(certificate_upload_access),
):
    t = await db.therapists.find_one({"id": therapist_id}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Therapist not found")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="File required")
    content = await file.read()
    if len(content) > 12 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 12MB)")
    ext = Path(file.filename).suffix.lower() or ".pdf"
    if ext not in (".pdf", ".png", ".jpg", ".jpeg", ".webp"):
        raise HTTPException(status_code=400, detail="PDF or image only")
    cert_id = str(uuid.uuid4())
    stored = f"cert_{cert_id}{ext}"
    file_data = _persist_upload(stored, content)
    doc = {
        "id": cert_id,
        "therapist_id": therapist_id,
        "therapist_name": t.get("name"),
        "course_name": course_name.strip(),
        "title": (title or course_name).strip(),
        "file_path": stored,
        "file_name": file.filename,
        "file_data": file_data,
        "issued_at": issued_at or now_iso()[:10],
        "uploaded_by": user.get("email") or user.get("name"),
        "created_at": now_iso(),
    }
    await db.therapist_certificates.insert_one(doc)
    if str(notify_trainee).lower() in ("true", "1", "yes", "on"):
        await _notify_therapist_certificate_ready(t, course_name.strip(), cert_id)
    doc.pop("file_data", None)
    doc.pop("_id", None)
    doc["download_url"] = f"/api/therapist-certificates/{cert_id}/file"
    return doc


@api.get("/therapist-certificates/{cert_id}/file")
async def get_therapist_certificate_file(cert_id: str, user=Depends(get_current_user)):
    cert = await db.therapist_certificates.find_one({"id": cert_id}, {"_id": 0})
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    tid = await _resolve_user_therapist_id(user)
    is_owner = tid and cert.get("therapist_id") == tid
    is_admin = False
    try:
        await center_test_results_access(user)
        is_admin = True
    except HTTPException:
        pass
    if not is_owner and not is_admin:
        raise HTTPException(status_code=403, detail="Forbidden")
    content = _load_upload(cert.get("file_path"), cert.get("file_data"))
    if not content:
        raise HTTPException(status_code=404, detail=FILE_UNAVAILABLE_DETAIL)
    return _bytes_file_response(content, cert.get("file_name") or "certificate.pdf")


app.include_router(api)

@app.on_event("shutdown")
async def shutdown():
    client.close()


# ------------------- Frontend (React build) — same host as /api -------------------
FRONTEND_DIR = ROOT_DIR / "static"

if FRONTEND_DIR.is_dir():
    _SPA_INDEX_HEADERS = {"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache"}
    _SPA_ASSET_HEADERS = {"Cache-Control": "public, max-age=31536000, immutable"}

    @app.get("/{spa_path:path}")
    async def serve_frontend(spa_path: str = ""):
        """Serve CRA build; unknown paths → index.html for client-side routing."""
        if spa_path.startswith("api") or spa_path.startswith("api/"):
            raise HTTPException(status_code=404, detail="Not found")
        if spa_path:
            asset = FRONTEND_DIR / spa_path
            if asset.is_file():
                headers = _SPA_ASSET_HEADERS if re.search(r"\.[a-f0-9]{8}\.(js|css)$", spa_path) else None
                return FileResponse(asset, headers=headers)
        index = FRONTEND_DIR / "index.html"
        if index.is_file():
            return FileResponse(index, headers=_SPA_INDEX_HEADERS)
        raise HTTPException(status_code=404, detail="Frontend not built — run build or deploy with Dockerfile")
